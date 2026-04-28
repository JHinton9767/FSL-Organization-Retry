from __future__ import annotations

import argparse
import hashlib
import inspect
import json
import math
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from time import perf_counter
from typing import Callable, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.config_loader import (
    APP_SETTINGS_PATH,
    DEFAULT_CHAPTER_GROUPS_PATH,
    EXAMPLE_CHAPTER_GROUPS_PATH,
    MANUAL_CHAPTER_ASSIGNMENTS_PATH,
    load_chapter_mapping,
    load_manual_chapter_assignments,
    load_settings,
)
from app.status_framework import build_outcome_resolution_fields
from src.build_master_roster import (
    DEFAULT_INPUT_ROOT,
    SUPPORTED_EXTENSIONS,
    build_individual_new_member_form_lookup,
    canonical_header,
    chapter_from_filename,
    detect_inline_chapter_label,
    find_header_row_in_rows,
    find_header_row,
    find_status_column_in_rows,
    find_status_column,
    get_cell,
    infer_chapter,
    is_individual_new_member_form_pdf,
    is_placeholder_sheet_name,
    normalize_banner_id,
    normalize_chapter_name,
    normalize_status,
    pdf_table_rows,
    source_file_format_priority,
    source_file_label,
    source_context_indicates_new_member,
    should_upgrade_to_new_member_status,
)
from src.shared_utils import bucket_30_hours, clean_text, coerce_numeric


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ROSTER_ROOT = DEFAULT_INPUT_ROOT
DEFAULT_ROSTER_INBOX = ROOT / "data" / "inbox" / "rosters"
DEFAULT_ACADEMIC_ROOT = ROOT / "data" / "inbox" / "academic"
DEFAULT_TRANSCRIPT_TEXT_ROOT = ROOT / "data" / "inbox" / "transcript_text"
DEFAULT_GRADUATION_ROOT = ROOT / "data" / "inbox" / "graduation"
DEFAULT_REFERENCE_DATA_ROOT = ROOT / "data" / "inbox" / "reference_data"
DEFAULT_MEMBERSHIP_REFERENCE_ROOT = ROOT / "data" / "inbox" / "membership_reference"
DEFAULT_GPA_REFERENCE_ROOT = ROOT / "data" / "inbox" / "gpa_reference"
DEFAULT_GPA_BENCHMARK_ROOT = ROOT / "data" / "inbox" / "gpa_benchmark_reference"
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "canonical"
DEFAULT_CACHE_ROOT = DEFAULT_OUTPUT_ROOT / "_source_cache"
TRANSCRIPT_TEXT_MANIFEST_PATH = ROOT / "config" / "transcript_text_manifest.csv"
SCHEMA_PATH = ROOT / "config" / "canonical_schema.json"
ROSTER_SOURCE_EXTENSIONS = SUPPORTED_EXTENSIONS.union({".csv", ".pdf"})
TABULAR_SOURCE_EXTENSIONS = SUPPORTED_EXTENSIONS.union({".csv"})
TEXT_SOURCE_EXTENSIONS = {".txt"}

TERM_RE = re.compile(r"(Winter|Spring|Summer|Fall)\s+(19\d{2}|20\d{2})", re.IGNORECASE)
TERM_CODE_RE = re.compile(r"^(19\d{2}|20\d{2})(WI|SP|SU|FA)$", re.IGNORECASE)
UPDATE_RE = re.compile(r"\((\d{1,2})\.(\d{1,2})\.(\d{2,4})\)")
MONTH_PATTERNS = [
    ("January", 1, r"\bjan(?:uary)?\b"),
    ("February", 2, r"\bfeb(?:ruary)?\b"),
    ("March", 3, r"\bmar(?:ch)?\b"),
    ("April", 4, r"\bapr(?:il)?\b"),
    ("May", 5, r"\bmay\b"),
    ("June", 6, r"\bjun(?:e)?\b"),
    ("July", 7, r"\bjul(?:y)?\b"),
    ("August", 8, r"\baug(?:ust)?\b"),
    ("September", 9, r"\bsep(?:t|tember)?\b"),
    ("October", 10, r"\boct(?:ober)?\b"),
    ("November", 11, r"\bnov(?:ember)?\b"),
    ("December", 12, r"\bdec(?:ember)?\b"),
]
SEASON_ORDER = {"WI": 0, "SP": 1, "SU": 2, "FA": 3}
SEASON_NAME = {"WI": "Winter", "SP": "Spring", "SU": "Summer", "FA": "Fall"}
UNRESOLVED_OUTCOMES = {"Active/Unknown", "No Further Observation", "Unknown", ""}

SNAPSHOT_ALIAS_GROUPS = {
    "Student ID": {"student id", "banner id", "banner", "student number", "PLID", "plid"},
    "First Name": {"first name", "firstname"},
    "Last Name": {"last name", "lastname"},
    "NetID": {"netid", "net id"},
    "High School GPA": {"high school gpa", "hs gpa"},
    "Overall GPA": {"overall gpa"},
    "Institutional GPA": {"institutional gpa", "txst gpa", "texas state gpa"},
    "Transfer GPA": {"transfer gpa"},
    "Total Credit Hours": {"total credit hours", "total hours", "overall credit hours"},
    "TXST Credit Hours": {"txst credit hours", "texas state credit hours", "institutional credit hours"},
    "Previous Semester GPA": {"previous semester gpa", "previous term gpa"},
    "Student Status": {"student status"},
    "Student Status (FT/PT)": {"student status ftpt", "student status ft pt", "student status (ft/pt)", "ft/pt"},
}

SNAPSHOT_COLUMNS = [
    "Student ID",
    "First Name",
    "Last Name",
    "NetID",
    "High School GPA",
    "Overall GPA",
    "Institutional GPA",
    "Transfer GPA",
    "Total Credit Hours",
    "TXST Credit Hours",
    "Previous Semester GPA",
    "Student Status",
    "Student Status (FT/PT)",
    "Snapshot Source File",
]

GRADUATION_ALIAS_GROUPS = {
    "Student ID": {"student id", "banner id", "banner", "student number", "PLID", "plid"},
    "First Name": {"first name", "firstname"},
    "Last Name": {"last name", "lastname"},
    "Graduation Term": {
        "graduation term",
        "graduation semester",
        "graduation date",
        "grad term",
        "grad semester",
        "graduated term",
    },
    "Outcome": {"status", "outcome", "graduation status", "degree status"},
}

GRADUATION_COLUMNS = [
    "Student ID",
    "First Name",
    "Last Name",
    "Graduation Term",
    "Outcome",
    "Graduation Source File",
]

QA_COLUMNS = ["Check Group", "Check", "Status", "Value", "Notes"]

GRADE_COLUMN_ALIASES = {
    "Banner ID": {"banner id", "student id", "banner", "student number", "PLID", "plid", "netid", "net id"},
    "Last Name": {"last name", "lastname"},
    "First Name": {"first name", "firstname"},
    "Email": {"email", "email address"},
    "Student Status": {"student status", "status"},
    "Major": {"major", "primary major"},
    "Semester Hours": {"semester hours", "credits attempted", "attempted hours", "hours attempted"},
    "Cumulative Hours": {"cumulative hours", "total credit hours", "total hours"},
    "Current Academic Standing": {"current academic standing", "academic standing", "standing"},
    "Texas State GPA": {"texas state gpa", "institutional gpa", "txst gpa"},
    "Overall GPA": {"overall gpa"},
    "Transfer GPA": {"transfer gpa"},
    "Term GPA": {"term gpa", "gpa"},
    "Term Passed Hours": {"term passed hours", "credits earned", "earned hours", "passed hours"},
    "TxState Cumulative GPA": {"txstate cumulative gpa", "texas state cumulative gpa", "institutional cumulative gpa"},
    "Overall Cumulative GPA": {"overall cumulative gpa"},
    "Graduation Term": {"graduation term", "graduation semester", "graduated term", "grad term"},
}

TRANSCRIPT_TERM_COLUMNS = [
    "student_id",
    "student_id_raw",
    "identity_resolution_basis",
    "identity_resolution_notes",
    "first_name",
    "last_name",
    "source_file",
    "term_code",
    "term_label",
    "term_year",
    "term_season",
    "summary_credits_earned",
    "summary_credit_completion_pct",
    "summary_term_gpa",
    "summary_cumulative_gpa",
    "summary_academic_standing",
    "summary_graduation_term_code",
    "summary_graduation_term_label",
    "summary_graduation_signal_text",
]

TRANSCRIPT_COURSE_COLUMNS = [
    "student_id",
    "student_id_raw",
    "first_name",
    "last_name",
    "source_file",
    "term_code",
    "term_label",
    "raw_course_line",
    "course_code",
    "section_type",
    "course_title",
    "grade",
    "transfer_flag",
    "credits_attempted",
    "credits_earned",
    "credit_token_raw",
]

TRANSCRIPT_AUDIT_COLUMNS = [
    "source_file",
    "student_id",
    "student_id_raw",
    "first_name",
    "last_name",
    "identity_resolution_basis",
    "parse_status",
    "term_count",
    "course_count",
    "warning_count",
    "warnings",
    "unmatched_lines",
]


@dataclass(frozen=True)
class CanonicalBuildResult:
    output_folder: Path
    files: Dict[str, Path]


@dataclass
class PerformanceStage:
    stage: str
    seconds: float
    cache_status: str
    rows: str
    notes: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build the canonical FSL analytics tables and QA outputs. "
            "Authoritative outputs: roster_term, academic_term, master_longitudinal, student_summary, cohort_metrics, qa_checks."
        )
    )
    parser.add_argument("--roster-root", default=str(DEFAULT_ROSTER_ROOT))
    parser.add_argument("--roster-inbox", default=str(DEFAULT_ROSTER_INBOX))
    parser.add_argument("--academic-root", default=str(DEFAULT_ACADEMIC_ROOT))
    parser.add_argument("--transcript-text-root", default=str(DEFAULT_TRANSCRIPT_TEXT_ROOT))
    parser.add_argument("--graduation-root", default=str(DEFAULT_GRADUATION_ROOT))
    parser.add_argument("--reference-data-root", default=str(DEFAULT_REFERENCE_DATA_ROOT))
    parser.add_argument("--membership-reference-root", default=str(DEFAULT_MEMBERSHIP_REFERENCE_ROOT))
    parser.add_argument("--gpa-reference-root", default=str(DEFAULT_GPA_REFERENCE_ROOT))
    parser.add_argument("--gpa-benchmark-root", default=str(DEFAULT_GPA_BENCHMARK_ROOT))
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--cache-root", default=str(DEFAULT_CACHE_ROOT))
    parser.add_argument("--refresh-source-cache", action="store_true", help="Force raw source files to be re-read instead of using cached normalized source tables.")
    return parser.parse_args()


def load_schema() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def ensure_columns(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    result = frame.copy()
    for column in columns:
        if column not in result.columns:
            result[column] = pd.NA
    return result.loc[:, list(columns)]


def ensure_text_columns(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    """Force columns that receive string updates to object dtype.

    Cached CSVs can read fully blank columns as float64. Pandas 3.x then rejects
    later string assignments such as term codes or provenance labels.
    """
    result = frame.copy()
    for column in columns:
        if column not in result.columns:
            result[column] = ""
        result[column] = result[column].astype("object").where(result[column].notna(), "")
    return result


def combine_reference_frames(
    frames: Sequence[pd.DataFrame],
    columns: Sequence[str],
    dedupe_subset: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    usable = [frame for frame in frames if frame is not None and not frame.empty]
    if not usable:
        return pd.DataFrame(columns=list(columns))
    combined = pd.concat([ensure_columns(frame, columns) for frame in usable], ignore_index=True)
    if dedupe_subset:
        combined = combined.drop_duplicates(subset=list(dedupe_subset), keep="first")
    return ensure_columns(combined, columns)


def read_cached_frame(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, low_memory=False)
    except pd.errors.EmptyDataError:
        return pd.DataFrame()


def write_cache_manifest(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def read_cache_manifest(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def file_signature(path: Path) -> dict:
    try:
        stat = path.stat()
        resolved = path.resolve()
    except Exception:
        stat = path.stat()
        resolved = path
    return {
        "path": str(resolved),
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def files_manifest(paths: Sequence[Path]) -> List[dict]:
    seen: set[str] = set()
    manifest: List[dict] = []
    for path in sorted(paths, key=lambda item: str(item).lower()):
        key = str(path.resolve()) if path.exists() else str(path)
        if key in seen:
            continue
        seen.add(key)
        if path.exists():
            manifest.append(file_signature(path))
    return manifest


def optional_files_manifest(paths: Sequence[Path]) -> List[dict]:
    return files_manifest([path for path in paths if path.exists()])


def source_cache_token(functions: Sequence[Callable[..., object]]) -> str:
    hasher = hashlib.sha256()
    for function in functions:
        try:
            source = inspect.getsource(function)
        except (OSError, TypeError):
            source = repr(function)
        hasher.update(source.encode("utf-8"))
    return hasher.hexdigest()


def load_or_build_cached_frames(
    *,
    cache_root: Path,
    cache_name: str,
    manifest: dict,
    builder: Callable[[], Tuple[pd.DataFrame, ...]],
    file_names: Sequence[str],
    refresh: bool,
) -> Tuple[Tuple[pd.DataFrame, ...], bool]:
    cache_dir = cache_root / cache_name
    cache_manifest_path = cache_dir / "manifest.json"
    cache_files = [cache_dir / file_name for file_name in file_names]
    cached_manifest = read_cache_manifest(cache_manifest_path)
    cache_ready = (
        not refresh
        and cached_manifest == manifest
        and all(path.exists() for path in cache_files)
    )
    if cache_ready:
        return tuple(read_cached_frame(path) for path in cache_files), True

    frames = builder()
    cache_dir.mkdir(parents=True, exist_ok=True)
    for frame, path in zip(frames, cache_files):
        write_frame(path, frame)
    write_cache_manifest(cache_manifest_path, manifest)
    return frames, False


def stage_rows_label(frames: Dict[str, pd.DataFrame]) -> str:
    parts: List[str] = []
    for label, frame in frames.items():
        if frame is None:
            continue
        parts.append(f"{label}={len(frame):,}")
    return ", ".join(parts)


def append_stage(performance: List[PerformanceStage], stage: str, started_at: float, cache_status: str, frames: Dict[str, pd.DataFrame], notes: str = "") -> None:
    performance.append(
        PerformanceStage(
            stage=stage,
            seconds=perf_counter() - started_at,
            cache_status=cache_status,
            rows=stage_rows_label(frames),
            notes=notes,
        )
    )


def write_performance_report(
    *,
    performance: Sequence[PerformanceStage],
    output_folder: Path,
    latest_folder: Path,
    previous_report_path: Path,
) -> Dict[str, Path]:
    total_seconds = sum(item.seconds for item in performance)
    baseline_total_seconds = None
    if previous_report_path.exists():
        try:
            baseline_payload = json.loads(previous_report_path.read_text(encoding="utf-8"))
            baseline_total_seconds = baseline_payload.get("total_seconds")
        except Exception:
            baseline_total_seconds = None

    performance_frame = pd.DataFrame(
        [
            {
                "stage": item.stage,
                "seconds": round(item.seconds, 6),
                "cache_status": item.cache_status,
                "rows": item.rows,
                "notes": item.notes,
            }
            for item in performance
        ]
    )
    performance_payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total_seconds": round(total_seconds, 6),
        "baseline_total_seconds": baseline_total_seconds,
        "delta_seconds_vs_baseline": None if baseline_total_seconds is None else round(total_seconds - float(baseline_total_seconds), 6),
        "stages": performance_frame.to_dict(orient="records"),
    }

    csv_path = output_folder / "performance_report.csv"
    json_path = output_folder / "performance_report.json"
    write_frame(csv_path, performance_frame)
    json_path.write_text(json.dumps(performance_payload, indent=2), encoding="utf-8")

    latest_csv = latest_folder / "performance_report.csv"
    latest_json = latest_folder / "performance_report.json"
    shutil.copyfile(csv_path, latest_csv)
    shutil.copyfile(json_path, latest_json)
    return {"performance_report_csv": csv_path, "performance_report_json": json_path}


@lru_cache(maxsize=None)
def _sort_term_code_cached(term_code: str) -> int:
    match = TERM_CODE_RE.fullmatch(term_code)
    if not match:
        return 999999
    return int(match.group(1)) * 10 + SEASON_ORDER.get(match.group(2).upper(), 9)


def sort_term_code(term_code: str) -> int:
    return _sort_term_code_cached(clean_text(term_code).upper())


@lru_cache(maxsize=None)
def _parse_term_code_cached(text: str) -> Tuple[str, str, object, str]:
    match = TERM_CODE_RE.fullmatch(text)
    if match:
        year = int(match.group(1))
        season_code = match.group(2).upper()
        return (
            f"{year}{season_code}",
            f"{SEASON_NAME[season_code]} {year}",
            year,
            SEASON_NAME[season_code],
        )

    match = TERM_RE.search(text)
    if match:
        season_name = match.group(1).title()
        year = int(match.group(2))
        season_code = next(code for code, label in SEASON_NAME.items() if label == season_name)
        return (
            f"{year}{season_code}",
            f"{season_name} {year}",
            year,
            season_name,
        )

    year_match = re.search(r"(19\d{2}|20\d{2})", text)
    if year_match:
        year = int(year_match.group(1))
        return (f"{year}UN", str(year), year, "Unknown")

    return ("", text, pd.NA, "Unknown")


def parse_term_code(value: object) -> Tuple[str, str, object, str]:
    return _parse_term_code_cached(clean_text(value))


def term_label_from_code(term_code: object) -> str:
    code, label, _, _ = parse_term_code(term_code)
    return label if code else clean_text(term_code)


def update_key_from_name(value: str) -> Tuple[int, int, int]:
    match = UPDATE_RE.search(clean_text(value))
    if not match:
        return (0, 0, 0)
    month = int(match.group(1))
    day = int(match.group(2))
    year = int(match.group(3))
    if year < 100:
        year += 2000
    return (year, month, day)


def roster_file_version_details(value: object) -> Tuple[str, float]:
    text = re.sub(r"[_\-.]+", " ", clean_text(value).lower())
    if re.search(r"\bfinal\b", text):
        return "Final", 3
    has_revised = bool(re.search(r"\brevised\b|\brevision\b|\brev\b", text))
    has_updated = bool(re.search(r"\bupdated\b|\bupdate\b", text))
    if has_revised and has_updated:
        return "Revised + Updated", 2.5
    if has_revised:
        return "Revised", 2
    if has_updated:
        return "Updated", 2
    if re.search(r"\binitial\b", text):
        return "Initial", 1
    return "Regular", 1


def roster_file_month_details(value: object) -> Tuple[str, int]:
    text = re.sub(r"[_\-.]+", " ", clean_text(value).lower())
    for month_name, month_number, pattern in MONTH_PATTERNS:
        if re.search(pattern, text):
            return month_name, month_number
    return "", 0


def path_term_candidates(path: Path) -> List[str]:
    candidates: List[str] = [clean_text(path.stem), clean_text(path.name)]
    for part in reversed(path.parts):
        text = clean_text(part)
        if text and text not in candidates:
            candidates.append(text)
    return candidates


def parse_grade_term(path: Path, sheet_name: object) -> str:
    for candidate in [clean_text(sheet_name), *path_term_candidates(path)]:
        term_code, term_label, _, _ = parse_term_code(candidate)
        if term_code:
            return term_label
    return ""


def canonical_snapshot_header(value: object) -> str:
    text = canonical_header(value)
    match = next(
        (
            standard
            for standard, aliases in SNAPSHOT_ALIAS_GROUPS.items()
            if text in aliases
        ),
        clean_text(value),
    )
    return match


def build_person_identity_key(frame: pd.DataFrame) -> pd.Series:
    student_id = frame.get("student_id", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip()
    email = frame.get("email", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    first_name = frame.get("first_name", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    last_name = frame.get("last_name", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    name_key = (last_name + "|" + first_name).where(last_name.ne("") | first_name.ne(""), "")
    return student_id.where(student_id.ne(""), email.where(email.ne(""), name_key))


def build_resolution_identity_key(frame: pd.DataFrame) -> pd.Series:
    student_id = frame.get("student_id", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    email = frame.get("email", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    first_name = frame.get("first_name", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    last_name = frame.get("last_name", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    id_key = ("id:" + student_id).where(student_id.ne(""), "")
    email_key = ("email:" + email).where(email.ne(""), "")
    name_key = ("name:" + last_name + "|" + first_name).where(last_name.ne("") | first_name.ne(""), "")
    return id_key.where(id_key.ne(""), email_key.where(email_key.ne(""), name_key))


def build_review_key(frame: pd.DataFrame, first_name_column: str = "first_name", last_name_column: str = "last_name") -> pd.Series:
    student_id = frame.get("student_id", pd.Series("", index=frame.index)).fillna("").astype(str).str.strip()
    first_name = frame.get(first_name_column, pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    last_name = frame.get(last_name_column, pd.Series("", index=frame.index)).fillna("").astype(str).str.strip().str.lower()
    name_key = ("name::" + last_name + "|" + first_name).where(last_name.ne("") | first_name.ne(""), "")
    return student_id.where(student_id.ne(""), name_key)


def detect_membership_reference_header_row(frame: pd.DataFrame) -> Tuple[Optional[int], Dict[int, Tuple[str, str]], int]:
    best_row: Optional[int] = None
    best_terms: Dict[int, Tuple[str, str]] = {}
    best_chapter_col = 0
    search_rows = min(len(frame.index), 25)
    search_cols = min(len(frame.columns), 40)
    for row_idx in range(search_rows):
        term_columns: Dict[int, Tuple[str, str]] = {}
        for col_idx in range(search_cols):
            term_code, term_label, _, _ = parse_term_code(frame.iat[row_idx, col_idx])
            if term_code:
                term_columns[col_idx] = (term_code, term_label)
        if len(term_columns) < 2:
            continue
        first_term_col = min(term_columns)
        chapter_col = 0
        for candidate in range(first_term_col):
            candidate_values = frame.iloc[row_idx + 1 :, candidate].fillna("").astype(str).map(clean_text)
            has_chapter_text = candidate_values.map(
                lambda value: bool(value)
                and not parse_term_code(value)[0]
                and not re.search(r"(average|total|council)", value, re.IGNORECASE)
            ).any()
            if has_chapter_text:
                chapter_col = candidate
                break
        if len(term_columns) > len(best_terms):
            best_row = row_idx
            best_terms = term_columns
            best_chapter_col = chapter_col
    return best_row, best_terms, best_chapter_col


def parse_membership_count_value(value: object) -> Optional[int]:
    text = clean_text(value)
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    return None


def parse_reference_gpa_value(value: object) -> Optional[float]:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_reference_numeric_entry(value: object) -> Optional[Tuple[float, bool]]:
    text = clean_text(value)
    if not text or text in {"-", "--"}:
        return None
    had_percent = "%" in text
    cleaned = text.replace(",", "").replace("%", "").strip()
    try:
        return float(cleaned), had_percent
    except ValueError:
        return None


def is_integer_like(value: float) -> bool:
    return abs(value - round(value)) < 1e-9


def classify_reference_row(
    label: str,
    source_file: Path,
    source_sheet: str,
    numeric_entries: Sequence[dict],
) -> Tuple[str, str, str, str]:
    normalized_chapter = normalize_chapter_name(label)
    entity_type = "chapter" if normalized_chapter else "benchmark"
    entity_label_normalized = normalized_chapter if normalized_chapter else clean_text(label)
    context = " ".join(
        [
            clean_text(label),
            clean_text(source_sheet),
            clean_text(source_file.stem),
            clean_text(source_file.name),
        ]
    ).lower()
    values = [float(item["reference_value"]) for item in numeric_entries]
    has_percent = any(bool(item.get("had_percent")) for item in numeric_entries)
    all_small = bool(values) and all(0 <= value <= 4.5 for value in values)
    all_integer = bool(values) and all(is_integer_like(value) for value in values)

    if any(token in context for token in ["new member", "new members", "associate member", "nm count", "new mem"]):
        return "new_member_count", entity_type, entity_label_normalized, "keyword:new_member"
    if any(token in context for token in ["gpa", "grade point"]):
        return "average_gpa", entity_type, entity_label_normalized, "keyword:gpa"
    if any(token in context for token in ["retention", "retained", "return rate", "returned", "continuation", "continued", "persistence", "persist"]):
        return "retention_rate", entity_type, entity_label_normalized, "keyword:retention"
    if all_small:
        return "average_gpa", entity_type, entity_label_normalized, "value-range:gpa"
    if has_percent:
        return "retention_rate", entity_type, entity_label_normalized, "value-format:percent"
    if any(token in context for token in ["membership", "member count", "members", "chapter size", "roster", "headcount", "count"]):
        return "membership_count", entity_type, entity_label_normalized, "keyword:membership"
    if all_integer and values and max(values) > 4:
        return "membership_count", entity_type, entity_label_normalized, "value-shape:count"
    return "unknown", entity_type, entity_label_normalized, "unclassified"


def list_reference_files(roots: Sequence[Path]) -> List[Path]:
    seen: set[str] = set()
    files: List[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
                continue
            key = str(path.resolve())
            if key in seen:
                continue
            seen.add(key)
            files.append(path)
    return files


def load_reference_inventory_table(roots: Sequence[Path]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    inventory_rows: List[dict] = []
    issue_rows: List[dict] = []
    inventory_columns = [
        "reference_type",
        "entity_type",
        "entity_label_raw",
        "entity_label_normalized",
        "term_code",
        "term_label",
        "reference_value",
        "classification_basis",
        "source_file",
        "source_sheet",
    ]
    issue_columns = ["exception_type", "source_file", "student_id", "term_code", "details"]

    for path in list_reference_files(roots):
        try:
            workbook = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as exc:
            issue_rows.append(
                {
                    "exception_type": "reference_inventory_unreadable",
                    "source_file": str(path),
                    "student_id": "",
                    "term_code": "",
                    "details": clean_text(exc),
                }
            )
            continue

        for sheet_name, raw_sheet in workbook.items():
            frame = raw_sheet.fillna("")
            header_row, term_columns, label_col = detect_membership_reference_header_row(frame)
            if header_row is None or not term_columns:
                continue
            for row_idx in range(header_row + 1, len(frame.index)):
                label = clean_text(frame.iat[row_idx, label_col]) if label_col < len(frame.columns) else ""
                if not label:
                    continue
                numeric_entries: List[dict] = []
                for col_idx, (term_code, term_label) in term_columns.items():
                    parsed = parse_reference_numeric_entry(frame.iat[row_idx, col_idx])
                    if parsed is None:
                        continue
                    numeric_value, had_percent = parsed
                    numeric_entries.append(
                        {
                            "term_code": term_code,
                            "term_label": term_label,
                            "reference_value": numeric_value,
                            "had_percent": had_percent,
                        }
                    )
                if not numeric_entries:
                    continue
                reference_type, entity_type, normalized_label, basis = classify_reference_row(
                    label,
                    path,
                    clean_text(sheet_name),
                    numeric_entries,
                )
                for entry in numeric_entries:
                    inventory_rows.append(
                        {
                            "reference_type": reference_type,
                            "entity_type": entity_type,
                            "entity_label_raw": label,
                            "entity_label_normalized": normalized_label,
                            "term_code": entry["term_code"],
                            "term_label": entry["term_label"],
                            "reference_value": entry["reference_value"],
                            "classification_basis": basis,
                            "source_file": str(path),
                            "source_sheet": clean_text(sheet_name),
                        }
                    )
                if reference_type == "unknown":
                    issue_rows.append(
                        {
                            "exception_type": "reference_inventory_unclassified_row",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {label}",
                        }
                    )

    inventory = pd.DataFrame(inventory_rows, columns=inventory_columns)
    if not inventory.empty:
        inventory = inventory.drop_duplicates(
            subset=["reference_type", "entity_type", "entity_label_raw", "term_code", "source_file", "source_sheet"],
            keep="first",
        ).reset_index(drop=True)
    issues = pd.DataFrame(issue_rows, columns=issue_columns)
    return inventory, issues


def build_reference_subset(
    inventory: pd.DataFrame,
    reference_type: str,
    entity_type: str,
    value_column: str,
    columns: Sequence[str],
    dedupe_subset: Sequence[str],
) -> pd.DataFrame:
    if inventory.empty:
        return pd.DataFrame(columns=list(columns))
    frame = inventory.loc[
        inventory["reference_type"].eq(reference_type)
        & inventory["entity_type"].eq(entity_type)
    ].copy()
    if frame.empty:
        return pd.DataFrame(columns=list(columns))
    if entity_type == "chapter":
        frame = frame.rename(
            columns={
                "entity_label_raw": "chapter_raw",
                "entity_label_normalized": "chapter",
                "reference_value": value_column,
            }
        )
    else:
        frame["benchmark_label"] = frame["entity_label_normalized"]
        frame = frame.rename(columns={"reference_value": value_column})
    subset = frame.loc[:, list(columns)].copy()
    if value_column in subset.columns:
        subset[value_column] = coerce_numeric(subset[value_column])
    return subset.drop_duplicates(subset=list(dedupe_subset), keep="first").reset_index(drop=True)


def load_membership_reference_table(root: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    reference_rows: List[dict] = []
    issue_rows: List[dict] = []
    columns = [
        "chapter",
        "chapter_raw",
        "term_code",
        "term_label",
        "membership_count_reference",
        "source_file",
        "source_sheet",
    ]
    empty_reference = pd.DataFrame(columns=columns)
    empty_issues = pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    if not root.exists():
        return empty_reference, empty_issues

    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        try:
            workbook = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as exc:
            issue_rows.append(
                {
                    "exception_type": "membership_reference_unreadable",
                    "source_file": str(path),
                    "student_id": "",
                    "term_code": "",
                    "details": clean_text(exc),
                }
            )
            continue

        for sheet_name, raw_sheet in workbook.items():
            frame = raw_sheet.fillna("")
            header_row, term_columns, chapter_col = detect_membership_reference_header_row(frame)
            if header_row is None or not term_columns:
                continue
            for row_idx in range(header_row + 1, len(frame.index)):
                chapter_raw = clean_text(frame.iat[row_idx, chapter_col]) if chapter_col < len(frame.columns) else ""
                if not chapter_raw:
                    continue
                if re.search(r"(average|total|council)", chapter_raw, re.IGNORECASE):
                    continue
                chapter = normalize_chapter_name(chapter_raw)
                if not chapter:
                    issue_rows.append(
                        {
                            "exception_type": "membership_reference_unmapped_chapter",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {chapter_raw}",
                        }
                    )
                    continue
                found_numeric_count = False
                for col_idx, (term_code, term_label) in term_columns.items():
                    count_value = parse_membership_count_value(frame.iat[row_idx, col_idx])
                    if count_value is None:
                        continue
                    found_numeric_count = True
                    reference_rows.append(
                        {
                            "chapter": chapter,
                            "chapter_raw": chapter_raw,
                            "term_code": term_code,
                            "term_label": term_label,
                            "membership_count_reference": count_value,
                            "source_file": str(path),
                            "source_sheet": clean_text(sheet_name),
                        }
                    )
                if not found_numeric_count:
                    issue_rows.append(
                        {
                            "exception_type": "membership_reference_row_without_counts",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {chapter_raw}",
                        }
                    )

    reference = pd.DataFrame(reference_rows, columns=columns)
    if not reference.empty:
        reference = (
            reference.sort_values(["chapter", "term_code", "source_file", "source_sheet"])
            .drop_duplicates(subset=["chapter", "term_code"], keep="first")
            .reset_index(drop=True)
        )
    issues = pd.DataFrame(issue_rows, columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    return reference, issues


def load_gpa_reference_table(root: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    reference_rows: List[dict] = []
    issue_rows: List[dict] = []
    columns = [
        "chapter",
        "chapter_raw",
        "term_code",
        "term_label",
        "chapter_average_gpa_reference",
        "source_file",
        "source_sheet",
    ]
    empty_reference = pd.DataFrame(columns=columns)
    empty_issues = pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    if not root.exists():
        return empty_reference, empty_issues

    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        try:
            workbook = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as exc:
            issue_rows.append(
                {
                    "exception_type": "gpa_reference_unreadable",
                    "source_file": str(path),
                    "student_id": "",
                    "term_code": "",
                    "details": clean_text(exc),
                }
            )
            continue

        for sheet_name, raw_sheet in workbook.items():
            frame = raw_sheet.fillna("")
            header_row, term_columns, chapter_col = detect_membership_reference_header_row(frame)
            if header_row is None or not term_columns:
                continue
            for row_idx in range(header_row + 1, len(frame.index)):
                chapter_raw = clean_text(frame.iat[row_idx, chapter_col]) if chapter_col < len(frame.columns) else ""
                if not chapter_raw:
                    continue
                if re.search(r"(average|total|council)", chapter_raw, re.IGNORECASE):
                    continue
                chapter = normalize_chapter_name(chapter_raw)
                if not chapter:
                    issue_rows.append(
                        {
                            "exception_type": "gpa_reference_unmapped_chapter",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {chapter_raw}",
                        }
                    )
                    continue
                found_numeric_gpa = False
                for col_idx, (term_code, term_label) in term_columns.items():
                    gpa_value = parse_reference_gpa_value(frame.iat[row_idx, col_idx])
                    if gpa_value is None:
                        continue
                    found_numeric_gpa = True
                    reference_rows.append(
                        {
                            "chapter": chapter,
                            "chapter_raw": chapter_raw,
                            "term_code": term_code,
                            "term_label": term_label,
                            "chapter_average_gpa_reference": gpa_value,
                            "source_file": str(path),
                            "source_sheet": clean_text(sheet_name),
                        }
                    )
                if not found_numeric_gpa:
                    issue_rows.append(
                        {
                            "exception_type": "gpa_reference_row_without_values",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {chapter_raw}",
                        }
                    )

    reference = pd.DataFrame(reference_rows, columns=columns)
    if not reference.empty:
        reference = (
            reference.sort_values(["chapter", "term_code", "source_file", "source_sheet"])
            .drop_duplicates(subset=["chapter", "term_code"], keep="first")
            .reset_index(drop=True)
        )
    issues = pd.DataFrame(issue_rows, columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    return reference, issues


def load_gpa_benchmark_reference_table(root: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    reference_rows: List[dict] = []
    issue_rows: List[dict] = []
    columns = [
        "benchmark_label",
        "term_code",
        "term_label",
        "benchmark_average_gpa_reference",
        "source_file",
        "source_sheet",
    ]
    empty_reference = pd.DataFrame(columns=columns)
    empty_issues = pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    if not root.exists():
        return empty_reference, empty_issues

    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        try:
            workbook = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as exc:
            issue_rows.append(
                {
                    "exception_type": "gpa_benchmark_unreadable",
                    "source_file": str(path),
                    "student_id": "",
                    "term_code": "",
                    "details": clean_text(exc),
                }
            )
            continue

        for sheet_name, raw_sheet in workbook.items():
            frame = raw_sheet.fillna("")
            header_row, term_columns, label_col = detect_membership_reference_header_row(frame)
            if header_row is None or not term_columns:
                continue
            for row_idx in range(header_row + 1, len(frame.index)):
                benchmark_label = clean_text(frame.iat[row_idx, label_col]) if label_col < len(frame.columns) else ""
                if not benchmark_label:
                    continue
                found_numeric_gpa = False
                for col_idx, (term_code, term_label) in term_columns.items():
                    gpa_value = parse_reference_gpa_value(frame.iat[row_idx, col_idx])
                    if gpa_value is None:
                        continue
                    found_numeric_gpa = True
                    reference_rows.append(
                        {
                            "benchmark_label": benchmark_label,
                            "term_code": term_code,
                            "term_label": term_label,
                            "benchmark_average_gpa_reference": gpa_value,
                            "source_file": str(path),
                            "source_sheet": clean_text(sheet_name),
                        }
                    )
                if not found_numeric_gpa:
                    issue_rows.append(
                        {
                            "exception_type": "gpa_benchmark_row_without_values",
                            "source_file": str(path),
                            "student_id": "",
                            "term_code": "",
                            "details": f"{sheet_name}: {benchmark_label}",
                        }
                    )

    reference = pd.DataFrame(reference_rows, columns=columns)
    if not reference.empty:
        reference = (
            reference.sort_values(["benchmark_label", "term_code", "source_file", "source_sheet"])
            .drop_duplicates(subset=["benchmark_label", "term_code"], keep="first")
            .reset_index(drop=True)
        )
    issues = pd.DataFrame(issue_rows, columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    return reference, issues


def build_membership_reference_validation(roster: pd.DataFrame, reference: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "chapter",
        "term_code",
        "term_label",
        "membership_count_reference",
        "membership_count_pipeline",
        "difference",
        "comparison_status",
        "source_file",
        "source_sheet",
    ]
    if reference.empty:
        return pd.DataFrame(columns=columns)

    roster_counts = roster.copy()
    roster_counts["_person_key"] = build_person_identity_key(roster_counts)
    roster_counts = roster_counts.loc[roster_counts["_person_key"].ne("")]
    pipeline_counts = (
        roster_counts.groupby(["chapter", "term_code"], dropna=False)["_person_key"]
        .nunique()
        .reset_index(name="membership_count_pipeline")
    )
    validation = reference.merge(pipeline_counts, on=["chapter", "term_code"], how="left")
    validation["term_label"] = validation["term_label"].fillna(validation["term_code"].map(term_label_from_code))
    validation["membership_count_pipeline"] = coerce_numeric(validation["membership_count_pipeline"])
    validation["membership_count_reference"] = coerce_numeric(validation["membership_count_reference"])
    validation["difference"] = validation["membership_count_pipeline"] - validation["membership_count_reference"]
    validation.loc[validation["membership_count_pipeline"].isna(), "difference"] = pd.NA
    validation["comparison_status"] = "Match"
    validation.loc[validation["membership_count_pipeline"].isna(), "comparison_status"] = "Reference Only"
    validation.loc[
        validation["membership_count_pipeline"].notna()
        & validation["difference"].fillna(0).ne(0),
        "comparison_status",
    ] = "Mismatch"
    return validation.loc[:, columns].sort_values(["chapter", "term_code"]).reset_index(drop=True)


def build_new_member_reference_validation(roster: pd.DataFrame, reference: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "chapter",
        "term_code",
        "term_label",
        "new_member_count_reference",
        "new_member_count_pipeline",
        "difference",
        "comparison_status",
        "source_file",
        "source_sheet",
    ]
    if reference.empty:
        return pd.DataFrame(columns=columns)

    roster_counts = roster.copy()
    roster_counts["_person_key"] = build_person_identity_key(roster_counts)
    roster_counts = roster_counts.loc[
        roster_counts["_person_key"].ne("")
        & roster_counts["new_member_flag"].fillna("").astype(str).eq("Yes")
    ]
    pipeline_counts = (
        roster_counts.groupby(["chapter", "term_code"], dropna=False)["_person_key"]
        .nunique()
        .reset_index(name="new_member_count_pipeline")
    )
    validation = reference.merge(pipeline_counts, on=["chapter", "term_code"], how="left")
    validation["term_label"] = validation["term_label"].fillna(validation["term_code"].map(term_label_from_code))
    validation["new_member_count_pipeline"] = coerce_numeric(validation["new_member_count_pipeline"])
    validation["new_member_count_reference"] = coerce_numeric(validation["new_member_count_reference"])
    validation["difference"] = validation["new_member_count_pipeline"] - validation["new_member_count_reference"]
    validation.loc[validation["new_member_count_pipeline"].isna(), "difference"] = pd.NA
    validation["comparison_status"] = "Match"
    validation.loc[validation["new_member_count_pipeline"].isna(), "comparison_status"] = "Reference Only"
    validation.loc[
        validation["new_member_count_pipeline"].notna()
        & validation["difference"].fillna(0).ne(0),
        "comparison_status",
    ] = "Mismatch"
    return validation.loc[:, columns].sort_values(["chapter", "term_code"]).reset_index(drop=True)


def build_gpa_reference_validation(master: pd.DataFrame, reference: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "chapter",
        "term_code",
        "term_label",
        "chapter_average_gpa_reference",
        "chapter_average_gpa_pipeline",
        "difference",
        "comparison_status",
        "source_file",
        "source_sheet",
    ]
    if reference.empty:
        return pd.DataFrame(columns=columns)

    pipeline_frame = master.copy()
    pipeline_frame["term_gpa_numeric"] = coerce_numeric(pipeline_frame["term_gpa"])
    pipeline_frame = pipeline_frame.loc[
        pipeline_frame["chapter"].fillna("").astype(str).str.strip().ne("")
        & pipeline_frame["academic_present"].fillna("").astype(str).eq("Yes")
        & pipeline_frame["term_gpa_numeric"].notna()
    ]
    pipeline_gpa = (
        pipeline_frame.groupby(["chapter", "term_code"], dropna=False)["term_gpa_numeric"]
        .mean()
        .reset_index(name="chapter_average_gpa_pipeline")
    )
    validation = reference.merge(pipeline_gpa, on=["chapter", "term_code"], how="left")
    validation["term_label"] = validation["term_label"].fillna(validation["term_code"].map(term_label_from_code))
    validation["chapter_average_gpa_pipeline"] = coerce_numeric(validation["chapter_average_gpa_pipeline"])
    validation["chapter_average_gpa_reference"] = coerce_numeric(validation["chapter_average_gpa_reference"])
    validation["difference"] = validation["chapter_average_gpa_pipeline"] - validation["chapter_average_gpa_reference"]
    validation.loc[validation["chapter_average_gpa_pipeline"].isna(), "difference"] = pd.NA
    validation["comparison_status"] = "Match"
    validation.loc[validation["chapter_average_gpa_pipeline"].isna(), "comparison_status"] = "Reference Only"
    validation.loc[
        validation["chapter_average_gpa_pipeline"].notna()
        & validation["difference"].abs().fillna(0).gt(0.01),
        "comparison_status",
    ] = "Mismatch"
    return validation.loc[:, columns].sort_values(["chapter", "term_code"]).reset_index(drop=True)


def compute_pipeline_gpa_benchmarks(master: pd.DataFrame, chapter_mapping: pd.DataFrame) -> pd.DataFrame:
    columns = ["benchmark_label", "term_code", "benchmark_average_gpa_pipeline"]
    if master.empty:
        return pd.DataFrame(columns=columns)

    base = master.copy()
    base["term_gpa_numeric"] = coerce_numeric(base["term_gpa"])
    base = base.loc[
        base["academic_present"].fillna("").astype(str).eq("Yes")
        & base["term_gpa_numeric"].notna()
        & base["chapter"].fillna("").astype(str).str.strip().ne("")
    ]
    if base.empty:
        return pd.DataFrame(columns=columns)

    mapping = chapter_mapping.copy()
    if not mapping.empty and "chapter" in mapping.columns:
        mapping["chapter"] = mapping["chapter"].fillna("").astype(str).str.strip()
        mapping["org_type"] = mapping.get("org_type", "").astype(str) if "org_type" in mapping.columns else ""
        base = base.merge(mapping[["chapter", "org_type"]].drop_duplicates(subset=["chapter"]), on="chapter", how="left")
    else:
        base["org_type"] = ""

    benchmark_frames: List[pd.DataFrame] = []

    all_greek = (
        base.groupby("term_code", dropna=False)["term_gpa_numeric"]
        .mean()
        .reset_index(name="benchmark_average_gpa_pipeline")
    )
    all_greek["benchmark_label"] = "All Greek Average"
    benchmark_frames.append(all_greek)

    org_type_series = base["org_type"].fillna("").astype(str).str.lower()
    sorority = base.loc[org_type_series.str.contains("soror")]
    if not sorority.empty:
        frame = sorority.groupby("term_code", dropna=False)["term_gpa_numeric"].mean().reset_index(name="benchmark_average_gpa_pipeline")
        frame["benchmark_label"] = "All Sorority Average"
        benchmark_frames.append(frame)

    fraternity = base.loc[org_type_series.str.contains("fratern")]
    if not fraternity.empty:
        frame = fraternity.groupby("term_code", dropna=False)["term_gpa_numeric"].mean().reset_index(name="benchmark_average_gpa_pipeline")
        frame["benchmark_label"] = "All Fraternity Average"
        benchmark_frames.append(frame)

    if not benchmark_frames:
        return pd.DataFrame(columns=columns)
    return pd.concat(benchmark_frames, ignore_index=True).loc[:, columns]


def build_gpa_benchmark_validation(master: pd.DataFrame, reference: pd.DataFrame, chapter_mapping: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "benchmark_label",
        "term_code",
        "term_label",
        "benchmark_average_gpa_reference",
        "benchmark_average_gpa_pipeline",
        "difference",
        "comparison_status",
        "source_file",
        "source_sheet",
    ]
    if reference.empty:
        return pd.DataFrame(columns=columns)

    pipeline_benchmarks = compute_pipeline_gpa_benchmarks(master, chapter_mapping)
    validation = reference.merge(pipeline_benchmarks, on=["benchmark_label", "term_code"], how="left")
    validation["term_label"] = validation["term_label"].fillna(validation["term_code"].map(term_label_from_code))
    validation["benchmark_average_gpa_pipeline"] = coerce_numeric(validation["benchmark_average_gpa_pipeline"])
    validation["benchmark_average_gpa_reference"] = coerce_numeric(validation["benchmark_average_gpa_reference"])
    validation["difference"] = validation["benchmark_average_gpa_pipeline"] - validation["benchmark_average_gpa_reference"]
    validation.loc[validation["benchmark_average_gpa_pipeline"].isna(), "difference"] = pd.NA
    validation["comparison_status"] = "Match"
    validation.loc[validation["benchmark_average_gpa_pipeline"].isna(), "comparison_status"] = "Reference Only"
    validation.loc[
        validation["benchmark_average_gpa_pipeline"].notna()
        & validation["difference"].abs().fillna(0).gt(0.01),
        "comparison_status",
    ] = "Mismatch"
    return validation.loc[:, columns].sort_values(["benchmark_label", "term_code"]).reset_index(drop=True)


def canonical_graduation_header(value: object) -> str:
    text = canonical_header(value)
    return next(
        (
            standard
            for standard, aliases in GRADUATION_ALIAS_GROUPS.items()
            if text in aliases
        ),
        clean_text(value),
    )


def is_snapshot_filename(path: Path) -> bool:
    return clean_text(path.stem).lower().startswith("new member")


def list_source_files(folder: Path, extensions: Optional[set[str]] = None) -> List[Path]:
    if not folder.exists():
        return []
    allowed = extensions or TABULAR_SOURCE_EXTENSIONS
    return sorted(path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in allowed)


def roster_files(roots: Sequence[Path]) -> List[Path]:
    files: List[Path] = []
    seen: set[Path] = set()
    for root in roots:
        for path in list_source_files(root, ROSTER_SOURCE_EXTENSIONS):
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                files.append(path)
    return sorted(files)


def academic_files(root: Path) -> List[Path]:
    return [path for path in list_source_files(root) if not is_snapshot_filename(path)]


def snapshot_files(root: Path) -> List[Path]:
    return [path for path in list_source_files(root) if is_snapshot_filename(path)]


def graduation_files(root: Path) -> List[Path]:
    if not root.exists():
        return []
    return sorted(path for path in list_source_files(root) if clean_text(path.stem).lower().find("graduat") >= 0)


def transcript_text_files(root: Path) -> List[Path]:
    if not root.exists():
        return []
    return sorted(path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in TEXT_SOURCE_EXTENSIONS)


def ensure_transcript_text_manifest_template(path: Path = TRANSCRIPT_TEXT_MANIFEST_PATH) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        columns=[
            "source_file",
            "student_id",
            "first_name",
            "last_name",
            "notes",
        ]
    ).to_csv(path, index=False)


def load_transcript_text_manifest(path: Path = TRANSCRIPT_TEXT_MANIFEST_PATH) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["source_file", "student_id", "first_name", "last_name", "notes"])
    frame = pd.read_csv(path)
    frame.columns = [canonical_header(column) for column in frame.columns]
    rename_map = {
        "source file": "source_file",
        "student id": "student_id",
        "first name": "first_name",
        "last name": "last_name",
        "notes": "notes",
    }
    frame = frame.rename(columns={column: rename_map.get(column, column) for column in frame.columns})
    for column in ["source_file", "student_id", "first_name", "last_name", "notes"]:
        if column not in frame.columns:
            frame[column] = ""
    frame = frame[["source_file", "student_id", "first_name", "last_name", "notes"]].copy()
    frame["source_file"] = frame["source_file"].fillna("").astype(str).str.strip()
    frame["student_id"] = frame["student_id"].map(normalize_banner_id)
    frame["first_name"] = frame["first_name"].map(clean_text)
    frame["last_name"] = frame["last_name"].map(clean_text)
    frame["notes"] = frame["notes"].map(clean_text)
    return frame.loc[frame["source_file"].ne("")].drop_duplicates(subset=["source_file"], keep="first").reset_index(drop=True)


def transcript_identity_from_filename(path: Path) -> dict[str, str]:
    stem = path.stem
    id_match = re.search(r"(?i)(A0?\d{7}|A\d{8}|\d{7,8})", stem)
    student_id_raw = clean_text(id_match.group(1)) if id_match else ""
    student_id = normalize_banner_id(student_id_raw)
    working = stem
    if id_match:
        working = (stem[: id_match.start()] + " " + stem[id_match.end() :]).strip()
    tokens = [token for token in re.split(r"[_\-\s,]+", working) if clean_text(token)]
    first_name = ""
    last_name = ""
    if len(tokens) >= 2:
        last_name = clean_text(tokens[0])
        first_name = clean_text(tokens[1])
    return {
        "student_id": student_id,
        "student_id_raw": student_id_raw,
        "first_name": first_name,
        "last_name": last_name,
    }


def resolve_transcript_identity(path: Path, manifest: pd.DataFrame) -> dict[str, str]:
    manifest_row = manifest.loc[manifest["source_file"].fillna("").astype(str).str.strip().eq(path.name)].head(1)
    if not manifest_row.empty:
        row = manifest_row.iloc[0]
        return {
            "student_id": normalize_banner_id(row.get("student_id", "")),
            "student_id_raw": clean_text(row.get("student_id", "")),
            "first_name": clean_text(row.get("first_name", "")),
            "last_name": clean_text(row.get("last_name", "")),
            "identity_resolution_basis": "transcript_manifest",
            "identity_resolution_notes": clean_text(row.get("notes", "")) or "Matched from config/transcript_text_manifest.csv.",
        }

    inferred = transcript_identity_from_filename(path)
    basis = "transcript_filename_name"
    notes = "Matched from transcript filename."
    if inferred["student_id"]:
        basis = "transcript_filename_student_id"
        notes = "Matched from transcript filename student ID."
    elif inferred["first_name"] or inferred["last_name"]:
        basis = "transcript_filename_name"
        notes = "Matched from transcript filename name pattern."
    else:
        basis = "transcript_unresolved"
        notes = "No manifest row or reliable filename identity pattern was found."
    return {
        **inferred,
        "identity_resolution_basis": basis,
        "identity_resolution_notes": notes,
    }


def parse_transcript_credit_token(value: str) -> Tuple[object, object]:
    token = clean_text(value)
    if not token or token == "--":
        return pd.NA, pd.NA
    paren_match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*\((\d+(?:\.\d+)?)\)", token)
    if paren_match:
        return float(paren_match.group(2)), float(paren_match.group(1))
    numeric_match = re.fullmatch(r"\d+(?:\.\d+)?", token)
    if numeric_match:
        number = float(numeric_match.group(0))
        return number, number
    return pd.NA, pd.NA


def parse_transcript_course_line(line: str) -> Optional[dict]:
    cleaned = clean_text(line)
    if not cleaned:
        return None
    match = re.match(r"^(?P<credit_token>--|\d+(?:\.\d+)?(?:\s*\(\d+(?:\.\d+)?\))?)\s+(?P<body>.+)$", cleaned)
    if not match:
        return None
    credit_token = clean_text(match.group("credit_token"))
    body = clean_text(match.group("body"))
    credits_attempted, credits_earned = parse_transcript_credit_token(credit_token)
    transfer_flag = "[TR]" in body.upper()
    grade_match = re.search(r"\s(?P<grade>A\+|A-|A|B\+|B-|B|C\+|C-|C|D\+|D-|D|F|W|I|P|NP|CR|NC|RW|S|U)\s*$", body, re.IGNORECASE)
    grade = clean_text(grade_match.group("grade")).upper() if grade_match else ""
    descriptor = clean_text(body[: grade_match.start()]) if grade_match else body
    if "|" in descriptor:
        course_code, remainder = descriptor.split("|", 1)
        course_code = clean_text(course_code)
        remainder_parts = clean_text(remainder).split(None, 1)
        section_type = clean_text(remainder_parts[0]) if remainder_parts else ""
        course_title = clean_text(remainder_parts[1]) if len(remainder_parts) > 1 else ""
    else:
        course_code = ""
        section_type = ""
        course_title = descriptor
    return {
        "raw_course_line": cleaned,
        "course_code": course_code,
        "section_type": section_type,
        "course_title": course_title,
        "grade": grade,
        "transfer_flag": "Yes" if transfer_flag else "No",
        "credits_attempted": credits_attempted,
        "credits_earned": credits_earned,
        "credit_token_raw": credit_token,
    }


def parse_transcript_numeric(value: str) -> object:
    text = clean_text(value).replace("%", "")
    if not text or text == "-":
        return pd.NA
    try:
        return float(text)
    except ValueError:
        return pd.NA


def extract_transcript_summary_pairs(lines: List[str]) -> Dict[str, str]:
    pairs: Dict[str, str] = {}
    idx = 0
    while idx < len(lines):
        label = clean_text(lines[idx])
        if not label:
            idx += 1
            continue
        if label.endswith(":"):
            value = ""
            look_ahead = idx + 1
            while look_ahead < len(lines):
                candidate = clean_text(lines[look_ahead])
                if not candidate:
                    look_ahead += 1
                    continue
                if candidate.endswith(":"):
                    break
                value = candidate
                break
            pairs[label[:-1].strip().lower()] = value
            idx = look_ahead + 1 if value else idx + 1
            continue
        idx += 1
    return pairs


def load_transcript_text_tables(root: Path, manifest: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    term_rows: List[dict] = []
    course_rows: List[dict] = []
    audit_rows: List[dict] = []
    issue_rows: List[dict] = []

    for path in transcript_text_files(root):
        identity = resolve_transcript_identity(path, manifest)
        warnings: List[str] = []
        unmatched_lines: List[str] = []
        file_term_rows = 0
        file_course_rows = 0

        try:
            raw_text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            raw_text = path.read_text(encoding="latin-1")
        except Exception as exc:
            issue_rows.append(
                {
                    "exception_type": "transcript_text_open_error",
                    "source_file": path.name,
                    "student_id": identity["student_id"],
                    "term_code": "",
                    "details": clean_text(exc),
                }
            )
            audit_rows.append(
                {
                    "source_file": path.name,
                    "student_id": identity["student_id"],
                    "student_id_raw": identity["student_id_raw"],
                    "first_name": identity["first_name"],
                    "last_name": identity["last_name"],
                    "identity_resolution_basis": identity["identity_resolution_basis"],
                    "parse_status": "error",
                    "term_count": 0,
                    "course_count": 0,
                    "warning_count": 1,
                    "warnings": clean_text(exc),
                    "unmatched_lines": "",
                }
            )
            continue

        lines = [line.rstrip() for line in raw_text.splitlines()]
        term_headers = [(idx, *parse_term_code(clean_text(line))) for idx, line in enumerate(lines) if TERM_RE.fullmatch(clean_text(line))]
        if not term_headers:
            warnings.append("No term headers were found in transcript text.")

        for position, (line_idx, term_code, term_label, term_year, term_season) in enumerate(term_headers):
            next_idx = term_headers[position + 1][0] if position + 1 < len(term_headers) else len(lines)
            block_lines = [clean_text(line) for line in lines[line_idx + 1 : next_idx]]
            pre_enrollment_idx = next(
                (idx for idx, value in enumerate(block_lines) if value.lower().startswith("pre-enrollment and progression")),
                None,
            )
            if pre_enrollment_idx is not None:
                block_lines = block_lines[:pre_enrollment_idx]
            at_a_glance_idx = next((idx for idx, value in enumerate(block_lines) if value.lower() == "term at a glance:"), None)
            course_lines = block_lines[:at_a_glance_idx] if at_a_glance_idx is not None else block_lines
            summary_lines = block_lines[at_a_glance_idx + 1 :] if at_a_glance_idx is not None else []
            if at_a_glance_idx is None:
                warnings.append(f"{term_label}: missing 'Term at a glance' block.")

            parsed_courses: List[dict] = []
            for line in course_lines:
                if not line:
                    continue
                course = parse_transcript_course_line(line)
                if course is None:
                    unmatched_lines.append(f"{term_label}: {line}")
                    continue
                parsed_courses.append(course)
                file_course_rows += 1
                course_rows.append(
                    {
                        "student_id": identity["student_id"],
                        "student_id_raw": identity["student_id_raw"],
                        "first_name": identity["first_name"],
                        "last_name": identity["last_name"],
                        "source_file": path.name,
                        "term_code": term_code,
                        "term_label": term_label,
                        **course,
                    }
                )

            summary_pairs = extract_transcript_summary_pairs(summary_lines)
            summary_credits = parse_transcript_numeric(summary_pairs.get("credits", ""))
            summary_completion = parse_transcript_numeric(summary_pairs.get("credit comp %", ""))
            summary_term_gpa = parse_transcript_numeric(summary_pairs.get("term gpa", ""))
            summary_cumulative_gpa = parse_transcript_numeric(summary_pairs.get("cum gpa", ""))
            summary_academic_standing = clean_text(summary_pairs.get("academic standing", ""))
            explicit_graduation_term_text = clean_text(summary_pairs.get("graduation term", ""))
            explicit_graduation_term_code, explicit_graduation_term_label, _, _ = parse_term_code(explicit_graduation_term_text)
            summary_graduation_signal_text = ""
            for label, value in summary_pairs.items():
                if "graduat" in label and clean_text(value):
                    summary_graduation_signal_text = clean_text(value)
                    break

            attempted_sum = pd.to_numeric(pd.Series([row["credits_attempted"] for row in parsed_courses]), errors="coerce").sum(min_count=1)
            earned_sum = pd.to_numeric(pd.Series([row["credits_earned"] for row in parsed_courses]), errors="coerce").sum(min_count=1)
            file_term_rows += 1
            term_rows.append(
                {
                    "student_id": identity["student_id"],
                    "student_id_raw": identity["student_id_raw"],
                    "identity_resolution_basis": identity["identity_resolution_basis"],
                    "identity_resolution_notes": identity["identity_resolution_notes"],
                    "first_name": identity["first_name"],
                    "last_name": identity["last_name"],
                    "source_file": path.name,
                    "term_code": term_code,
                    "term_label": term_label,
                    "term_year": term_year,
                    "term_season": term_season,
                    "summary_credits_earned": summary_credits if not pd.isna(summary_credits) else earned_sum,
                    "summary_credit_completion_pct": summary_completion,
                    "summary_term_gpa": summary_term_gpa,
                    "summary_cumulative_gpa": summary_cumulative_gpa,
                    "summary_academic_standing": summary_academic_standing,
                    "summary_graduation_term_code": explicit_graduation_term_code,
                    "summary_graduation_term_label": explicit_graduation_term_label,
                    "summary_graduation_signal_text": summary_graduation_signal_text,
                }
            )

            if summary_academic_standing == "":
                warnings.append(f"{term_label}: missing academic standing in transcript summary block.")
            if pd.isna(summary_term_gpa):
                warnings.append(f"{term_label}: missing term GPA in transcript summary block.")
            if pd.isna(summary_credits) and pd.isna(earned_sum):
                warnings.append(f"{term_label}: missing credits in transcript summary and course lines.")

        parse_status = "parsed"
        if not term_headers:
            parse_status = "warning"
        if unmatched_lines or warnings:
            parse_status = "warning" if parse_status != "error" else parse_status

        audit_rows.append(
            {
                "source_file": path.name,
                "student_id": identity["student_id"],
                "student_id_raw": identity["student_id_raw"],
                "first_name": identity["first_name"],
                "last_name": identity["last_name"],
                "identity_resolution_basis": identity["identity_resolution_basis"],
                "parse_status": parse_status,
                "term_count": file_term_rows,
                "course_count": file_course_rows,
                "warning_count": len(warnings) + len(unmatched_lines),
                "warnings": " | ".join(warnings),
                "unmatched_lines": " | ".join(unmatched_lines[:100]),
            }
        )

    transcript_terms = ensure_columns(pd.DataFrame(term_rows), TRANSCRIPT_TERM_COLUMNS)
    transcript_courses = ensure_columns(pd.DataFrame(course_rows), TRANSCRIPT_COURSE_COLUMNS)
    transcript_audit = ensure_columns(pd.DataFrame(audit_rows), TRANSCRIPT_AUDIT_COLUMNS)
    transcript_issues = ensure_columns(pd.DataFrame(issue_rows), ["exception_type", "source_file", "student_id", "term_code", "details"])
    return transcript_terms, transcript_courses, transcript_audit, transcript_issues


def transcript_terms_to_academic_rows(transcript_terms: pd.DataFrame, transcript_courses: pd.DataFrame) -> pd.DataFrame:
    if transcript_terms.empty:
        return pd.DataFrame(columns=load_schema()["tables"]["academic_term"])

    course_rollup = pd.DataFrame(columns=["source_file", "term_code", "attempted_hours_term", "earned_hours_from_courses"])
    if not transcript_courses.empty:
        course_rollup = (
            transcript_courses.groupby(["source_file", "term_code"], dropna=False)
            .agg(
                attempted_hours_term=("credits_attempted", lambda values: pd.to_numeric(pd.Series(values), errors="coerce").sum(min_count=1)),
                earned_hours_from_courses=("credits_earned", lambda values: pd.to_numeric(pd.Series(values), errors="coerce").sum(min_count=1)),
            )
            .reset_index()
        )

    merged = transcript_terms.merge(course_rollup, on=["source_file", "term_code"], how="left")
    academic = pd.DataFrame(
        {
            "student_id": merged["student_id"],
            "student_id_raw": merged["student_id_raw"],
            "identity_resolution_basis": merged["identity_resolution_basis"],
            "identity_resolution_notes": merged["identity_resolution_notes"],
            "first_name": merged["first_name"],
            "last_name": merged["last_name"],
            "email": "",
            "source_file": merged["source_file"],
            "source_sheet": "transcript_text",
            "term_code": merged["term_code"],
            "term_label": merged["term_label"],
            "term_year": merged["term_year"],
            "term_season": merged["term_season"],
            "term_source_basis": "transcript_text",
            "academic_status_raw": merged["summary_graduation_signal_text"].where(
                merged["summary_graduation_signal_text"].fillna("").astype(str).str.strip().ne(""),
                "",
            ),
            "major": "",
            "term_gpa": pd.to_numeric(merged["summary_term_gpa"], errors="coerce"),
            "institutional_cumulative_gpa": pd.to_numeric(merged["summary_cumulative_gpa"], errors="coerce"),
            "overall_cumulative_gpa": pd.to_numeric(merged["summary_cumulative_gpa"], errors="coerce"),
            "transfer_gpa": pd.NA,
            "attempted_hours_term": pd.to_numeric(merged["attempted_hours_term"], errors="coerce"),
            "earned_hours_term": pd.to_numeric(merged["summary_credits_earned"], errors="coerce").where(
                pd.to_numeric(merged["summary_credits_earned"], errors="coerce").notna(),
                pd.to_numeric(merged["earned_hours_from_courses"], errors="coerce"),
            ),
            "institutional_cumulative_hours": pd.NA,
            "total_cumulative_hours": pd.NA,
            "academic_standing_raw": merged["summary_academic_standing"],
            "academic_standing_bucket": merged["summary_academic_standing"].map(standing_bucket),
            "graduation_term_code": merged["summary_graduation_term_code"],
            "graduation_term_label": merged["summary_graduation_term_label"],
        }
    )
    return ensure_columns(academic, load_schema()["tables"]["academic_term"])


def build_transcript_text_cache_bundle(root: Path, manifest: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    transcript_terms, transcript_courses, transcript_audit, transcript_issues = load_transcript_text_tables(root, manifest)
    transcript_academic_term = transcript_terms_to_academic_rows(transcript_terms, transcript_courses)
    return transcript_terms, transcript_courses, transcript_audit, transcript_issues, transcript_academic_term


def source_label_for_roster_path(path: Path, roots: Sequence[Path]) -> str:
    for root in roots:
        try:
            return source_file_label(path, root)
        except Exception:
            continue
    return source_file_label(path)


def normalize_email(value: object) -> str:
    return clean_text(value).lower()


def person_name_key(first_name: object, last_name: object) -> Tuple[str, str]:
    return clean_text(first_name).lower(), clean_text(last_name).lower()


def chapter_is_missing(value: object) -> bool:
    text = clean_text(value)
    if not text:
        return True
    normalized = normalize_chapter_name(text)
    return not normalized or normalized == "Unknown"


def secondary_organization_set(settings: Dict[str, object]) -> set[str]:
    values = settings.get("secondary_organizations", []) if isinstance(settings, dict) else []
    normalized: set[str] = set()
    for value in values:
        chapter = normalize_chapter_name(value)
        if chapter and chapter != "Unknown":
            normalized.add(chapter)
    return normalized


def is_secondary_organization(value: object, settings: Dict[str, object]) -> bool:
    normalized = normalize_chapter_name(clean_text(value))
    return bool(normalized) and normalized in secondary_organization_set(settings)


def choose_preferred_roster_rows(roster: pd.DataFrame, settings: Dict[str, object]) -> pd.DataFrame:
    if roster.empty:
        return roster

    secondary_orgs = secondary_organization_set(settings)
    working = roster.copy()
    working["_identity_key"] = build_resolution_identity_key(working)
    working["_chapter_missing"] = working["chapter"].map(chapter_is_missing).astype(int)
    working["_secondary_org"] = working["chapter"].map(lambda value: normalize_chapter_name(clean_text(value)) in secondary_orgs).astype(int)
    working["_source_version_priority"] = coerce_numeric(working.get("roster_file_version_priority", pd.Series([1] * len(working), index=working.index))).fillna(1)
    working["_source_month_priority"] = coerce_numeric(working.get("roster_file_month_priority", pd.Series([0] * len(working), index=working.index))).fillna(0)
    working["_source_format_priority"] = working.get("source_file", pd.Series([""] * len(working), index=working.index)).map(source_file_format_priority)
    working["_assignment_rank"] = working["chapter_assignment_source"].fillna("").astype(str).map(
        {
            "manual_override": 0,
            "matched_by_id_name": 1,
            "matched_by_id": 2,
            "original": 3,
            "inferred_from_sheet_name": 4,
            "inferred_from_file_name": 5,
            "unresolved": 6,
        }
    ).fillna(9)
    preferred = (
        working.sort_values(
            by=["_identity_key", "term_code", "_chapter_missing", "_secondary_org", "_source_version_priority", "_source_month_priority", "_source_format_priority", "_assignment_rank", "chapter", "source_file", "source_sheet"],
            ascending=[True, True, True, True, False, False, False, True, True, True, True],
            na_position="last",
        )
        .drop_duplicates(subset=["_identity_key", "term_code"], keep="first")
        .reset_index(drop=True)
    )
    return preferred.drop(columns=["_identity_key", "_chapter_missing", "_secondary_org", "_source_version_priority", "_source_month_priority", "_source_format_priority", "_assignment_rank"], errors="ignore").reset_index(drop=True)


def chapter_assignment_details(
    path: Path,
    sheet_name: str,
    chapter_raw: object,
    fallback_raw: object,
    fallback_source: str,
) -> Tuple[str, str, str, str]:
    explicit_chapter = normalize_chapter_name(chapter_raw)
    if explicit_chapter and explicit_chapter != "Unknown":
        return explicit_chapter, "original", "high", "Loaded from source chapter field."

    fallback_chapter = normalize_chapter_name(fallback_raw)
    if fallback_chapter and fallback_chapter != "Unknown" and fallback_source == "original":
        return fallback_chapter, "original", "high", "Loaded from inline chapter label in source sheet."

    inferred_sheet = normalize_chapter_name(sheet_name)
    inferred_file = chapter_from_filename(path)
    inferred = infer_chapter(path, sheet_name) or fallback_chapter

    if inferred and inferred != "Unknown":
        if inferred_sheet and inferred == inferred_sheet and not is_placeholder_sheet_name(sheet_name):
            return inferred, "inferred_from_sheet_name", "medium", "Inferred from workbook sheet name."
        if inferred_file and inferred == inferred_file:
            return inferred, "inferred_from_file_name", "medium", "Inferred from source file name."
        if fallback_source in {"inferred_from_sheet_name", "inferred_from_file_name"}:
            return inferred, fallback_source, "medium", f"Inferred from {fallback_source.replace('_', ' ')}."

    return "", "unresolved", "low", "No reliable chapter assignment was found in the source row, sheet, or file name."


def map_grade_headers(headers: Sequence[object]) -> Dict[str, int]:
    mapped: Dict[str, int] = {}
    canon_headers = [canonical_header(value) for value in headers]
    for idx, header in enumerate(canon_headers):
        for target, aliases in GRADE_COLUMN_ALIASES.items():
            if header in aliases and target not in mapped:
                mapped[target] = idx
    return mapped


def roster_status_bucket(raw_status: object, raw_position: object) -> str:
    status = normalize_status(clean_text(raw_status))
    status_text = status.upper()
    combined = f"{status} {clean_text(raw_position)}".upper()
    if has_explicit_roster_graduation_status(raw_status, status):
        return "Graduated"
    if "SUSPEND" in status_text:
        return "Suspended"
    if "TRANSFER" in status_text:
        return "Transfer"
    if "REVOK" in status_text:
        return "Revoked"
    if "RESIGN" in status_text:
        return "Resigned"
    if "INACTIVE" in status_text or "DROP" in status_text or "REMOVE" in status_text:
        return "Inactive"
    if "NEW MEMBER" in combined:
        return "New Member"
    if "ACTIVE" in combined or "MEMBER" in combined or "COUNCIL" in combined:
        return "Active"
    return status or "Unknown"


def has_explicit_roster_graduation_status(raw_status: object, normalized_status: object = "") -> bool:
    raw_text = clean_text(raw_status).upper()
    normalized_text = clean_text(normalized_status).upper()
    if normalized_text == "GRADUATED":
        return True
    return raw_text in {"G", "GRAD", "GRADUATED"}


def outcome_bucket_from_signals(status_bucket: str, academic_status_raw: str, snapshot_status_raw: str) -> Tuple[str, str]:
    signals = " ".join([status_bucket, clean_text(academic_status_raw), clean_text(snapshot_status_raw)]).upper()
    if "SUSPEND" in signals:
        return "Suspended", "Explicit suspension signal"
    if "TRANSFER" in signals:
        return "Transfer", "Explicit transfer signal"
    if any(token in signals for token in ["INACTIVE", "DROP", "RESIGN", "REVOK", "REMOVE", "WITHDRAW", "TERMINAT", "DISMISS", "EXPEL"]):
        return "Dropped/Resigned/Revoked/Inactive", "Explicit non-graduate exit signal"
    if any(token in signals for token in ["ACTIVE", "CURRENT", "MEMBER", "NEW MEMBER", "COUNCIL", "ENROLLED"]):
        return "Active/Unknown", "Current or active signal only"
    return "Unknown", "No explicit outcome evidence"


def has_confirmed_graduation_text(value: object) -> bool:
    text = clean_text(value).upper()
    if not text:
        return False
    if any(token in text for token in ["DEGREE SEEK", "SEEKING DEGREE", "NON-DEGREE", "NON DEGREE"]):
        return False
    return any(
        [
            bool(re.search(r"\bGRADUATED\b", text)),
            bool(re.search(r"\bGRAD\b", text)),
            "DEGREE AWARDED" in text,
            "AWARDED DEGREE" in text,
            "DEGREE CONFER" in text,
            "CONFERRED DEGREE" in text,
        ]
    )


def standing_bucket(value: object) -> str:
    text = clean_text(value).upper()
    if not text:
        return "Unknown"
    if any(token in text for token in ["GOOD", "CLEAR", "SATISFACTORY"]):
        return "Good Standing"
    if any(token in text for token in ["PROBATION", "WARNING"]):
        return "Probation/Warning"
    if "SUSPEND" in text:
        return "Suspended"
    if any(token in text for token in ["DISMISS", "SEPARAT", "EXPEL", "TERMINAT"]):
        return "Dismissed/Separated"
    return "Other/Unmapped"


def metric_row(metric_group: str, metric_label: str, cohort: str, eligible: int, numerator: Optional[int] = None, rate: Optional[float] = None, average: Optional[float] = None, dimension: str = "", value: str = "", notes: str = "") -> dict:
    return {
        "Metric Group": metric_group,
        "Metric Label": metric_label,
        "Cohort": cohort,
        "Dimension": dimension,
        "Value": value,
        "Eligible Students": eligible,
        "Student Count": numerator if numerator is not None else "",
        "Rate": rate if rate is not None else "",
        "Average Value": average if average is not None else "",
        "Notes": notes,
    }


def choose_best_snapshot_rows(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    ranked = frame.copy()
    ranked["_filled_fields"] = ranked.notna().sum(axis=1)
    ranked["_status_present"] = ranked["Student Status"].fillna("").astype(str).str.strip().ne("").astype(int)
    ranked = ranked.sort_values(
        by=["Student ID", "_filled_fields", "_status_present", "Last Name", "First Name"],
        ascending=[True, False, False, True, True],
    )
    ranked = ranked.drop_duplicates(subset=["Student ID"], keep="first")
    return ranked.drop(columns=["_filled_fields", "_status_present"])


def load_snapshot_table(root: Path) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for path in snapshot_files(root):
        if path.suffix.lower() == ".csv":
            frame = pd.read_csv(path)
            frame.columns = [canonical_snapshot_header(column) for column in frame.columns]
        else:
            selected: Optional[pd.DataFrame] = None
            for _, sheet in pd.read_excel(path, sheet_name=None).items():
                candidate = sheet.copy()
                candidate.columns = [canonical_snapshot_header(column) for column in candidate.columns]
                if {"Student ID", "First Name", "Last Name"}.issubset(set(candidate.columns)):
                    selected = candidate
                    break
            if selected is None:
                continue
            frame = selected

        for column in SNAPSHOT_COLUMNS:
            if column not in frame.columns:
                frame[column] = ""
        frame = frame[SNAPSHOT_COLUMNS[:-1]].copy()
        frame["Student ID"] = frame["Student ID"].map(normalize_banner_id)
        frame["First Name"] = frame["First Name"].map(clean_text)
        frame["Last Name"] = frame["Last Name"].map(clean_text)
        frame["NetID"] = frame["NetID"].map(clean_text)
        frame["Student Status"] = frame["Student Status"].map(clean_text)
        frame["Student Status (FT/PT)"] = frame["Student Status (FT/PT)"].map(clean_text)
        frame["Snapshot Source File"] = path.name
        frame = frame.loc[frame["Student ID"].ne("")]
        frames.append(frame)

    if not frames:
        return pd.DataFrame(columns=SNAPSHOT_COLUMNS)

    combined = pd.concat(frames, ignore_index=True)
    combined = choose_best_snapshot_rows(combined)
    return combined.reset_index(drop=True)


def load_graduation_table(root: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows: List[dict] = []
    exceptions: List[dict] = []

    for path in graduation_files(root):
        try:
            if path.suffix.lower() == ".csv":
                frame = pd.read_csv(path)
                frame.columns = [canonical_graduation_header(column) for column in frame.columns]
                frames = [frame]
            else:
                workbook_frames = []
                for _, sheet in pd.read_excel(path, sheet_name=None).items():
                    candidate = sheet.copy()
                    candidate.columns = [canonical_graduation_header(column) for column in candidate.columns]
                    workbook_frames.append(candidate)
                frames = workbook_frames
        except Exception as exc:
            exceptions.append(
                {
                    "exception_type": "graduation_open_error",
                    "source_file": path.name,
                    "student_id": "",
                    "term_code": "",
                    "details": str(exc),
                }
            )
            continue

        for frame in frames:
            if "Student ID" not in frame.columns and not {"First Name", "Last Name"}.issubset(set(frame.columns)):
                continue
            for column in GRADUATION_COLUMNS:
                if column not in frame.columns:
                    frame[column] = ""
            subset = frame[GRADUATION_COLUMNS[:-1]].copy()
            subset["Student ID"] = subset["Student ID"].map(normalize_banner_id)
            subset["First Name"] = subset["First Name"].map(clean_text)
            subset["Last Name"] = subset["Last Name"].map(clean_text)
            subset["Graduation Term"] = subset["Graduation Term"].map(clean_text)
            subset["Outcome"] = subset["Outcome"].map(clean_text)
            subset["Graduation Source File"] = path.name
            rows.extend(subset.to_dict("records"))

    if not rows:
        return pd.DataFrame(columns=GRADUATION_COLUMNS), pd.DataFrame(exceptions)

    graduation = pd.DataFrame(rows)
    graduation = graduation.loc[
        graduation["Student ID"].fillna("").astype(str).str.strip().ne("")
        | (
            graduation["First Name"].fillna("").astype(str).str.strip().ne("")
            | graduation["Last Name"].fillna("").astype(str).str.strip().ne("")
        )
    ].copy()
    return graduation.reset_index(drop=True), pd.DataFrame(exceptions)


def load_roster_term_table(roots: Sequence[Path]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows: List[dict] = []
    exceptions: List[dict] = []
    schema_columns = load_schema()["tables"]["roster_term"]
    all_roster_paths = roster_files(roots)
    new_member_form_lookup = build_individual_new_member_form_lookup(all_roster_paths)

    def process_roster_table(path: Path, source_label: str, sheet_name: str, table_rows: List[Tuple[object, ...]]) -> None:
        header_row_idx, header_map = find_header_row_in_rows(table_rows)
        if header_row_idx is None:
            exceptions.append(
                {
                    "exception_type": "roster_header_missing",
                    "source_file": source_label,
                    "student_id": "",
                    "term_code": "",
                    "details": f"Sheet {sheet_name} skipped because no usable header row was found.",
                }
            )
            return

        status_row_idx, status_col_idx = find_status_column_in_rows(table_rows)
        if "status" not in header_map and status_col_idx is not None:
            header_map["status"] = status_col_idx
        data_start_index = max(header_row_idx, status_row_idx or header_row_idx)
        source_is_new_member = source_context_indicates_new_member(path, sheet_name)

        default_chapter = infer_chapter(path, sheet_name) or normalize_chapter_name(sheet_name or path.stem) or "Unknown"
        current_chapter_raw = sheet_name
        current_chapter = default_chapter
        if default_chapter and default_chapter != "Unknown":
            if normalize_chapter_name(sheet_name) == default_chapter and not is_placeholder_sheet_name(sheet_name):
                current_chapter_source = "inferred_from_sheet_name"
            elif chapter_from_filename(path) == default_chapter:
                current_chapter_source = "inferred_from_file_name"
            else:
                current_chapter_source = "inferred_from_sheet_name"
        else:
            current_chapter_source = "unresolved"

        term_label = ""
        for candidate in path_term_candidates(path):
            code, label, _, _ = parse_term_code(candidate)
            if code:
                term_label = label
                break
        term_code, term_label, term_year, term_season = parse_term_code(term_label)
        if not term_code:
            exceptions.append(
                {
                    "exception_type": "roster_term_unparsed",
                    "source_file": source_label,
                    "student_id": "",
                    "term_code": "",
                    "details": f"Could not infer term for {source_label}::{sheet_name}",
                }
            )

        version_context = " ".join([str(part) for part in path.parts])
        roster_file_version, roster_file_version_priority = roster_file_version_details(version_context)
        roster_file_month, roster_file_month_priority = roster_file_month_details(version_context)

        for row in table_rows[data_start_index:]:
            inline_chapter_raw = detect_inline_chapter_label(row, header_map)
            if inline_chapter_raw:
                current_chapter_raw = inline_chapter_raw
                current_chapter = normalize_chapter_name(inline_chapter_raw) or default_chapter
                current_chapter_source = "original" if current_chapter and current_chapter != "Unknown" else current_chapter_source
                continue

            last_name = clean_text(get_cell(row, header_map.get("last_name")))
            first_name = clean_text(get_cell(row, header_map.get("first_name")))
            if not last_name and not first_name:
                continue

            banner_raw = clean_text(get_cell(row, header_map.get("banner_id")))
            email = normalize_email(get_cell(row, header_map.get("email")))
            chapter_raw = clean_text(get_cell(row, header_map.get("chapter")))
            status_raw = clean_text(get_cell(row, header_map.get("status")))
            position_raw = clean_text(get_cell(row, header_map.get("position")))
            semester_joined_raw = clean_text(get_cell(row, header_map.get("semester_joined")))
            chapter, chapter_assignment_source, chapter_assignment_confidence, chapter_assignment_notes = chapter_assignment_details(
                path,
                sheet_name,
                chapter_raw,
                current_chapter_raw or sheet_name,
                current_chapter_source,
            )
            status_bucket = roster_status_bucket(status_raw, position_raw)
            form_key = (str(term_year).lower(), term_label.lower(), first_name.lower(), last_name.lower())
            has_form_evidence = bool(term_label and new_member_form_lookup.get(form_key))
            if should_upgrade_to_new_member_status(status_raw, position_raw, source_is_new_member, has_form_evidence):
                status_bucket = "New Member"
                semester_joined_raw = semester_joined_raw or term_label

            rows.append(
                {
                    "student_id": normalize_banner_id(banner_raw),
                    "student_id_raw": banner_raw,
                    "identity_resolution_basis": "source_banner_id" if banner_raw else "",
                    "identity_resolution_notes": "",
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "source_file": source_label,
                    "source_sheet": sheet_name,
                    "roster_file_version": roster_file_version,
                    "roster_file_version_priority": roster_file_version_priority,
                    "roster_file_month": roster_file_month,
                    "roster_file_month_priority": roster_file_month_priority,
                    "term_code": term_code,
                    "term_label": term_label,
                    "term_year": term_year,
                    "term_season": term_season,
                    "term_source_basis": "folder_or_filename",
                    "chapter": chapter,
                    "chapter_raw": chapter_raw or current_chapter_raw or sheet_name,
                    "chapter_assignment_source": chapter_assignment_source,
                    "chapter_assignment_confidence": chapter_assignment_confidence,
                    "chapter_assignment_notes": chapter_assignment_notes,
                    "org_status_raw": status_raw,
                    "org_status_bucket": status_bucket,
                    "org_position_raw": position_raw,
                    "semester_joined_raw": semester_joined_raw,
                    "new_member_flag": "Yes" if status_bucket == "New Member" else "No",
                    "org_entry_term_code": "",
                    "org_entry_term_basis": "",
                }
            )

    for path in all_roster_paths:
        source_label = source_label_for_roster_path(path, roots)
        if is_individual_new_member_form_pdf(path):
            continue
        if path.suffix.lower() == ".pdf":
            table_sources, pdf_issues = pdf_table_rows(path)
            for issue in pdf_issues:
                exceptions.append(
                    {
                        "exception_type": "roster_pdf_issue",
                        "source_file": source_label,
                        "student_id": "",
                        "term_code": "",
                        "details": issue,
                    }
                )
            for sheet_name, table_rows in table_sources:
                process_roster_table(path, source_label, sheet_name, table_rows)
            continue

        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            exceptions.append(
                {
                    "exception_type": "roster_open_error",
                    "source_file": source_label,
                    "student_id": "",
                    "term_code": "",
                    "details": str(exc),
                }
            )
            continue

        try:
            for ws in workbook.worksheets:
                table_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
                process_roster_table(path, source_label, ws.title, table_rows)
        finally:
            workbook.close()

    roster = pd.DataFrame(rows)
    if roster.empty:
        roster = pd.DataFrame(columns=schema_columns)
    return ensure_columns(roster, schema_columns), pd.DataFrame(exceptions)


def load_academic_term_table(root: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows: List[dict] = []
    exceptions: List[dict] = []
    schema_columns = load_schema()["tables"]["academic_term"]

    for path in academic_files(root):
        if path.suffix.lower() == ".csv":
            raw = pd.read_csv(path)
            raw.columns = [canonical_header(column) for column in raw.columns]
            inferred_term = ""
            for candidate in path_term_candidates(path):
                code, label, _, _ = parse_term_code(candidate)
                if code:
                    inferred_term = label
                    break
            term_code, term_label, term_year, term_season = parse_term_code(inferred_term or path.stem)
            if not term_code:
                exceptions.append(
                    {
                        "exception_type": "academic_term_unparsed",
                        "source_file": path.name,
                        "student_id": "",
                        "term_code": "",
                        "details": f"Could not infer term for {path.name}",
                    }
                )
            rename_map = {
                "student id": "Banner ID",
                "banner id": "Banner ID",
                "last name": "Last Name",
                "first name": "First Name",
                "email": "Email",
                "student status": "Student Status",
                "major": "Major",
                "semester hours": "Semester Hours",
                "credits attempted": "Semester Hours",
                "cumulative hours": "Cumulative Hours",
                "current academic standing": "Current Academic Standing",
                "academic standing": "Current Academic Standing",
                "texas state gpa": "Texas State GPA",
                "overall gpa": "Overall GPA",
                "transfer gpa": "Transfer GPA",
                "term gpa": "Term GPA",
                "gpa": "Term GPA",
                "term passed hours": "Term Passed Hours",
                "credits earned": "Term Passed Hours",
                "txstate cumulative gpa": "TxState Cumulative GPA",
                "overall cumulative gpa": "Overall Cumulative GPA",
                "graduation term": "Graduation Term",
            }
            frame = raw.rename(columns={column: rename_map.get(column, column) for column in raw.columns})
            for record in frame.to_dict("records"):
                rows.append(
                    {
                        "student_id": normalize_banner_id(record.get("Banner ID", "")),
                        "student_id_raw": clean_text(record.get("Banner ID", "")),
                        "identity_resolution_basis": "source_student_id" if clean_text(record.get("Banner ID", "")) else "",
                        "identity_resolution_notes": "",
                        "first_name": clean_text(record.get("First Name", "")),
                        "last_name": clean_text(record.get("Last Name", "")),
                        "email": normalize_email(record.get("Email", "")),
                        "source_file": path.name,
                        "source_sheet": "csv",
                        "term_code": term_code,
                        "term_label": term_label,
                        "term_year": term_year,
                        "term_season": term_season,
                        "term_source_basis": "filename",
                        "academic_status_raw": clean_text(record.get("Student Status", "")),
                        "major": clean_text(record.get("Major", "")),
                        "term_gpa": record.get("Term GPA", ""),
                        "institutional_cumulative_gpa": record.get("TxState Cumulative GPA", record.get("Texas State GPA", "")),
                        "overall_cumulative_gpa": record.get("Overall Cumulative GPA", record.get("Overall GPA", "")),
                        "transfer_gpa": record.get("Transfer GPA", ""),
                        "attempted_hours_term": record.get("Semester Hours", ""),
                        "earned_hours_term": record.get("Term Passed Hours", ""),
                        "institutional_cumulative_hours": record.get("Cumulative Hours", ""),
                        "total_cumulative_hours": record.get("Cumulative Hours", ""),
                        "academic_standing_raw": clean_text(record.get("Current Academic Standing", "")),
                        "academic_standing_bucket": standing_bucket(record.get("Current Academic Standing", "")),
                        "graduation_term_code": parse_term_code(record.get("Graduation Term", ""))[0],
                        "graduation_term_label": parse_term_code(record.get("Graduation Term", ""))[1],
                    }
                )
            continue

        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in workbook.worksheets:
                term_label = parse_grade_term(path, ws.title)
                if not term_label:
                    continue
                term_code, term_label, term_year, term_season = parse_term_code(term_label)
                sheet_rows = list(ws.iter_rows(values_only=True))
                if not sheet_rows:
                    continue
                header_row_idx = None
                header_map: Dict[str, int] = {}
                best_score = 0
                for idx, candidate_row in enumerate(sheet_rows[:25]):
                    candidate_map = map_grade_headers(candidate_row)
                    score = len(candidate_map)
                    if {"Last Name", "First Name"}.issubset(set(candidate_map)):
                        score += 2
                    if "Banner ID" in candidate_map or "Email" in candidate_map:
                        score += 1
                    if score > best_score:
                        best_score = score
                        header_row_idx = idx
                        header_map = candidate_map
                required = {"Last Name", "First Name"}
                if header_row_idx is None or not required.issubset(set(header_map)):
                    continue
                for row in sheet_rows[header_row_idx + 1:]:
                    first_name = get_cell(row, header_map.get("First Name"))
                    last_name = get_cell(row, header_map.get("Last Name"))
                    if not first_name and not last_name:
                        continue
                    banner_raw = get_cell(row, header_map.get("Banner ID"))
                    rows.append(
                        {
                            "student_id": normalize_banner_id(banner_raw),
                            "student_id_raw": clean_text(banner_raw),
                            "identity_resolution_basis": "source_student_id" if clean_text(banner_raw) else "",
                            "identity_resolution_notes": "",
                            "first_name": clean_text(first_name),
                            "last_name": clean_text(last_name),
                            "email": normalize_email(get_cell(row, header_map.get("Email"))),
                            "source_file": path.name,
                            "source_sheet": ws.title,
                            "term_code": term_code,
                            "term_label": term_label,
                            "term_year": term_year,
                            "term_season": term_season,
                            "term_source_basis": "filename_or_sheet",
                            "academic_status_raw": clean_text(get_cell(row, header_map.get("Student Status"))),
                            "major": clean_text(get_cell(row, header_map.get("Major"))),
                            "term_gpa": get_cell(row, header_map.get("Term GPA")),
                            "institutional_cumulative_gpa": get_cell(row, header_map.get("TxState Cumulative GPA")) or get_cell(row, header_map.get("Texas State GPA")),
                            "overall_cumulative_gpa": get_cell(row, header_map.get("Overall Cumulative GPA")) or get_cell(row, header_map.get("Overall GPA")),
                            "transfer_gpa": get_cell(row, header_map.get("Transfer GPA")),
                            "attempted_hours_term": get_cell(row, header_map.get("Semester Hours")),
                            "earned_hours_term": get_cell(row, header_map.get("Term Passed Hours")),
                            "institutional_cumulative_hours": get_cell(row, header_map.get("Cumulative Hours")),
                            "total_cumulative_hours": get_cell(row, header_map.get("Cumulative Hours")),
                            "academic_standing_raw": clean_text(get_cell(row, header_map.get("Current Academic Standing"))),
                            "academic_standing_bucket": standing_bucket(get_cell(row, header_map.get("Current Academic Standing"))),
                            "graduation_term_code": "",
                            "graduation_term_label": "",
                        }
                    )
        finally:
            workbook.close()

    academic = pd.DataFrame(rows)
    if academic.empty:
        academic = pd.DataFrame(columns=schema_columns)
    else:
        for column in [
            "term_gpa",
            "institutional_cumulative_gpa",
            "overall_cumulative_gpa",
            "transfer_gpa",
            "attempted_hours_term",
            "earned_hours_term",
            "institutional_cumulative_hours",
            "total_cumulative_hours",
        ]:
            academic[column] = coerce_numeric(academic[column])
    return ensure_columns(academic, schema_columns), pd.DataFrame(exceptions)


def build_identity_maps(
    roster: pd.DataFrame,
    academic: pd.DataFrame,
    snapshot: pd.DataFrame,
    graduation: pd.DataFrame,
) -> Tuple[Dict[str, str], Dict[Tuple[str, str], str], pd.DataFrame]:
    email_candidates: Dict[str, set[str]] = defaultdict(set)
    name_candidates: Dict[Tuple[str, str], set[str]] = defaultdict(set)
    exceptions: List[dict] = []

    source_frames = [
        roster[["student_id", "email", "first_name", "last_name"]].copy(),
        academic[["student_id", "email", "first_name", "last_name"]].copy(),
    ]
    if not snapshot.empty:
        snap = pd.DataFrame(
            {
                "student_id": snapshot["Student ID"].map(normalize_banner_id),
                "email": pd.Series("", index=snapshot.index, dtype="object"),
                "first_name": snapshot["First Name"].map(clean_text),
                "last_name": snapshot["Last Name"].map(clean_text),
            }
        )
        source_frames.append(snap)
    if not graduation.empty:
        grad = pd.DataFrame(
            {
                "student_id": graduation["Student ID"].map(normalize_banner_id),
                "email": pd.Series("", index=graduation.index, dtype="object"),
                "first_name": graduation["First Name"].map(clean_text),
                "last_name": graduation["Last Name"].map(clean_text),
            }
        )
        source_frames.append(grad)

    combined = pd.concat(source_frames, ignore_index=True)
    combined = combined.loc[combined["student_id"].fillna("").astype(str).str.strip().ne("")]
    for row in combined.itertuples(index=False):
        if clean_text(row.email):
            email_candidates[clean_text(row.email).lower()].add(clean_text(row.student_id))
        if clean_text(row.first_name) or clean_text(row.last_name):
            name_candidates[person_name_key(row.first_name, row.last_name)].add(clean_text(row.student_id))

    email_map: Dict[str, str] = {}
    name_map: Dict[Tuple[str, str], str] = {}
    for email, ids in email_candidates.items():
        if len(ids) == 1:
            email_map[email] = next(iter(ids))
        else:
            exceptions.append({"exception_type": "ambiguous_email_match", "source_file": "", "student_id": "", "term_code": "", "details": f"Email {email} matched multiple student IDs: {', '.join(sorted(ids))}"})
    for key, ids in name_candidates.items():
        if len(ids) == 1:
            name_map[key] = next(iter(ids))
        else:
            exceptions.append({"exception_type": "ambiguous_name_match", "source_file": "", "student_id": "", "term_code": "", "details": f"Name {key[0]} {key[1]} matched multiple student IDs: {', '.join(sorted(ids))}"})

    return email_map, name_map, pd.DataFrame(exceptions)


def resolve_missing_ids(frame: pd.DataFrame, email_map: Dict[str, str], name_map: Dict[Tuple[str, str], str], source_label: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    result = ensure_text_columns(
        frame,
        [
            "student_id",
            "identity_resolution_basis",
            "identity_resolution_notes",
        ],
    )
    if result.empty:
        return result, pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    missing_mask = result["student_id"].fillna("").astype(str).str.strip().eq("")
    if not missing_mask.any():
        return result, pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    missing = result.loc[missing_mask].copy()
    missing["_email_key"] = missing["email"].fillna("").astype(str).str.strip().str.lower()
    missing["_name_first"] = missing["first_name"].fillna("").astype(str).str.strip().str.lower()
    missing["_name_last"] = missing["last_name"].fillna("").astype(str).str.strip().str.lower()
    missing["_name_key"] = list(zip(missing["_name_first"], missing["_name_last"]))
    missing["_matched_id"] = missing["_email_key"].map(email_map)
    name_match_mask = missing["_matched_id"].isna() & (missing["_name_first"].ne("") | missing["_name_last"].ne(""))
    if name_match_mask.any():
        missing.loc[name_match_mask, "_matched_id"] = missing.loc[name_match_mask, "_name_key"].map(name_map)

    matched_mask = missing["_matched_id"].fillna("").astype(str).str.strip().ne("")
    if matched_mask.any():
        matched = missing.loc[matched_mask].copy()
        email_match_mask = matched["_email_key"].map(email_map).fillna("").astype(str).str.strip().ne("")
        matched["identity_resolution_basis"] = email_match_mask.map(lambda flag: "unique_email_match" if flag else "unique_name_match")
        matched["identity_resolution_notes"] = matched["identity_resolution_basis"].map(
            {
                "unique_email_match": "Resolved from unique email match.",
                "unique_name_match": "Resolved from unique exact name match.",
            }
        )
        result.loc[matched.index, "student_id"] = matched["_matched_id"].astype(str)
        result.loc[matched.index, "identity_resolution_basis"] = matched["identity_resolution_basis"]
        result.loc[matched.index, "identity_resolution_notes"] = matched["identity_resolution_notes"]

    unresolved = missing.loc[~matched_mask].copy()
    if unresolved.empty:
        return result, pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    exceptions = pd.DataFrame(
        {
            "exception_type": f"{source_label}_missing_student_id",
            "source_file": unresolved["source_file"].fillna("").astype(str).str.strip(),
            "student_id": "",
            "term_code": unresolved["term_code"].fillna("").astype(str).str.strip(),
            "details": (
                unresolved["first_name"].fillna("").astype(str).str.strip()
                + " "
                + unresolved["last_name"].fillna("").astype(str).str.strip()
            ).str.strip(),
        }
    )
    exceptions["details"] = exceptions["details"].where(
        exceptions["details"].str.strip().ne(""),
        unresolved["_email_key"].where(unresolved["_email_key"].ne(""), "Unidentified row"),
    )
    return result, exceptions.reset_index(drop=True)


def resolve_missing_roster_chapters(roster: pd.DataFrame, settings: Dict[str, object]) -> pd.DataFrame:
    if roster.empty:
        return roster

    result = ensure_text_columns(
        roster,
        [
            "chapter",
            "chapter_assignment_source",
            "chapter_assignment_confidence",
            "chapter_assignment_notes",
        ],
    )
    secondary_orgs = secondary_organization_set(settings)
    known = result.loc[~result["chapter"].map(chapter_is_missing)].copy()
    if known.empty:
        known = pd.DataFrame(columns=result.columns)
    else:
        known = known.loc[
            ~known["chapter"].fillna("").astype(str).map(lambda value: normalize_chapter_name(value) in secondary_orgs)
        ].copy()
    known["_id_key"] = known["student_id"].fillna("").astype(str).str.strip()
    known["_name_first"] = known["first_name"].fillna("").astype(str).str.strip().str.lower()
    known["_name_last"] = known["last_name"].fillna("").astype(str).str.strip().str.lower()
    known["_id_name_key"] = (
        known["_id_key"]
        + "||"
        + known["_name_first"]
        + "||"
        + known["_name_last"]
    ).where(
        known["_id_key"].ne("") & (known["_name_first"].ne("") | known["_name_last"].ne("")),
        "",
    )

    id_name_lookup = (
        known.loc[known["_id_name_key"].ne(""), ["_id_name_key", "chapter"]]
        .drop_duplicates()
        .groupby("_id_name_key", dropna=False)["chapter"]
        .agg(list)
    )
    id_name_lookup = id_name_lookup.loc[id_name_lookup.map(len).eq(1)].map(lambda values: values[0]).to_dict()

    id_lookup = (
        known.loc[known["_id_key"].ne(""), ["_id_key", "chapter"]]
        .drop_duplicates()
        .groupby("_id_key", dropna=False)["chapter"]
        .agg(list)
    )
    id_lookup = id_lookup.loc[id_lookup.map(len).eq(1)].map(lambda values: values[0]).to_dict()

    eligible_mask = result["chapter"].map(chapter_is_missing) | result["chapter_assignment_source"].fillna("").astype(str).isin(
        ["inferred_from_sheet_name", "inferred_from_file_name", "unresolved"]
    )
    if not eligible_mask.any():
        return result

    eligible = result.loc[eligible_mask].copy()
    eligible["_id_key"] = eligible["student_id"].fillna("").astype(str).str.strip()
    eligible["_name_first"] = eligible["first_name"].fillna("").astype(str).str.strip().str.lower()
    eligible["_name_last"] = eligible["last_name"].fillna("").astype(str).str.strip().str.lower()
    eligible["_id_name_key"] = (
        eligible["_id_key"]
        + "||"
        + eligible["_name_first"]
        + "||"
        + eligible["_name_last"]
    ).where(
        eligible["_id_key"].ne("") & (eligible["_name_first"].ne("") | eligible["_name_last"].ne("")),
        "",
    )
    eligible["_matched_chapter"] = eligible["_id_name_key"].map(id_name_lookup)
    eligible["_chapter_source"] = eligible["_matched_chapter"].notna().map(lambda flag: "matched_by_id_name" if flag else "")
    eligible["_chapter_confidence"] = eligible["_matched_chapter"].notna().map(lambda flag: "high" if flag else "")
    eligible["_chapter_notes"] = eligible["_matched_chapter"].notna().map(
        lambda flag: "Backfilled from another roster row with matching student ID and exact name." if flag else ""
    )

    id_only_mask = eligible["_matched_chapter"].isna() & eligible["_id_key"].ne("")
    if id_only_mask.any():
        eligible.loc[id_only_mask, "_matched_chapter"] = eligible.loc[id_only_mask, "_id_key"].map(id_lookup)
        new_id_mask = id_only_mask & eligible["_matched_chapter"].notna()
        eligible.loc[new_id_mask, "_chapter_source"] = "matched_by_id"
        eligible.loc[new_id_mask, "_chapter_confidence"] = "medium"
        eligible.loc[new_id_mask, "_chapter_notes"] = "Backfilled from another roster row with matching student ID."

    matched_mask = eligible["_matched_chapter"].fillna("").astype(str).str.strip().ne("")
    if matched_mask.any():
        matched = eligible.loc[matched_mask]
        result.loc[matched.index, "chapter"] = matched["_matched_chapter"].astype(str)
        result.loc[matched.index, "chapter_assignment_source"] = matched["_chapter_source"].astype(str)
        result.loc[matched.index, "chapter_assignment_confidence"] = matched["_chapter_confidence"].astype(str)
        result.loc[matched.index, "chapter_assignment_notes"] = matched["_chapter_notes"].astype(str)

    unresolved = eligible.loc[~matched_mask].copy()
    for idx, row in unresolved.iterrows():
        existing_source = clean_text(row.get("chapter_assignment_source", "")) or "unresolved"
        inferred, inferred_source, inferred_confidence, inferred_notes = chapter_assignment_details(
            Path(clean_text(row.get("source_file", "")) or "unknown.xlsx"),
            clean_text(row.get("source_sheet", "")),
            row.get("chapter_raw", ""),
            row.get("source_sheet", ""),
            existing_source,
        )
        result.at[idx, "chapter"] = inferred
        result.at[idx, "chapter_assignment_source"] = inferred_source or "unresolved"
        result.at[idx, "chapter_assignment_confidence"] = inferred_confidence or "low"
        result.at[idx, "chapter_assignment_notes"] = inferred_notes or "Chapter assignment remained unresolved."

    return result


def ensure_manual_chapter_assignment_template(path: Path = MANUAL_CHAPTER_ASSIGNMENTS_PATH) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        columns=[
            "student_id",
            "first_name",
            "last_name",
            "chapter_override",
            "notes",
        ]
    ).to_csv(path, index=False)


def apply_manual_chapter_assignments(roster: pd.DataFrame, overrides: pd.DataFrame) -> pd.DataFrame:
    if roster.empty or overrides.empty:
        return roster

    result = ensure_text_columns(
        roster,
        [
            "chapter",
            "chapter_assignment_source",
            "chapter_assignment_confidence",
            "chapter_assignment_notes",
        ],
    )
    override_by_id: Dict[str, dict] = {}
    override_by_name: Dict[Tuple[str, str], dict] = {}

    for row in overrides.itertuples(index=False):
        override = {
            "chapter": normalize_chapter_name(getattr(row, "chapter_override", "")),
            "notes": clean_text(getattr(row, "notes", "")),
        }
        student_id = normalize_banner_id(getattr(row, "student_id", ""))
        first_name = clean_text(getattr(row, "first_name", ""))
        last_name = clean_text(getattr(row, "last_name", ""))
        if student_id:
            override_by_id[student_id] = override
        elif first_name or last_name:
            override_by_name[person_name_key(first_name, last_name)] = override

    result["_manual_id_key"] = result["student_id"].map(normalize_banner_id)
    result["_manual_name_key"] = list(
        zip(
            result["first_name"].fillna("").astype(str).str.strip().str.lower(),
            result["last_name"].fillna("").astype(str).str.strip().str.lower(),
        )
    )
    result["_manual_override"] = result["_manual_id_key"].map(override_by_id)
    name_override_mask = result["_manual_override"].isna()
    if name_override_mask.any():
        result.loc[name_override_mask, "_manual_override"] = result.loc[name_override_mask, "_manual_name_key"].map(override_by_name)

    override_mask = result["_manual_override"].notna()
    if override_mask.any():
        override_values = result.loc[override_mask, "_manual_override"]
        result.loc[override_mask, "chapter"] = override_values.map(lambda value: value.get("chapter", ""))
        result.loc[override_mask, "chapter_assignment_source"] = "manual_override"
        result.loc[override_mask, "chapter_assignment_confidence"] = "manual"
        result.loc[override_mask, "chapter_assignment_notes"] = override_values.map(
            lambda value: value.get("notes", "") or "Applied from config/manual_chapter_assignments.csv."
        )

    return result.drop(columns=["_manual_id_key", "_manual_name_key", "_manual_override"], errors="ignore")


def build_unresolved_chapter_review(
    roster: pd.DataFrame,
    academic: pd.DataFrame,
    summary: pd.DataFrame,
) -> pd.DataFrame:
    if roster.empty:
        return pd.DataFrame(
            columns=[
                "review_key",
                "student_id",
                "first_name",
                "last_name",
                "student_name",
                "chapter_assignment_source",
                "chapter_assignment_confidence",
                "chapter_assignment_notes",
                "candidate_chapters_seen",
                "terms_seen",
                "roster_files_seen",
                "roster_sheets_seen",
                "academic_files_seen",
                "academic_sheets_seen",
                "manual_override_path",
            ]
        )

    roster_lookup = roster.copy()
    roster_lookup["_review_key"] = build_review_key(roster_lookup)
    unresolved = roster_lookup.loc[
        roster_lookup["chapter_assignment_source"].fillna("").astype(str).isin(["unresolved", "inferred_from_file_name", "inferred_from_sheet_name"])
    ].copy()
    if unresolved.empty:
        return pd.DataFrame(
            columns=[
                "review_key",
                "student_id",
                "first_name",
                "last_name",
                "student_name",
                "chapter_assignment_source",
                "chapter_assignment_confidence",
                "chapter_assignment_notes",
                "candidate_chapters_seen",
                "terms_seen",
                "roster_files_seen",
                "roster_sheets_seen",
                "academic_files_seen",
                "academic_sheets_seen",
                "manual_override_path",
            ]
        )

    academic_lookup = academic.copy()
    if not academic_lookup.empty:
        academic_lookup["_review_key"] = build_review_key(academic_lookup)
    summary_lookup = summary.copy() if not summary.empty else pd.DataFrame()
    if not summary_lookup.empty:
        summary_lookup["_review_key"] = summary_lookup["student_id"].fillna("").astype(str).str.strip()
        summary_lookup["_review_key"] = summary_lookup["_review_key"].where(
            summary_lookup["_review_key"].ne(""),
            "name::" + summary_lookup["student_name"].fillna("").astype(str).str.strip().str.lower(),
        )

    roster_key_chapters = (
        roster_lookup.loc[roster_lookup["_review_key"].ne("") & ~roster_lookup["chapter"].map(chapter_is_missing), ["_review_key", "chapter"]]
        .drop_duplicates()
        .groupby("_review_key", dropna=False)["chapter"]
        .agg(lambda values: " | ".join(sorted({clean_text(value) for value in values if clean_text(value)})))
        .to_dict()
    )
    academic_files_by_key = (
        academic_lookup.loc[academic_lookup["_review_key"].ne(""), ["_review_key", "source_file"]]
        .drop_duplicates()
        .groupby("_review_key", dropna=False)["source_file"]
        .agg(lambda values: " | ".join(sorted({clean_text(value) for value in values if clean_text(value)})))
        .to_dict()
        if not academic_lookup.empty
        else {}
    )
    academic_sheets_by_key = (
        academic_lookup.loc[academic_lookup["_review_key"].ne(""), ["_review_key", "source_sheet"]]
        .drop_duplicates()
        .groupby("_review_key", dropna=False)["source_sheet"]
        .agg(lambda values: " | ".join(sorted({clean_text(value) for value in values if clean_text(value)})))
        .to_dict()
        if not academic_lookup.empty
        else {}
    )
    summary_by_key = (
        summary_lookup.drop_duplicates(subset=["_review_key"], keep="first").set_index("_review_key")
        if not summary_lookup.empty
        else pd.DataFrame()
    )

    rows: List[dict] = []
    for review_key, group in unresolved.groupby("_review_key", dropna=False):
        roster_files_seen = " | ".join(sorted({clean_text(value) for value in group["source_file"] if clean_text(value)}))
        roster_sheets_seen = " | ".join(sorted({clean_text(value) for value in group["source_sheet"] if clean_text(value)}))
        terms_seen = " | ".join(sorted({clean_text(value) for value in group["term_label"] if clean_text(value)}))
        candidate_chapters_seen = roster_key_chapters.get(review_key, "")
        academic_files_seen = academic_files_by_key.get(review_key, "")
        academic_sheets_seen = academic_sheets_by_key.get(review_key, "")

        first_row = group.iloc[0]
        summary_match = summary_by_key.loc[review_key] if not summary_by_key.empty and review_key in summary_by_key.index else None
        rows.append(
            {
                "review_key": review_key,
                "student_id": clean_text(first_row.get("student_id", "")),
                "first_name": clean_text(first_row.get("first_name", "")),
                "last_name": clean_text(first_row.get("last_name", "")),
                "student_name": clean_text(summary_match.get("student_name", "")) if summary_match is not None else f"{clean_text(first_row.get('first_name', ''))} {clean_text(first_row.get('last_name', ''))}".strip(),
                "chapter_assignment_source": clean_text(first_row.get("chapter_assignment_source", "")),
                "chapter_assignment_confidence": clean_text(first_row.get("chapter_assignment_confidence", "")),
                "chapter_assignment_notes": clean_text(first_row.get("chapter_assignment_notes", "")),
                "candidate_chapters_seen": candidate_chapters_seen,
                "terms_seen": terms_seen,
                "roster_files_seen": roster_files_seen,
                "roster_sheets_seen": roster_sheets_seen,
                "academic_files_seen": academic_files_seen,
                "academic_sheets_seen": academic_sheets_seen,
                "manual_override_path": str(MANUAL_CHAPTER_ASSIGNMENTS_PATH),
            }
        )

    return pd.DataFrame(rows).sort_values(["student_id", "last_name", "first_name"], na_position="last").reset_index(drop=True)


def build_reference_derivatives(reference_inventory: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    membership_reference = build_reference_subset(
        reference_inventory,
        "membership_count",
        "chapter",
        "membership_count_reference",
        ["chapter", "chapter_raw", "term_code", "term_label", "membership_count_reference", "source_file", "source_sheet"],
        dedupe_subset=["chapter", "term_code"],
    )
    gpa_reference = build_reference_subset(
        reference_inventory,
        "average_gpa",
        "chapter",
        "chapter_average_gpa_reference",
        ["chapter", "chapter_raw", "term_code", "term_label", "chapter_average_gpa_reference", "source_file", "source_sheet"],
        dedupe_subset=["chapter", "term_code"],
    )
    gpa_benchmark_reference = build_reference_subset(
        reference_inventory,
        "average_gpa",
        "benchmark",
        "benchmark_average_gpa_reference",
        ["benchmark_label", "term_code", "term_label", "benchmark_average_gpa_reference", "source_file", "source_sheet"],
        dedupe_subset=["benchmark_label", "term_code"],
    )
    new_member_reference = build_reference_subset(
        reference_inventory,
        "new_member_count",
        "chapter",
        "new_member_count_reference",
        ["chapter", "chapter_raw", "term_code", "term_label", "new_member_count_reference", "source_file", "source_sheet"],
        dedupe_subset=["chapter", "term_code"],
    )
    retention_reference = ensure_columns(
        reference_inventory.loc[reference_inventory["reference_type"].eq("retention_rate")].rename(
            columns={
                "entity_label_raw": "entity_label_raw",
                "entity_label_normalized": "entity_label_normalized",
                "reference_value": "retention_rate_reference",
            }
        ),
        [
            "entity_type",
            "entity_label_raw",
            "entity_label_normalized",
            "term_code",
            "term_label",
            "retention_rate_reference",
            "source_file",
            "source_sheet",
        ],
    ).drop_duplicates(
        subset=["entity_type", "entity_label_normalized", "term_code", "source_file", "source_sheet"],
        keep="first",
    ).reset_index(drop=True)
    reference_unclassified_rows = ensure_columns(
        reference_inventory.loc[reference_inventory["reference_type"].eq("unknown")].copy(),
        [
            "reference_type",
            "entity_type",
            "entity_label_raw",
            "entity_label_normalized",
            "term_code",
            "term_label",
            "reference_value",
            "classification_basis",
            "source_file",
            "source_sheet",
        ],
    )
    return (
        membership_reference,
        gpa_reference,
        gpa_benchmark_reference,
        new_member_reference,
        retention_reference,
        reference_unclassified_rows,
    )


def prepare_canonical_sources(
    roster_term: pd.DataFrame,
    academic_term: pd.DataFrame,
    snapshot: pd.DataFrame,
    graduation: pd.DataFrame,
    settings: Dict[str, object],
    manual_chapter_assignments: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    empty_exception_frame = pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    email_map, name_map, identity_map_issues = build_identity_maps(roster_term, academic_term, snapshot, graduation)
    roster_term, roster_id_issues = resolve_missing_ids(roster_term, email_map, name_map, "roster")
    academic_term, academic_id_issues = resolve_missing_ids(academic_term, email_map, name_map, "academic")
    roster_term = resolve_missing_roster_chapters(roster_term, settings)
    roster_term = apply_manual_chapter_assignments(roster_term, manual_chapter_assignments)

    roster_term, roster_dup_issues = dedupe_table(roster_term, ["student_id", "term_code", "chapter"], "roster")
    academic_term, academic_dup_issues = dedupe_table(academic_term, ["student_id", "term_code"], "academic")
    roster_term, roster_conflicts = resolve_roster_conflicts(roster_term, settings)
    roster_term = attach_org_entry_terms(roster_term, settings)

    identity_exceptions = pd.concat(
        [frame for frame in [identity_map_issues, roster_id_issues, academic_id_issues] if not frame.empty],
        ignore_index=True,
    ) if any(not frame.empty for frame in [identity_map_issues, roster_id_issues, academic_id_issues]) else empty_exception_frame
    term_exceptions = pd.concat(
        [frame for frame in [roster_dup_issues, academic_dup_issues] if not frame.empty],
        ignore_index=True,
    ) if any(not frame.empty for frame in [roster_dup_issues, academic_dup_issues]) else empty_exception_frame
    status_exceptions = build_status_exceptions(roster_term, academic_term)
    return roster_term, academic_term, identity_exceptions, term_exceptions, status_exceptions, roster_conflicts


SUMMARY_TO_MASTER_COLUMNS = [
    "latest_outcome_bucket",
    "exit_reason_code",
    "graduation_term_code",
    "resolved_outcome_flag",
    "outcome_evidence_source",
    "graduation_evidence_confirmed",
    "graduation_status_without_evidence",
    "graduation_status_corrected_flag",
    "graduation_status_correction_reason",
    "school_entry_term_code",
    "school_entry_term_basis",
    "outcome_resolution_group",
    "is_resolved_outcome",
    "is_active_outcome",
    "is_unknown_outcome",
    "is_graduated",
    "is_known_non_graduate_exit",
]


def enrich_master_longitudinal_with_summary(master_longitudinal: pd.DataFrame, student_summary: pd.DataFrame) -> pd.DataFrame:
    if master_longitudinal.empty or student_summary.empty:
        return master_longitudinal

    available_columns = [column for column in SUMMARY_TO_MASTER_COLUMNS if column in student_summary.columns]
    if not available_columns:
        return master_longitudinal

    summary_subset = (
        student_summary[["student_id", *available_columns]]
        .drop_duplicates(subset=["student_id"], keep="first")
        .reset_index(drop=True)
    )
    result = master_longitudinal.drop(columns=available_columns, errors="ignore").merge(summary_subset, on="student_id", how="left")
    return result


def build_canonical_core_tables(
    roster_term: pd.DataFrame,
    academic_term: pd.DataFrame,
    snapshot: pd.DataFrame,
    graduation: pd.DataFrame,
    settings: Dict[str, object],
    chapter_mapping: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    master_longitudinal = build_master_longitudinal(roster_term, academic_term, settings)
    student_summary, summary_qa, outcome_issues = build_student_summary(master_longitudinal, snapshot, graduation, settings, chapter_mapping)
    if not student_summary.empty:
        student_summary = student_summary.loc[:, ~student_summary.columns.duplicated()].copy()
    unresolved_chapter_review = build_unresolved_chapter_review(roster_term, academic_term, student_summary)
    master_longitudinal = enrich_master_longitudinal_with_summary(master_longitudinal, student_summary)
    return master_longitudinal, student_summary, summary_qa, outcome_issues, unresolved_chapter_review


def dedupe_table(frame: pd.DataFrame, unique_keys: Sequence[str], source_label: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if frame.empty:
        return frame, pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    ranked = frame.copy()
    if "student_id" in unique_keys:
        ranked["_identity_key"] = build_resolution_identity_key(ranked)
        effective_keys = ["_identity_key" if key == "student_id" else key for key in unique_keys]
    else:
        effective_keys = list(unique_keys)
    ranked["_source_version_priority"] = (
        coerce_numeric(ranked["roster_file_version_priority"]).fillna(1)
        if source_label == "roster" and "roster_file_version_priority" in ranked.columns
        else 0
    )
    ranked["_source_month_priority"] = (
        coerce_numeric(ranked["roster_file_month_priority"]).fillna(0)
        if source_label == "roster" and "roster_file_month_priority" in ranked.columns
        else 0
    )
    ranked["_source_format_priority"] = (
        ranked.get("source_file", pd.Series([""] * len(ranked), index=ranked.index)).map(source_file_format_priority)
        if source_label == "roster"
        else 0
    )
    ranked["_completeness"] = ranked.notna().sum(axis=1)
    ranked["_update_key"] = ranked["source_file"].map(update_key_from_name) if "source_file" in ranked.columns else [(0, 0, 0)] * len(ranked)
    ranked = ranked.sort_values(by=effective_keys + ["_source_version_priority", "_source_month_priority", "_source_format_priority", "_completeness", "_update_key"], ascending=[True] * len(effective_keys) + [False, False, False, False, False])
    duplicate_mask = ranked.duplicated(subset=effective_keys, keep="first")
    exceptions = ranked.loc[duplicate_mask].copy()
    deduped = ranked.drop_duplicates(subset=effective_keys, keep="first").drop(columns=["_source_version_priority", "_source_month_priority", "_source_format_priority", "_completeness", "_update_key"] + (["_identity_key"] if "_identity_key" in ranked.columns else []))
    if exceptions.empty:
        return deduped.reset_index(drop=True), pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    exception_rows = []
    for row in exceptions.itertuples(index=False):
        exception_rows.append(
            {
                "exception_type": f"{source_label}_duplicate_removed",
                "source_file": clean_text(getattr(row, "source_file", "")),
                "student_id": clean_text(getattr(row, "student_id", "")),
                "term_code": clean_text(getattr(row, "term_code", "")),
                "details": f"Duplicate row removed for keys {', '.join(str(getattr(row, key, '')) for key in unique_keys)}",
            }
        )
    return deduped.reset_index(drop=True), pd.DataFrame(exception_rows)


def resolve_roster_conflicts(roster: pd.DataFrame, settings: Dict[str, object]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if roster.empty:
        return roster, pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    exceptions: List[dict] = []
    resolved_rows: List[pd.Series] = []
    secondary_orgs = secondary_organization_set(settings)
    roster = roster.copy()
    roster["_identity_key"] = build_resolution_identity_key(roster)
    for (_, _), group in roster.groupby(["_identity_key", "term_code"], dropna=False):
        non_resigned_revoked = group.loc[~group["org_status_bucket"].fillna("").astype(str).isin(["Resigned", "Revoked"])].copy()
        effective_group = non_resigned_revoked if not non_resigned_revoked.empty else group
        chapter_values = sorted(
            {
                clean_text(value)
                for value in effective_group["chapter"]
                if clean_text(value) and normalize_chapter_name(clean_text(value)) not in secondary_orgs
            }
        )
        status_values = sorted({clean_text(value) for value in effective_group["org_status_bucket"] if clean_text(value)})
        if len(chapter_values) > 1:
            exceptions.append(
                {
                    "exception_type": "chapter_conflict_same_term",
                    "source_file": " | ".join(sorted({clean_text(value) for value in group["source_file"] if clean_text(value)})),
                    "student_id": clean_text(group["student_id"].iloc[0]),
                    "term_code": clean_text(group["term_code"].iloc[0]),
                    "details": "Multiple chapters in same term: " + ", ".join(chapter_values),
                }
            )
        if len(status_values) > 1:
            exceptions.append(
                {
                    "exception_type": "roster_status_conflict_same_term",
                    "source_file": " | ".join(sorted({clean_text(value) for value in group["source_file"] if clean_text(value)})),
                    "student_id": clean_text(group["student_id"].iloc[0]),
                    "term_code": clean_text(group["term_code"].iloc[0]),
                    "details": "Multiple roster statuses in same term: " + ", ".join(status_values),
                }
            )
        ranked = group.copy()
        ranked["_resigned_or_revoked"] = ranked["org_status_bucket"].fillna("").astype(str).isin(["Resigned", "Revoked"]).astype(int)
        ranked["_secondary_org"] = ranked["chapter"].map(lambda value: normalize_chapter_name(clean_text(value)) in secondary_orgs).astype(int)
        ranked["_source_version_priority"] = coerce_numeric(ranked.get("roster_file_version_priority", pd.Series([1] * len(ranked), index=ranked.index))).fillna(1)
        ranked["_source_month_priority"] = coerce_numeric(ranked.get("roster_file_month_priority", pd.Series([0] * len(ranked), index=ranked.index))).fillna(0)
        ranked["_source_format_priority"] = ranked.get("source_file", pd.Series([""] * len(ranked), index=ranked.index)).map(source_file_format_priority)
        ranked["_new_member"] = ranked["new_member_flag"].eq("Yes").astype(int)
        ranked["_known_id"] = ranked["student_id"].fillna("").astype(str).str.strip().ne("").astype(int)
        ranked["_status_priority"] = ranked["org_status_bucket"].map(
            {
                "Graduated": 90,
                "Suspended": 80,
                "Transfer": 70,
                "Revoked": 65,
                "Resigned": 60,
                "Inactive": 55,
                "New Member": 54,
                "Active": 50,
            }
        ).fillna(0)
        ranked = ranked.sort_values(by=["_resigned_or_revoked", "_secondary_org", "_source_version_priority", "_source_month_priority", "_source_format_priority", "_status_priority", "_new_member", "_known_id"], ascending=[True, True, False, False, False, False, False, False])
        chosen_row = ranked.iloc[0].copy()
        discarded_resigned_revoked = group.loc[group["org_status_bucket"].fillna("").astype(str).isin(["Resigned", "Revoked"])]
        if (
            not discarded_resigned_revoked.empty
            and chosen_row["_resigned_or_revoked"] == 0
            and discarded_resigned_revoked["chapter"].fillna("").astype(str).str.strip().ne(chosen_row.get("chapter", "")).any()
        ):
            existing_notes = clean_text(chosen_row.get("chapter_assignment_notes", ""))
            rsrv_chapters = ", ".join(
                sorted(
                    {
                        clean_text(value)
                        for value in discarded_resigned_revoked["chapter"]
                        if clean_text(value)
                    }
                )
            )
            note = f"Ignored same-term resigned/revoked roster row(s) from: {rsrv_chapters}."
            chosen_row["chapter_assignment_notes"] = f"{existing_notes} {note}".strip() if existing_notes else note
        resolved_rows.append(chosen_row.drop(labels=["_resigned_or_revoked", "_secondary_org", "_source_version_priority", "_source_month_priority", "_new_member", "_known_id", "_status_priority", "_identity_key"]))
    resolved = pd.DataFrame(resolved_rows).reset_index(drop=True)
    return resolved, pd.DataFrame(exceptions)


def attach_org_entry_terms(roster: pd.DataFrame, settings: Dict[str, object]) -> pd.DataFrame:
    if roster.empty:
        return roster
    result = ensure_text_columns(roster, ["org_entry_term_code", "org_entry_term_basis"])
    secondary_orgs = secondary_organization_set(settings)
    result["_term_sort"] = result["term_code"].map(sort_term_code)
    for student_id, group in result.groupby("student_id", dropna=False):
        if not clean_text(student_id):
            continue
        primary_group = group.loc[~group["chapter"].map(lambda value: normalize_chapter_name(clean_text(value)) in secondary_orgs)].copy()
        working_group = primary_group if not primary_group.empty else group
        explicit = working_group.loc[working_group["new_member_flag"].eq("Yes")].sort_values("_term_sort")
        if not explicit.empty:
            entry_code = clean_text(explicit.iloc[0]["term_code"])
            basis = "Explicit New Member"
        else:
            ordered = working_group.sort_values("_term_sort")
            entry_code = clean_text(ordered.iloc[0]["term_code"])
            basis = "First Observed Roster"
        if not primary_group.empty and len(primary_group) != len(group):
            basis = f"{basis} (Primary Organization Preferred)"
        result.loc[group.index, "org_entry_term_code"] = entry_code
        result.loc[group.index, "org_entry_term_basis"] = basis
    return result.drop(columns=["_term_sort"])


def build_master_longitudinal(roster: pd.DataFrame, academic: pd.DataFrame, settings: Dict[str, object]) -> pd.DataFrame:
    preferred_roster = choose_preferred_roster_rows(roster, settings)
    roster_base = preferred_roster[
        [
            "student_id",
            "term_code",
            "first_name",
            "last_name",
            "email",
            "org_entry_term_code",
            "org_entry_term_basis",
            "chapter",
            "chapter_raw",
            "chapter_assignment_source",
            "chapter_assignment_confidence",
            "chapter_assignment_notes",
            "org_status_raw",
            "org_status_bucket",
            "org_position_raw",
            "new_member_flag",
        ]
    ].copy() if not preferred_roster.empty else pd.DataFrame(columns=[
        "student_id",
        "term_code",
        "first_name",
        "last_name",
        "email",
        "org_entry_term_code",
        "org_entry_term_basis",
        "chapter",
        "chapter_raw",
        "chapter_assignment_source",
        "chapter_assignment_confidence",
        "chapter_assignment_notes",
        "org_status_raw",
        "org_status_bucket",
        "org_position_raw",
        "new_member_flag",
    ])
    if not roster_base.empty:
        roster_base["roster_present_marker"] = "Yes"
    academic_base = academic[
        [
            "student_id",
            "term_code",
            "first_name",
            "last_name",
            "email",
            "major",
            "term_gpa",
            "institutional_cumulative_gpa",
            "overall_cumulative_gpa",
            "attempted_hours_term",
            "earned_hours_term",
            "institutional_cumulative_hours",
            "total_cumulative_hours",
            "academic_status_raw",
            "academic_standing_raw",
            "academic_standing_bucket",
            "graduation_term_code",
        ]
    ].copy() if not academic.empty else pd.DataFrame(columns=[
        "student_id",
        "term_code",
        "first_name",
        "last_name",
        "email",
        "major",
        "term_gpa",
        "institutional_cumulative_gpa",
        "overall_cumulative_gpa",
        "attempted_hours_term",
        "earned_hours_term",
        "institutional_cumulative_hours",
        "total_cumulative_hours",
        "academic_status_raw",
        "academic_standing_raw",
        "academic_standing_bucket",
        "graduation_term_code",
    ])
    if not academic_base.empty:
        academic_base["academic_present_marker"] = "Yes"

    master = roster_base.merge(
        academic_base,
        on=["student_id", "term_code"],
        how="outer",
        suffixes=("_roster", "_academic"),
    )
    if master.empty:
        return pd.DataFrame(columns=load_schema()["tables"]["master_longitudinal"])

    master["first_name"] = master["first_name_roster"].fillna("").astype(str).str.strip().where(
        master["first_name_roster"].fillna("").astype(str).str.strip().ne(""),
        master["first_name_academic"].fillna("").astype(str).str.strip(),
    )
    master["last_name"] = master["last_name_roster"].fillna("").astype(str).str.strip().where(
        master["last_name_roster"].fillna("").astype(str).str.strip().ne(""),
        master["last_name_academic"].fillna("").astype(str).str.strip(),
    )
    master["email"] = master["email_roster"].fillna("").astype(str).str.strip().str.lower().where(
        master["email_roster"].fillna("").astype(str).str.strip().ne(""),
        master["email_academic"].fillna("").astype(str).str.strip().str.lower(),
    )
    master["term_label"] = master["term_code"].map(term_label_from_code)
    master["observed_year"] = master["term_code"].map(lambda value: parse_term_code(value)[2])
    master["observed_term_sort"] = master["term_code"].map(sort_term_code)
    master["join_term_code"] = master["org_entry_term_code"].fillna("").astype(str).str.strip()
    master["join_term"] = master["join_term_code"].map(term_label_from_code)
    master["join_year"] = master["join_term_code"].map(lambda value: parse_term_code(value)[2] if clean_text(value) else pd.NA)
    master["roster_present"] = master.get("roster_present_marker", pd.Series(pd.NA, index=master.index)).fillna("No")
    master["academic_present"] = master.get("academic_present_marker", pd.Series(pd.NA, index=master.index)).fillna("No")
    master["cumulative_gpa"] = coerce_numeric(master["overall_cumulative_gpa"]).where(
        coerce_numeric(master["overall_cumulative_gpa"]).notna(),
        coerce_numeric(master["institutional_cumulative_gpa"]),
    )
    master["final_outcome_bucket"] = ""
    master["exit_reason_code"] = ""
    master["resolved_outcome_flag"] = ""
    master["outcome_evidence_source"] = ""
    master["school_entry_term_code"] = ""
    master["school_entry_term_basis"] = ""

    master = master.sort_values(["student_id", "observed_term_sort", "term_code"], na_position="last").reset_index(drop=True)
    master["_join_term_sort"] = master["join_term_code"].map(lambda value: sort_term_code(value) if clean_text(value) else 999999)
    master["_within_join_window"] = master["_join_term_sort"].lt(999999) & master["observed_term_sort"].ge(master["_join_term_sort"])
    master["_relative_counter"] = master["_within_join_window"].astype(int).groupby(master["student_id"], dropna=False).cumsum()
    master["relative_term_index"] = (master["_relative_counter"] - 1).where(master["_within_join_window"], pd.NA)
    master = master.drop(
        columns=[
            "first_name_roster",
            "last_name_roster",
            "email_roster",
            "first_name_academic",
            "last_name_academic",
            "email_academic",
            "roster_present_marker",
            "academic_present_marker",
            "_join_term_sort",
            "_within_join_window",
            "_relative_counter",
        ],
        errors="ignore",
    )

    return ensure_columns(master, load_schema()["tables"]["master_longitudinal"])


def attach_snapshot_fields(summary: pd.DataFrame, snapshot: pd.DataFrame, longitudinal: pd.DataFrame) -> pd.DataFrame:
    result = summary.copy()
    if snapshot.empty:
        result["snapshot_matched"] = "No"
        result["current_total_hours"] = pd.NA
        result["estimated_pre_org_hours_txst"] = pd.NA
        result["estimated_pre_org_stage_txst"] = "Unknown"
        return result

    snap = snapshot.copy()
    snap["Student ID"] = snap["Student ID"].map(normalize_banner_id)
    snap = snap.drop_duplicates(subset=["Student ID"], keep="first")
    snap = snap.rename(
        columns={
            "Student ID": "student_id",
            "Total Credit Hours": "snapshot_total_hours",
            "TXST Credit Hours": "snapshot_txst_hours",
            "Overall GPA": "snapshot_overall_gpa",
            "Institutional GPA": "snapshot_institutional_gpa",
            "Student Status": "snapshot_student_status",
            "Student Status (FT/PT)": "snapshot_student_status_ftpt",
        }
    )
    result = result.merge(
        snap[
            [
                "student_id",
                "snapshot_total_hours",
                "snapshot_txst_hours",
                "snapshot_overall_gpa",
                "snapshot_institutional_gpa",
                "snapshot_student_status",
                "snapshot_student_status_ftpt",
            ]
        ],
        on="student_id",
        how="left",
    )
    observed_hours = (
        longitudinal.groupby("student_id", dropna=False)["earned_hours_term"]
        .sum(min_count=1)
        .rename("observed_earned_hours_since_org")
        .reset_index()
    )
    result = result.merge(observed_hours, on="student_id", how="left")
    result["snapshot_matched"] = result["snapshot_total_hours"].notna().map(lambda flag: "Yes" if flag else "No")
    result["current_total_hours"] = result["snapshot_total_hours"].where(result["snapshot_total_hours"].notna(), result["total_cumulative_hours"])
    result["estimated_pre_org_hours_txst"] = (coerce_numeric(result["snapshot_txst_hours"]) - coerce_numeric(result["observed_earned_hours_since_org"])).clip(lower=0)
    result["estimated_pre_org_stage_txst"] = result["estimated_pre_org_hours_txst"].map(bucket_30_hours)
    result["average_cumulative_gpa"] = result["average_cumulative_gpa"].where(
        result["average_cumulative_gpa"].notna(),
        coerce_numeric(result["snapshot_overall_gpa"]).where(
            coerce_numeric(result["snapshot_overall_gpa"]).notna(),
            coerce_numeric(result["snapshot_institutional_gpa"]),
        ),
    )
    result["latest_snapshot_student_status"] = result["snapshot_student_status"].fillna("").astype(str)
    return result


def build_graduation_maps(graduation: pd.DataFrame) -> Tuple[Dict[str, Tuple[str, str]], Dict[Tuple[str, str], Tuple[str, str]]]:
    id_map: Dict[str, Tuple[str, str]] = {}
    name_map: Dict[Tuple[str, str], Tuple[str, str]] = {}
    if graduation.empty:
        return id_map, name_map

    ranked = graduation.copy()
    ranked["_term_sort"] = ranked["Graduation Term"].map(lambda value: sort_term_code(parse_term_code(value)[0]))
    ranked = ranked.sort_values(by=["_term_sort", "Graduation Source File"], ascending=[True, True])

    for row in ranked.itertuples(index=False):
        student_id = normalize_banner_id(getattr(row, "Student ID", ""))
        first_name = clean_text(getattr(row, "First Name", ""))
        last_name = clean_text(getattr(row, "Last Name", ""))
        grad_term_code = parse_term_code(getattr(row, "Graduation Term", ""))[0]
        source = clean_text(getattr(row, "Graduation Source File", "")) or "Graduation List"
        if student_id and student_id not in id_map:
            id_map[student_id] = (grad_term_code, f"Graduation List ID match: {source}")
        name_key = person_name_key(first_name, last_name)
        if (first_name or last_name) and name_key not in name_map:
            name_map[name_key] = (grad_term_code, f"Graduation List name match: {source}")

    return id_map, name_map


def explicit_exit_reason(latest_outcome_bucket: str, latest_status_bucket: str) -> str:
    outcome = clean_text(latest_outcome_bucket)
    status = clean_text(latest_status_bucket)
    if outcome == "Dropped/Resigned/Revoked/Inactive":
        for candidate in ["Revoked", "Resigned", "Inactive"]:
            if candidate == status:
                return candidate
        return "Dropped/Resigned/Revoked/Inactive"
    if outcome in {"Graduated", "Suspended", "Transfer"}:
        return outcome
    return ""


def build_student_summary(
    master: pd.DataFrame,
    snapshot: pd.DataFrame,
    graduation: pd.DataFrame,
    settings: dict[str, object],
    chapter_mapping: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if master.empty:
        empty = pd.DataFrame(columns=load_schema()["tables"]["student_summary"])
        return empty, pd.DataFrame(columns=QA_COLUMNS), pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    summary_rows: List[dict] = []
    qa_rows: List[dict] = []
    outcome_exceptions: List[dict] = []
    max_term_sort = int(master["observed_term_sort"].dropna().max()) if master["observed_term_sort"].dropna().shape[0] else 0
    graduation_by_id, graduation_by_name = build_graduation_maps(graduation)
    sorted_master = master.sort_values(["student_id", "observed_term_sort", "term_code"], na_position="last")
    snapshot_status_by_id: Dict[str, str] = {}
    if not snapshot.empty:
        snapshot_lookup = snapshot.copy()
        snapshot_lookup["_student_id"] = snapshot_lookup["Student ID"].map(normalize_banner_id)
        snapshot_lookup["Student Status"] = snapshot_lookup["Student Status"].fillna("").astype(str)
        snapshot_status_by_id = (
            snapshot_lookup.loc[snapshot_lookup["_student_id"].fillna("").astype(str).str.strip().ne("")]
            .groupby("_student_id", dropna=False)["Student Status"]
            .agg(lambda values: " ".join(value for value in values if clean_text(value)))
            .to_dict()
        )

    for student_id, group in sorted_master.groupby("student_id", dropna=False, sort=False):
        ordered = group.sort_values("observed_term_sort")
        roster_rows = ordered.loc[ordered["roster_present"].eq("Yes")]
        academic_rows = ordered.loc[ordered["academic_present"].eq("Yes")]
        first_row = ordered.iloc[0]
        join_term_code = clean_text(first_row["join_term_code"])
        join_term = term_label_from_code(join_term_code)
        join_year = parse_term_code(join_term_code)[2] if join_term_code else pd.NA

        explicit_grad_term = ""
        evidence_source = ""
        graduation_confirmed = False
        graduation_status_corrected = False
        graduation_status_correction_reason = ""
        snapshot_status_text = snapshot_status_by_id.get(student_id, "")
        graduation_list_term = ""
        if student_id in graduation_by_id:
            graduation_list_term, _ = graduation_by_id[student_id]
        else:
            name_key = person_name_key(first_row["first_name"], first_row["last_name"])
            if name_key in graduation_by_name:
                graduation_list_term, _ = graduation_by_name[name_key]
        if (roster_rows["org_status_bucket"].fillna("").eq("Graduated")).any():
            explicit_grad_term = clean_text(roster_rows.loc[roster_rows["org_status_bucket"].fillna("").eq("Graduated"), "term_code"].iloc[-1])
            evidence_source = "Roster status"
            graduation_confirmed = True
        elif graduation_list_term:
            evidence_source = "Graduation list only; no Copy of Rosters confirmation"

        latest_status_bucket = clean_text(roster_rows["org_status_bucket"].iloc[-1]) if not roster_rows.empty else "Unknown"
        derived_outcome_bucket, derived_evidence_source = outcome_bucket_from_signals(
            " ".join(roster_rows["org_status_bucket"].fillna("").astype(str).tolist()),
            " ".join(academic_rows["academic_status_raw"].fillna("").astype(str).tolist()),
            snapshot_status_text,
        )
        latest_outcome_bucket = derived_outcome_bucket
        evidence_source = evidence_source or derived_evidence_source
        if graduation_confirmed:
            latest_outcome_bucket = "Graduated"
        elif derived_outcome_bucket == "Graduated" or graduation_list_term:
            latest_outcome_bucket = "Unknown"
            evidence_source = evidence_source or "No explicit outcome evidence"
            graduation_status_corrected = True
            if graduation_list_term:
                graduation_status_correction_reason = (
                    "Graduation list matched, but Copy of Rosters did not mark the student as graduated."
                )
            else:
                graduation_status_correction_reason = "Removed graduation classification because no confirmed graduation evidence was present."
        if latest_outcome_bucket == "Unknown":
            outcome_exceptions.append(
                {
                    "exception_type": "unresolved_outcome",
                    "source_file": "",
                    "student_id": student_id,
                    "term_code": clean_text(ordered["term_code"].iloc[-1]),
                    "details": "No explicit outcome evidence; student remains unresolved.",
                }
            )

        entry_row = None
        if join_term_code and not academic_rows.empty:
            entry_candidates = academic_rows.loc[academic_rows["term_code"].eq(join_term_code)]
            if not entry_candidates.empty:
                entry_row = entry_candidates.iloc[0]

        first_ac_term = clean_text(academic_rows["term_code"].iloc[0]) if not academic_rows.empty else ""
        last_ac_term = clean_text(academic_rows["term_code"].iloc[-1]) if not academic_rows.empty else ""
        first_org_term = clean_text(roster_rows["term_code"].iloc[0]) if not roster_rows.empty else ""
        last_org_term = clean_text(roster_rows["term_code"].iloc[-1]) if not roster_rows.empty else ""
        gpa_values = coerce_numeric(academic_rows["term_gpa"]).dropna()
        first_term_gpa = gpa_values.iloc[0] if gpa_values.shape[0] else pd.NA
        second_term_gpa = gpa_values.iloc[1] if gpa_values.shape[0] > 1 else pd.NA
        first_year_window = academic_rows.loc[academic_rows["relative_term_index"].fillna(-1).astype(float).between(0, 2, inclusive="both")]
        first_year_avg_gpa = coerce_numeric(first_year_window["term_gpa"]).mean() if not first_year_window.empty else pd.NA
        latest_overall_cum = coerce_numeric(academic_rows["overall_cumulative_gpa"]).dropna().iloc[-1] if coerce_numeric(academic_rows["overall_cumulative_gpa"]).dropna().shape[0] else pd.NA
        latest_txstate_cum = coerce_numeric(academic_rows["institutional_cumulative_gpa"]).dropna().iloc[-1] if coerce_numeric(academic_rows["institutional_cumulative_gpa"]).dropna().shape[0] else pd.NA
        latest_cumulative_hours = coerce_numeric(academic_rows["total_cumulative_hours"]).dropna().iloc[-1] if coerce_numeric(academic_rows["total_cumulative_hours"]).dropna().shape[0] else pd.NA
        entry_hours = coerce_numeric(pd.Series([entry_row["institutional_cumulative_hours"] if entry_row is not None else pd.NA])).iloc[0]
        first_passed = coerce_numeric(first_year_window["earned_hours_term"]).dropna().iloc[0] if coerce_numeric(first_year_window["earned_hours_term"]).dropna().shape[0] else pd.NA
        first_year_passed = coerce_numeric(first_year_window["earned_hours_term"]).sum(min_count=1) if not first_year_window.empty else pd.NA

        join_sort = sort_term_code(join_term_code) if join_term_code else None
        next_term_sort = None
        next_fall_sort = None
        one_year_sort = None
        if join_sort is not None and join_sort < 999999:
            join_year_value = int(join_term_code[:4])
            join_season_code = join_term_code[-2:]
            if join_season_code == "FA":
                next_term_sort = (join_year_value + 1) * 10 + SEASON_ORDER["SP"]
            elif join_season_code == "SP":
                next_term_sort = join_year_value * 10 + SEASON_ORDER["SU"]
            elif join_season_code == "SU":
                next_term_sort = join_year_value * 10 + SEASON_ORDER["FA"]
            elif join_season_code == "WI":
                next_term_sort = join_year_value * 10 + SEASON_ORDER["SP"]
            next_fall_sort = (join_year_value + (1 if join_season_code == "FA" else 0)) * 10 + SEASON_ORDER["FA"]
            one_year_sort = (join_year_value + 1) * 10 + SEASON_ORDER.get(join_season_code, 9)

        roster_term_sorts = set(coerce_numeric(roster_rows["observed_term_sort"]).dropna().astype(int).tolist()) if not roster_rows.empty else set()
        academic_term_sorts = set(coerce_numeric(academic_rows["observed_term_sort"]).dropna().astype(int).tolist()) if not academic_rows.empty else set()

        def has_term(term_sorts: set[int], target_sort: Optional[int]) -> str:
            if target_sort is None or target_sort > max_term_sort:
                return ""
            return "Yes" if target_sort in term_sorts else "No"

        def measurable_for_years(years: int) -> str:
            if not join_term_code:
                return ""
            target_sort = sort_term_code(f"{int(join_term_code[:4]) + years}{join_term_code[-2:]}")
            return "Yes" if max_term_sort >= target_sort else "No"

        graduated_eventual = "Yes" if latest_outcome_bucket == "Graduated" else "No"
        graduated_eventual_measurable = "Yes" if join_term_code else "No"
        graduated_4yr_measurable = measurable_for_years(4)
        graduated_6yr_measurable = measurable_for_years(6)
        grad_4_target = sort_term_code(f"{int(join_term_code[:4]) + 4}{join_term_code[-2:]}") if join_term_code else 999999
        grad_6_target = sort_term_code(f"{int(join_term_code[:4]) + 6}{join_term_code[-2:]}") if join_term_code else 999999
        graduated_4yr = "Yes" if latest_outcome_bucket == "Graduated" and graduated_4yr_measurable == "Yes" and explicit_grad_term and sort_term_code(explicit_grad_term) <= grad_4_target else "No" if graduated_4yr_measurable == "Yes" else ""
        graduated_6yr = "Yes" if latest_outcome_bucket == "Graduated" and graduated_6yr_measurable == "Yes" and explicit_grad_term and sort_term_code(explicit_grad_term) <= grad_6_target else "No" if graduated_6yr_measurable == "Yes" else ""

        first_standing = clean_text(academic_rows["academic_standing_bucket"].iloc[0]) if not academic_rows.empty else "Unknown"
        latest_standing = clean_text(academic_rows["academic_standing_bucket"].iloc[-1]) if not academic_rows.empty else "Unknown"
        chapter_source_row = roster_rows.iloc[0] if not roster_rows.empty else None
        if not roster_rows.empty:
            chapter_source_candidates = roster_rows.loc[~roster_rows["chapter"].map(chapter_is_missing)]
            if not chapter_source_candidates.empty:
                chapter_source_row = chapter_source_candidates.iloc[0]

        summary_rows.append(
            {
                "student_id": student_id,
                "student_name": f"{clean_text(first_row['first_name'])} {clean_text(first_row['last_name'])}".strip(),
                "chapter": clean_text(roster_rows["chapter"].iloc[0]) if not roster_rows.empty else "",
                "initial_chapter": clean_text(roster_rows["chapter"].iloc[0]) if not roster_rows.empty else "",
                "latest_chapter": clean_text(roster_rows["chapter"].iloc[-1]) if not roster_rows.empty else "",
                "chapter_assignment_source": clean_text(chapter_source_row["chapter_assignment_source"]) if chapter_source_row is not None else "unresolved",
                "chapter_assignment_confidence": clean_text(chapter_source_row["chapter_assignment_confidence"]) if chapter_source_row is not None else "low",
                "chapter_assignment_notes": clean_text(chapter_source_row["chapter_assignment_notes"]) if chapter_source_row is not None else "No chapter assignment was available.",
                "join_term_code": join_term_code,
                "join_term": join_term,
                "join_year": join_year,
                "org_entry_cohort": join_term,
                "org_entry_term_basis": clean_text(roster_rows["org_entry_term_basis"].iloc[0]) if not roster_rows.empty else "",
                "school_entry_term_code": clean_text(first_row["school_entry_term_code"]),
                "school_entry_term": term_label_from_code(first_row["school_entry_term_code"]),
                "school_entry_term_basis": clean_text(first_row["school_entry_term_basis"]),
                "first_observed_org_term_code": first_org_term,
                "first_observed_org_term": term_label_from_code(first_org_term),
                "last_observed_org_term_code": last_org_term,
                "last_observed_org_term": term_label_from_code(last_org_term),
                "first_observed_academic_term_code": first_ac_term,
                "first_observed_academic_term": term_label_from_code(first_ac_term),
                "last_observed_academic_term_code": last_ac_term,
                "last_observed_academic_term": term_label_from_code(last_ac_term),
                "latest_outcome_bucket": latest_outcome_bucket,
                "exit_reason_code": explicit_exit_reason(latest_outcome_bucket, latest_status_bucket),
                "graduation_term_code": explicit_grad_term,
                "resolved_outcome_flag": "Yes" if latest_outcome_bucket not in UNRESOLVED_OUTCOMES else "No",
                "resolved_outcome_excluded_flag": "Yes" if latest_outcome_bucket in UNRESOLVED_OUTCOMES else "No",
                "resolved_outcome_exclusion_reason": latest_outcome_bucket if latest_outcome_bucket in UNRESOLVED_OUTCOMES else "",
                "outcome_evidence_source": evidence_source,
                "graduation_evidence_confirmed": "Yes" if graduation_confirmed else "No",
                "graduation_status_corrected_flag": "Yes" if graduation_status_corrected else "No",
                "graduation_status_correction_reason": graduation_status_correction_reason,
                "latest_roster_status_bucket": latest_status_bucket or "Unknown",
                "initial_roster_status_bucket": clean_text(roster_rows["org_status_bucket"].iloc[0]) if not roster_rows.empty else "Unknown",
                "active_flag": "Yes" if latest_status_bucket in {"Active", "New Member"} else "No",
                "major": clean_text(academic_rows["major"].iloc[-1]) if not academic_rows.empty else "",
                "pell_flag": "",
                "transfer_flag": "Yes" if latest_outcome_bucket == "Transfer" else "No" if latest_outcome_bucket else "",
                "graduation_term": term_label_from_code(explicit_grad_term),
                "graduation_year": parse_term_code(explicit_grad_term)[2] if explicit_grad_term else pd.NA,
                "entry_cumulative_hours": entry_hours,
                "entry_hours_bucket": bucket_30_hours(entry_hours),
                "estimated_pre_org_hours_txst": pd.NA,
                "estimated_pre_org_stage_txst": "Unknown",
                "current_total_hours": latest_cumulative_hours,
                "total_cumulative_hours": latest_cumulative_hours,
                "first_term_gpa": first_term_gpa,
                "second_term_gpa": second_term_gpa,
                "first_year_avg_term_gpa": first_year_avg_gpa,
                "average_term_gpa": first_year_avg_gpa if not pd.isna(first_year_avg_gpa) else coerce_numeric(academic_rows["term_gpa"]).mean(),
                "gpa_change": (second_term_gpa - first_term_gpa) if not pd.isna(first_term_gpa) and not pd.isna(second_term_gpa) else pd.NA,
                "latest_overall_cumulative_gpa": latest_overall_cum,
                "latest_txstate_cumulative_gpa": latest_txstate_cum,
                "average_cumulative_gpa": latest_overall_cum if not pd.isna(latest_overall_cum) else latest_txstate_cum,
                "first_term_passed_hours": first_passed,
                "first_year_passed_hours": first_year_passed,
                "graduated_eventual": graduated_eventual,
                "graduated_eventual_measurable": graduated_eventual_measurable,
                "graduated_4yr": graduated_4yr,
                "graduated_4yr_measurable": graduated_4yr_measurable,
                "graduated_6yr": graduated_6yr,
                "graduated_6yr_measurable": graduated_6yr_measurable,
                "retained_next_term": has_term(roster_term_sorts, next_term_sort),
                "retained_next_term_measurable": "Yes" if next_term_sort and next_term_sort <= max_term_sort else "",
                "retained_next_fall": has_term(roster_term_sorts, next_fall_sort),
                "retained_next_fall_measurable": "Yes" if next_fall_sort and next_fall_sort <= max_term_sort else "",
                "retained_one_year": has_term(roster_term_sorts, one_year_sort),
                "retained_one_year_measurable": "Yes" if one_year_sort and one_year_sort <= max_term_sort else "",
                "continued_next_term": has_term(academic_term_sorts, next_term_sort),
                "continued_next_term_measurable": "Yes" if next_term_sort and next_term_sort <= max_term_sort else "",
                "continued_next_fall": has_term(academic_term_sorts, next_fall_sort),
                "continued_next_fall_measurable": "Yes" if next_fall_sort and next_fall_sort <= max_term_sort else "",
                "continued_one_year": has_term(academic_term_sorts, one_year_sort),
                "continued_one_year_measurable": "Yes" if one_year_sort and one_year_sort <= max_term_sort else "",
                "low_gpa_2_0_flag": "Yes" if not pd.isna(first_term_gpa) and float(first_term_gpa) < 2.0 else "No" if not pd.isna(first_term_gpa) else "",
                "low_gpa_2_5_flag": "Yes" if not pd.isna(first_term_gpa) and float(first_term_gpa) < 2.5 else "No" if not pd.isna(first_term_gpa) else "",
                "first_year_low_gpa_2_0_flag": "Yes" if not pd.isna(first_year_avg_gpa) and float(first_year_avg_gpa) < 2.0 else "No" if not pd.isna(first_year_avg_gpa) else "",
                "first_year_low_gpa_2_5_flag": "Yes" if not pd.isna(first_year_avg_gpa) and float(first_year_avg_gpa) < 2.5 else "No" if not pd.isna(first_year_avg_gpa) else "",
                "good_standing_first_term": "Yes" if first_standing == "Good Standing" else "No" if first_standing != "Unknown" else "",
                "probation_warning_first_year": "Yes" if (first_year_window["academic_standing_bucket"].fillna("").eq("Probation/Warning")).any() else "No" if not first_year_window.empty else "",
                "academic_standing_suspended_ever": "Yes" if (academic_rows["academic_standing_bucket"].fillna("").eq("Suspended")).any() else "No",
                "first_academic_standing_bucket": first_standing,
                "latest_academic_standing_bucket": latest_standing,
                "snapshot_matched": "No",
                "source_logic": "canonical_pipeline",
            }
        )

    summary = pd.DataFrame(summary_rows)
    summary = attach_snapshot_fields(summary, snapshot, master)
    summary["chapter"] = summary["initial_chapter"].where(summary["initial_chapter"].fillna("").astype(str).str.strip().ne(""), summary["latest_chapter"])
    summary["is_fsl_member"] = summary["chapter"].fillna("").astype(str).str.strip().ne("")
    summary["chapter_size"] = summary.groupby("chapter", dropna=False)["student_id"].transform("nunique")

    def chapter_band(value: object) -> str:
        if pd.isna(value):
            return "Unknown"
        number = float(value)
        for band in settings.get("chapter_size_bands", []):
            lower = float(band.get("min", 0))
            upper = band.get("max")
            if number >= lower and (upper is None or number <= float(upper)):
                return str(band["label"])
        return "Unknown"

    summary["chapter_size_band"] = summary["chapter_size"].map(chapter_band)
    summary["high_hours_flag"] = coerce_numeric(summary["total_cumulative_hours"]).ge(settings.get("high_hours_threshold", 60))
    summary["high_hours_group"] = summary["high_hours_flag"].map(lambda value: "High Hours" if value else "Lower Hours" if value is not pd.NA and not pd.isna(value) else "Unknown")
    summary["active_membership_group"] = summary["active_flag"].map(lambda value: "Active" if value == "Yes" else "Inactive/Other" if value == "No" else "Unknown")
    summary["pell_group"] = summary["pell_flag"].map(lambda value: "Pell" if value == "Yes" else "Non-Pell" if value == "No" else "Unknown")
    summary["transfer_group"] = summary["transfer_flag"].map(lambda value: "Transfer" if value == "Yes" else "Non-Transfer" if value == "No" else "Unknown")
    summary["snapshot_group"] = summary["snapshot_matched"].map(lambda value: "Snapshot Matched" if value == "Yes" else "No Snapshot Match" if value == "No" else "Unknown")
    summary["status_group"] = summary["latest_outcome_bucket"].replace("", "Unknown")
    summary["major_group"] = summary["major"].replace("", "Unknown")

    if not chapter_mapping.empty:
        mapping = chapter_mapping.copy()
        mapping["_chapter_key"] = mapping["chapter"].fillna("").astype(str).str.strip().str.lower()
        summary["_chapter_key"] = summary["chapter"].fillna("").astype(str).str.strip().str.lower()
        summary = summary.merge(mapping[["_chapter_key", "chapter_group", "council", "org_type", "family", "custom_group"]], on="_chapter_key", how="left")
        summary = summary.drop(columns=["_chapter_key"])
    for column, default in {
        "chapter_group": "Unassigned",
        "council": "Unknown",
        "org_type": "Unknown",
        "family": "Unknown",
        "custom_group": "Unassigned",
    }.items():
        if column not in summary.columns:
            summary[column] = default
        summary[column] = summary[column].fillna("").astype(str).replace("", default)

    roster_present_mask = master["roster_present"].fillna("").astype(str).str.strip().str.lower().isin(["true", "yes", "1"])
    current_active_fields = build_current_active_fields(summary, master.loc[roster_present_mask].copy(), chapter_mapping, settings)
    for column in current_active_fields.columns:
        summary[column] = current_active_fields[column]

    resolution_fields = build_outcome_resolution_fields(summary, settings.get("outcome_resolution", {}))
    for column in resolution_fields.columns:
        summary[column] = resolution_fields[column]
    if "graduation_status_without_evidence" in summary.columns:
        corrected_mask = summary["graduation_status_without_evidence"].fillna(False).astype(bool)
        summary["graduation_status_corrected_flag"] = summary["graduation_status_corrected_flag"].where(
            summary["graduation_status_corrected_flag"].fillna("").astype(str).str.strip().eq("Yes") | ~corrected_mask,
            "Yes",
        )
        summary["graduation_status_correction_reason"] = summary["graduation_status_correction_reason"].where(
            summary["graduation_status_correction_reason"].fillna("").astype(str).str.strip().ne("") | ~corrected_mask,
            "Graduation claim was present, but no confirmed graduation evidence source was available.",
        )
        summary.loc[corrected_mask, "latest_outcome_bucket"] = "Unknown"
    confirmed_grad_mask = summary["is_graduated"].fillna(False).astype(bool)
    summary.loc[~confirmed_grad_mask, "graduated_eventual"] = "No"
    summary.loc[~confirmed_grad_mask & summary["graduated_4yr_measurable"].fillna("").astype(str).eq("Yes"), "graduated_4yr"] = "No"
    summary.loc[~confirmed_grad_mask & summary["graduated_6yr_measurable"].fillna("").astype(str).eq("Yes"), "graduated_6yr"] = "No"
    summary["status_group"] = summary["latest_outcome_bucket"].replace("", "Unknown")
    summary["resolved_outcome_flag"] = summary["is_resolved_outcome"].fillna(False).map(lambda value: "Yes" if bool(value) else "No")
    summary["resolved_outcome_excluded_flag"] = (~summary["is_resolved_outcome"].fillna(False)).map(lambda value: "Yes" if bool(value) else "No")
    summary["resolved_outcome_exclusion_reason"] = summary["outcome_resolution_group"].where(~summary["is_resolved_outcome"].fillna(False), "")

    completeness_fields = [field for field in settings.get("completeness_fields", []) if field in summary.columns]
    if completeness_fields:
        present = summary[completeness_fields].notna() & summary[completeness_fields].astype(str).ne("")
        summary["data_completeness_rate"] = present.sum(axis=1) / len(completeness_fields)
    else:
        summary["data_completeness_rate"] = pd.NA

    resolved_mask = boolish_series(summary.get("is_resolved_outcome", pd.Series(False, index=summary.index)))
    active_mask = boolish_series(summary.get("is_active_outcome", pd.Series(False, index=summary.index)))
    unknown_mask = boolish_series(summary.get("is_unknown_outcome", pd.Series(False, index=summary.index)))
    qa_rows.extend(
        [
            {"Check Group": "Coverage", "Check": "Unique students", "Status": "Pass", "Value": int(summary["student_id"].nunique()), "Notes": ""},
            {"Check Group": "Coverage", "Check": "Resolved outcomes", "Status": "Pass", "Value": int(resolved_mask.sum()), "Notes": ""},
            {"Check Group": "Coverage", "Check": "Still active outcomes", "Status": "Pass", "Value": int(active_mask.sum()), "Notes": ""},
            {"Check Group": "Coverage", "Check": "Truly unknown / unresolved outcomes", "Status": "Pass", "Value": int(unknown_mask.sum()), "Notes": ""},
            {"Check Group": "Coverage", "Check": "Other / unmapped outcomes", "Status": "Pass", "Value": int((~resolved_mask & ~active_mask & ~unknown_mask).sum()), "Notes": ""},
            {"Check Group": "Coverage", "Check": "Current active students (most recent roster only)", "Status": "Pass", "Value": int(summary["current_active_flag"].fillna("").astype(str).eq("Yes").sum()), "Notes": f"Most recent roster term: {clean_text(summary['current_active_roster_term'].iloc[0]) or clean_text(summary['current_active_roster_term_code'].iloc[0])}"},
            {"Check Group": "Coverage", "Check": "Historical latest-status active students", "Status": "Pass", "Value": int(summary["active_flag"].fillna("").astype(str).eq("Yes").sum()), "Notes": "Retained for historical outcome logic only; not used for current active headcounts."},
            {"Check Group": "Coverage", "Check": "Unresolved chapter assignments", "Status": "Pass", "Value": int(summary["chapter_assignment_source"].fillna("").astype(str).eq("unresolved").sum()), "Notes": ""},
        ]
    )

    summary_columns = load_schema()["tables"]["student_summary"] + [
        "is_fsl_member",
        "chapter_size",
        "chapter_size_band",
        "high_hours_flag",
        "high_hours_group",
        "active_membership_group",
        "current_active_flag",
        "current_active_membership_group",
        "current_active_chapter",
        "current_active_chapter_group",
        "current_active_council",
        "current_active_org_type",
        "current_active_family",
        "current_active_custom_group",
        "current_active_chapter_size",
        "current_active_chapter_size_band",
        "current_active_roster_term_code",
        "current_active_roster_term",
        "current_active_source_file",
        "current_active_source_sheet",
        "pell_group",
        "transfer_group",
        "snapshot_group",
        "status_group",
        "major_group",
        "chapter_group",
        "council",
        "org_type",
        "family",
        "custom_group",
        "resolved_outcomes_only_flag",
        "data_completeness_rate",
        "latest_snapshot_student_status",
    ]
    summary_columns = list(dict.fromkeys(summary_columns))
    return ensure_columns(summary, summary_columns), pd.DataFrame(qa_rows), pd.DataFrame(outcome_exceptions)


def build_current_active_fields(
    summary: pd.DataFrame,
    roster: pd.DataFrame,
    chapter_mapping: pd.DataFrame,
    settings: dict[str, object],
) -> pd.DataFrame:
    columns = [
        "current_active_flag",
        "current_active_membership_group",
        "current_active_chapter",
        "current_active_chapter_group",
        "current_active_council",
        "current_active_org_type",
        "current_active_family",
        "current_active_custom_group",
        "current_active_chapter_size",
        "current_active_chapter_size_band",
        "current_active_roster_term_code",
        "current_active_roster_term",
        "current_active_source_file",
        "current_active_source_sheet",
    ]
    result = pd.DataFrame(index=summary.index)
    for column in columns:
        result[column] = ""

    if summary.empty:
        return result

    result["current_active_flag"] = "No"
    result["current_active_membership_group"] = "Not Current Active"
    result["current_active_chapter_size"] = pd.NA

    if roster.empty or "term_code" not in roster.columns:
        return result

    roster_working = roster.copy()
    roster_working["term_code"] = roster_working["term_code"].fillna("").astype(str).str.strip()
    roster_working = roster_working.loc[roster_working["term_code"].ne("")].copy()
    if roster_working.empty:
        return result

    roster_working["_term_sort"] = roster_working["term_code"].map(sort_term_code)
    latest_term_sort = roster_working["_term_sort"].max()
    latest_roster = roster_working.loc[roster_working["_term_sort"].eq(latest_term_sort)].copy()
    if latest_roster.empty:
        return result

    latest_term_code = clean_text(latest_roster["term_code"].iloc[0])
    latest_term_label = term_label_from_code(latest_term_code)
    result["current_active_roster_term_code"] = latest_term_code
    result["current_active_roster_term"] = latest_term_label

    active_latest = latest_roster.loc[
        latest_roster["org_status_bucket"].fillna("").astype(str).isin(["Active", "New Member"])
    ].copy()
    if active_latest.empty:
        return result

    active_latest["student_id"] = active_latest["student_id"].fillna("").astype(str).str.strip()
    active_latest = active_latest.loc[active_latest["student_id"].ne("")].copy()
    for column in ["source_file", "source_sheet", "chapter"]:
        if column not in active_latest.columns:
            active_latest[column] = ""
    active_latest["_source_format_priority"] = active_latest["source_file"].map(source_file_format_priority)
    active_latest = active_latest.sort_values(
        ["student_id", "_source_format_priority", "source_file", "source_sheet", "chapter"],
        ascending=[True, False, True, True, True],
        na_position="last",
    ).drop_duplicates(subset=["student_id"], keep="first")

    chapter_lookup = dict(zip(active_latest["student_id"], active_latest["chapter"].fillna("").astype(str).str.strip()))
    source_file_lookup = dict(zip(active_latest["student_id"], active_latest["source_file"].fillna("").astype(str).str.strip()))
    source_sheet_lookup = dict(zip(active_latest["student_id"], active_latest["source_sheet"].fillna("").astype(str).str.strip()))
    active_ids = set(chapter_lookup.keys())
    summary_ids = summary["student_id"].fillna("").astype(str).str.strip()
    active_mask = summary_ids.isin(active_ids)

    result.loc[active_mask, "current_active_flag"] = "Yes"
    result.loc[active_mask, "current_active_membership_group"] = "Current Active"
    result["current_active_chapter"] = summary_ids.map(chapter_lookup).fillna("")
    result["current_active_source_file"] = summary_ids.map(source_file_lookup).fillna("")
    result["current_active_source_sheet"] = summary_ids.map(source_sheet_lookup).fillna("")

    if not chapter_mapping.empty:
        mapping = chapter_mapping.copy()
        mapping["_chapter_key"] = mapping["chapter"].fillna("").astype(str).str.strip().str.lower()
        result["_chapter_key"] = result["current_active_chapter"].fillna("").astype(str).str.strip().str.lower()
        mapped = result.merge(
            mapping[["_chapter_key", "chapter_group", "council", "org_type", "family", "custom_group"]],
            on="_chapter_key",
            how="left",
            suffixes=("", "_mapped"),
        )
        result = mapped.drop(columns=["_chapter_key"])

    mapped_defaults = {
        "current_active_chapter_group": "Unassigned",
        "current_active_council": "Unknown",
        "current_active_org_type": "Unknown",
        "current_active_family": "Unknown",
        "current_active_custom_group": "Unassigned",
    }
    mapped_aliases = {
        "current_active_chapter_group": "chapter_group",
        "current_active_council": "council",
        "current_active_org_type": "org_type",
        "current_active_family": "family",
        "current_active_custom_group": "custom_group",
    }
    for output_column, source_column in mapped_aliases.items():
        if output_column not in result.columns and source_column in result.columns:
            result[output_column] = result[source_column]
        if output_column not in result.columns:
            result[output_column] = ""
        current_mask = result["current_active_flag"].eq("Yes")
        result.loc[current_mask, output_column] = (
            result.loc[current_mask, output_column]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", mapped_defaults[output_column])
        )
        result.loc[~current_mask, output_column] = ""

    active_chapter_counts = (
        result.loc[
            result["current_active_flag"].eq("Yes")
            & result["current_active_chapter"].fillna("").astype(str).str.strip().ne("")
        ]
        .groupby("current_active_chapter", dropna=False)["current_active_flag"]
        .size()
        .rename("current_active_chapter_size")
    )
    if not active_chapter_counts.empty:
        result["current_active_chapter_size"] = result["current_active_chapter"].map(active_chapter_counts)

    def chapter_band(value: object) -> str:
        if pd.isna(value):
            return ""
        number = float(value)
        for band in settings.get("chapter_size_bands", []):
            lower = float(band.get("min", 0))
            upper = band.get("max")
            if number >= lower and (upper is None or number <= float(upper)):
                return str(band["label"])
        return ""

    result["current_active_chapter_size_band"] = result["current_active_chapter_size"].map(chapter_band)
    result.loc[~result["current_active_flag"].eq("Yes"), "current_active_chapter_size_band"] = ""
    return result


def build_cohort_metrics(summary: pd.DataFrame) -> pd.DataFrame:
    rows: List[dict] = []
    if summary.empty:
        return pd.DataFrame(columns=load_schema()["tables"]["cohort_metrics"])

    cohorts = ["Overall"] + sorted(value for value in summary["org_entry_cohort"].fillna("").astype(str).unique().tolist() if value)
    for cohort in cohorts:
        frame = summary if cohort == "Overall" else summary.loc[summary["org_entry_cohort"].eq(cohort)].copy()
        if frame.empty:
            continue

        def rate_row(label: str, numerator_col: str, denominator_col: str, group: str, notes: str = "") -> None:
            rate_frame = frame.drop_duplicates(subset=["student_id"], keep="first") if group == "Graduation" and "student_id" in frame.columns else frame
            eligible = int(rate_frame[denominator_col].fillna("").astype(str).eq("Yes").sum())
            numerator = int((rate_frame[numerator_col].fillna("").astype(str).eq("Yes") & rate_frame[denominator_col].fillna("").astype(str).eq("Yes")).sum())
            rows.append(metric_row(group, label, cohort, eligible, numerator=numerator, rate=(numerator / eligible) if eligible else None, notes=notes))

        def mean_row(label: str, value_col: str, group: str) -> None:
            values = coerce_numeric(frame[value_col]).dropna()
            rows.append(metric_row(group, label, cohort, int(values.shape[0]), average=float(values.mean()) if not values.empty else None))

        rows.append(metric_row("Coverage", "Students", cohort, int(frame["student_id"].nunique()), numerator=int(frame["student_id"].nunique()), rate=1.0))
        rate_row("Observed Eventual Graduation Rate", "graduated_eventual", "graduated_eventual_measurable", "Graduation")
        rate_row("Observed 4-Year Graduation Rate", "graduated_4yr", "graduated_4yr_measurable", "Graduation")
        rate_row("Observed 6-Year Graduation Rate", "graduated_6yr", "graduated_6yr_measurable", "Graduation")
        rate_row("Organization Retention To Next Term", "retained_next_term", "retained_next_term_measurable", "Retention")
        rate_row("Organization Retention To Next Fall", "retained_next_fall", "retained_next_fall_measurable", "Retention")
        rate_row("Academic Continuation To Next Fall", "continued_next_fall", "continued_next_fall_measurable", "Retention")
        mean_row("Average First-Year Term GPA", "first_year_avg_term_gpa", "GPA")
        mean_row("Average Cumulative GPA", "average_cumulative_gpa", "GPA")

        first_term_hours = coerce_numeric(frame["first_term_passed_hours"]).dropna()
        rows.append(metric_row("Credit Momentum", "First-Term 15+ Passed Hours Rate", cohort, int(first_term_hours.shape[0]), numerator=int((first_term_hours >= 15).sum()), rate=((first_term_hours >= 15).sum() / first_term_hours.shape[0]) if first_term_hours.shape[0] else None))
        first_year_hours = coerce_numeric(frame["first_year_passed_hours"]).dropna()
        rows.append(metric_row("Credit Momentum", "First-Year 30+ Passed Hours Rate", cohort, int(first_year_hours.shape[0]), numerator=int((first_year_hours >= 30).sum()), rate=((first_year_hours >= 30).sum() / first_year_hours.shape[0]) if first_year_hours.shape[0] else None))

    return ensure_columns(pd.DataFrame(rows), load_schema()["tables"]["cohort_metrics"])


def build_status_exceptions(roster: pd.DataFrame, academic: pd.DataFrame) -> pd.DataFrame:
    rows: List[dict] = []
    valid_roster_buckets = {
        "Graduated",
        "Suspended",
        "Transfer",
        "Revoked",
        "Resigned",
        "Inactive",
        "New Member",
        "Active",
        "Alumni",
        "Unknown",
    }
    if not roster.empty:
        unmapped_roster = roster.loc[
            roster["org_status_raw"].fillna("").astype(str).str.strip().ne("")
            & ~roster["org_status_bucket"].fillna("").astype(str).isin(valid_roster_buckets)
        ]
        for row in unmapped_roster.itertuples(index=False):
            rows.append(
                {
                    "exception_type": "unmapped_roster_status",
                    "source_file": clean_text(getattr(row, "source_file", "")),
                    "student_id": clean_text(getattr(row, "student_id", "")),
                    "term_code": clean_text(getattr(row, "term_code", "")),
                    "details": clean_text(getattr(row, "org_status_raw", "")),
                }
            )
    if not academic.empty:
        unmapped_standing = academic.loc[
            academic["academic_standing_raw"].fillna("").astype(str).str.strip().ne("")
            & academic["academic_standing_bucket"].fillna("").astype(str).eq("Other/Unmapped")
        ]
        for row in unmapped_standing.itertuples(index=False):
            rows.append(
                {
                    "exception_type": "unmapped_academic_standing",
                    "source_file": clean_text(getattr(row, "source_file", "")),
                    "student_id": clean_text(getattr(row, "student_id", "")),
                    "term_code": clean_text(getattr(row, "term_code", "")),
                    "details": clean_text(getattr(row, "academic_standing_raw", "")),
                }
            )
    return pd.DataFrame(rows, columns=["exception_type", "source_file", "student_id", "term_code", "details"])


def build_spring_coverage_checks(frame: pd.DataFrame, label: str) -> List[dict]:
    rows: List[dict] = []
    if frame.empty or "term_year" not in frame.columns or "term_season" not in frame.columns:
        return rows
    years = sorted({int(value) for value in coerce_numeric(frame["term_year"]).dropna().tolist()})
    for year in years:
        year_frame = frame.loc[coerce_numeric(frame["term_year"]).eq(year)]
        spring_count = int(year_frame["term_season"].fillna("").astype(str).str.strip().eq("Spring").sum())
        rows.append(
            {
                "Check Group": "Coverage",
                "Check": f"{label} spring coverage {year}",
                "Status": "Pass" if spring_count > 0 else "Review",
                "Value": spring_count,
                "Notes": "" if spring_count > 0 else f"No spring {label.lower()} rows found for {year}.",
            }
        )
    return rows


def build_measurable_window_checks(summary: pd.DataFrame) -> List[dict]:
    rows: List[dict] = []
    if summary.empty:
        return rows
    invalid_4yr = summary.loc[
        summary["graduated_4yr_measurable"].fillna("").astype(str).ne("Yes")
        & summary["graduated_4yr"].fillna("").astype(str).str.strip().ne("")
    ]
    invalid_6yr = summary.loc[
        summary["graduated_6yr_measurable"].fillna("").astype(str).ne("Yes")
        & summary["graduated_6yr"].fillna("").astype(str).str.strip().ne("")
    ]
    rows.append(
        {
            "Check Group": "Windows",
            "Check": "4-year graduation window enforcement",
            "Status": "Pass" if invalid_4yr.empty else "Fail",
            "Value": int(len(invalid_4yr)),
            "Notes": "" if invalid_4yr.empty else "Non-measurable students have populated 4-year graduation values.",
        }
    )
    rows.append(
        {
            "Check Group": "Windows",
            "Check": "6-year graduation window enforcement",
            "Status": "Pass" if invalid_6yr.empty else "Fail",
            "Value": int(len(invalid_6yr)),
            "Notes": "" if invalid_6yr.empty else "Non-measurable students have populated 6-year graduation values.",
        }
    )
    return rows


def boolish_series(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.lower().isin({"true", "1", "yes", "y"})


def build_graduation_status_audit(summary: pd.DataFrame) -> pd.DataFrame:
    columns = ["Audit Section", "Measure", "Value", "Status", "Notes"]
    if summary.empty:
        return pd.DataFrame(columns=columns)

    total_rows = int(len(summary))
    total_unique = int(summary["student_id"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    duplicate_student_rows = max(total_rows - total_unique, 0)
    graduated = boolish_series(summary.get("is_graduated", pd.Series(False, index=summary.index)))
    confirmed = boolish_series(summary.get("graduation_evidence_confirmed", pd.Series(False, index=summary.index)))
    corrected = summary.get("graduation_status_corrected_flag", pd.Series("", index=summary.index)).fillna("").astype(str).str.strip().eq("Yes")
    active = boolish_series(summary.get("is_active_outcome", pd.Series(False, index=summary.index)))
    unknown = boolish_series(summary.get("is_unknown_outcome", pd.Series(False, index=summary.index)))
    non_grad = boolish_series(summary.get("is_known_non_graduate_exit", pd.Series(False, index=summary.index)))
    numerator = summary.get("graduated_eventual", pd.Series("", index=summary.index)).fillna("").astype(str).str.strip().eq("Yes")
    without_evidence = boolish_series(summary.get("graduation_status_without_evidence", pd.Series(False, index=summary.index)))

    rows = [
        {"Audit Section": "Summary", "Measure": "Total unique students used for graduation calculations", "Value": total_unique, "Status": "Pass", "Notes": "Graduation rates are student-level calculations."},
        {"Audit Section": "Summary", "Measure": "Duplicate student rows in summary", "Value": duplicate_student_rows, "Status": "Pass" if duplicate_student_rows == 0 else "Fail", "Notes": "Should be 0 for student-level graduation metrics."},
        {"Audit Section": "Summary", "Measure": "Students marked Graduated", "Value": int(graduated.sum()), "Status": "Pass", "Notes": ""},
        {"Audit Section": "Summary", "Measure": "Students with confirmed graduation evidence", "Value": int(confirmed.sum()), "Status": "Pass", "Notes": ""},
        {"Audit Section": "Summary", "Measure": "Graduation numerator", "Value": int(numerator.sum()), "Status": "Pass" if int(numerator.sum()) <= int(confirmed.sum()) else "Fail", "Notes": "Graduation numerator must not exceed confirmed evidence count."},
        {"Audit Section": "Summary", "Measure": "Graduation claims corrected to non-graduated outcome", "Value": int(corrected.sum()), "Status": "Review" if int(corrected.sum()) else "Pass", "Notes": "These rows claimed graduation somewhere but lacked confirmed graduation evidence."},
        {"Audit Section": "Summary", "Measure": "Still active students", "Value": int(active.sum()), "Status": "Pass", "Notes": ""},
        {"Audit Section": "Summary", "Measure": "Truly unknown / unresolved students", "Value": int(unknown.sum()), "Status": "Pass", "Notes": "Disappearance without confirmed graduation remains unknown."},
        {"Audit Section": "Summary", "Measure": "Resolved non-graduate exits", "Value": int(non_grad.sum()), "Status": "Pass", "Notes": ""},
        {"Audit Section": "Summary", "Measure": "Graduated without confirmed evidence", "Value": int((graduated & ~confirmed).sum()), "Status": "Pass" if int((graduated & ~confirmed).sum()) == 0 else "Fail", "Notes": "No student should be classified as Graduated without evidence."},
        {"Audit Section": "Summary", "Measure": "Graduation claim without evidence", "Value": int(without_evidence.sum()), "Status": "Review" if int(without_evidence.sum()) else "Pass", "Notes": "These claims are not counted as confirmed graduation."},
    ]

    if total_unique:
        full_population_rate = float(numerator.sum()) / float(total_unique)
        rows.append(
            {
                "Audit Section": "Warning",
                "Measure": "Full-population graduation rate sanity check",
                "Value": full_population_rate,
                "Status": "Review" if total_unique >= 20 and full_population_rate > 0.95 else "Pass",
                "Notes": "Review if the full-population graduation rate is near 100%.",
            }
        )

    if "chapter" in summary.columns:
        chapter_rates = []
        for chapter, frame in summary.groupby("chapter", dropna=False):
            if not clean_text(chapter):
                continue
            resolved = boolish_series(frame.get("is_resolved_outcome", pd.Series(False, index=frame.index)))
            resolved_n = int(resolved.sum())
            if resolved_n < 5:
                continue
            chapter_grad = int(boolish_series(frame.get("is_graduated", pd.Series(False, index=frame.index))).sum())
            chapter_rates.append(chapter_grad / resolved_n if resolved_n else 0)
        near_perfect = sum(1 for rate in chapter_rates if rate > 0.95)
        rows.append(
            {
                "Audit Section": "Warning",
                "Measure": "Chapters with resolved graduation rate above 95%",
                "Value": near_perfect,
                "Status": "Review" if near_perfect > max(3, int(len(chapter_rates) * 0.25)) else "Pass",
                "Notes": "A high count may indicate graduation overclassification.",
            }
        )

    if "outcome_evidence_source" in summary.columns:
        source_counts = (
            summary.loc[graduated, "outcome_evidence_source"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Missing evidence source")
            .value_counts()
        )
        for source, count in source_counts.items():
            rows.append(
                {
                    "Audit Section": "Graduation Evidence Source",
                    "Measure": source,
                    "Value": int(count),
                    "Status": "Pass" if source != "Missing evidence source" else "Fail",
                    "Notes": "Counts students classified as Graduated by evidence source.",
                }
            )

    if corrected.any():
        for row in summary.loc[corrected, ["student_id", "student_name", "latest_outcome_bucket", "outcome_evidence_source", "graduation_status_correction_reason"]].itertuples(index=False):
            rows.append(
                {
                    "Audit Section": "Corrected Student",
                    "Measure": clean_text(getattr(row, "student_id", "")),
                    "Value": 1,
                    "Status": "Review",
                    "Notes": f"{clean_text(getattr(row, 'student_name', ''))}: {clean_text(getattr(row, 'graduation_status_correction_reason', ''))}",
                }
            )

    return pd.DataFrame(rows, columns=columns)


def build_qa_checks(
    roster: pd.DataFrame,
    academic: pd.DataFrame,
    master: pd.DataFrame,
    summary: pd.DataFrame,
    settings: Dict[str, object],
    issue_frames: Dict[str, pd.DataFrame],
    membership_reference_validation: pd.DataFrame,
    new_member_reference_validation: pd.DataFrame,
    gpa_reference_validation: pd.DataFrame,
    gpa_benchmark_validation: pd.DataFrame,
    reference_inventory: pd.DataFrame,
    reference_unclassified_rows: pd.DataFrame,
    retention_reference: pd.DataFrame,
) -> pd.DataFrame:
    secondary_orgs = secondary_organization_set(settings)
    primary_roster = roster.loc[~roster["chapter"].map(lambda value: normalize_chapter_name(clean_text(value)) in secondary_orgs)].copy() if not roster.empty else roster
    resolved_mask = boolish_series(summary.get("is_resolved_outcome", pd.Series(False, index=summary.index)))
    active_mask = boolish_series(summary.get("is_active_outcome", pd.Series(False, index=summary.index)))
    unknown_mask = boolish_series(summary.get("is_unknown_outcome", pd.Series(False, index=summary.index)))
    rows: List[dict] = [
        {"Check Group": "Schema", "Check": "Authoritative tables built", "Status": "Pass", "Value": 6, "Notes": "roster_term, academic_term, master_longitudinal, student_summary, cohort_metrics, qa_checks"},
        {"Check Group": "Duplicates", "Check": "Roster duplicate student-term rows (primary chapters only)", "Status": "Pass" if primary_roster.duplicated(subset=["student_id", "term_code"]).sum() == 0 else "Fail", "Value": int(primary_roster.duplicated(subset=["student_id", "term_code"]).sum()), "Notes": "Secondary organizations are ignored for this duplicate check."},
        {"Check Group": "Duplicates", "Check": "Academic duplicate student-term rows", "Status": "Pass" if academic.duplicated(subset=["student_id", "term_code"]).sum() == 0 else "Fail", "Value": int(academic.duplicated(subset=["student_id", "term_code"]).sum()), "Notes": ""},
        {"Check Group": "Duplicates", "Check": "Master duplicate student-term rows", "Status": "Pass" if master.duplicated(subset=["student_id", "term_code"]).sum() == 0 else "Fail", "Value": int(master.duplicated(subset=["student_id", "term_code"]).sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Students with roster but no academics", "Status": "Pass", "Value": int(summary.loc[summary["first_observed_org_term_code"].ne("") & summary["first_observed_academic_term_code"].eq("")].shape[0]), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Students with academics but no roster", "Status": "Pass", "Value": int(summary.loc[summary["first_observed_org_term_code"].eq("") & summary["first_observed_academic_term_code"].ne("")].shape[0]), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Resolved outcomes", "Status": "Pass", "Value": int(resolved_mask.sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Still active outcomes", "Status": "Pass", "Value": int(active_mask.sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Truly unknown / unresolved outcomes", "Status": "Pass", "Value": int(unknown_mask.sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Other / unmapped outcomes", "Status": "Pass", "Value": int((~resolved_mask & ~active_mask & ~unknown_mask).sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Secondary organization roster rows", "Status": "Pass", "Value": int(roster["chapter"].map(lambda value: normalize_chapter_name(clean_text(value)) in secondary_orgs).sum()) if not roster.empty else 0, "Notes": "These rows are preserved but ignored when choosing the primary chapter for analytics."},
        {"Check Group": "Coverage", "Check": "Unresolved chapter assignments", "Status": "Pass", "Value": int(summary["chapter_assignment_source"].fillna("").astype(str).eq("unresolved").sum()), "Notes": ""},
        {"Check Group": "Coverage", "Check": "Manual chapter overrides applied", "Status": "Pass", "Value": int(summary["chapter_assignment_source"].fillna("").astype(str).eq("manual_override").sum()), "Notes": ""},
    ]
    rows.extend(build_spring_coverage_checks(roster, "Roster"))
    rows.extend(build_spring_coverage_checks(academic, "Academic"))
    rows.extend(build_measurable_window_checks(summary))
    graduation_audit = build_graduation_status_audit(summary)
    if not graduation_audit.empty:
        for measure in [
            "Graduation numerator",
            "Graduated without confirmed evidence",
            "Graduation claim without evidence",
            "Full-population graduation rate sanity check",
            "Chapters with resolved graduation rate above 95%",
        ]:
            audit_row = graduation_audit.loc[graduation_audit["Measure"].eq(measure)]
            if audit_row.empty:
                continue
            first = audit_row.iloc[0]
            rows.append(
                {
                    "Check Group": "Graduation Evidence",
                    "Check": clean_text(first["Measure"]),
                    "Status": clean_text(first["Status"]),
                    "Value": first["Value"],
                    "Notes": clean_text(first["Notes"]),
                }
            )
    if reference_inventory.empty:
        rows.append(
            {
                "Check Group": "Reference Validation",
                "Check": "Reference inventory rows loaded",
                "Status": "Review",
                "Value": 0,
                "Notes": "No numeric reference rows were cataloged from the reference workbooks.",
            }
        )
    else:
        rows.extend(
            [
                {
                    "Check Group": "Reference Validation",
                    "Check": "Reference inventory rows loaded",
                    "Status": "Pass",
                    "Value": int(len(reference_inventory)),
                    "Notes": "",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Reference inventory unclassified rows",
                    "Status": "Pass" if reference_unclassified_rows.empty else "Review",
                    "Value": int(len(reference_unclassified_rows)),
                    "Notes": "" if reference_unclassified_rows.empty else "See reference_unclassified_rows.csv for rows that need manual interpretation.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Retention reference rows cataloged",
                    "Status": "Pass" if not retention_reference.empty else "Review",
                    "Value": int(len(retention_reference)),
                    "Notes": "" if not retention_reference.empty else "No retention reference rows were cataloged.",
                },
            ]
        )
    if membership_reference_validation.empty:
        rows.append(
            {
                "Check Group": "Reference Validation",
                "Check": "Supplemental membership reference rows loaded",
                "Status": "Review",
                "Value": 0,
                "Notes": "No supplemental membership reference workbook rows were loaded.",
            }
        )
    else:
        match_count = int(membership_reference_validation["comparison_status"].eq("Match").sum())
        mismatch_count = int(membership_reference_validation["comparison_status"].eq("Mismatch").sum())
        reference_only_count = int(membership_reference_validation["comparison_status"].eq("Reference Only").sum())
        rows.extend(
            [
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental membership reference rows loaded",
                    "Status": "Pass",
                    "Value": int(len(membership_reference_validation)),
                    "Notes": "",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental membership count matches",
                    "Status": "Pass" if mismatch_count == 0 and reference_only_count == 0 else "Review",
                    "Value": match_count,
                    "Notes": "" if mismatch_count == 0 and reference_only_count == 0 else "See membership_reference_validation.csv for non-matching rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental membership count mismatches",
                    "Status": "Pass" if mismatch_count == 0 else "Review",
                    "Value": mismatch_count,
                    "Notes": "" if mismatch_count == 0 else "Reference and pipeline counts differ for these chapter-term rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental reference-only chapter-term rows",
                    "Status": "Pass" if reference_only_count == 0 else "Review",
                    "Value": reference_only_count,
                    "Notes": "" if reference_only_count == 0 else "Reference workbook contains chapter-term counts not present in the rebuilt roster data.",
                },
            ]
        )
    if new_member_reference_validation.empty:
        rows.append(
            {
                "Check Group": "Reference Validation",
                "Check": "Supplemental new-member reference rows loaded",
                "Status": "Review",
                "Value": 0,
                "Notes": "No supplemental new-member reference rows were loaded.",
            }
        )
    else:
        new_member_match_count = int(new_member_reference_validation["comparison_status"].eq("Match").sum())
        new_member_mismatch_count = int(new_member_reference_validation["comparison_status"].eq("Mismatch").sum())
        new_member_reference_only_count = int(new_member_reference_validation["comparison_status"].eq("Reference Only").sum())
        rows.extend(
            [
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental new-member reference rows loaded",
                    "Status": "Pass",
                    "Value": int(len(new_member_reference_validation)),
                    "Notes": "",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental new-member count matches",
                    "Status": "Pass" if new_member_mismatch_count == 0 and new_member_reference_only_count == 0 else "Review",
                    "Value": new_member_match_count,
                    "Notes": "" if new_member_mismatch_count == 0 and new_member_reference_only_count == 0 else "See new_member_reference_validation.csv for non-matching rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental new-member count mismatches",
                    "Status": "Pass" if new_member_mismatch_count == 0 else "Review",
                    "Value": new_member_mismatch_count,
                    "Notes": "" if new_member_mismatch_count == 0 else "Reference and pipeline new-member counts differ for these chapter-term rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental new-member reference-only rows",
                    "Status": "Pass" if new_member_reference_only_count == 0 else "Review",
                    "Value": new_member_reference_only_count,
                    "Notes": "" if new_member_reference_only_count == 0 else "Reference workbook contains chapter-term new-member rows not present in the rebuilt roster data.",
                },
            ]
        )
    if gpa_reference_validation.empty:
        rows.append(
            {
                "Check Group": "Reference Validation",
                "Check": "Supplemental GPA reference rows loaded",
                "Status": "Review",
                "Value": 0,
                "Notes": "No supplemental GPA reference workbook rows were loaded.",
            }
        )
    else:
        gpa_match_count = int(gpa_reference_validation["comparison_status"].eq("Match").sum())
        gpa_mismatch_count = int(gpa_reference_validation["comparison_status"].eq("Mismatch").sum())
        gpa_reference_only_count = int(gpa_reference_validation["comparison_status"].eq("Reference Only").sum())
        rows.extend(
            [
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA reference rows loaded",
                    "Status": "Pass",
                    "Value": int(len(gpa_reference_validation)),
                    "Notes": "",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental chapter GPA matches",
                    "Status": "Pass" if gpa_mismatch_count == 0 and gpa_reference_only_count == 0 else "Review",
                    "Value": gpa_match_count,
                    "Notes": "" if gpa_mismatch_count == 0 and gpa_reference_only_count == 0 else "See gpa_reference_validation.csv for non-matching rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental chapter GPA mismatches",
                    "Status": "Pass" if gpa_mismatch_count == 0 else "Review",
                    "Value": gpa_mismatch_count,
                    "Notes": "" if gpa_mismatch_count == 0 else "Reference and pipeline chapter GPAs differ for these chapter-term rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA reference-only chapter-term rows",
                    "Status": "Pass" if gpa_reference_only_count == 0 else "Review",
                    "Value": gpa_reference_only_count,
                    "Notes": "" if gpa_reference_only_count == 0 else "Reference workbook contains chapter-term GPA rows not present in the rebuilt pipeline data.",
                },
            ]
        )
    if gpa_benchmark_validation.empty:
        rows.append(
            {
                "Check Group": "Reference Validation",
                "Check": "Supplemental GPA benchmark rows loaded",
                "Status": "Review",
                "Value": 0,
                "Notes": "No supplemental GPA benchmark workbook rows were loaded.",
            }
        )
    else:
        benchmark_match_count = int(gpa_benchmark_validation["comparison_status"].eq("Match").sum())
        benchmark_mismatch_count = int(gpa_benchmark_validation["comparison_status"].eq("Mismatch").sum())
        benchmark_reference_only_count = int(gpa_benchmark_validation["comparison_status"].eq("Reference Only").sum())
        rows.extend(
            [
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA benchmark rows loaded",
                    "Status": "Pass",
                    "Value": int(len(gpa_benchmark_validation)),
                    "Notes": "",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA benchmark matches",
                    "Status": "Pass" if benchmark_mismatch_count == 0 and benchmark_reference_only_count == 0 else "Review",
                    "Value": benchmark_match_count,
                    "Notes": "" if benchmark_mismatch_count == 0 and benchmark_reference_only_count == 0 else "See gpa_benchmark_validation.csv for non-matching rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA benchmark mismatches",
                    "Status": "Pass" if benchmark_mismatch_count == 0 else "Review",
                    "Value": benchmark_mismatch_count,
                    "Notes": "" if benchmark_mismatch_count == 0 else "Reference and pipeline benchmark GPAs differ for these term rows.",
                },
                {
                    "Check Group": "Reference Validation",
                    "Check": "Supplemental GPA benchmark reference-only rows",
                    "Status": "Pass" if benchmark_reference_only_count == 0 else "Review",
                    "Value": benchmark_reference_only_count,
                    "Notes": "" if benchmark_reference_only_count == 0 else "Benchmark workbook contains rows the rebuilt pipeline cannot compute directly, such as TXST-wide averages.",
                },
            ]
        )
    for name, frame in issue_frames.items():
        rows.append(
            {
                "Check Group": "Exceptions",
                "Check": name,
                "Status": "Pass" if frame.empty else "Review",
                "Value": int(len(frame)),
                "Notes": "Manual review required." if not frame.empty else "",
            }
        )
    return ensure_columns(pd.DataFrame(rows), load_schema()["tables"]["qa_checks"])


def write_frame(path: Path, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False)


def build_canonical_pipeline(
    roster_root: Path,
    roster_inbox: Path,
    academic_root: Path,
    transcript_text_root: Path,
    graduation_root: Path,
    reference_data_root: Path,
    membership_reference_root: Path,
    gpa_reference_root: Path,
    gpa_benchmark_root: Path,
    output_root: Path,
    cache_root: Path,
    refresh_source_cache: bool = False,
) -> CanonicalBuildResult:
    pipeline_started = perf_counter()
    performance: List[PerformanceStage] = []
    schema = load_schema()
    settings = load_settings()
    chapter_mapping = load_chapter_mapping()
    ensure_manual_chapter_assignment_template()
    transcript_text_root.mkdir(parents=True, exist_ok=True)
    ensure_transcript_text_manifest_template()
    manual_chapter_assignments = load_manual_chapter_assignments()
    transcript_text_manifest = load_transcript_text_manifest()
    config_manifest = optional_files_manifest(
        [
            APP_SETTINGS_PATH,
            DEFAULT_CHAPTER_GROUPS_PATH,
            EXAMPLE_CHAPTER_GROUPS_PATH,
            MANUAL_CHAPTER_ASSIGNMENTS_PATH,
            TRANSCRIPT_TEXT_MANIFEST_PATH,
            SCHEMA_PATH,
        ]
    )

    discovery_started = perf_counter()
    roster_source_paths = roster_files([roster_root, roster_inbox])
    academic_source_paths = academic_files(academic_root)
    snapshot_source_paths = snapshot_files(academic_root)
    transcript_text_source_paths = transcript_text_files(transcript_text_root)
    graduation_source_paths = graduation_files(graduation_root)
    reference_source_paths = []
    for reference_root in [reference_data_root, membership_reference_root, gpa_reference_root, gpa_benchmark_root]:
        reference_source_paths.extend(list_source_files(reference_root, TABULAR_SOURCE_EXTENSIONS))
    append_stage(
        performance,
        "source_discovery",
        discovery_started,
        "rebuilt",
        {},
        notes=f"roster_files={len(roster_source_paths)}, academic_files={len(academic_source_paths)}, snapshot_files={len(snapshot_source_paths)}, transcript_text_files={len(transcript_text_source_paths)}, graduation_files={len(graduation_source_paths)}, reference_files={len(reference_source_paths)}",
    )

    roster_manifest = {
        "files": files_manifest(roster_source_paths),
        "loader_token": source_cache_token(
            [
                load_roster_term_table,
                pdf_table_rows,
                find_header_row,
                find_header_row_in_rows,
                find_status_column,
                find_status_column_in_rows,
                detect_inline_chapter_label,
                get_cell,
                infer_chapter,
                normalize_banner_id,
                normalize_chapter_name,
                normalize_status,
                source_file_label,
                roster_file_version_details,
                roster_file_month_details,
                chapter_assignment_details,
                roster_status_bucket,
            ]
        ),
    }
    academic_manifest = {
        "files": files_manifest(academic_source_paths),
        "loader_token": source_cache_token(
            [
                load_academic_term_table,
                map_grade_headers,
                normalize_banner_id,
                standing_bucket,
            ]
        ),
    }
    snapshot_manifest = {
        "files": files_manifest(snapshot_source_paths),
        "loader_token": source_cache_token(
            [
                load_snapshot_table,
                choose_best_snapshot_rows,
                normalize_banner_id,
            ]
        ),
    }
    graduation_manifest = {
        "files": files_manifest(graduation_source_paths),
        "loader_token": source_cache_token(
            [
                load_graduation_table,
                normalize_banner_id,
            ]
        ),
    }
    transcript_text_manifest_payload = {
        "files": files_manifest(transcript_text_source_paths),
        "manifest_file": optional_files_manifest([TRANSCRIPT_TEXT_MANIFEST_PATH]),
        "loader_token": source_cache_token(
            [
                load_transcript_text_manifest,
                resolve_transcript_identity,
                parse_transcript_course_line,
                load_transcript_text_tables,
                transcript_terms_to_academic_rows,
                build_transcript_text_cache_bundle,
            ]
        ),
    }
    reference_manifest = {
        "files": files_manifest(reference_source_paths),
        "loader_token": source_cache_token(
            [
                load_reference_inventory_table,
                detect_membership_reference_header_row,
                classify_reference_row,
            ]
        ),
    }

    source_stage_started = perf_counter()
    (roster_term, roster_load_issues), roster_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="roster_sources",
        manifest=roster_manifest,
        builder=lambda: load_roster_term_table([roster_root, roster_inbox]),
        file_names=["roster_term.csv", "roster_load_issues.csv"],
        refresh=refresh_source_cache,
    )
    (academic_term, academic_load_issues), academic_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="academic_sources",
        manifest=academic_manifest,
        builder=lambda: load_academic_term_table(academic_root),
        file_names=["academic_term.csv", "academic_load_issues.csv"],
        refresh=refresh_source_cache,
    )
    (snapshot,), snapshot_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="snapshot_sources",
        manifest=snapshot_manifest,
        builder=lambda: (load_snapshot_table(academic_root),),
        file_names=["snapshot.csv"],
        refresh=refresh_source_cache,
    )
    (graduation, graduation_load_issues), graduation_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="graduation_sources",
        manifest=graduation_manifest,
        builder=lambda: load_graduation_table(graduation_root),
        file_names=["graduation.csv", "graduation_load_issues.csv"],
        refresh=refresh_source_cache,
    )
    (
        transcript_term_summary,
        transcript_course_detail,
        transcript_parse_audit,
        transcript_parse_issues,
        transcript_academic_term,
    ), transcript_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="transcript_text_sources",
        manifest=transcript_text_manifest_payload,
        builder=lambda: build_transcript_text_cache_bundle(transcript_text_root, transcript_text_manifest),
        file_names=[
            "transcript_term_summary.csv",
            "transcript_course_detail.csv",
            "transcript_parse_audit.csv",
            "transcript_parse_issues.csv",
            "transcript_academic_term.csv",
        ],
        refresh=refresh_source_cache,
    )
    (reference_inventory, reference_inventory_issues), reference_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="reference_sources",
        manifest=reference_manifest,
        builder=lambda: load_reference_inventory_table([reference_data_root, membership_reference_root, gpa_reference_root, gpa_benchmark_root]),
        file_names=["reference_inventory.csv", "reference_inventory_issues.csv"],
        refresh=refresh_source_cache,
    )
    if not transcript_academic_term.empty:
        academic_term = pd.concat([academic_term, transcript_academic_term], ignore_index=True)
        academic_term = ensure_columns(academic_term, schema["tables"]["academic_term"])
    print(f"Source cache - roster: {'hit' if roster_cache_hit else 'rebuilt'}")
    print(f"Source cache - academic: {'hit' if academic_cache_hit else 'rebuilt'}")
    print(f"Source cache - snapshot: {'hit' if snapshot_cache_hit else 'rebuilt'}")
    print(f"Source cache - transcript text: {'hit' if transcript_cache_hit else 'rebuilt'}")
    print(f"Source cache - graduation: {'hit' if graduation_cache_hit else 'rebuilt'}")
    print(f"Source cache - reference: {'hit' if reference_cache_hit else 'rebuilt'}")
    append_stage(
        performance,
        "source_ingest_and_normalization",
        source_stage_started,
        f"roster={'hit' if roster_cache_hit else 'rebuilt'}, academic={'hit' if academic_cache_hit else 'rebuilt'}, snapshot={'hit' if snapshot_cache_hit else 'rebuilt'}, transcript_text={'hit' if transcript_cache_hit else 'rebuilt'}, graduation={'hit' if graduation_cache_hit else 'rebuilt'}, reference={'hit' if reference_cache_hit else 'rebuilt'}",
        {
            "roster_term": roster_term,
            "academic_term": academic_term,
            "snapshot": snapshot,
            "transcript_term_summary": transcript_term_summary,
            "graduation": graduation,
            "reference_inventory": reference_inventory,
        },
    )

    reference_stage_started = perf_counter()
    (
        membership_reference,
        gpa_reference,
        gpa_benchmark_reference,
        new_member_reference,
        retention_reference,
        reference_unclassified_rows,
    ), reference_derivatives_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="reference_derivatives",
        manifest={
            "reference_manifest": reference_manifest,
            "config_manifest": config_manifest,
            "loader_token": source_cache_token([build_reference_derivatives, build_reference_subset, ensure_columns]),
        },
        builder=lambda: build_reference_derivatives(reference_inventory),
        file_names=[
            "membership_reference.csv",
            "gpa_reference.csv",
            "gpa_benchmark_reference.csv",
            "new_member_reference.csv",
            "retention_reference.csv",
            "reference_unclassified_rows.csv",
        ],
        refresh=refresh_source_cache,
    )
    append_stage(
        performance,
        "reference_derivatives",
        reference_stage_started,
        "hit" if reference_derivatives_cache_hit else "rebuilt",
        {
            "membership_reference": membership_reference,
            "gpa_reference": gpa_reference,
            "gpa_benchmark_reference": gpa_benchmark_reference,
            "new_member_reference": new_member_reference,
            "retention_reference": retention_reference,
            "reference_unclassified_rows": reference_unclassified_rows,
        },
    )

    prepared_stage_started = perf_counter()
    (
        roster_term,
        academic_term,
        identity_exceptions,
        prepared_term_exceptions,
        status_exceptions,
        roster_conflicts,
    ), prepared_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="prepared_sources",
        manifest={
            "roster_manifest": roster_manifest,
            "academic_manifest": academic_manifest,
            "transcript_text_manifest": transcript_text_manifest_payload,
            "snapshot_manifest": snapshot_manifest,
            "graduation_manifest": graduation_manifest,
            "config_manifest": config_manifest,
            "loader_token": source_cache_token(
                [
                    prepare_canonical_sources,
                    build_identity_maps,
                    resolve_missing_ids,
                    resolve_missing_roster_chapters,
                    apply_manual_chapter_assignments,
                    dedupe_table,
                    resolve_roster_conflicts,
                    attach_org_entry_terms,
                    build_status_exceptions,
                ]
            ),
        },
        builder=lambda: prepare_canonical_sources(
            roster_term,
            academic_term,
            snapshot,
            graduation,
            settings,
            manual_chapter_assignments,
        ),
        file_names=[
            "roster_term_prepared.csv",
            "academic_term_prepared.csv",
            "identity_exceptions.csv",
            "term_exceptions.csv",
            "status_exceptions.csv",
            "chapter_conflicts.csv",
        ],
        refresh=refresh_source_cache,
    )
    append_stage(
        performance,
        "identity_chapter_and_dedup_resolution",
        prepared_stage_started,
        "hit" if prepared_cache_hit else "rebuilt",
        {
            "roster_term": roster_term,
            "academic_term": academic_term,
            "identity_exceptions": identity_exceptions,
            "term_exceptions": prepared_term_exceptions,
            "status_exceptions": status_exceptions,
            "chapter_conflicts": roster_conflicts,
        },
    )

    core_stage_started = perf_counter()
    (
        master_longitudinal,
        student_summary,
        summary_qa,
        outcome_issues,
        unresolved_chapter_review,
    ), canonical_core_cache_hit = load_or_build_cached_frames(
        cache_root=cache_root,
        cache_name="canonical_core",
        manifest={
            "roster_manifest": roster_manifest,
            "academic_manifest": academic_manifest,
            "transcript_text_manifest": transcript_text_manifest_payload,
            "snapshot_manifest": snapshot_manifest,
            "graduation_manifest": graduation_manifest,
            "config_manifest": config_manifest,
            "prepared_loader_token": source_cache_token(
                [
                    prepare_canonical_sources,
                    build_identity_maps,
                    resolve_missing_ids,
                    resolve_missing_roster_chapters,
                    apply_manual_chapter_assignments,
                    dedupe_table,
                    resolve_roster_conflicts,
                    attach_org_entry_terms,
                ]
            ),
            "loader_token": source_cache_token(
                [
                    build_canonical_core_tables,
                    build_master_longitudinal,
                    choose_preferred_roster_rows,
                    build_student_summary,
                    attach_snapshot_fields,
                    build_current_active_fields,
                    build_outcome_resolution_fields,
                    build_unresolved_chapter_review,
                    enrich_master_longitudinal_with_summary,
                ]
            ),
        },
        builder=lambda: build_canonical_core_tables(
            roster_term,
            academic_term,
            snapshot,
            graduation,
            settings,
            chapter_mapping,
        ),
        file_names=[
            "master_longitudinal.csv",
            "student_summary.csv",
            "summary_qa.csv",
            "outcome_issues.csv",
            "unresolved_chapter_review.csv",
        ],
        refresh=refresh_source_cache,
    )
    append_stage(
        performance,
        "canonical_core_build",
        core_stage_started,
        "hit" if canonical_core_cache_hit else "rebuilt",
        {
            "master_longitudinal": master_longitudinal,
            "student_summary": student_summary,
            "summary_qa": summary_qa,
            "outcome_issues": outcome_issues,
            "unresolved_chapter_review": unresolved_chapter_review,
        },
    )

    analytics_stage_started = perf_counter()
    cohort_metrics = build_cohort_metrics(student_summary)
    graduation_status_audit = build_graduation_status_audit(student_summary)
    membership_reference_validation = build_membership_reference_validation(roster_term, membership_reference)
    new_member_reference_validation = build_new_member_reference_validation(roster_term, new_member_reference)
    gpa_reference_validation = build_gpa_reference_validation(master_longitudinal, gpa_reference)
    gpa_benchmark_validation = build_gpa_benchmark_validation(master_longitudinal, gpa_benchmark_reference, chapter_mapping)
    empty_exception_frame = pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])
    term_exceptions = pd.concat(
        [frame for frame in [roster_load_issues, academic_load_issues, transcript_parse_issues, graduation_load_issues, reference_inventory_issues, prepared_term_exceptions] if not frame.empty],
        ignore_index=True,
    ) if any(not frame.empty for frame in [roster_load_issues, academic_load_issues, transcript_parse_issues, graduation_load_issues, reference_inventory_issues, prepared_term_exceptions]) else empty_exception_frame
    missing_evidence_cases = student_summary.loc[
        (~boolish_series(student_summary.get("is_resolved_outcome", pd.Series(False, index=student_summary.index))))
        & (~boolish_series(student_summary.get("is_active_outcome", pd.Series(False, index=student_summary.index)))),
        ["student_id", "student_name", "join_term", "outcome_resolution_group", "outcome_evidence_source"],
    ].rename(
        columns={
            "student_name": "details",
            "join_term": "term_code",
            "outcome_resolution_group": "exception_type",
            "outcome_evidence_source": "source_file",
        }
    ) if not student_summary.empty else pd.DataFrame(columns=["exception_type", "source_file", "student_id", "term_code", "details"])

    issue_frames = {
        "identity_exceptions": identity_exceptions,
        "term_exceptions": term_exceptions,
        "status_exceptions": status_exceptions,
        "chapter_conflicts": roster_conflicts,
        "outcome_exceptions": outcome_issues,
        "missing_evidence_cases": missing_evidence_cases,
        "unresolved_chapter_review": unresolved_chapter_review,
    }
    qa_checks = build_qa_checks(
        roster_term,
        academic_term,
        master_longitudinal,
        student_summary,
        settings,
        issue_frames,
        membership_reference_validation,
        new_member_reference_validation,
        gpa_reference_validation,
        gpa_benchmark_validation,
        reference_inventory,
        reference_unclassified_rows,
        retention_reference,
    )
    transcript_qa_rows = []
    if not transcript_parse_audit.empty:
        transcript_qa_rows.extend(
            [
                {
                    "Check Group": "Transcript Text",
                    "Check": "Transcript text files scanned",
                    "Status": "Pass",
                    "Value": int(len(transcript_parse_audit)),
                    "Notes": "",
                },
                {
                    "Check Group": "Transcript Text",
                    "Check": "Transcript text files with warnings",
                    "Status": "Review" if int(transcript_parse_audit["warning_count"].fillna(0).astype(int).gt(0).sum()) else "Pass",
                    "Value": int(transcript_parse_audit["warning_count"].fillna(0).astype(int).gt(0).sum()),
                    "Notes": "",
                },
                {
                    "Check Group": "Transcript Text",
                    "Check": "Transcript term rows parsed",
                    "Status": "Pass",
                    "Value": int(len(transcript_term_summary)),
                    "Notes": "",
                },
                {
                    "Check Group": "Transcript Text",
                    "Check": "Transcript course rows parsed",
                    "Status": "Pass",
                    "Value": int(len(transcript_course_detail)),
                    "Notes": "",
                },
            ]
        )
    if not summary_qa.empty:
        qa_checks = pd.concat([qa_checks, ensure_columns(summary_qa, QA_COLUMNS)], ignore_index=True)
    if transcript_qa_rows:
        qa_checks = pd.concat([qa_checks, ensure_columns(pd.DataFrame(transcript_qa_rows), QA_COLUMNS)], ignore_index=True)
    append_stage(
        performance,
        "metrics_and_qa",
        analytics_stage_started,
        "rebuilt",
        {
            "cohort_metrics": cohort_metrics,
            "graduation_status_audit": graduation_status_audit,
            "membership_reference_validation": membership_reference_validation,
            "new_member_reference_validation": new_member_reference_validation,
            "gpa_reference_validation": gpa_reference_validation,
            "gpa_benchmark_validation": gpa_benchmark_validation,
            "qa_checks": qa_checks,
        },
    )

    export_stage_started = perf_counter()
    timestamp = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    output_folder = output_root / timestamp
    output_folder.mkdir(parents=True, exist_ok=True)
    latest_folder = output_root / "latest"
    latest_folder.mkdir(parents=True, exist_ok=True)
    previous_performance_report = latest_folder / "performance_report.json"

    files = {
        "roster_term": output_folder / "roster_term.csv",
        "academic_term": output_folder / "academic_term.csv",
        "master_longitudinal": output_folder / "master_longitudinal.csv",
        "student_summary": output_folder / "student_summary.csv",
        "cohort_metrics": output_folder / "cohort_metrics.csv",
        "qa_checks": output_folder / "qa_checks.csv",
        "graduation_status_audit": output_folder / "graduation_status_audit.csv",
        "reference_inventory": output_folder / "reference_inventory.csv",
        "reference_unclassified_rows": output_folder / "reference_unclassified_rows.csv",
        "membership_reference_counts": output_folder / "membership_reference_counts.csv",
        "membership_reference_validation": output_folder / "membership_reference_validation.csv",
        "new_member_reference_values": output_folder / "new_member_reference_values.csv",
        "new_member_reference_validation": output_folder / "new_member_reference_validation.csv",
        "gpa_reference_values": output_folder / "gpa_reference_values.csv",
        "gpa_reference_validation": output_folder / "gpa_reference_validation.csv",
        "gpa_benchmark_reference_values": output_folder / "gpa_benchmark_reference_values.csv",
        "gpa_benchmark_validation": output_folder / "gpa_benchmark_validation.csv",
        "retention_reference_values": output_folder / "retention_reference_values.csv",
        "transcript_term_summary": output_folder / "transcript_term_summary.csv",
        "transcript_course_detail": output_folder / "transcript_course_detail.csv",
        "transcript_parse_audit": output_folder / "transcript_parse_audit.csv",
        "transcript_parse_issues": output_folder / "transcript_parse_issues.csv",
        "schema": output_folder / "canonical_schema.json",
        "identity_exceptions": output_folder / "identity_exceptions.csv",
        "term_exceptions": output_folder / "term_exceptions.csv",
        "status_exceptions": output_folder / "status_exceptions.csv",
        "chapter_conflicts": output_folder / "chapter_conflicts.csv",
        "outcome_exceptions": output_folder / "outcome_exceptions.csv",
        "missing_evidence_cases": output_folder / "missing_evidence_cases.csv",
        "unresolved_chapter_review": output_folder / "unresolved_chapter_review.csv",
    }

    write_frame(files["roster_term"], roster_term)
    write_frame(files["academic_term"], academic_term)
    write_frame(files["master_longitudinal"], ensure_columns(master_longitudinal, schema["tables"]["master_longitudinal"]))
    write_frame(files["student_summary"], student_summary)
    write_frame(files["cohort_metrics"], cohort_metrics)
    write_frame(files["qa_checks"], qa_checks)
    write_frame(files["graduation_status_audit"], graduation_status_audit)
    write_frame(files["reference_inventory"], reference_inventory)
    write_frame(files["reference_unclassified_rows"], reference_unclassified_rows)
    write_frame(files["membership_reference_counts"], membership_reference)
    write_frame(files["membership_reference_validation"], membership_reference_validation)
    write_frame(files["new_member_reference_values"], new_member_reference)
    write_frame(files["new_member_reference_validation"], new_member_reference_validation)
    write_frame(files["gpa_reference_values"], gpa_reference)
    write_frame(files["gpa_reference_validation"], gpa_reference_validation)
    write_frame(files["gpa_benchmark_reference_values"], gpa_benchmark_reference)
    write_frame(files["gpa_benchmark_validation"], gpa_benchmark_validation)
    write_frame(files["retention_reference_values"], retention_reference)
    write_frame(files["transcript_term_summary"], transcript_term_summary)
    write_frame(files["transcript_course_detail"], transcript_course_detail)
    write_frame(files["transcript_parse_audit"], transcript_parse_audit)
    write_frame(files["transcript_parse_issues"], transcript_parse_issues)
    write_frame(files["unresolved_chapter_review"], unresolved_chapter_review)
    files["schema"].write_text(json.dumps(schema, indent=2), encoding="utf-8")
    for key in ["identity_exceptions", "term_exceptions", "status_exceptions", "chapter_conflicts", "outcome_exceptions", "missing_evidence_cases"]:
        write_frame(files[key], ensure_columns(issue_frames.get(key, pd.DataFrame()), empty_exception_frame.columns))

    for key, path in files.items():
        target = latest_folder / path.name
        shutil.copyfile(path, target)

    append_stage(
        performance,
        "export_write",
        export_stage_started,
        "rebuilt",
        {
            "roster_term": roster_term,
            "academic_term": academic_term,
            "master_longitudinal": master_longitudinal,
            "student_summary": student_summary,
            "cohort_metrics": cohort_metrics,
            "qa_checks": qa_checks,
        },
        notes=f"output_folder={output_folder.name}",
    )
    performance_paths = write_performance_report(
        performance=performance,
        output_folder=output_folder,
        latest_folder=latest_folder,
        previous_report_path=previous_performance_report,
    )
    files.update(performance_paths)
    total_runtime = perf_counter() - pipeline_started
    print(f"Total runtime: {total_runtime:,.2f} seconds")
    for item in performance:
        print(f"Stage timing - {item.stage}: {item.seconds:,.2f}s [{item.cache_status}] {item.rows}")

    return CanonicalBuildResult(output_folder=output_folder, files=files)


def main() -> None:
    args = parse_args()
    result = build_canonical_pipeline(
        roster_root=Path(args.roster_root).expanduser().resolve(),
        roster_inbox=Path(args.roster_inbox).expanduser().resolve(),
        academic_root=Path(args.academic_root).expanduser().resolve(),
        transcript_text_root=Path(args.transcript_text_root).expanduser().resolve(),
        graduation_root=Path(args.graduation_root).expanduser().resolve(),
        reference_data_root=Path(args.reference_data_root).expanduser().resolve(),
        membership_reference_root=Path(args.membership_reference_root).expanduser().resolve(),
        gpa_reference_root=Path(args.gpa_reference_root).expanduser().resolve(),
        gpa_benchmark_root=Path(args.gpa_benchmark_root).expanduser().resolve(),
        output_root=Path(args.output_root).expanduser().resolve(),
        cache_root=Path(args.cache_root).expanduser().resolve(),
        refresh_source_cache=args.refresh_source_cache,
    )
    print(f"Canonical outputs written to: {result.output_folder}")
    for key, path in result.files.items():
        print(f"{key}: {path}")
    print(f"manual_chapter_assignments_template: {MANUAL_CHAPTER_ASSIGNMENTS_PATH}")


if __name__ == "__main__":
    main()
