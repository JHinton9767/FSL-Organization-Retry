from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.excel_utils import autosize_columns, safe_filename, safe_sheet_name
from src.build_master_roster import is_excluded_chapter
from src.canonical_bundle import DEFAULT_CANONICAL_ROOT, load_canonical_bundle
from src.shared_utils import adjusted_grad_rate as adjusted_graduation_rate, clean_text, coerce_numeric, mean_or_blank, simple_rate as rate, unique_non_blank_count, yes_mask


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "chapter_history"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)

YEAR_DETAIL_COLUMNS = [
    "Term",
    "Student ID",
    "Student Name",
    "Email",
    "Roster Present",
    "Academic Present",
    "Organization Entry Term",
    "Organization Entry Cohort",
    "Relative Term Index From Org Entry",
    "Roster Status Raw",
    "Roster Status Bucket",
    "Roster Position",
    "New Member Marker",
    "Latest Known Outcome Bucket",
    "Major",
    "Attempted Hours",
    "Passed Hours",
    "Total Cumulative Hours",
    "Academic Standing Raw",
    "Academic Standing Bucket",
    "Term GPA",
    "Institutional Cumulative GPA",
    "Overall Cumulative GPA",
]
PERCENT_HEADERS = {
    "Observed Graduation Rate (Resolved)",
    "Observed 4-Year Graduation Rate (Resolved)",
    "Observed 6-Year Graduation Rate (Resolved)",
    "Retained In Chapter To Next Term",
    "Retained In Chapter To Next Fall",
    "Stayed In School To Next Term",
    "Stayed In School To Next Fall",
    "Share Of Entry Students",
    "Good Standing Rate",
}
DECIMAL_HEADERS = {
    "Average First-Term GPA",
    "Average First-Year GPA",
    "Average Latest Cumulative GPA",
    "Average Term GPA",
    "Average Cumulative GPA",
    "Average Passed Hours",
}


@dataclass(frozen=True)
class ChapterBuildResult:
    output_folder: Path
    workbook_folder: Path
    index_workbook: Path
    index_csv: Path
    readme_path: Path
    chapters_written: int
    source_folder: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build one workbook per chapter from the canonical analytics bundle. "
            "Each chapter workbook contains a summary sheet plus year-by-year detail sheets."
        )
    )
    parser.add_argument("--canonical-root", default=str(DEFAULT_CANONICAL_ROOT))
    parser.add_argument("--canonical-folder", default="")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    return parser.parse_args()

def style_row_as_header(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def write_section_title(ws, row_idx: int, title: str, note: str = "") -> int:
    ws.cell(row=row_idx, column=1, value=title).font = SECTION_FONT
    if note:
        ws.cell(row=row_idx + 1, column=1, value=note)
        return row_idx + 2
    return row_idx + 1


def format_table_columns(ws, header_row: int, start_data_row: int, end_data_row: int) -> None:
    if end_data_row < start_data_row:
        return
    headers = [clean_text(cell.value) for cell in ws[header_row]]
    for col_idx, header in enumerate(headers, start=1):
        if header in PERCENT_HEADERS:
            for row_idx in range(start_data_row, end_data_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"
        elif header in DECIMAL_HEADERS:
            for row_idx in range(start_data_row, end_data_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00"


def append_table(ws, start_row: int, title: str, note: str, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> int:
    row_idx = write_section_title(ws, start_row, title, note)
    header_row = row_idx
    ws.append(list(headers))
    style_row_as_header(ws, header_row)
    data_start = header_row + 1
    data_end = header_row
    for row in rows:
        ws.append(list(row))
        data_end = ws.max_row
    format_table_columns(ws, header_row, data_start, data_end)
    return ws.max_row + 2

def prepare_summary(summary: pd.DataFrame) -> pd.DataFrame:
    frame = summary.copy()
    for column in [
        "student_id",
        "student_name",
        "chapter",
        "initial_chapter",
        "latest_chapter",
        "org_entry_cohort",
        "latest_outcome_bucket",
        "first_academic_standing_bucket",
    ]:
        if column in frame.columns:
            frame[column] = frame[column].fillna("").astype(str).str.strip()
    return frame


def prepare_longitudinal(longitudinal: pd.DataFrame) -> pd.DataFrame:
    frame = longitudinal.copy()
    for column in [
        "chapter",
        "term_label",
        "student_id",
        "first_name",
        "last_name",
        "email",
    ]:
        if column in frame.columns:
            frame[column] = frame[column].fillna("").astype(str).str.strip()
    frame["student_name"] = (
        frame["first_name"].fillna("").astype(str).str.strip()
        + " "
        + frame["last_name"].fillna("").astype(str).str.strip()
    ).str.strip()
    return frame


def filter_to_chapter_entry(summary: pd.DataFrame, chapter: str) -> pd.DataFrame:
    target = clean_text(chapter).lower()
    return summary.loc[summary["initial_chapter"].fillna("").astype(str).str.strip().str.lower().eq(target)].copy()


def filter_to_chapter_rows(longitudinal: pd.DataFrame, chapter: str) -> pd.DataFrame:
    target = clean_text(chapter).lower()
    return longitudinal.loc[longitudinal["chapter"].fillna("").astype(str).str.strip().str.lower().eq(target)].copy()


def build_overview_rows(entry_students: pd.DataFrame, chapter_rows: pd.DataFrame) -> List[List[object]]:
    grad_rate, grad_n = adjusted_graduation_rate(entry_students, "graduated_eventual")
    grad4_rate, grad4_n = adjusted_graduation_rate(entry_students, "graduated_4yr", measurable_field="graduated_4yr_measurable")
    grad6_rate, grad6_n = adjusted_graduation_rate(entry_students, "graduated_6yr", measurable_field="graduated_6yr_measurable")
    org_next_term_rate, org_next_term_n = rate(entry_students, "retained_next_term", measurable_field="retained_next_term_measurable")
    org_next_fall_rate, org_next_fall_n = rate(entry_students, "retained_next_fall", measurable_field="retained_next_fall_measurable")
    acad_next_term_rate, acad_next_term_n = rate(entry_students, "continued_next_term", measurable_field="continued_next_term_measurable")
    acad_next_fall_rate, acad_next_fall_n = rate(entry_students, "continued_next_fall", measurable_field="continued_next_fall_measurable")
    return [
        ["Distinct students ever observed in this chapter", unique_non_blank_count(chapter_rows["student_id"]) if not chapter_rows.empty else 0, "Counts anyone observed in this chapter in any term."],
        ["Students with observed entry into this chapter", len(entry_students), "These students start their observed organization history in this chapter."],
        ["Organization-entry cohorts covered", unique_non_blank_count(entry_students["org_entry_cohort"]) if not entry_students.empty else 0, "Counts distinct observed entry cohorts tied to this chapter."],
        ["Observed eventual graduation rate from chapter entry (excluding unresolved outcomes)", grad_rate, f"Uses {grad_n} resolved students from this chapter's observed entry group."],
        ["Observed 4-year graduation rate from chapter entry (excluding unresolved outcomes)", grad4_rate, f"Uses {grad4_n} measurable and resolved students."],
        ["Observed 6-year graduation rate from chapter entry (excluding unresolved outcomes)", grad6_rate, f"Uses {grad6_n} measurable and resolved students."],
        ["Retained in chapter to next observed term", org_next_term_rate, f"Uses {org_next_term_n} students whose next-term follow-up is measurable."],
        ["Retained in chapter to next fall", org_next_fall_rate, f"Uses {org_next_fall_n} students whose next-fall follow-up is measurable."],
        ["Stayed in school to next observed term", acad_next_term_rate, f"Uses {acad_next_term_n} students whose next-term academic follow-up is measurable."],
        ["Stayed in school to next fall", acad_next_fall_rate, f"Uses {acad_next_fall_n} students whose next-fall academic follow-up is measurable."],
        ["Average first-term GPA after chapter entry", mean_or_blank(entry_students["first_term_gpa"]) if not entry_students.empty else "", "Averages the first academic term GPA observed after chapter entry."],
        ["Average first-year GPA after chapter entry", mean_or_blank(entry_students["first_year_avg_term_gpa"]) if not entry_students.empty else "", "Averages the first-year GPA across observed post-entry terms."],
        ["Average latest cumulative GPA", mean_or_blank(entry_students["average_cumulative_gpa"]) if not entry_students.empty else "", "Uses the latest available cumulative GPA field."],
    ]


def build_cohort_rows(entry_students: pd.DataFrame) -> List[List[object]]:
    rows: List[List[object]] = []
    cohorts = sorted({clean_text(value) for value in entry_students["org_entry_cohort"].tolist() if clean_text(value)})
    for cohort in cohorts:
        frame = entry_students.loc[entry_students["org_entry_cohort"].eq(cohort)].copy()
        grad_rate, grad_n = adjusted_graduation_rate(frame, "graduated_eventual")
        grad4_rate, _ = adjusted_graduation_rate(frame, "graduated_4yr", measurable_field="graduated_4yr_measurable")
        grad6_rate, _ = adjusted_graduation_rate(frame, "graduated_6yr", measurable_field="graduated_6yr_measurable")
        org_next_term_rate, _ = rate(frame, "retained_next_term", measurable_field="retained_next_term_measurable")
        org_next_fall_rate, _ = rate(frame, "retained_next_fall", measurable_field="retained_next_fall_measurable")
        acad_next_term_rate, _ = rate(frame, "continued_next_term", measurable_field="continued_next_term_measurable")
        acad_next_fall_rate, _ = rate(frame, "continued_next_fall", measurable_field="continued_next_fall_measurable")
        rows.append(
            [
                cohort,
                len(frame),
                grad_n,
                grad_rate,
                grad4_rate,
                grad6_rate,
                org_next_term_rate,
                org_next_fall_rate,
                acad_next_term_rate,
                acad_next_fall_rate,
                mean_or_blank(frame["first_term_gpa"]),
                mean_or_blank(frame["average_cumulative_gpa"]),
            ]
        )
    return rows


def build_outcome_rows(entry_students: pd.DataFrame) -> List[List[object]]:
    counts = Counter(clean_text(value) or "Unknown" for value in entry_students["latest_outcome_bucket"].tolist())
    total = sum(counts.values())
    rows: List[List[object]] = []
    for bucket in [
        "Graduated",
        "Suspended",
        "Transfer",
        "Dropped/Resigned/Revoked/Inactive",
        "No Further Observation",
        "Active/Unknown",
        "Unknown",
    ]:
        count = counts.get(bucket, 0)
        if count:
            rows.append([bucket, count, float(count) / float(total) if total else ""])
    return rows


def build_yearly_trend_rows(chapter_rows: pd.DataFrame) -> List[List[object]]:
    rows: List[List[object]] = []
    if chapter_rows.empty:
        return rows
    grouped = chapter_rows.loc[coerce_numeric(chapter_rows["observed_year"]).notna()].groupby("observed_year")
    for year_value, frame in sorted(grouped, key=lambda item: int(item[0])):
        academic_only = frame.loc[yes_mask(frame["academic_present"])]
        standing_known = academic_only.loc[academic_only["academic_standing_bucket"].fillna("").astype(str).str.strip().ne("")]
        good_standing_rate = ""
        if not standing_known.empty:
            good_standing_rate = float(standing_known["academic_standing_bucket"].eq("Good Standing").sum()) / float(len(standing_known))
        rows.append(
            [
                int(year_value),
                unique_non_blank_count(frame["student_id"]),
                int(yes_mask(frame["roster_present"]).sum()),
                int(yes_mask(frame["academic_present"]).sum()),
                mean_or_blank(frame["term_gpa"]),
                mean_or_blank(frame["cumulative_gpa"]),
                good_standing_rate,
                mean_or_blank(frame["earned_hours_term"]),
            ]
        )
    return rows


def build_relative_term_gpa_rows(chapter_rows: pd.DataFrame) -> List[List[object]]:
    rows: List[List[object]] = []
    if chapter_rows.empty:
        return rows
    eligible = chapter_rows.loc[
        yes_mask(chapter_rows["academic_present"])
        & coerce_numeric(chapter_rows["relative_term_index"]).notna()
        & coerce_numeric(chapter_rows["relative_term_index"]).ge(0)
    ].copy()
    grouped = eligible.groupby("relative_term_index")
    for relative_term, frame in sorted(grouped, key=lambda item: int(float(item[0]))):
        rows.append(
            [
                int(float(relative_term)),
                len(frame),
                unique_non_blank_count(frame["student_id"]),
                mean_or_blank(frame["term_gpa"]),
                mean_or_blank(frame["institutional_cumulative_gpa"]),
                mean_or_blank(frame["overall_cumulative_gpa"]),
            ]
        )
    return rows


def build_year_sheet_rows(frame: pd.DataFrame) -> List[List[object]]:
    sorted_frame = frame.sort_values(by=["observed_term_sort", "last_name", "first_name", "student_id"], ascending=[True, True, True, True], na_position="last")
    rows: List[List[object]] = []
    for _, row in sorted_frame.iterrows():
        rows.append(
            [
                row.get("term_label", ""),
                row.get("student_id", ""),
                row.get("student_name", ""),
                row.get("email", ""),
                row.get("roster_present", ""),
                row.get("academic_present", ""),
                row.get("join_term", ""),
                row.get("join_term", ""),
                row.get("relative_term_index", ""),
                row.get("org_status_raw", ""),
                row.get("org_status_bucket", ""),
                row.get("org_position_raw", ""),
                row.get("new_member_flag", ""),
                row.get("final_outcome_bucket", ""),
                row.get("major", ""),
                row.get("attempted_hours_term", ""),
                row.get("earned_hours_term", ""),
                row.get("total_cumulative_hours", ""),
                row.get("academic_standing_raw", ""),
                row.get("academic_standing_bucket", ""),
                row.get("term_gpa", ""),
                row.get("institutional_cumulative_gpa", ""),
                row.get("overall_cumulative_gpa", ""),
            ]
        )
    return rows


def create_index_row(chapter: str, entry_students: pd.DataFrame, chapter_rows: pd.DataFrame, workbook_path: Path) -> List[object]:
    grad_rate, _ = adjusted_graduation_rate(entry_students, "graduated_eventual")
    org_next_fall_rate, _ = rate(entry_students, "retained_next_fall", measurable_field="retained_next_fall_measurable")
    acad_next_fall_rate, _ = rate(entry_students, "continued_next_fall", measurable_field="continued_next_fall_measurable")
    return [
        chapter,
        workbook_path.name,
        unique_non_blank_count(chapter_rows["student_id"]) if not chapter_rows.empty else 0,
        len(entry_students),
        unique_non_blank_count(entry_students["org_entry_cohort"]) if not entry_students.empty else 0,
        grad_rate,
        org_next_fall_rate,
        acad_next_fall_rate,
        mean_or_blank(entry_students["first_term_gpa"]) if not entry_students.empty else "",
        mean_or_blank(entry_students["average_cumulative_gpa"]) if not entry_students.empty else "",
    ]


def write_summary_sheet(ws, chapter: str, entry_students: pd.DataFrame, chapter_rows: pd.DataFrame, source_folder: Path) -> None:
    ws.title = "Summary"
    ws["A1"] = f"{chapter} Chapter History"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "This workbook flips the yearly roster builder into one workbook per chapter. "
        "Summary rates are based on students whose observed organization entry begins in this chapter."
    )
    ws["A3"] = f"Canonical analytics source folder: {source_folder}"

    row_idx = 5
    row_idx = append_table(
        ws,
        row_idx,
        "Overview",
        "What this tells us: these are the headline counts and chapter-level rates built from observed organization entry and follow-up data.",
        ["Metric", "Value", "How To Read This"],
        build_overview_rows(entry_students, chapter_rows),
    )
    row_idx = append_table(
        ws,
        row_idx,
        "Entry Cohort Rates",
        "How to read this: each row is one observed chapter-entry cohort. Retention uses measurable follow-up windows only, and graduation excludes unresolved outcomes.",
        [
            "Organization Entry Cohort",
            "Students",
            "Resolved Outcome Students",
            "Observed Graduation Rate (Resolved)",
            "Observed 4-Year Graduation Rate (Resolved)",
            "Observed 6-Year Graduation Rate (Resolved)",
            "Retained In Chapter To Next Term",
            "Retained In Chapter To Next Fall",
            "Stayed In School To Next Term",
            "Stayed In School To Next Fall",
            "Average First-Term GPA",
            "Average Latest Cumulative GPA",
        ],
        build_cohort_rows(entry_students),
    )
    row_idx = append_table(
        ws,
        row_idx,
        "Latest Outcome Breakdown",
        "Why this matters: this shows the latest observed outcomes for students whose observed chapter entry starts here.",
        ["Latest Known Outcome", "Students", "Share Of Entry Students"],
        build_outcome_rows(entry_students),
    )
    row_idx = append_table(
        ws,
        row_idx,
        "GPA Trend After Chapter Entry",
        "What this tells us: this averages academic performance by relative term after observed chapter entry for students seen in this chapter.",
        [
            "Relative Term After Entry",
            "Academic Records",
            "Distinct Students",
            "Average Term GPA",
            "Average Institutional Cumulative GPA",
            "Average Overall Cumulative GPA",
        ],
        build_relative_term_gpa_rows(chapter_rows),
    )
    append_table(
        ws,
        row_idx,
        "Yearly Trend",
        "How to read this: this summarizes the chapter's observed student-term records by year, including GPA and standing patterns.",
        [
            "Year",
            "Distinct Students",
            "Roster Rows",
            "Academic Rows",
            "Average Term GPA",
            "Average Cumulative GPA",
            "Good Standing Rate",
            "Average Passed Hours",
        ],
        build_yearly_trend_rows(chapter_rows),
    )
    ws.freeze_panes = "A5"
    autosize_columns(ws)


def write_year_sheet(wb: Workbook, chapter: str, year_label: str, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=safe_sheet_name(year_label))
    ws["A1"] = f"{chapter} - {year_label}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "One row per observed student-term for this chapter and year. "
        "Rows include both roster and academic fields when available."
    )
    ws.append(list(YEAR_DETAIL_COLUMNS))
    style_row_as_header(ws, 3)
    for row in build_year_sheet_rows(frame):
        ws.append(row)
    ws.freeze_panes = "A4"
    autosize_columns(ws)


def write_chapter_workbook(chapter: str, entry_students: pd.DataFrame, chapter_rows: pd.DataFrame, output_path: Path, source_folder: Path) -> None:
    wb = Workbook()
    summary_ws = wb.active
    write_summary_sheet(summary_ws, chapter, entry_students, chapter_rows, source_folder)
    if not chapter_rows.empty:
        grouped = chapter_rows.loc[coerce_numeric(chapter_rows["observed_year"]).notna()].groupby("observed_year")
        for year_value, frame in sorted(grouped, key=lambda item: int(item[0])):
            write_year_sheet(wb, chapter, str(int(year_value)), frame.copy())
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_index_workbook(index_rows: List[List[object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chapters"
    ws.append(
        [
            "Chapter",
            "Workbook File",
            "Distinct Students Ever Observed",
            "Students With Observed Chapter Entry",
            "Entry Cohorts Covered",
            "Observed Graduation Rate (Resolved)",
            "Retained In Chapter To Next Fall",
            "Stayed In School To Next Fall",
            "Average First-Term GPA",
            "Average Latest Cumulative GPA",
        ]
    )
    style_row_as_header(ws, 1)
    for row in index_rows:
        ws.append(row)
    format_table_columns(ws, 1, 2, ws.max_row)
    ws.freeze_panes = "A2"
    autosize_columns(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_readme(result: ChapterBuildResult) -> None:
    lines = [
        "# Chapter History Workbooks",
        "",
        "This folder contains one workbook per chapter, built additively from the canonical analytics bundle.",
        "",
        "## Files",
        "",
        "- `Chapter_History_Index.xlsx`: one-row summary per chapter",
        "- `chapter_history_index.csv`: CSV copy of the chapter index",
        "- `workbooks/*.xlsx`: one workbook per chapter",
        "",
        "## Workbook layout",
        "",
        "- `Summary`: chapter-level graduation, retention, school continuation, GPA, outcome, and yearly trend tables",
        "- `YYYY`: one row per observed student-term for that chapter in that year",
        "",
        "## Source bundle",
        "",
        f"- `{result.source_folder}`",
        "",
        "## Important note",
        "",
        "Graduation rates in these workbooks exclude unresolved outcomes such as `Active/Unknown` and `No Further Observation` and count graduation only when confirmed evidence exists.",
    ]
    result.readme_path.write_text("\n".join(lines), encoding="utf-8")


def build_chapter_history_workbooks(canonical_root: Path, explicit_folder: Path | None, output_root: Path) -> ChapterBuildResult:
    bundle = load_canonical_bundle(canonical_root=canonical_root, explicit_folder=explicit_folder)
    summary = prepare_summary(bundle.tables["student_summary"])
    longitudinal = prepare_longitudinal(bundle.tables["master_longitudinal"])

    chapter_names = sorted(
        {
            clean_text(value)
            for value in list(summary["initial_chapter"].tolist()) + list(summary["latest_chapter"].tolist()) + list(longitudinal["chapter"].tolist())
            if clean_text(value) and clean_text(value).lower() != "unknown" and not is_excluded_chapter(clean_text(value))
        }
    )
    if not chapter_names:
        raise FileNotFoundError("No usable chapters were found in the canonical analytics bundle.")

    timestamp = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    output_folder = output_root / timestamp
    workbook_folder = output_folder / "workbooks"
    workbook_folder.mkdir(parents=True, exist_ok=True)

    index_rows: List[List[object]] = []
    chapters_written = 0
    for chapter in chapter_names:
        entry_students = filter_to_chapter_entry(summary, chapter)
        chapter_rows = filter_to_chapter_rows(longitudinal, chapter)
        if entry_students.empty and chapter_rows.empty:
            continue
        workbook_path = workbook_folder / f"{safe_filename(chapter)}.xlsx"
        write_chapter_workbook(chapter, entry_students, chapter_rows, workbook_path, bundle.output_folder)
        index_rows.append(create_index_row(chapter, entry_students, chapter_rows, workbook_path))
        chapters_written += 1

    if not chapters_written:
        raise FileNotFoundError("No chapter workbooks were written because no chapter rows were available.")

    index_rows.sort(key=lambda row: clean_text(row[0]).lower())
    index_workbook = output_folder / "Chapter_History_Index.xlsx"
    index_csv = output_folder / "chapter_history_index.csv"
    readme_path = output_folder / "README.md"

    write_index_workbook(index_rows, index_workbook)
    pd.DataFrame(
        index_rows,
        columns=[
            "Chapter",
            "Workbook File",
            "Distinct Students Ever Observed",
            "Students With Observed Chapter Entry",
            "Entry Cohorts Covered",
            "Observed Graduation Rate (Resolved)",
            "Retained In Chapter To Next Fall",
            "Stayed In School To Next Fall",
            "Average First-Term GPA",
            "Average Latest Cumulative GPA",
        ],
    ).to_csv(index_csv, index=False)

    result = ChapterBuildResult(
        output_folder=output_folder,
        workbook_folder=workbook_folder,
        index_workbook=index_workbook,
        index_csv=index_csv,
        readme_path=readme_path,
        chapters_written=chapters_written,
        source_folder=bundle.output_folder,
    )
    write_readme(result)
    return result


def main() -> None:
    args = parse_args()
    result = build_chapter_history_workbooks(
        canonical_root=Path(args.canonical_root).expanduser().resolve(),
        explicit_folder=Path(args.canonical_folder).expanduser().resolve() if args.canonical_folder else None,
        output_root=Path(args.output_root).expanduser().resolve(),
    )
    print(f"Chapter history workbooks created in: {result.output_folder}")
    print(f"Chapter workbooks: {result.workbook_folder}")
    print(f"Chapter index workbook: {result.index_workbook}")


if __name__ == "__main__":
    main()
