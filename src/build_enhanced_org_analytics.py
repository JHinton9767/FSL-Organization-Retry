from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.build_master_roster import (
    autosize_columns,
    clean_text,
    normalize_banner_id,
    style_header,
)
from src.build_master_roster_grades import DEFAULT_OUTPUT_WORKBOOK as DEFAULT_MERGED_WORKBOOK


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "enhanced_metrics"
DEFAULT_SEGMENT_MIN_SIZE = 5
MAX_EXCEL_ROWS = 1_000_000

ROSTER_REQUIRED_COLUMNS = {
    "Academic Year",
    "Term",
    "Source File",
    "Chapter",
    "Last Name",
    "First Name",
    "Banner ID",
    "Email",
    "Status",
    "Semester Joined",
    "Position",
}

UNMATCHED_GRADE_REQUIRED_COLUMNS = {
    "Term",
    "Source File",
    "Last Name",
    "First Name",
    "Banner ID",
    "Email",
}

GRADE_FIELDS = [
    "Grade Student Status",
    "Major",
    "Semester Hours",
    "Cumulative Hours",
    "Current Academic Standing",
    "Texas State GPA",
    "Overall GPA",
    "Transfer GPA",
    "Term GPA",
    "Term Passed Hours",
    "TxState Cumulative GPA",
    "Overall Cumulative GPA",
]

SEASON_ORDER = {
    "Winter": 0,
    "Spring": 1,
    "Summer": 2,
    "Fall": 3,
    "Unknown": 9,
}

SEASON_CODES = {
    "Winter": "WI",
    "Spring": "SP",
    "Summer": "SU",
    "Fall": "FA",
    "Unknown": "UN",
}

TERM_RE = re.compile(r"(Winter|Spring|Summer|Fall)\s+(19\d{2}|20\d{2})", re.IGNORECASE)
UPDATE_RE = re.compile(r"\((\d{1,2})\.(\d{1,2})\.(\d{2,4})\)")

ROSTER_STATUS_PRIORITY = {
    "Graduated": 90,
    "Alumni": 85,
    "Suspended": 80,
    "Transfer": 70,
    "Dropped/Resigned/Revoked/Inactive": 60,
    "Active": 50,
    "Unknown": 0,
}

PRIMARY_STATUS_BUCKETS = [
    "Graduated",
    "Alumni",
    "Suspended",
    "Transfer",
    "Dropped/Resigned/Revoked/Inactive",
    "Active",
    "Unknown",
]

OUTCOME_BUCKETS = [
    "Graduated",
    "Suspended",
    "Transfer",
    "Dropped/Resigned/Revoked/Inactive",
    "No Further Observation",
    "Active/Unknown",
]

STANDING_BUCKETS = [
    "Good Standing",
    "Probation/Warning",
    "Suspended",
    "Dismissed/Separated",
    "Other/Unmapped",
    "Unknown",
]

ENTRY_HOURS_BUCKET_SIZE = 30

MASTER_LONGITUDINAL_COLUMNS = [
    "Student ID",
    "Identity Resolution Basis",
    "Original Banner ID Present",
    "Last Name",
    "First Name",
    "Email",
    "Term",
    "Term Code",
    "Term Year",
    "Term Season",
    "Roster Present",
    "Academic Present",
    "Organization Entry Term",
    "Organization Entry Cohort",
    "First Observed Academic Term",
    "First Observed Term Overall",
    "Relative Term Index From Org Entry",
    "Relative Academic Year Window",
    "Is Organization Entry Term",
    "Is Within First Academic Year After Org Entry",
    "Chapter",
    "Roster Source File",
    "Roster Status Raw",
    "Roster Status Bucket",
    "Roster Position",
    "Roster Semester Joined",
    "New Member Marker",
    "Roster Row Count Same Term",
    "Roster Statuses Same Term",
    "Roster Status Conflict Same Term",
    "Roster Chapters Same Term",
    "Roster Chapter Conflict Same Term",
    "Academic Source File",
    "Academic Student Status Raw",
    "Academic Status Signal Bucket",
    "Major",
    "Semester Hours",
    "Term Passed Hours",
    "Cumulative Hours",
    "Entry Cumulative Hours Bucket",
    "Current Academic Standing Raw",
    "Academic Standing Bucket",
    "Academic Row Count Same Term",
    "Academic Standing Conflict Same Term",
    "Term GPA",
    "Texas State GPA",
    "Overall GPA",
    "Transfer GPA",
    "TxState Cumulative GPA",
    "Overall Cumulative GPA",
    "Latest Known Outcome Bucket",
    "No Further Observation Flag",
]

STUDENT_SUMMARY_COLUMNS = [
    "Student ID",
    "Identity Resolution Basis Used",
    "Preferred Last Name",
    "Preferred First Name",
    "Preferred Email",
    "First Observed Organization Term",
    "Last Observed Organization Term",
    "First Observed Academic Term",
    "Last Observed Academic Term",
    "First Observed Term Overall",
    "Organization Entry Cohort",
    "Initial Chapter",
    "Initial Roster Status Raw",
    "Initial Roster Status Bucket",
    "Latest Chapter",
    "Latest Known Roster Status Raw",
    "Latest Known Roster Status Bucket",
    "Latest Known Outcome Bucket",
    "Ever Graduated Flag",
    "Ever Suspended Flag",
    "Ever Dropped/Inactive/Resigned/Revoked Flag",
    "Ever Transfer Flag",
    "No Further Observation Flag",
    "Observed Graduation Term",
    "First Post-Entry Academic Term",
    "Second Post-Entry Academic Term",
    "First Major After Org Entry",
    "Latest Major",
    "First Academic Standing After Org Entry",
    "Latest Academic Standing",
    "Entry Cumulative Hours",
    "Entry Cumulative Hours Bucket",
    "Roster Terms Observed",
    "Academic Terms Observed",
    "Total Terms Observed Overall",
    "Organization Next Observed Term Measurable",
    "Organization Next Observed Term Category",
    "Retained In Organization To Next Observed Term",
    "Organization Next Fall Measurable",
    "Organization Next Fall Category",
    "Retained In Organization To Next Fall",
    "Organization One-Year Same-Season Measurable",
    "Organization One-Year Same-Season Category",
    "Retained In Organization One Year After Entry",
    "Academic Next Observed Term Measurable",
    "Academic Next Observed Term Category",
    "Continued Academically To Next Observed Term",
    "Academic Next Fall Measurable",
    "Academic Next Fall Category",
    "Continued Academically To Next Fall",
    "Academic One-Year Same-Season Measurable",
    "Academic One-Year Same-Season Category",
    "Continued Academically One Year After Entry",
    "Eventual Observed Graduation From Org Entry",
    "Observed Graduation Within 4 Years Of Org Entry Measurable",
    "Observed Graduation Within 4 Years Of Org Entry",
    "Observed Graduation Within 6 Years Of Org Entry Measurable",
    "Observed Graduation Within 6 Years Of Org Entry",
    "Eventual Observed Graduation From First Observed Academic Term",
    "First Post-Entry Term GPA",
    "Second Post-Entry Term GPA",
    "First-Year Average Term GPA After Org Entry",
    "Change In Term GPA First To Second Term",
    "First Post-Entry TxState Cumulative GPA",
    "Latest TxState Cumulative GPA",
    "First Post-Entry Overall Cumulative GPA",
    "Latest Overall Cumulative GPA",
    "First Post-Entry Passed Hours",
    "First-Year Passed Hours After Org Entry",
    "First-Term 12+ Passed Hours Flag",
    "First-Term 15+ Passed Hours Flag",
    "First-Year 24+ Passed Hours Flag",
    "First-Year 30+ Passed Hours Flag",
    "First-Term GPA Below 2.0 Flag",
    "First-Term GPA Below 2.5 Flag",
    "First-Year Average GPA Below 2.0 Flag",
    "First-Year Average GPA Below 2.5 Flag",
    "First-Term Good Standing Flag",
    "First-Year Probation/Warning Flag",
    "Academic Standing Suspended Ever Flag",
    "Roster Status Transition Count",
    "Academic Standing Transition Count",
]

COHORT_METRIC_COLUMNS = [
    "Metric Group",
    "Metric Label",
    "Cohort",
    "Dimension",
    "Value",
    "Eligible Students",
    "Student Count",
    "Rate",
    "Average Value",
    "Notes",
]

TRANSITION_COLUMNS = [
    "Transition Family",
    "Cohort",
    "From Value",
    "To Value",
    "From Relative Term Index",
    "Transition Count",
    "Distinct Students",
]

SEGMENT_COLUMNS = [
    "Cohort",
    "Dimension",
    "Value",
    "Group Size",
    "Observed Eventual Graduation Rate From Org Entry",
    "Observed Graduation Within 4 Years Of Org Entry",
    "Observed Graduation Within 6 Years Of Org Entry",
    "Retained In Organization To Next Fall",
    "Continued Academically To Next Fall",
    "Average First-Year Term GPA After Org Entry",
    "Average First-Year Passed Hours After Org Entry",
    "Share First-Term GPA Below 2.0",
    "Share First-Term GPA Below 2.5",
]

QA_COLUMNS = [
    "Check",
    "Status",
    "Value",
    "Notes",
]

STATUS_MAPPING_COLUMNS = [
    "Field",
    "Derived Bucket",
    "Rule",
    "Notes",
]

CHANGE_LOG_COLUMNS = [
    "Component",
    "Type",
    "Description",
]

METRIC_DEFINITION_COLUMNS = [
    "Metric Group",
    "Metric Label",
    "Definition",
    "Window / Denominator",
    "Limitations / Notes",
]

EXCLUDED_ROW_COLUMNS = [
    "Row Type",
    "Reason",
    "Source Sheet",
    "Source File",
    "Term",
    "Banner ID",
    "Email",
    "Last Name",
    "First Name",
]


@dataclass(frozen=True)
class TermInfo:
    label: str
    code: str
    year: int
    season: str
    sort_key: Tuple[int, int, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build additive organization-entry analytics from Master_Roster_Grades.xlsx without "
            "overwriting any existing outputs."
        )
    )
    parser.add_argument(
        "--merged-workbook",
        default=str(DEFAULT_MERGED_WORKBOOK),
        help="Path to Master_Roster_Grades.xlsx. Default: Master_Roster_Grades.xlsx next to the code.",
    )
    parser.add_argument(
        "--output-root",
        default=str(DEFAULT_OUTPUT_ROOT),
        help="Root folder for versioned additive outputs. Default: output\\enhanced_metrics",
    )
    parser.add_argument(
        "--segment-min-size",
        type=int,
        default=DEFAULT_SEGMENT_MIN_SIZE,
        help="Minimum group size for segmentation reporting. Default: 5",
    )
    return parser.parse_args()


def canonical_text(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value).lower()).strip()


def get_cell(row: Tuple[object, ...], header_map: Dict[str, int], column: str) -> str:
    idx = header_map.get(column)
    if idx is None or idx >= len(row):
        return ""
    return clean_text(row[idx])


def yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def to_float(value: object) -> Optional[float]:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def average(values: Iterable[Optional[float]]) -> Optional[float]:
    usable = [value for value in values if value is not None]
    if not usable:
        return None
    return sum(usable) / len(usable)


def unique_sorted(values: Iterable[str]) -> List[str]:
    return sorted({clean_text(value) for value in values if clean_text(value)})


def parse_term_info(term_label: str, academic_year: str = "") -> Optional[TermInfo]:
    cleaned = clean_text(term_label)
    match = TERM_RE.search(cleaned)
    if match:
        season = match.group(1).title()
        year = int(match.group(2))
        return TermInfo(
            label=f"{season} {year}",
            code=f"{year}{SEASON_CODES[season]}",
            year=year,
            season=season,
            sort_key=(year, SEASON_ORDER.get(season, 9), cleaned.lower()),
        )
    if academic_year.isdigit() and len(academic_year) == 4:
        year = int(academic_year)
        return TermInfo(
            label=cleaned or academic_year,
            code=f"{year}{SEASON_CODES['Unknown']}",
            year=year,
            season="Unknown",
            sort_key=(year, SEASON_ORDER["Unknown"], cleaned.lower()),
        )
    return None


def parse_term_code(term_code: str) -> Optional[TermInfo]:
    cleaned = clean_text(term_code).upper()
    match = re.fullmatch(r"(19\d{2}|20\d{2})(WI|SP|SU|FA|UN)", cleaned)
    if not match:
        return None
    year = int(match.group(1))
    season_lookup = {value: key for key, value in SEASON_CODES.items()}
    season = season_lookup.get(match.group(2), "Unknown")
    return TermInfo(
        label=f"{season} {year}" if season != "Unknown" else str(year),
        code=cleaned,
        year=year,
        season=season,
        sort_key=(year, SEASON_ORDER.get(season, 9), cleaned),
    )


def term_code_of(value: str) -> str:
    cleaned = clean_text(value)
    info = parse_term_code(cleaned)
    if info is not None:
        return info.code
    parsed = parse_term_info(cleaned)
    return parsed.code if parsed is not None else ""


def term_sort_tuple(term_code: str) -> Tuple[int, int, str]:
    info = parse_term_code(term_code_of(term_code))
    if info is None:
        return (9999, 9, clean_text(term_code).lower())
    return info.sort_key


def sort_term_codes(term_codes: Iterable[str]) -> List[str]:
    return sorted(
        {clean_text(term_code) for term_code in term_codes if clean_text(term_code)},
        key=term_sort_tuple,
    )


def next_observed_term(term_code: str, ordered_codes: Sequence[str]) -> Optional[str]:
    try:
        idx = ordered_codes.index(term_code)
    except ValueError:
        return None
    if idx + 1 >= len(ordered_codes):
        return None
    return ordered_codes[idx + 1]


def next_fall_code(term_code: str) -> Optional[str]:
    info = parse_term_code(term_code_of(term_code))
    if info is None:
        return None
    target_year = info.year + 1 if info.season == "Fall" else info.year
    return f"{target_year}{SEASON_CODES['Fall']}"


def same_season_years_later(term_code: str, years: int) -> Optional[str]:
    info = parse_term_code(term_code_of(term_code))
    if info is None:
        return None
    return f"{info.year + years}{SEASON_CODES.get(info.season, 'UN')}"


def is_new_member_marker(status: str, position: str) -> bool:
    values = [canonical_text(status), canonical_text(position)]
    return any(value == "new member" or "new member" in value or value == "nm" for value in values)


def roster_status_bucket(status: str, position: str) -> str:
    status_text = canonical_text(status)
    position_text = canonical_text(position)
    combined = " ".join(part for part in [status_text, position_text] if part)
    if not combined:
        return "Unknown"
    if "graduated" in status_text:
        return "Graduated"
    if "alumni" in status_text:
        return "Alumni"
    if "suspend" in status_text:
        return "Suspended"
    if "transfer" in status_text:
        return "Transfer"
    if any(word in status_text for word in ["inactive", "resign", "revok", "drop", "removed"]):
        return "Dropped/Resigned/Revoked/Inactive"
    if "active" in combined or "member" in combined or is_new_member_marker(status, position):
        return "Active"
    return "Unknown"


def academic_status_signal_bucket(value: str) -> str:
    text = canonical_text(value)
    if not text:
        return "Unknown"
    if any(word in text for word in ["graduated", "degree awarded", "awarded degree", "degree conferred", "alumni"]):
        return "Graduated"
    if "suspend" in text:
        return "Suspended"
    if "transfer" in text:
        return "Transfer"
    if any(word in text for word in ["inactive", "withdraw", "drop", "dismiss", "separat"]):
        return "Dropped/Resigned/Revoked/Inactive"
    return "Active"


def academic_standing_bucket(value: str) -> str:
    text = canonical_text(value)
    if not text:
        return "Unknown"
    if "good" in text and "not" not in text:
        return "Good Standing"
    if "probation" in text or "warning" in text or "alert" in text:
        return "Probation/Warning"
    if "suspend" in text:
        return "Suspended"
    if any(word in text for word in ["dismiss", "drop", "separat"]):
        return "Dismissed/Separated"
    return "Other/Unmapped"


def entry_hours_bucket(hours: Optional[float]) -> str:
    if hours is None or hours < 0:
        return "Unknown"
    lower = int(hours // ENTRY_HOURS_BUCKET_SIZE) * ENTRY_HOURS_BUCKET_SIZE
    upper = lower + ENTRY_HOURS_BUCKET_SIZE - 1
    return f"{lower}-{upper}"


def entry_hours_bucket_sort_key(bucket: str) -> Tuple[int, int, str]:
    if bucket == "Unknown":
        return (1, 999999, bucket)
    lower = bucket.split("-", 1)[0]
    if lower.isdigit():
        return (0, int(lower), bucket)
    return (1, 999999, bucket)


def followup_category(
    target_code: Optional[str],
    available_target_codes: Sequence[str],
    observed_codes: Sequence[str],
    explicit_graduation_sort: Optional[Tuple[int, int, str]],
    explicit_terminal_sort: Optional[Tuple[int, int, str]],
    last_observed_sort: Optional[Tuple[int, int, str]],
) -> Tuple[str, str]:
    if not target_code or target_code not in available_target_codes:
        return "No", "Not measurable"
    target_sort = term_sort_tuple(target_code)
    if target_code in observed_codes:
        return "Yes", "Observed at follow-up"
    if explicit_graduation_sort is not None and explicit_graduation_sort < target_sort:
        return "No", "Graduated before follow-up"
    if explicit_terminal_sort is not None and explicit_terminal_sort < target_sort:
        return "No", "Terminal status before follow-up"
    if last_observed_sort is not None and last_observed_sort < target_sort:
        return "No", "No further observation before follow-up"
    return "No", "Not observed at follow-up"


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]+", " ", clean_text(title))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:31] or "Sheet"


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers))
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def write_records_sheet(
    wb: Workbook,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Dict[str, object]],
) -> None:
    ws = wb.create_sheet(title=safe_sheet_title(title))
    ws.append(list(headers))
    style_header(ws)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_chunked_records_sheet(
    wb: Workbook,
    base_title: str,
    headers: Sequence[str],
    rows: Sequence[Dict[str, object]],
) -> None:
    if not rows:
        write_records_sheet(wb, base_title, headers, [])
        return
    for start in range(0, len(rows), MAX_EXCEL_ROWS):
        end = min(start + MAX_EXCEL_ROWS, len(rows))
        chunk_index = (start // MAX_EXCEL_ROWS) + 1
        title = base_title if len(rows) <= MAX_EXCEL_ROWS else f"{base_title}_{chunk_index}"
        write_records_sheet(wb, title, headers, rows[start:end])


def write_overview_sheet(
    wb: Workbook,
    merged_workbook: Path,
    output_folder: Path,
    run_timestamp: str,
    total_students: int,
    total_longitudinal_rows: int,
    qa_rows: Sequence[Dict[str, object]],
) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    style_header(ws)
    failed_checks = sum(1 for row in qa_rows if row.get("Status") == "Flag")
    for metric, value in [
        ("Run timestamp", run_timestamp),
        ("Merged workbook used", str(merged_workbook)),
        ("Output folder", str(output_folder)),
        ("Distinct students in additive analysis", total_students),
        ("Student-term longitudinal rows", total_longitudinal_rows),
        ("QA checks flagged", failed_checks),
        ("Method", "Additive analysis only; existing builders and outputs left untouched"),
        (
            "Primary caution",
            "Metrics are labeled from first observed organization or academic terms, not true first-time-in-college entry",
        ),
    ]:
        ws.append([metric, value])

    ws.append([])
    ws.append(["Existing Components Left Untouched"])
    ws[ws.max_row][0].font = Font(bold=True)
    for item in [
        "src/greek_life_pipeline.py",
        "src/build_master_roster.py",
        "src/build_master_roster_grades.py",
        "src/build_member_tenure_report.py",
        "src/build_yearly_chapter_rosters.py",
        "powerquery/AcademicFolderTransform.pq",
        "powerquery/RosterFolderTransform.pq",
        "powerquery/MasterDataset.pq",
    ]:
        ws.append([item])
    autosize_columns(ws)


def load_merged_workbook_observations(
    merged_workbook: Path,
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    if not merged_workbook.exists():
        raise FileNotFoundError(
            f"Merged workbook not found at {merged_workbook}. Run py run_master_roster_grades.py first."
        )

    roster_rows: List[Dict[str, object]] = []
    academic_rows: List[Dict[str, object]] = []
    excluded_rows: List[Dict[str, object]] = []

    wb = load_workbook(merged_workbook, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue

            header_values = [clean_text(value) for value in sheet_rows[0]]
            header_map = {header: idx for idx, header in enumerate(header_values)}
            title_lower = ws.title.lower()

            if title_lower == "summary":
                continue

            if title_lower == "unmatched grades" and UNMATCHED_GRADE_REQUIRED_COLUMNS.issubset(set(header_values)):
                for row in sheet_rows[1:]:
                    term_label = get_cell(row, header_map, "Term")
                    info = parse_term_info(term_label)
                    academic_row = {
                        "source_sheet": ws.title,
                        "source_file": get_cell(row, header_map, "Source File"),
                        "academic_source_file": get_cell(row, header_map, "Source File"),
                        "term": term_label,
                        "term_code": info.code if info else "",
                        "academic_year": str(info.year) if info else "",
                        "banner_id": normalize_banner_id(get_cell(row, header_map, "Banner ID")),
                        "email": get_cell(row, header_map, "Email").lower(),
                        "last_name": get_cell(row, header_map, "Last Name"),
                        "first_name": get_cell(row, header_map, "First Name"),
                        "grade_student_status": get_cell(row, header_map, "Student Status"),
                        "major": get_cell(row, header_map, "Major"),
                        "semester_hours": get_cell(row, header_map, "Semester Hours"),
                        "cumulative_hours": get_cell(row, header_map, "Cumulative Hours"),
                        "current_academic_standing": get_cell(row, header_map, "Current Academic Standing"),
                        "texas_state_gpa": get_cell(row, header_map, "Texas State GPA"),
                        "overall_gpa": get_cell(row, header_map, "Overall GPA"),
                        "transfer_gpa": get_cell(row, header_map, "Transfer GPA"),
                        "term_gpa": get_cell(row, header_map, "Term GPA"),
                        "term_passed_hours": get_cell(row, header_map, "Term Passed Hours"),
                        "txstate_cumulative_gpa": get_cell(row, header_map, "TxState Cumulative GPA"),
                        "overall_cumulative_gpa": get_cell(row, header_map, "Overall Cumulative GPA"),
                    }
                    if not any(
                        [
                            academic_row["banner_id"],
                            academic_row["email"],
                            academic_row["last_name"],
                            academic_row["first_name"],
                        ]
                    ):
                        continue
                    if info is None:
                        excluded_rows.append(
                            {
                                "Row Type": "Academic",
                                "Reason": "Unparseable term",
                                "Source Sheet": ws.title,
                                "Source File": academic_row["source_file"],
                                "Term": academic_row["term"],
                                "Banner ID": academic_row["banner_id"],
                                "Email": academic_row["email"],
                                "Last Name": academic_row["last_name"],
                                "First Name": academic_row["first_name"],
                            }
                        )
                        continue
                    academic_rows.append(academic_row)
                continue

            if not ROSTER_REQUIRED_COLUMNS.issubset(set(header_values)):
                continue

            for row in sheet_rows[1:]:
                term_label = get_cell(row, header_map, "Term")
                academic_year = get_cell(row, header_map, "Academic Year")
                info = parse_term_info(term_label, academic_year=academic_year)
                roster_row = {
                    "source_sheet": ws.title,
                    "source_file": get_cell(row, header_map, "Source File"),
                    "term": term_label,
                    "term_code": info.code if info else "",
                    "academic_year": academic_year,
                    "banner_id": normalize_banner_id(get_cell(row, header_map, "Banner ID")),
                    "email": get_cell(row, header_map, "Email").lower(),
                    "last_name": get_cell(row, header_map, "Last Name"),
                    "first_name": get_cell(row, header_map, "First Name"),
                    "chapter": get_cell(row, header_map, "Chapter"),
                    "raw_status": get_cell(row, header_map, "Status"),
                    "position": get_cell(row, header_map, "Position"),
                    "semester_joined": get_cell(row, header_map, "Semester Joined"),
                }
                if not any(
                    [
                        roster_row["banner_id"],
                        roster_row["email"],
                        roster_row["last_name"],
                        roster_row["first_name"],
                    ]
                ):
                    continue
                if info is None:
                    excluded_rows.append(
                        {
                            "Row Type": "Roster",
                            "Reason": "Unparseable term",
                            "Source Sheet": ws.title,
                            "Source File": roster_row["source_file"],
                            "Term": roster_row["term"],
                            "Banner ID": roster_row["banner_id"],
                            "Email": roster_row["email"],
                            "Last Name": roster_row["last_name"],
                            "First Name": roster_row["first_name"],
                        }
                    )
                    continue
                roster_rows.append(roster_row)

                has_academic_values = any(get_cell(row, header_map, field) for field in GRADE_FIELDS)
                if has_academic_values:
                    academic_rows.append(
                        {
                            "source_sheet": ws.title,
                            "source_file": get_cell(row, header_map, "Source File"),
                            "academic_source_file": "",
                            "term": term_label,
                            "term_code": info.code,
                            "academic_year": academic_year,
                            "banner_id": roster_row["banner_id"],
                            "email": roster_row["email"],
                            "last_name": roster_row["last_name"],
                            "first_name": roster_row["first_name"],
                            "grade_student_status": get_cell(row, header_map, "Grade Student Status"),
                            "major": get_cell(row, header_map, "Major"),
                            "semester_hours": get_cell(row, header_map, "Semester Hours"),
                            "cumulative_hours": get_cell(row, header_map, "Cumulative Hours"),
                            "current_academic_standing": get_cell(row, header_map, "Current Academic Standing"),
                            "texas_state_gpa": get_cell(row, header_map, "Texas State GPA"),
                            "overall_gpa": get_cell(row, header_map, "Overall GPA"),
                            "transfer_gpa": get_cell(row, header_map, "Transfer GPA"),
                            "term_gpa": get_cell(row, header_map, "Term GPA"),
                            "term_passed_hours": get_cell(row, header_map, "Term Passed Hours"),
                            "txstate_cumulative_gpa": get_cell(row, header_map, "TxState Cumulative GPA"),
                            "overall_cumulative_gpa": get_cell(row, header_map, "Overall Cumulative GPA"),
                        }
                    )
    finally:
        wb.close()

    return roster_rows, academic_rows, excluded_rows


def apply_identity_resolution(rows: Sequence[Dict[str, object]]) -> Tuple[List[Dict[str, object]], int]:
    email_to_ids: Dict[str, set[str]] = defaultdict(set)
    for row in rows:
        banner_id = clean_text(row.get("banner_id"))
        email = clean_text(row.get("email")).lower()
        if banner_id and email:
            email_to_ids[email].add(banner_id)

    resolved_rows: List[Dict[str, object]] = []
    derived_count = 0
    for row in rows:
        updated = dict(row)
        banner_id = clean_text(updated.get("banner_id"))
        email = clean_text(updated.get("email")).lower()
        if banner_id:
            updated["resolved_student_id"] = banner_id
            updated["identity_basis"] = "Banner ID"
        elif email and len(email_to_ids.get(email, set())) == 1:
            updated["resolved_student_id"] = next(iter(email_to_ids[email]))
            updated["identity_basis"] = "Derived from unique email"
            derived_count += 1
        else:
            updated["resolved_student_id"] = ""
            updated["identity_basis"] = "Missing Student ID"
        resolved_rows.append(updated)
    return resolved_rows, derived_count


def roster_row_score(row: Dict[str, object]) -> Tuple[int, int, int, int, str]:
    bucket = clean_text(row.get("roster_status_bucket"))
    return (
        1 if clean_text(row.get("resolved_student_id")) else 0,
        ROSTER_STATUS_PRIORITY.get(bucket, 0),
        1 if row.get("new_member_marker") == "Yes" else 0,
        1 if clean_text(row.get("chapter")) and clean_text(row.get("chapter")) != "Unknown" else 0,
        clean_text(row.get("source_file")).lower(),
    )


def grade_update_key(source_file: str) -> Tuple[int, int, int]:
    match = UPDATE_RE.search(clean_text(source_file))
    if not match:
        return (0, 0, 0)
    month = int(match.group(1))
    day = int(match.group(2))
    year = int(match.group(3))
    if year < 100:
        year += 2000
    return (year, month, day)


def academic_row_score(row: Dict[str, object]) -> Tuple[Tuple[int, int, int], int, int, int]:
    completeness = sum(
        1
        for field in [
            "grade_student_status",
            "major",
            "semester_hours",
            "cumulative_hours",
            "current_academic_standing",
            "term_gpa",
            "term_passed_hours",
            "txstate_cumulative_gpa",
            "overall_cumulative_gpa",
        ]
        if clean_text(row.get(field))
    )
    return (
        grade_update_key(clean_text(row.get("academic_source_file") or row.get("source_file"))),
        1 if clean_text(row.get("term_gpa")) else 0,
        1 if clean_text(row.get("term_passed_hours")) else 0,
        completeness,
    )


def dedupe_roster_rows(
    rows: Sequence[Dict[str, object]],
    excluded_rows: List[Dict[str, object]],
) -> Tuple[Dict[Tuple[str, str], Dict[str, object]], Dict[str, int]]:
    groups: Dict[Tuple[str, str], List[Dict[str, object]]] = defaultdict(list)
    stats = {
        "raw_rows": len(rows),
        "missing_student_id_rows": 0,
        "duplicate_student_term_groups": 0,
        "status_conflict_groups": 0,
        "chapter_conflict_groups": 0,
    }

    for row in rows:
        student_id = clean_text(row.get("resolved_student_id"))
        term_code = clean_text(row.get("term_code"))
        if not student_id:
            stats["missing_student_id_rows"] += 1
            excluded_rows.append(
                {
                    "Row Type": "Roster",
                    "Reason": "Missing resolved Student ID",
                    "Source Sheet": clean_text(row.get("source_sheet")),
                    "Source File": clean_text(row.get("source_file")),
                    "Term": clean_text(row.get("term")),
                    "Banner ID": clean_text(row.get("banner_id")),
                    "Email": clean_text(row.get("email")),
                    "Last Name": clean_text(row.get("last_name")),
                    "First Name": clean_text(row.get("first_name")),
                }
            )
            continue
        groups[(student_id, term_code)].append(row)

    deduped: Dict[Tuple[str, str], Dict[str, object]] = {}
    for key, group in groups.items():
        primary = max(group, key=roster_row_score)
        statuses = unique_sorted(row.get("raw_status", "") for row in group)
        buckets = unique_sorted(row.get("roster_status_bucket", "") for row in group)
        chapters = unique_sorted(row.get("chapter", "") for row in group)
        if len(group) > 1:
            stats["duplicate_student_term_groups"] += 1
        if len(buckets) > 1:
            stats["status_conflict_groups"] += 1
        if len(chapters) > 1:
            stats["chapter_conflict_groups"] += 1
        aggregated = dict(primary)
        aggregated["roster_row_count_same_term"] = len(group)
        aggregated["roster_statuses_same_term"] = " | ".join(statuses)
        aggregated["roster_status_conflict_same_term"] = yes_no(len(buckets) > 1)
        aggregated["roster_chapters_same_term"] = " | ".join(chapters)
        aggregated["roster_chapter_conflict_same_term"] = yes_no(len(chapters) > 1)
        deduped[key] = aggregated
    return deduped, stats


def dedupe_academic_rows(
    rows: Sequence[Dict[str, object]],
    excluded_rows: List[Dict[str, object]],
) -> Tuple[Dict[Tuple[str, str], Dict[str, object]], Dict[str, int]]:
    groups: Dict[Tuple[str, str], List[Dict[str, object]]] = defaultdict(list)
    stats = {
        "raw_rows": len(rows),
        "missing_student_id_rows": 0,
        "duplicate_student_term_groups": 0,
        "standing_conflict_groups": 0,
    }

    for row in rows:
        student_id = clean_text(row.get("resolved_student_id"))
        term_code = clean_text(row.get("term_code"))
        if not student_id:
            stats["missing_student_id_rows"] += 1
            excluded_rows.append(
                {
                    "Row Type": "Academic",
                    "Reason": "Missing resolved Student ID",
                    "Source Sheet": clean_text(row.get("source_sheet")),
                    "Source File": clean_text(row.get("academic_source_file") or row.get("source_file")),
                    "Term": clean_text(row.get("term")),
                    "Banner ID": clean_text(row.get("banner_id")),
                    "Email": clean_text(row.get("email")),
                    "Last Name": clean_text(row.get("last_name")),
                    "First Name": clean_text(row.get("first_name")),
                }
            )
            continue
        groups[(student_id, term_code)].append(row)

    deduped: Dict[Tuple[str, str], Dict[str, object]] = {}
    for key, group in groups.items():
        primary = max(group, key=academic_row_score)
        standing_buckets = unique_sorted(row.get("academic_standing_bucket", "") for row in group)
        if len(group) > 1:
            stats["duplicate_student_term_groups"] += 1
        if len(standing_buckets) > 1:
            stats["standing_conflict_groups"] += 1
        aggregated = dict(primary)
        aggregated["academic_row_count_same_term"] = len(group)
        aggregated["academic_standing_conflict_same_term"] = yes_no(len(standing_buckets) > 1)
        deduped[key] = aggregated
    return deduped, stats


def build_master_longitudinal(
    roster_map: Dict[Tuple[str, str], Dict[str, object]],
    academic_map: Dict[Tuple[str, str], Dict[str, object]],
) -> Tuple[List[Dict[str, object]], List[str], List[str], List[str]]:
    all_keys = sorted(
        set(roster_map.keys()) | set(academic_map.keys()),
        key=lambda item: (term_sort_tuple(item[1]), item[0]),
    )
    roster_term_codes = sort_term_codes(term for _, term in roster_map.keys())
    academic_term_codes = sort_term_codes(term for _, term in academic_map.keys())
    all_term_codes = sort_term_codes(term for _, term in all_keys)

    master_rows: List[Dict[str, object]] = []
    for student_id, term_code in all_keys:
        roster = roster_map.get((student_id, term_code))
        academic = academic_map.get((student_id, term_code))
        info = parse_term_code(term_code)
        cumulative_hours = to_float(academic.get("cumulative_hours")) if academic else None
        master_rows.append(
            {
                "Student ID": student_id,
                "Identity Resolution Basis": clean_text((roster or academic or {}).get("identity_basis")),
                "Original Banner ID Present": yes_no(bool(clean_text((roster or academic or {}).get("banner_id")))),
                "Last Name": clean_text((roster or academic or {}).get("last_name")),
                "First Name": clean_text((roster or academic or {}).get("first_name")),
                "Email": clean_text((roster or academic or {}).get("email")).lower(),
                "Term": info.label if info else clean_text((roster or academic or {}).get("term")),
                "Term Code": term_code,
                "Term Year": info.year if info else "",
                "Term Season": info.season if info else "",
                "Roster Present": yes_no(roster is not None),
                "Academic Present": yes_no(academic is not None),
                "Organization Entry Term": "",
                "Organization Entry Cohort": "",
                "First Observed Academic Term": "",
                "First Observed Term Overall": "",
                "Relative Term Index From Org Entry": "",
                "Relative Academic Year Window": "",
                "Is Organization Entry Term": "No",
                "Is Within First Academic Year After Org Entry": "No",
                "Chapter": clean_text(roster.get("chapter")) if roster else "",
                "Roster Source File": clean_text(roster.get("source_file")) if roster else "",
                "Roster Status Raw": clean_text(roster.get("raw_status")) if roster else "",
                "Roster Status Bucket": clean_text(roster.get("roster_status_bucket")) if roster else "",
                "Roster Position": clean_text(roster.get("position")) if roster else "",
                "Roster Semester Joined": clean_text(roster.get("semester_joined")) if roster else "",
                "New Member Marker": roster.get("new_member_marker") if roster else "No",
                "Roster Row Count Same Term": roster.get("roster_row_count_same_term") if roster else "",
                "Roster Statuses Same Term": roster.get("roster_statuses_same_term") if roster else "",
                "Roster Status Conflict Same Term": roster.get("roster_status_conflict_same_term") if roster else "No",
                "Roster Chapters Same Term": roster.get("roster_chapters_same_term") if roster else "",
                "Roster Chapter Conflict Same Term": roster.get("roster_chapter_conflict_same_term") if roster else "No",
                "Academic Source File": clean_text(academic.get("academic_source_file") or academic.get("source_file")) if academic else "",
                "Academic Student Status Raw": clean_text(academic.get("grade_student_status")) if academic else "",
                "Academic Status Signal Bucket": clean_text(academic.get("academic_status_signal_bucket")) if academic else "",
                "Major": clean_text(academic.get("major")) if academic else "",
                "Semester Hours": to_float(academic.get("semester_hours")) if academic else "",
                "Term Passed Hours": to_float(academic.get("term_passed_hours")) if academic else "",
                "Cumulative Hours": cumulative_hours if academic else "",
                "Entry Cumulative Hours Bucket": entry_hours_bucket(cumulative_hours) if academic else "Unknown",
                "Current Academic Standing Raw": clean_text(academic.get("current_academic_standing")) if academic else "",
                "Academic Standing Bucket": clean_text(academic.get("academic_standing_bucket")) if academic else "",
                "Academic Row Count Same Term": academic.get("academic_row_count_same_term") if academic else "",
                "Academic Standing Conflict Same Term": academic.get("academic_standing_conflict_same_term") if academic else "No",
                "Term GPA": to_float(academic.get("term_gpa")) if academic else "",
                "Texas State GPA": to_float(academic.get("texas_state_gpa")) if academic else "",
                "Overall GPA": to_float(academic.get("overall_gpa")) if academic else "",
                "Transfer GPA": to_float(academic.get("transfer_gpa")) if academic else "",
                "TxState Cumulative GPA": to_float(academic.get("txstate_cumulative_gpa")) if academic else "",
                "Overall Cumulative GPA": to_float(academic.get("overall_cumulative_gpa")) if academic else "",
                "Latest Known Outcome Bucket": "",
                "No Further Observation Flag": "",
            }
        )

    return master_rows, all_term_codes, roster_term_codes, academic_term_codes


def sort_summary_rows(rows: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    return sorted(
        rows,
        key=lambda row: (
            clean_text(row.get("Organization Entry Cohort")).lower(),
            clean_text(row.get("Preferred Last Name")).lower(),
            clean_text(row.get("Preferred First Name")).lower(),
            clean_text(row.get("Student ID")).lower(),
        ),
    )


def build_student_summary(
    master_rows: List[Dict[str, object]],
    all_term_codes: Sequence[str],
    roster_term_codes: Sequence[str],
    academic_term_codes: Sequence[str],
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    rows_by_student: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in master_rows:
        rows_by_student[clean_text(row.get("Student ID"))].append(row)

    summary_rows: List[Dict[str, object]] = []
    status_transition_rows: List[Dict[str, object]] = []
    standing_transition_rows: List[Dict[str, object]] = []
    max_overall_sort = term_sort_tuple(all_term_codes[-1]) if all_term_codes else None

    for student_id, student_rows in rows_by_student.items():
        ordered_rows = sorted(student_rows, key=lambda item: term_sort_tuple(clean_text(item.get("Term Code"))))
        roster_rows = [row for row in ordered_rows if row.get("Roster Present") == "Yes"]
        academic_rows = [row for row in ordered_rows if row.get("Academic Present") == "Yes"]

        first_org_term = clean_text(roster_rows[0].get("Term Code")) if roster_rows else ""
        last_org_term = clean_text(roster_rows[-1].get("Term Code")) if roster_rows else ""
        first_academic_term = clean_text(academic_rows[0].get("Term Code")) if academic_rows else ""
        last_academic_term = clean_text(academic_rows[-1].get("Term Code")) if academic_rows else ""
        first_overall_term = clean_text(ordered_rows[0].get("Term Code")) if ordered_rows else ""

        roster_status_sequence = [
            clean_text(row.get("Roster Status Bucket"))
            for row in roster_rows
            if clean_text(row.get("Roster Status Bucket"))
        ]
        standing_sequence = [
            clean_text(row.get("Academic Standing Bucket"))
            for row in academic_rows
            if clean_text(row.get("Academic Standing Bucket"))
        ]

        explicit_graduation_rows = [
            row
            for row in ordered_rows
            if clean_text(row.get("Roster Status Bucket")) in {"Graduated", "Alumni"}
            or clean_text(row.get("Academic Status Signal Bucket")) == "Graduated"
        ]
        observed_graduation_term = clean_text(explicit_graduation_rows[0].get("Term Code")) if explicit_graduation_rows else ""
        explicit_graduation_sort = term_sort_tuple(observed_graduation_term) if observed_graduation_term else None

        explicit_terminal_roster_rows = [
            row
            for row in roster_rows
            if clean_text(row.get("Roster Status Bucket"))
            in {"Suspended", "Transfer", "Dropped/Resigned/Revoked/Inactive", "Graduated", "Alumni"}
        ]
        explicit_terminal_term = clean_text(explicit_terminal_roster_rows[0].get("Term Code")) if explicit_terminal_roster_rows else ""
        explicit_terminal_sort = term_sort_tuple(explicit_terminal_term) if explicit_terminal_term else None

        latest_roster_status_bucket = clean_text(roster_rows[-1].get("Roster Status Bucket")) if roster_rows else ""
        latest_roster_status_raw = clean_text(roster_rows[-1].get("Roster Status Raw")) if roster_rows else ""
        latest_outcome_bucket = "Active/Unknown"
        if explicit_graduation_rows:
            latest_outcome_bucket = "Graduated"
        elif latest_roster_status_bucket == "Suspended":
            latest_outcome_bucket = "Suspended"
        elif latest_roster_status_bucket == "Transfer":
            latest_outcome_bucket = "Transfer"
        elif latest_roster_status_bucket == "Dropped/Resigned/Revoked/Inactive":
            latest_outcome_bucket = "Dropped/Resigned/Revoked/Inactive"
        else:
            last_overall_sort = term_sort_tuple(clean_text(ordered_rows[-1].get("Term Code"))) if ordered_rows else None
            if max_overall_sort is not None and last_overall_sort is not None and last_overall_sort < max_overall_sort:
                latest_outcome_bucket = "No Further Observation"

        first_post_entry_academic_rows = [
            row
            for row in academic_rows
            if first_org_term and term_sort_tuple(clean_text(row.get("Term Code"))) >= term_sort_tuple(first_org_term)
        ]
        first_post_entry_academic = first_post_entry_academic_rows[0] if first_post_entry_academic_rows else None
        second_post_entry_academic = first_post_entry_academic_rows[1] if len(first_post_entry_academic_rows) > 1 else None

        one_year_target = same_season_years_later(first_org_term, 1) if first_org_term else None
        first_year_rows = [
            row
            for row in first_post_entry_academic_rows
            if one_year_target
            and term_sort_tuple(clean_text(row.get("Term Code"))) < term_sort_tuple(one_year_target)
        ]
        if not one_year_target:
            first_year_rows = []

        first_term_gpa = to_float(first_post_entry_academic.get("Term GPA")) if first_post_entry_academic else None
        second_term_gpa = to_float(second_post_entry_academic.get("Term GPA")) if second_post_entry_academic else None
        first_year_average_gpa = average(to_float(row.get("Term GPA")) for row in first_year_rows)

        first_year_passed_values = [
            value
            for value in (to_float(row.get("Term Passed Hours")) for row in first_year_rows)
            if value is not None
        ]
        first_year_passed_hours = sum(first_year_passed_values) if first_year_passed_values else None

        first_term_passed_hours = to_float(first_post_entry_academic.get("Term Passed Hours")) if first_post_entry_academic else None
        entry_cumulative_hours = to_float(first_post_entry_academic.get("Cumulative Hours")) if first_post_entry_academic else None
        entry_cumulative_bucket = entry_hours_bucket(entry_cumulative_hours)

        first_term_txstate_cum = to_float(first_post_entry_academic.get("TxState Cumulative GPA")) if first_post_entry_academic else None
        first_term_overall_cum = to_float(first_post_entry_academic.get("Overall Cumulative GPA")) if first_post_entry_academic else None
        latest_txstate_cum = next(
            (
                to_float(row.get("TxState Cumulative GPA"))
                for row in reversed(first_post_entry_academic_rows)
                if to_float(row.get("TxState Cumulative GPA")) is not None
            ),
            None,
        )
        latest_overall_cum = next(
            (
                to_float(row.get("Overall Cumulative GPA"))
                for row in reversed(first_post_entry_academic_rows)
                if to_float(row.get("Overall Cumulative GPA")) is not None
            ),
            None,
        )

        org_observed_codes = [clean_text(row.get("Term Code")) for row in roster_rows]
        academic_observed_codes = [clean_text(row.get("Term Code")) for row in academic_rows]
        all_observed_codes = [clean_text(row.get("Term Code")) for row in ordered_rows]

        org_next_term_target = next_observed_term(first_org_term, roster_term_codes) if first_org_term else None
        org_next_fall_target = next_fall_code(first_org_term) if first_org_term else None
        org_year_two_target = same_season_years_later(first_org_term, 1) if first_org_term else None
        acad_next_term_target = next_observed_term(first_org_term, academic_term_codes) if first_org_term else None
        acad_next_fall_target = next_fall_code(first_org_term) if first_org_term else None
        acad_year_two_target = same_season_years_later(first_org_term, 1) if first_org_term else None

        last_observed_sort = term_sort_tuple(all_observed_codes[-1]) if all_observed_codes else None

        org_next_term_flag, org_next_term_category = followup_category(
            org_next_term_target,
            roster_term_codes,
            org_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )
        org_next_fall_flag, org_next_fall_category = followup_category(
            org_next_fall_target,
            roster_term_codes,
            org_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )
        org_year_two_flag, org_year_two_category = followup_category(
            org_year_two_target,
            roster_term_codes,
            org_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )
        acad_next_term_flag, acad_next_term_category = followup_category(
            acad_next_term_target,
            academic_term_codes,
            academic_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )
        acad_next_fall_flag, acad_next_fall_category = followup_category(
            acad_next_fall_target,
            academic_term_codes,
            academic_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )
        acad_year_two_flag, acad_year_two_category = followup_category(
            acad_year_two_target,
            academic_term_codes,
            academic_observed_codes,
            explicit_graduation_sort,
            explicit_terminal_sort,
            last_observed_sort,
        )

        grad_4_year_target = same_season_years_later(first_org_term, 4) if first_org_term else None
        grad_6_year_target = same_season_years_later(first_org_term, 6) if first_org_term else None
        grad_4_measurable = yes_no(
            bool(grad_4_year_target)
            and bool(all_term_codes)
            and term_sort_tuple(all_term_codes[-1]) >= term_sort_tuple(grad_4_year_target)
        )
        grad_6_measurable = yes_no(
            bool(grad_6_year_target)
            and bool(all_term_codes)
            and term_sort_tuple(all_term_codes[-1]) >= term_sort_tuple(grad_6_year_target)
        )
        grad_4_flag = ""
        grad_6_flag = ""
        if grad_4_measurable == "Yes":
            grad_4_flag = yes_no(
                explicit_graduation_sort is not None
                and explicit_graduation_sort <= term_sort_tuple(grad_4_year_target)
            )
        if grad_6_measurable == "Yes":
            grad_6_flag = yes_no(
                explicit_graduation_sort is not None
                and explicit_graduation_sort <= term_sort_tuple(grad_6_year_target)
            )

        summary = {
            "Student ID": student_id,
            "Identity Resolution Basis Used": " | ".join(
                unique_sorted(row.get("Identity Resolution Basis", "") for row in ordered_rows)
            ),
            "Preferred Last Name": clean_text(ordered_rows[-1].get("Last Name")),
            "Preferred First Name": clean_text(ordered_rows[-1].get("First Name")),
            "Preferred Email": clean_text(ordered_rows[-1].get("Email")).lower(),
            "First Observed Organization Term": parse_term_code(first_org_term).label if first_org_term else "",
            "Last Observed Organization Term": parse_term_code(last_org_term).label if last_org_term else "",
            "First Observed Academic Term": parse_term_code(first_academic_term).label if first_academic_term else "",
            "Last Observed Academic Term": parse_term_code(last_academic_term).label if last_academic_term else "",
            "First Observed Term Overall": parse_term_code(first_overall_term).label if first_overall_term else "",
            "Organization Entry Cohort": parse_term_code(first_org_term).label if first_org_term else "",
            "Initial Chapter": clean_text(roster_rows[0].get("Chapter")) if roster_rows else "",
            "Initial Roster Status Raw": clean_text(roster_rows[0].get("Roster Status Raw")) if roster_rows else "",
            "Initial Roster Status Bucket": clean_text(roster_rows[0].get("Roster Status Bucket")) if roster_rows else "",
            "Latest Chapter": clean_text(roster_rows[-1].get("Chapter")) if roster_rows else "",
            "Latest Known Roster Status Raw": latest_roster_status_raw,
            "Latest Known Roster Status Bucket": latest_roster_status_bucket,
            "Latest Known Outcome Bucket": latest_outcome_bucket,
            "Ever Graduated Flag": yes_no(bool(explicit_graduation_rows)),
            "Ever Suspended Flag": yes_no(
                any(
                    clean_text(row.get("Roster Status Bucket")) == "Suspended"
                    or clean_text(row.get("Academic Status Signal Bucket")) == "Suspended"
                    or clean_text(row.get("Academic Standing Bucket")) == "Suspended"
                    for row in ordered_rows
                )
            ),
            "Ever Dropped/Inactive/Resigned/Revoked Flag": yes_no(
                any(
                    clean_text(row.get("Roster Status Bucket")) == "Dropped/Resigned/Revoked/Inactive"
                    or clean_text(row.get("Academic Status Signal Bucket")) == "Dropped/Resigned/Revoked/Inactive"
                    for row in ordered_rows
                )
            ),
            "Ever Transfer Flag": yes_no(
                any(
                    clean_text(row.get("Roster Status Bucket")) == "Transfer"
                    or clean_text(row.get("Academic Status Signal Bucket")) == "Transfer"
                    for row in ordered_rows
                )
            ),
            "No Further Observation Flag": yes_no(latest_outcome_bucket == "No Further Observation"),
            "Observed Graduation Term": parse_term_code(observed_graduation_term).label if observed_graduation_term else "",
            "First Post-Entry Academic Term": clean_text(first_post_entry_academic.get("Term")) if first_post_entry_academic else "",
            "Second Post-Entry Academic Term": clean_text(second_post_entry_academic.get("Term")) if second_post_entry_academic else "",
            "First Major After Org Entry": clean_text(first_post_entry_academic.get("Major")) if first_post_entry_academic else "",
            "Latest Major": clean_text(academic_rows[-1].get("Major")) if academic_rows else "",
            "First Academic Standing After Org Entry": clean_text(first_post_entry_academic.get("Academic Standing Bucket")) if first_post_entry_academic else "",
            "Latest Academic Standing": clean_text(academic_rows[-1].get("Academic Standing Bucket")) if academic_rows else "",
            "Entry Cumulative Hours": entry_cumulative_hours if entry_cumulative_hours is not None else "",
            "Entry Cumulative Hours Bucket": entry_cumulative_bucket,
            "Roster Terms Observed": len(roster_rows),
            "Academic Terms Observed": len(academic_rows),
            "Total Terms Observed Overall": len(ordered_rows),
            "Organization Next Observed Term Measurable": yes_no(
                org_next_term_target in roster_term_codes if org_next_term_target else False
            ),
            "Organization Next Observed Term Category": org_next_term_category,
            "Retained In Organization To Next Observed Term": org_next_term_flag,
            "Organization Next Fall Measurable": yes_no(
                org_next_fall_target in roster_term_codes if org_next_fall_target else False
            ),
            "Organization Next Fall Category": org_next_fall_category,
            "Retained In Organization To Next Fall": org_next_fall_flag,
            "Organization One-Year Same-Season Measurable": yes_no(
                org_year_two_target in roster_term_codes if org_year_two_target else False
            ),
            "Organization One-Year Same-Season Category": org_year_two_category,
            "Retained In Organization One Year After Entry": org_year_two_flag,
            "Academic Next Observed Term Measurable": yes_no(
                acad_next_term_target in academic_term_codes if acad_next_term_target else False
            ),
            "Academic Next Observed Term Category": acad_next_term_category,
            "Continued Academically To Next Observed Term": acad_next_term_flag,
            "Academic Next Fall Measurable": yes_no(
                acad_next_fall_target in academic_term_codes if acad_next_fall_target else False
            ),
            "Academic Next Fall Category": acad_next_fall_category,
            "Continued Academically To Next Fall": acad_next_fall_flag,
            "Academic One-Year Same-Season Measurable": yes_no(
                acad_year_two_target in academic_term_codes if acad_year_two_target else False
            ),
            "Academic One-Year Same-Season Category": acad_year_two_category,
            "Continued Academically One Year After Entry": acad_year_two_flag,
            "Eventual Observed Graduation From Org Entry": yes_no(bool(explicit_graduation_rows)) if first_org_term else "",
            "Observed Graduation Within 4 Years Of Org Entry Measurable": grad_4_measurable,
            "Observed Graduation Within 4 Years Of Org Entry": grad_4_flag,
            "Observed Graduation Within 6 Years Of Org Entry Measurable": grad_6_measurable,
            "Observed Graduation Within 6 Years Of Org Entry": grad_6_flag,
            "Eventual Observed Graduation From First Observed Academic Term": (
                yes_no(bool(explicit_graduation_rows)) if first_academic_term else ""
            ),
            "First Post-Entry Term GPA": first_term_gpa if first_term_gpa is not None else "",
            "Second Post-Entry Term GPA": second_term_gpa if second_term_gpa is not None else "",
            "First-Year Average Term GPA After Org Entry": first_year_average_gpa if first_year_average_gpa is not None else "",
            "Change In Term GPA First To Second Term": (
                second_term_gpa - first_term_gpa
                if first_term_gpa is not None and second_term_gpa is not None
                else ""
            ),
            "First Post-Entry TxState Cumulative GPA": first_term_txstate_cum if first_term_txstate_cum is not None else "",
            "Latest TxState Cumulative GPA": latest_txstate_cum if latest_txstate_cum is not None else "",
            "First Post-Entry Overall Cumulative GPA": first_term_overall_cum if first_term_overall_cum is not None else "",
            "Latest Overall Cumulative GPA": latest_overall_cum if latest_overall_cum is not None else "",
            "First Post-Entry Passed Hours": first_term_passed_hours if first_term_passed_hours is not None else "",
            "First-Year Passed Hours After Org Entry": first_year_passed_hours if first_year_passed_hours is not None else "",
            "First-Term 12+ Passed Hours Flag": yes_no(first_term_passed_hours is not None and first_term_passed_hours >= 12),
            "First-Term 15+ Passed Hours Flag": yes_no(first_term_passed_hours is not None and first_term_passed_hours >= 15),
            "First-Year 24+ Passed Hours Flag": yes_no(first_year_passed_hours is not None and first_year_passed_hours >= 24),
            "First-Year 30+ Passed Hours Flag": yes_no(first_year_passed_hours is not None and first_year_passed_hours >= 30),
            "First-Term GPA Below 2.0 Flag": yes_no(first_term_gpa is not None and first_term_gpa < 2.0),
            "First-Term GPA Below 2.5 Flag": yes_no(first_term_gpa is not None and first_term_gpa < 2.5),
            "First-Year Average GPA Below 2.0 Flag": yes_no(first_year_average_gpa is not None and first_year_average_gpa < 2.0),
            "First-Year Average GPA Below 2.5 Flag": yes_no(first_year_average_gpa is not None and first_year_average_gpa < 2.5),
            "First-Term Good Standing Flag": yes_no(
                first_post_entry_academic is not None
                and clean_text(first_post_entry_academic.get("Academic Standing Bucket")) == "Good Standing"
            ),
            "First-Year Probation/Warning Flag": yes_no(
                any(clean_text(row.get("Academic Standing Bucket")) == "Probation/Warning" for row in first_year_rows)
            ),
            "Academic Standing Suspended Ever Flag": yes_no(
                any(clean_text(row.get("Academic Standing Bucket")) == "Suspended" for row in academic_rows)
            ),
            "Roster Status Transition Count": sum(
                1
                for idx in range(1, len(roster_status_sequence))
                if roster_status_sequence[idx] != roster_status_sequence[idx - 1]
            ),
            "Academic Standing Transition Count": sum(
                1
                for idx in range(1, len(standing_sequence))
                if standing_sequence[idx] != standing_sequence[idx - 1]
            ),
        }
        summary_rows.append(summary)

        if first_org_term:
            entry_index = all_term_codes.index(first_org_term) if first_org_term in all_term_codes else None
            for row in ordered_rows:
                term_code = clean_text(row.get("Term Code"))
                row["Organization Entry Term"] = parse_term_code(first_org_term).label if first_org_term else ""
                row["Organization Entry Cohort"] = parse_term_code(first_org_term).label if first_org_term else ""
                row["First Observed Academic Term"] = parse_term_code(first_academic_term).label if first_academic_term else ""
                row["First Observed Term Overall"] = parse_term_code(first_overall_term).label if first_overall_term else ""
                row["Latest Known Outcome Bucket"] = latest_outcome_bucket
                row["No Further Observation Flag"] = yes_no(latest_outcome_bucket == "No Further Observation")
                row["Entry Cumulative Hours Bucket"] = entry_cumulative_bucket
                if entry_index is not None and term_code in all_term_codes:
                    relative_index = all_term_codes.index(term_code) - entry_index
                    row["Relative Term Index From Org Entry"] = relative_index
                    row["Is Organization Entry Term"] = yes_no(relative_index == 0)
                    entry_info = parse_term_code(first_org_term)
                    current_info = parse_term_code(term_code)
                    if entry_info and current_info:
                        relative_year_window = current_info.year - entry_info.year
                        if current_info.season != "Unknown" and entry_info.season != "Unknown":
                            if SEASON_ORDER[current_info.season] < SEASON_ORDER[entry_info.season]:
                                relative_year_window -= 1
                        row["Relative Academic Year Window"] = max(relative_year_window, 0)
                    if one_year_target and term_sort_tuple(term_code) < term_sort_tuple(one_year_target):
                        row["Is Within First Academic Year After Org Entry"] = "Yes"

        for idx in range(1, len(roster_rows)):
            from_bucket = clean_text(roster_rows[idx - 1].get("Roster Status Bucket"))
            to_bucket = clean_text(roster_rows[idx].get("Roster Status Bucket"))
            if not from_bucket or not to_bucket:
                continue
            from_term_code = clean_text(roster_rows[idx - 1].get("Term Code"))
            from_relative = ""
            if first_org_term and from_term_code in all_term_codes and first_org_term in all_term_codes:
                from_relative = all_term_codes.index(from_term_code) - all_term_codes.index(first_org_term)
            status_transition_rows.append(
                {
                    "Student ID": student_id,
                    "Transition Family": "Roster Status",
                    "Cohort": parse_term_code(first_org_term).label if first_org_term else "",
                    "From Value": from_bucket,
                    "To Value": to_bucket,
                    "From Relative Term Index": from_relative,
                    "Transition Count": 1,
                    "Distinct Students": 1,
                }
            )

        for idx in range(1, len(academic_rows)):
            from_bucket = clean_text(academic_rows[idx - 1].get("Academic Standing Bucket"))
            to_bucket = clean_text(academic_rows[idx].get("Academic Standing Bucket"))
            if not from_bucket or not to_bucket:
                continue
            from_term_code = clean_text(academic_rows[idx - 1].get("Term Code"))
            from_relative = ""
            if first_org_term and from_term_code in all_term_codes and first_org_term in all_term_codes:
                from_relative = all_term_codes.index(from_term_code) - all_term_codes.index(first_org_term)
            standing_transition_rows.append(
                {
                    "Student ID": student_id,
                    "Transition Family": "Academic Standing",
                    "Cohort": parse_term_code(first_org_term).label if first_org_term else "",
                    "From Value": from_bucket,
                    "To Value": to_bucket,
                    "From Relative Term Index": from_relative,
                    "Transition Count": 1,
                    "Distinct Students": 1,
                }
            )

    return sort_summary_rows(summary_rows), status_transition_rows, standing_transition_rows


def count_yes(rows: Sequence[Dict[str, object]], field: str) -> int:
    return sum(1 for row in rows if clean_text(row.get(field)) == "Yes")


def average_numeric(rows: Sequence[Dict[str, object]], field: str) -> Optional[float]:
    return average(to_float(row.get(field)) for row in rows)


def add_metric_row(
    rows: List[Dict[str, object]],
    metric_group: str,
    metric_label: str,
    cohort: str,
    dimension: str,
    value: str,
    eligible_students: object = "",
    student_count: object = "",
    rate: object = "",
    average_value: object = "",
    notes: str = "",
) -> None:
    rows.append(
        {
            "Metric Group": metric_group,
            "Metric Label": metric_label,
            "Cohort": cohort,
            "Dimension": dimension,
            "Value": value,
            "Eligible Students": eligible_students,
            "Student Count": student_count,
            "Rate": rate,
            "Average Value": average_value,
            "Notes": notes,
        }
    )


def build_cohort_metrics(summary_rows: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    metrics: List[Dict[str, object]] = []
    groups: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in summary_rows:
        cohort = clean_text(row.get("Organization Entry Cohort")) or "Unknown"
        groups[cohort].append(row)
        groups["Overall"].append(row)

    for cohort, rows in sorted(groups.items(), key=lambda item: ("ZZZZ" if item[0] == "Overall" else item[0])):
        cohort_size = len(rows)
        add_metric_row(metrics, "Cohort Counts", "Cohort size", cohort, "All", "Cohort size", student_count=cohort_size)

        initial_status_counts = Counter(clean_text(row.get("Initial Roster Status Bucket")) or "Unknown" for row in rows)
        for bucket in PRIMARY_STATUS_BUCKETS:
            count = initial_status_counts.get(bucket, 0)
            add_metric_row(
                metrics,
                "Cohort Counts",
                "Initial roster status bucket",
                cohort,
                "Initial Status Bucket",
                bucket,
                student_count=count,
                rate=(count / cohort_size) if cohort_size else "",
            )

        latest_outcome_counts = Counter(clean_text(row.get("Latest Known Outcome Bucket")) or "Active/Unknown" for row in rows)
        for bucket in OUTCOME_BUCKETS:
            count = latest_outcome_counts.get(bucket, 0)
            add_metric_row(
                metrics,
                "Cohort Counts",
                "Latest known outcome bucket",
                cohort,
                "Latest Outcome Bucket",
                bucket,
                student_count=count,
                rate=(count / cohort_size) if cohort_size else "",
            )

        for label, measurable_field, category_field, retained_field in [
            (
                "Next observed term after first observed organization term",
                "Organization Next Observed Term Measurable",
                "Organization Next Observed Term Category",
                "Retained In Organization To Next Observed Term",
            ),
            (
                "Next fall after first observed organization term",
                "Organization Next Fall Measurable",
                "Organization Next Fall Category",
                "Retained In Organization To Next Fall",
            ),
            (
                "Same-season term one year after first observed organization term",
                "Organization One-Year Same-Season Measurable",
                "Organization One-Year Same-Season Category",
                "Retained In Organization One Year After Entry",
            ),
        ]:
            eligible = [row for row in rows if clean_text(row.get(measurable_field)) == "Yes"]
            category_counts = Counter(clean_text(row.get(category_field)) or "Unknown" for row in eligible)
            for category in [
                "Observed at follow-up",
                "Graduated before follow-up",
                "Terminal status before follow-up",
                "No further observation before follow-up",
                "Not observed at follow-up",
            ]:
                count = category_counts.get(category, 0)
                add_metric_row(
                    metrics,
                    "Organization Retention",
                    label,
                    cohort,
                    "Follow-Up Category",
                    category,
                    eligible_students=len(eligible),
                    student_count=count,
                    rate=(count / len(eligible)) if eligible else "",
                    notes="Graduates and terminal statuses are separated from simple non-retention.",
                )
            retained_count = count_yes(eligible, retained_field)
            add_metric_row(
                metrics,
                "Organization Retention",
                label,
                cohort,
                "Retained Flag",
                "Yes",
                eligible_students=len(eligible),
                student_count=retained_count,
                rate=(retained_count / len(eligible)) if eligible else "",
            )

        for label, measurable_field, category_field, retained_field in [
            (
                "Next observed academic term after first observed organization term",
                "Academic Next Observed Term Measurable",
                "Academic Next Observed Term Category",
                "Continued Academically To Next Observed Term",
            ),
            (
                "Next fall after first observed organization term",
                "Academic Next Fall Measurable",
                "Academic Next Fall Category",
                "Continued Academically To Next Fall",
            ),
            (
                "Same-season term one year after first observed organization term",
                "Academic One-Year Same-Season Measurable",
                "Academic One-Year Same-Season Category",
                "Continued Academically One Year After Entry",
            ),
        ]:
            eligible = [row for row in rows if clean_text(row.get(measurable_field)) == "Yes"]
            category_counts = Counter(clean_text(row.get(category_field)) or "Unknown" for row in eligible)
            for category in [
                "Observed at follow-up",
                "Graduated before follow-up",
                "Terminal status before follow-up",
                "No further observation before follow-up",
                "Not observed at follow-up",
            ]:
                count = category_counts.get(category, 0)
                add_metric_row(
                    metrics,
                    "Institutional Continuation",
                    label,
                    cohort,
                    "Follow-Up Category",
                    category,
                    eligible_students=len(eligible),
                    student_count=count,
                    rate=(count / len(eligible)) if eligible else "",
                )
            continued_count = count_yes(eligible, retained_field)
            add_metric_row(
                metrics,
                "Institutional Continuation",
                label,
                cohort,
                "Continuation Flag",
                "Yes",
                eligible_students=len(eligible),
                student_count=continued_count,
                rate=(continued_count / len(eligible)) if eligible else "",
            )

        org_entry_students = [row for row in rows if clean_text(row.get("First Observed Organization Term"))]
        eventual_grad_count = count_yes(org_entry_students, "Eventual Observed Graduation From Org Entry")
        add_metric_row(
            metrics,
            "Graduation Outcomes",
            "Observed eventual graduation from first observed organization term",
            cohort,
            "Observed Graduation",
            "Yes",
            eligible_students=len(org_entry_students),
            student_count=eventual_grad_count,
            rate=(eventual_grad_count / len(org_entry_students)) if org_entry_students else "",
            notes="Observed graduation only; not a true first-time-in-college graduation rate.",
        )

        grad4_eligible = [
            row
            for row in rows
            if clean_text(row.get("Observed Graduation Within 4 Years Of Org Entry Measurable")) == "Yes"
        ]
        grad4_count = count_yes(grad4_eligible, "Observed Graduation Within 4 Years Of Org Entry")
        add_metric_row(
            metrics,
            "Graduation Outcomes",
            "Graduated within 4 years of first observed organization term",
            cohort,
            "Observed Graduation",
            "Yes",
            eligible_students=len(grad4_eligible),
            student_count=grad4_count,
            rate=(grad4_count / len(grad4_eligible)) if grad4_eligible else "",
            notes="Right-censored to cohorts with enough elapsed observed time.",
        )

        grad6_eligible = [
            row
            for row in rows
            if clean_text(row.get("Observed Graduation Within 6 Years Of Org Entry Measurable")) == "Yes"
        ]
        grad6_count = count_yes(grad6_eligible, "Observed Graduation Within 6 Years Of Org Entry")
        add_metric_row(
            metrics,
            "Graduation Outcomes",
            "Graduated within 6 years of first observed organization term",
            cohort,
            "Observed Graduation",
            "Yes",
            eligible_students=len(grad6_eligible),
            student_count=grad6_count,
            rate=(grad6_count / len(grad6_eligible)) if grad6_eligible else "",
            notes="Right-censored to cohorts with enough elapsed observed time.",
        )

        academic_entry_students = [row for row in rows if clean_text(row.get("First Observed Academic Term"))]
        eventual_grad_academic_count = count_yes(
            academic_entry_students, "Eventual Observed Graduation From First Observed Academic Term"
        )
        add_metric_row(
            metrics,
            "Graduation Outcomes",
            "Observed eventual graduation from first observed academic term",
            cohort,
            "Observed Graduation",
            "Yes",
            eligible_students=len(academic_entry_students),
            student_count=eventual_grad_academic_count,
            rate=(eventual_grad_academic_count / len(academic_entry_students)) if academic_entry_students else "",
        )

        first_term_hours_rows = [row for row in rows if to_float(row.get("First Post-Entry Passed Hours")) is not None]
        add_metric_row(
            metrics,
            "Credit Momentum",
            "Average passed hours in first academic term after organization entry",
            cohort,
            "Average",
            "Passed Hours",
            eligible_students=len(first_term_hours_rows),
            average_value=average_numeric(first_term_hours_rows, "First Post-Entry Passed Hours") or "",
        )

        first_year_hours_rows = [row for row in rows if to_float(row.get("First-Year Passed Hours After Org Entry")) is not None]
        add_metric_row(
            metrics,
            "Credit Momentum",
            "Average passed hours in first academic year after organization entry",
            cohort,
            "Average",
            "Passed Hours",
            eligible_students=len(first_year_hours_rows),
            average_value=average_numeric(first_year_hours_rows, "First-Year Passed Hours After Org Entry") or "",
        )

        for label, field, source_field in [
            ("Passed 12+ hours in first academic term after organization entry", "First-Term 12+ Passed Hours Flag", "First Post-Entry Passed Hours"),
            ("Passed 15+ hours in first academic term after organization entry", "First-Term 15+ Passed Hours Flag", "First Post-Entry Passed Hours"),
            ("Passed 24+ hours in first academic year after organization entry", "First-Year 24+ Passed Hours Flag", "First-Year Passed Hours After Org Entry"),
            ("Passed 30+ hours in first academic year after organization entry", "First-Year 30+ Passed Hours Flag", "First-Year Passed Hours After Org Entry"),
        ]:
            eligible = [row for row in rows if to_float(row.get(source_field)) is not None]
            count = count_yes(eligible, field)
            add_metric_row(
                metrics,
                "Credit Momentum",
                label,
                cohort,
                "Flag",
                "Yes",
                eligible_students=len(eligible),
                student_count=count,
                rate=(count / len(eligible)) if eligible else "",
            )

        for label, field in [
            ("Average first-term GPA after organization entry", "First Post-Entry Term GPA"),
            ("Average second-term GPA after organization entry", "Second Post-Entry Term GPA"),
            ("Average first-year term GPA after organization entry", "First-Year Average Term GPA After Org Entry"),
            ("Average GPA change from first to second observed academic term", "Change In Term GPA First To Second Term"),
            ("Average latest TxState cumulative GPA after organization entry", "Latest TxState Cumulative GPA"),
            ("Average latest overall cumulative GPA after organization entry", "Latest Overall Cumulative GPA"),
        ]:
            eligible = [row for row in rows if to_float(row.get(field)) is not None]
            add_metric_row(
                metrics,
                "GPA and Academic Progress",
                label,
                cohort,
                "Average",
                field,
                eligible_students=len(eligible),
                average_value=average_numeric(eligible, field) or "",
            )

        for label, field, source_field in [
            ("First-term GPA below 2.0 after organization entry", "First-Term GPA Below 2.0 Flag", "First Post-Entry Term GPA"),
            ("First-term GPA below 2.5 after organization entry", "First-Term GPA Below 2.5 Flag", "First Post-Entry Term GPA"),
            ("First-year average GPA below 2.0 after organization entry", "First-Year Average GPA Below 2.0 Flag", "First-Year Average Term GPA After Org Entry"),
            ("First-year average GPA below 2.5 after organization entry", "First-Year Average GPA Below 2.5 Flag", "First-Year Average Term GPA After Org Entry"),
        ]:
            eligible = [row for row in rows if to_float(row.get(source_field)) is not None]
            count = count_yes(eligible, field)
            add_metric_row(
                metrics,
                "GPA and Academic Progress",
                label,
                cohort,
                "Flag",
                "Yes",
                eligible_students=len(eligible),
                student_count=count,
                rate=(count / len(eligible)) if eligible else "",
            )

        standing_rows = [row for row in rows if clean_text(row.get("First Academic Standing After Org Entry"))]
        for bucket in STANDING_BUCKETS:
            count = sum(1 for row in standing_rows if clean_text(row.get("First Academic Standing After Org Entry")) == bucket)
            add_metric_row(
                metrics,
                "Academic Standing",
                "First observed academic standing after organization entry",
                cohort,
                "Standing Bucket",
                bucket,
                eligible_students=len(standing_rows),
                student_count=count,
                rate=(count / len(standing_rows)) if standing_rows else "",
            )

        first_term_standing_rows = [row for row in rows if clean_text(row.get("First Academic Standing After Org Entry"))]
        first_term_good_count = count_yes(first_term_standing_rows, "First-Term Good Standing Flag")
        add_metric_row(
            metrics,
            "Academic Standing",
            "Good standing in first observed academic term after organization entry",
            cohort,
            "Flag",
            "Yes",
            eligible_students=len(first_term_standing_rows),
            student_count=first_term_good_count,
            rate=(first_term_good_count / len(first_term_standing_rows)) if first_term_standing_rows else "",
        )

        first_year_observed_rows = [
            row
            for row in rows
            if to_float(row.get("First-Year Average Term GPA After Org Entry")) is not None
            or to_float(row.get("First-Year Passed Hours After Org Entry")) is not None
            or clean_text(row.get("First-Year Probation/Warning Flag")) == "Yes"
        ]
        first_year_probation_count = count_yes(first_year_observed_rows, "First-Year Probation/Warning Flag")
        add_metric_row(
            metrics,
            "Academic Standing",
            "Probation/warning during first academic year after organization entry",
            cohort,
            "Flag",
            "Yes",
            eligible_students=len(first_year_observed_rows),
            student_count=first_year_probation_count,
            rate=(first_year_probation_count / len(first_year_observed_rows)) if first_year_observed_rows else "",
        )

        standing_suspended_count = count_yes(rows, "Academic Standing Suspended Ever Flag")
        add_metric_row(
            metrics,
            "Academic Standing",
            "Ever observed with suspended academic standing",
            cohort,
            "Flag",
            "Yes",
            eligible_students=cohort_size,
            student_count=standing_suspended_count,
            rate=(standing_suspended_count / cohort_size) if cohort_size else "",
        )

    return metrics


def aggregate_transition_rows(rows: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    grouped: Dict[Tuple[str, str, str, str, object], Dict[str, object]] = {}
    for row in rows:
        key = (
            clean_text(row.get("Transition Family")),
            clean_text(row.get("Cohort")),
            clean_text(row.get("From Value")),
            clean_text(row.get("To Value")),
            row.get("From Relative Term Index"),
        )
        if key not in grouped:
            grouped[key] = {"count": 0, "students": set()}
        grouped[key]["count"] += 1
        grouped[key]["students"].add(clean_text(row.get("Student ID")))

    aggregated: List[Dict[str, object]] = []
    for (family, cohort, from_value, to_value, from_relative), value in grouped.items():
        aggregated.append(
            {
                "Transition Family": family,
                "Cohort": cohort,
                "From Value": from_value,
                "To Value": to_value,
                "From Relative Term Index": from_relative,
                "Transition Count": value["count"],
                "Distinct Students": len({student for student in value["students"] if student}),
            }
        )

    return sorted(
        aggregated,
        key=lambda row: (
            clean_text(row.get("Transition Family")).lower(),
            clean_text(row.get("Cohort")).lower(),
            row.get("From Relative Term Index") if row.get("From Relative Term Index") != "" else 999,
            clean_text(row.get("From Value")).lower(),
            clean_text(row.get("To Value")).lower(),
        ),
    )


def build_outcome_segments(
    summary_rows: Sequence[Dict[str, object]],
    min_size: int,
) -> List[Dict[str, object]]:
    segment_rows: List[Dict[str, object]] = []
    dimensions = {
        "Initial Roster Status Bucket": "Initial Roster Status Bucket",
        "Latest Known Outcome Bucket": "Latest Known Outcome Bucket",
        "First Major After Org Entry": "First Major After Org Entry",
        "First Academic Standing After Org Entry": "First Academic Standing After Org Entry",
        "Entry Cumulative Hours Bucket": "Entry Cumulative Hours Bucket",
    }

    cohort_groups: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in summary_rows:
        cohort = clean_text(row.get("Organization Entry Cohort")) or "Unknown"
        cohort_groups["Overall"].append(row)
        cohort_groups[cohort].append(row)

    for cohort, rows in sorted(cohort_groups.items(), key=lambda item: ("ZZZZ" if item[0] == "Overall" else item[0])):
        for dimension_label, field in dimensions.items():
            grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
            for row in rows:
                grouped[clean_text(row.get(field)) or "Unknown"].append(row)

            for value, group in sorted(grouped.items(), key=lambda item: (len(item[1]) * -1, item[0].lower())):
                if len(group) < min_size:
                    continue
                grad4_eligible = [
                    row
                    for row in group
                    if clean_text(row.get("Observed Graduation Within 4 Years Of Org Entry Measurable")) == "Yes"
                ]
                grad6_eligible = [
                    row
                    for row in group
                    if clean_text(row.get("Observed Graduation Within 6 Years Of Org Entry Measurable")) == "Yes"
                ]
                next_fall_org_eligible = [
                    row for row in group if clean_text(row.get("Organization Next Fall Measurable")) == "Yes"
                ]
                next_fall_acad_eligible = [
                    row for row in group if clean_text(row.get("Academic Next Fall Measurable")) == "Yes"
                ]
                segment_rows.append(
                    {
                        "Cohort": cohort,
                        "Dimension": dimension_label,
                        "Value": value,
                        "Group Size": len(group),
                        "Observed Eventual Graduation Rate From Org Entry": (
                            count_yes(group, "Eventual Observed Graduation From Org Entry") / len(group)
                        ),
                        "Observed Graduation Within 4 Years Of Org Entry": (
                            count_yes(grad4_eligible, "Observed Graduation Within 4 Years Of Org Entry") / len(grad4_eligible)
                            if grad4_eligible
                            else ""
                        ),
                        "Observed Graduation Within 6 Years Of Org Entry": (
                            count_yes(grad6_eligible, "Observed Graduation Within 6 Years Of Org Entry") / len(grad6_eligible)
                            if grad6_eligible
                            else ""
                        ),
                        "Retained In Organization To Next Fall": (
                            count_yes(next_fall_org_eligible, "Retained In Organization To Next Fall") / len(next_fall_org_eligible)
                            if next_fall_org_eligible
                            else ""
                        ),
                        "Continued Academically To Next Fall": (
                            count_yes(next_fall_acad_eligible, "Continued Academically To Next Fall") / len(next_fall_acad_eligible)
                            if next_fall_acad_eligible
                            else ""
                        ),
                        "Average First-Year Term GPA After Org Entry": average_numeric(
                            group, "First-Year Average Term GPA After Org Entry"
                        )
                        or "",
                        "Average First-Year Passed Hours After Org Entry": average_numeric(
                            group, "First-Year Passed Hours After Org Entry"
                        )
                        or "",
                        "Share First-Term GPA Below 2.0": count_yes(group, "First-Term GPA Below 2.0 Flag") / len(group),
                        "Share First-Term GPA Below 2.5": count_yes(group, "First-Term GPA Below 2.5 Flag") / len(group),
                    }
                )

    return segment_rows


def build_status_mapping_rows() -> List[Dict[str, object]]:
    return [
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Graduated",
            "Rule": "Contains graduate / graduated",
            "Notes": "Counts as observed graduation.",
        },
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Alumni",
            "Rule": "Contains alumni",
            "Notes": "Tracked separately in status mapping but counted with graduation outcomes where appropriate.",
        },
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Suspended",
            "Rule": "Contains suspend",
            "Notes": "",
        },
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Transfer",
            "Rule": "Contains transfer",
            "Notes": "",
        },
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Dropped/Resigned/Revoked/Inactive",
            "Rule": "Contains inactive, resign, revoked, drop, removed",
            "Notes": "Kept additive; raw status text is preserved separately.",
        },
        {
            "Field": "Roster Status / Position",
            "Derived Bucket": "Active",
            "Rule": "Contains active, member, or new member marker",
            "Notes": "",
        },
        {
            "Field": "Academic Standing",
            "Derived Bucket": "Good Standing",
            "Rule": "Contains good standing",
            "Notes": "",
        },
        {
            "Field": "Academic Standing",
            "Derived Bucket": "Probation/Warning",
            "Rule": "Contains probation, warning, or alert",
            "Notes": "",
        },
        {
            "Field": "Academic Standing",
            "Derived Bucket": "Suspended",
            "Rule": "Contains suspend",
            "Notes": "",
        },
        {
            "Field": "Academic Standing",
            "Derived Bucket": "Dismissed/Separated",
            "Rule": "Contains dismiss, drop, or separate",
            "Notes": "",
        },
        {
            "Field": "Identity Resolution",
            "Derived Bucket": "Derived from unique email",
            "Rule": "Missing Banner ID but email mapped to exactly one known Student ID",
            "Notes": "Used only to preserve additive rows without inventing non-unique IDs.",
        },
    ]


def build_metric_definition_rows() -> List[Dict[str, object]]:
    return [
        {
            "Metric Group": "Cohort Counts",
            "Metric Label": "Cohort size",
            "Definition": "Count of distinct students with a first observed organization term in the cohort.",
            "Window / Denominator": "All students in the cohort.",
            "Limitations / Notes": "Organization-entry cohort is based on first observed organization term, not true first institution term.",
        },
        {
            "Metric Group": "Organization Retention",
            "Metric Label": "Retained in organization to next observed term",
            "Definition": "Student appears on an organization roster in the next observed roster term after first observed organization term.",
            "Window / Denominator": "Only students with a measurable next observed roster term.",
            "Limitations / Notes": "Graduates and other terminal statuses before the follow-up term are reported separately.",
        },
        {
            "Metric Group": "Organization Retention",
            "Metric Label": "Retained in organization to next fall",
            "Definition": "Student appears on an organization roster in the first fall term after first observed organization term.",
            "Window / Denominator": "Only cohorts where the follow-up fall exists in the roster calendar.",
            "Limitations / Notes": "",
        },
        {
            "Metric Group": "Institutional Continuation",
            "Metric Label": "Continued academically to next observed term / next fall / one-year same-season",
            "Definition": "Student has any academic record at the follow-up point after first observed organization term.",
            "Window / Denominator": "Only students with a measurable follow-up point in the academic calendar.",
            "Limitations / Notes": "This is continuation after organization entry, not institutional first-time retention.",
        },
        {
            "Metric Group": "Graduation Outcomes",
            "Metric Label": "Observed eventual graduation from first observed organization term",
            "Definition": "Student is ever observed with graduation/alumni signals after organization entry.",
            "Window / Denominator": "All students with a first observed organization term.",
            "Limitations / Notes": "Observed graduation only; not a true institutional first-time graduation rate.",
        },
        {
            "Metric Group": "Graduation Outcomes",
            "Metric Label": "Graduated within 4 / 6 years of first observed organization term",
            "Definition": "Observed graduation occurs on or before the same-season term 4 or 6 years after organization entry.",
            "Window / Denominator": "Only students whose observation window is long enough.",
            "Limitations / Notes": "Right-censored for recent cohorts.",
        },
        {
            "Metric Group": "Credit Momentum",
            "Metric Label": "Passed hours in first term and first academic year after organization entry",
            "Definition": "Uses Term Passed Hours where available.",
            "Window / Denominator": "First post-entry academic term or first observed academic year after organization entry.",
            "Limitations / Notes": "Does not infer attempted hours or attempted-vs-earned ratios.",
        },
        {
            "Metric Group": "GPA and Academic Progress",
            "Metric Label": "Primary GPA metrics",
            "Definition": "Uses Term GPA for term performance; TxState Cumulative GPA and Overall Cumulative GPA for cumulative context.",
            "Window / Denominator": "First term, second term, and first observed academic year after organization entry.",
            "Limitations / Notes": "Multiple GPA fields are preserved in the longitudinal table; primary reporting uses the most defensible term-level field.",
        },
        {
            "Metric Group": "Academic Standing",
            "Metric Label": "Academic standing distribution and transitions",
            "Definition": "Maps raw academic standing text into broad additive buckets and counts transitions over time.",
            "Window / Denominator": "Observed academic terms only.",
            "Limitations / Notes": "Raw standing text is preserved separately.",
        },
        {
            "Metric Group": "Outcome Segmentation",
            "Metric Label": "Segmented metrics by status, major, standing, and entry-hours bucket",
            "Definition": "Computes the same observed outcome and GPA metrics for non-trivial subgroups.",
            "Window / Denominator": "Groups meeting the minimum sample-size threshold.",
            "Limitations / Notes": "Small groups are suppressed to avoid overinterpreting tiny samples.",
        },
    ]


def build_change_log_rows(output_folder: Path) -> List[Dict[str, object]]:
    return [
        {
            "Component": "output/enhanced_metrics/<timestamp>/organization_entry_analytics_enhanced.xlsx",
            "Type": "New additive workbook",
            "Description": f"Versioned workbook written to {output_folder}",
        },
        {
            "Component": "master_longitudinal.csv",
            "Type": "New additive table",
            "Description": "One row per resolved Student ID + term combining organization and academic observations.",
        },
        {
            "Component": "student_summary.csv",
            "Type": "New additive table",
            "Description": "Person-level observed outcomes, follow-up flags, GPA, credit momentum, and standing summaries.",
        },
        {
            "Component": "cohort_metrics.csv",
            "Type": "New additive table",
            "Description": "Long-form cohort metrics covering counts, retention, continuation, graduation, GPA, credit momentum, and standing.",
        },
        {
            "Component": "status_mapping.csv",
            "Type": "New documentation table",
            "Description": "Documents additive status, standing, and identity-resolution mapping logic.",
        },
        {
            "Component": "qa_checks.csv",
            "Type": "New QA table",
            "Description": "Duplicate, missing-ID, conflict, censoring, and sanity-check outputs.",
        },
        {
            "Component": "methodology.md",
            "Type": "New documentation file",
            "Description": "Documents assumptions, windows, and limitations of the additive analysis.",
        },
        {
            "Component": "CHANGELOG.md",
            "Type": "New documentation file",
            "Description": "Lists the additive components generated by this run.",
        },
    ]


def build_qa_rows(
    master_rows: Sequence[Dict[str, object]],
    summary_rows: Sequence[Dict[str, object]],
    roster_stats: Dict[str, int],
    academic_stats: Dict[str, int],
    excluded_rows: Sequence[Dict[str, object]],
) -> List[Dict[str, object]]:
    duplicate_student_terms = len(master_rows) - len(
        {(clean_text(row.get("Student ID")), clean_text(row.get("Term Code"))) for row in master_rows}
    )
    impossible_term_ordering = sum(
        1
        for row in summary_rows
        if clean_text(row.get("First Observed Organization Term"))
        and clean_text(row.get("Last Observed Organization Term"))
        and term_sort_tuple(
            same_season_years_later(clean_text(row.get("First Observed Organization Term")), 0) or ""
        )
        > term_sort_tuple(
            same_season_years_later(clean_text(row.get("Last Observed Organization Term")), 0) or ""
        )
    )
    roster_no_acad = sum(
        1
        for row in summary_rows
        if clean_text(row.get("First Observed Organization Term")) and not clean_text(row.get("First Observed Academic Term"))
    )
    acad_no_roster = sum(
        1
        for row in summary_rows
        if clean_text(row.get("First Observed Academic Term")) and not clean_text(row.get("First Observed Organization Term"))
    )
    grad_before_observation = sum(
        1
        for row in summary_rows
        if clean_text(row.get("Observed Graduation Term"))
        and clean_text(row.get("First Observed Term Overall"))
        and term_sort_tuple(
            same_season_years_later(clean_text(row.get("Observed Graduation Term")), 0) or ""
        )
        < term_sort_tuple(
            same_season_years_later(clean_text(row.get("First Observed Term Overall")), 0) or ""
        )
    )
    bad_grad4 = sum(
        1
        for row in summary_rows
        if clean_text(row.get("Observed Graduation Within 4 Years Of Org Entry Measurable")) != "Yes"
        and clean_text(row.get("Observed Graduation Within 4 Years Of Org Entry"))
    )
    bad_grad6 = sum(
        1
        for row in summary_rows
        if clean_text(row.get("Observed Graduation Within 6 Years Of Org Entry Measurable")) != "Yes"
        and clean_text(row.get("Observed Graduation Within 6 Years Of Org Entry"))
    )

    checks = [
        ("Duplicate Student ID + Term combinations after merge", duplicate_student_terms == 0, duplicate_student_terms, "Should be 0 after additive dedupe."),
        ("Raw roster duplicate Student ID + Term groups", True, roster_stats["duplicate_student_term_groups"], "Non-zero is allowed and expected in source data; additive logic dedupes them."),
        ("Raw academic duplicate Student ID + Term groups", True, academic_stats["duplicate_student_term_groups"], "Non-zero is allowed and expected in source data; additive logic dedupes them."),
        ("Missing resolved Student ID in excluded roster rows", True, roster_stats["missing_student_id_rows"], "Excluded from student-level metrics."),
        ("Missing resolved Student ID in excluded academic rows", True, academic_stats["missing_student_id_rows"], "Excluded from student-level metrics."),
        ("Rows excluded for unparseable terms or IDs", True, len(excluded_rows), "See Excluded_Rows tab for details."),
        ("Impossible first/last term ordering in student summaries", impossible_term_ordering == 0, impossible_term_ordering, "Should be 0."),
        ("Students with roster but no academics", True, roster_no_acad, "Useful for data completeness review."),
        ("Students with academics but no roster", True, acad_no_roster, "Useful for data completeness review."),
        ("Roster status conflicts within the same term", True, roster_stats["status_conflict_groups"], "Flagged and preserved in the longitudinal table."),
        ("Roster chapter conflicts within the same term", True, roster_stats["chapter_conflict_groups"], "Flagged and preserved in the longitudinal table."),
        ("Academic standing conflicts within the same term", True, academic_stats["standing_conflict_groups"], "Flagged and preserved in the longitudinal table."),
        ("Graduation appearing before first observation", grad_before_observation == 0, grad_before_observation, "Should be 0."),
        ("Non-measurable students populated in 4-year graduation field", bad_grad4 == 0, bad_grad4, "Should be 0."),
        ("Non-measurable students populated in 6-year graduation field", bad_grad6 == 0, bad_grad6, "Should be 0."),
        ("Total unique students", True, len(summary_rows), "Sanity-check count."),
        ("Total unique observed terms", True, len({clean_text(row.get('Term Code')) for row in master_rows}), "Sanity-check count."),
        ("Total students ever graduated", True, count_yes(summary_rows, "Ever Graduated Flag"), "Observed graduation only."),
        ("Total students ever suspended", True, count_yes(summary_rows, "Ever Suspended Flag"), ""),
        ("Total students with unknown / no further observation exits", True, count_yes(summary_rows, "No Further Observation Flag"), ""),
    ]

    return [
        {
            "Check": check,
            "Status": "Pass" if passed else "Flag",
            "Value": value,
            "Notes": notes,
        }
        for check, passed, value, notes in checks
    ]


def write_markdown_files(
    output_folder: Path,
    merged_workbook: Path,
    qa_rows: Sequence[Dict[str, object]],
    change_log_rows: Sequence[Dict[str, object]],
) -> None:
    methodology_path = output_folder / "methodology.md"
    qa_flag_count = sum(1 for row in qa_rows if row.get("Status") == "Flag")
    methodology = f"""# Additive Organization-Entry Analytics

## Source

- Merged workbook: `{merged_workbook}`
- Analysis style: additive only

## What this run does

- Builds a one-row-per-student-term longitudinal table using resolved Student ID plus term.
- Preserves existing project behavior by creating new outputs in a separate timestamped folder.
- Labels outcomes from first observed organization or academic terms rather than claiming true first-time-in-college rates.

## Validity rules used

- Student ID is the primary person key.
- Rows without a Banner ID are only retained when the email maps uniquely to one known Student ID.
- Graduation, retention, continuation, GPA, and credit-momentum metrics are labeled as observed outcomes after organization entry.
- Four-year and six-year graduation metrics are right-censored to students with enough observed elapsed time.
- Passed-hours metrics use `Term Passed Hours` and do not fabricate attempted-vs-earned ratios.

## Existing components left untouched

- `src/greek_life_pipeline.py`
- `src/build_master_roster.py`
- `src/build_master_roster_grades.py`
- `src/build_member_tenure_report.py`
- `src/build_yearly_chapter_rosters.py`
- `powerquery/*.pq`

## QA

- QA flags raised in this run: {qa_flag_count}
- Full details are written to `qa_checks.csv` and the `QA_Checks` worksheet.
"""
    methodology_path.write_text(methodology, encoding="utf-8")

    changelog_path = output_folder / "CHANGELOG.md"
    changelog_lines = ["# Change Log", ""]
    for row in change_log_rows:
        changelog_lines.append(f"- {row['Component']}: {row['Description']}")
    changelog_path.write_text("\n".join(changelog_lines) + "\n", encoding="utf-8")


def build_enhanced_org_analytics(
    merged_workbook: Path,
    output_root: Path,
    segment_min_size: int,
) -> Tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = output_root / f"run_{timestamp}"
    output_folder.mkdir(parents=True, exist_ok=False)
    workbook_path = output_folder / f"organization_entry_analytics_enhanced_{timestamp}.xlsx"

    raw_roster_rows, raw_academic_rows, excluded_rows = load_merged_workbook_observations(merged_workbook)
    resolved_roster_rows, _ = apply_identity_resolution(raw_roster_rows)
    resolved_academic_rows, _ = apply_identity_resolution(raw_academic_rows)

    for row in resolved_roster_rows:
        row["roster_status_bucket"] = roster_status_bucket(
            clean_text(row.get("raw_status")),
            clean_text(row.get("position")),
        )
        row["new_member_marker"] = yes_no(
            is_new_member_marker(clean_text(row.get("raw_status")), clean_text(row.get("position")))
        )

    for row in resolved_academic_rows:
        row["academic_status_signal_bucket"] = academic_status_signal_bucket(
            clean_text(row.get("grade_student_status"))
        )
        row["academic_standing_bucket"] = academic_standing_bucket(
            clean_text(row.get("current_academic_standing"))
        )

    roster_map, roster_stats = dedupe_roster_rows(resolved_roster_rows, excluded_rows)
    academic_map, academic_stats = dedupe_academic_rows(resolved_academic_rows, excluded_rows)
    master_longitudinal_rows, all_term_codes, roster_term_codes, academic_term_codes = build_master_longitudinal(
        roster_map, academic_map
    )
    summary_rows, status_transition_rows, standing_transition_rows = build_student_summary(
        master_longitudinal_rows,
        all_term_codes,
        roster_term_codes,
        academic_term_codes,
    )
    transition_rows = aggregate_transition_rows(status_transition_rows + standing_transition_rows)
    cohort_metrics_rows = build_cohort_metrics(summary_rows)
    segment_rows = build_outcome_segments(summary_rows, min_size=segment_min_size)
    qa_rows = build_qa_rows(
        master_longitudinal_rows,
        summary_rows,
        roster_stats,
        academic_stats,
        excluded_rows,
    )
    status_mapping_rows = build_status_mapping_rows()
    metric_definition_rows = build_metric_definition_rows()
    change_log_rows = build_change_log_rows(output_folder)

    csv_tables = [
        ("master_longitudinal.csv", MASTER_LONGITUDINAL_COLUMNS, master_longitudinal_rows),
        ("student_summary.csv", STUDENT_SUMMARY_COLUMNS, summary_rows),
        ("cohort_metrics.csv", COHORT_METRIC_COLUMNS, cohort_metrics_rows),
        ("organization_retention_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Organization Retention"]),
        ("institutional_continuation_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Institutional Continuation"]),
        ("graduation_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Graduation Outcomes"]),
        ("credit_momentum_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Credit Momentum"]),
        ("gpa_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "GPA and Academic Progress"]),
        ("academic_standing_metrics.csv", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Academic Standing"]),
        ("status_transitions.csv", TRANSITION_COLUMNS, transition_rows),
        ("outcome_segments.csv", SEGMENT_COLUMNS, segment_rows),
        ("metric_definitions.csv", METRIC_DEFINITION_COLUMNS, metric_definition_rows),
        ("status_mapping.csv", STATUS_MAPPING_COLUMNS, status_mapping_rows),
        ("qa_checks.csv", QA_COLUMNS, qa_rows),
        ("change_log.csv", CHANGE_LOG_COLUMNS, change_log_rows),
        ("excluded_rows.csv", EXCLUDED_ROW_COLUMNS, excluded_rows),
    ]
    for filename, headers, rows in csv_tables:
        write_csv(output_folder / filename, headers, rows)

    write_markdown_files(output_folder, merged_workbook, qa_rows, change_log_rows)

    wb = Workbook()
    write_overview_sheet(
        wb,
        merged_workbook=merged_workbook,
        output_folder=output_folder,
        run_timestamp=timestamp,
        total_students=len(summary_rows),
        total_longitudinal_rows=len(master_longitudinal_rows),
        qa_rows=qa_rows,
    )
    write_chunked_records_sheet(wb, "Master_Longitudinal", MASTER_LONGITUDINAL_COLUMNS, master_longitudinal_rows)
    write_records_sheet(wb, "Student_Summary", STUDENT_SUMMARY_COLUMNS, summary_rows)
    write_records_sheet(wb, "Cohort_Metrics", COHORT_METRIC_COLUMNS, cohort_metrics_rows)
    write_records_sheet(wb, "Org_Retention", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Organization Retention"])
    write_records_sheet(wb, "Institutional_Cont", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Institutional Continuation"])
    write_records_sheet(wb, "Graduation_Metrics", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Graduation Outcomes"])
    write_records_sheet(wb, "Credit_Momentum", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Credit Momentum"])
    write_records_sheet(wb, "GPA_Metrics", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "GPA and Academic Progress"])
    write_records_sheet(wb, "Standing_Metrics", COHORT_METRIC_COLUMNS, [row for row in cohort_metrics_rows if row.get("Metric Group") == "Academic Standing"])
    write_records_sheet(wb, "Status_Transitions", TRANSITION_COLUMNS, transition_rows)
    write_records_sheet(wb, "Outcome_Segments", SEGMENT_COLUMNS, segment_rows)
    write_records_sheet(wb, "Metric_Definitions", METRIC_DEFINITION_COLUMNS, metric_definition_rows)
    write_records_sheet(wb, "Status_Mapping", STATUS_MAPPING_COLUMNS, status_mapping_rows)
    write_records_sheet(wb, "QA_Checks", QA_COLUMNS, qa_rows)
    write_records_sheet(wb, "Change_Log", CHANGE_LOG_COLUMNS, change_log_rows)
    write_records_sheet(wb, "Excluded_Rows", EXCLUDED_ROW_COLUMNS, excluded_rows)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    return output_folder, workbook_path


def main() -> None:
    args = parse_args()
    merged_workbook = Path(args.merged_workbook).expanduser().resolve()
    output_root = Path(args.output_root).expanduser().resolve()
    output_folder, workbook_path = build_enhanced_org_analytics(
        merged_workbook=merged_workbook,
        output_root=output_root,
        segment_min_size=args.segment_min_size,
    )
    print(f"Enhanced additive analytics created: {workbook_path}")
    print(f"Output folder: {output_folder}")


if __name__ == "__main__":
    main()
