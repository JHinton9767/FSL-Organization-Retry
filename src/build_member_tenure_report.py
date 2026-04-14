from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.build_master_roster import autosize_columns, style_header
from src.canonical_bundle import DEFAULT_CANONICAL_ROOT, load_canonical_bundle


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_WORKBOOK = ROOT / "Member_Tenure_Report.xlsx"
OUTCOME_ORDER = [
    "Graduated",
    "Dropped/Resigned/Revoked/Inactive",
    "Suspended",
    "Transfer",
    "Active/Unknown",
    "No Further Observation",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build Member_Tenure_Report.xlsx as a downstream export from the canonical analytics bundle. "
            "This workbook is no longer an upstream dependency for analytics."
        )
    )
    parser.add_argument("--canonical-root", default=str(DEFAULT_CANONICAL_ROOT))
    parser.add_argument("--canonical-folder", default="")
    parser.add_argument("-o", "--output", default=str(DEFAULT_OUTPUT_WORKBOOK))
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def extract_year(term_label: object) -> int | None:
    text = clean_text(term_label)
    for token in text.split():
        if token.isdigit() and len(token) == 4:
            return int(token)
    return None


def bucket_hours(value: object) -> str:
    if value in ("", None) or pd.isna(value):
        return "Unknown"
    number = float(value)
    lower = int(number // 30) * 30
    upper = lower + 29
    return f"{lower}-{upper}"


def semester_sort_key(label: str):
    parts = clean_text(label).split()
    year = next((int(part) for part in parts if part.isdigit() and len(part) == 4), 9999)
    season = clean_text(parts[0]).lower() if parts else ""
    order = {"winter": 0, "spring": 1, "summer": 2, "fall": 3}.get(season, 9)
    return (year, order, label)


def build_new_member_frame(summary: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    result = summary.copy()
    result = result.loc[result["join_term"].fillna("").astype(str).str.strip().ne("")]
    result = result.loc[result["join_year"].fillna(0).astype(float) >= 2015].copy()
    if result.empty:
        return result

    grouped = master.groupby("student_id", dropna=False)
    histories = grouped.apply(
        lambda frame: pd.Series(
            {
                "term_history": " | ".join(frame.sort_values("observed_term_sort")["term_label"].fillna("").astype(str).tolist()),
                "status_history": " | ".join(
                    f"{clean_text(term)}: {clean_text(status)}"
                    for term, status in zip(
                        frame.sort_values("observed_term_sort")["term_label"].fillna("").astype(str),
                        frame.sort_values("observed_term_sort")["org_status_bucket"].fillna("").astype(str),
                    )
                ),
                "semester_count": int(frame["term_code"].nunique()),
                "semesters_from_new_member": int(frame.loc[coerce_numeric(frame["relative_term_index"]).notna(), "term_code"].nunique()),
                "returned_later": "No",
            }
        )
    ).reset_index()
    result = result.merge(histories, on="student_id", how="left")
    result["join_cumulative_hours_bucket"] = result["entry_hours_bucket"].where(
        result["entry_hours_bucket"].fillna("").astype(str).str.strip().ne(""),
        result["entry_cumulative_hours"].map(bucket_hours),
    )
    result["confirmed_join_within_window"] = result["org_entry_term_basis"].fillna("").astype(str).eq("Explicit New Member").map(lambda value: "Yes" if value else "No")
    result["start_term"] = result["join_term"]
    result["start_basis"] = result["org_entry_term_basis"]
    result["first_new_member_term"] = result["join_term"].where(result["confirmed_join_within_window"].eq("Yes"), "")
    result["last_observed_term"] = result["last_observed_org_term"].where(
        result["last_observed_org_term"].fillna("").astype(str).str.strip().ne(""),
        result["last_observed_academic_term"],
    )
    result["left_term"] = result["graduation_term"].where(
        result["graduation_term"].fillna("").astype(str).str.strip().ne(""),
        result["last_observed_term"].where(result["latest_outcome_bucket"].fillna("").astype(str).isin(["Graduated", "Suspended", "Transfer", "Dropped/Resigned/Revoked/Inactive"]), ""),
    )
    result["final_status"] = result["latest_roster_status_bucket"].where(
        result["latest_roster_status_bucket"].fillna("").astype(str).str.strip().ne(""),
        result["latest_outcome_bucket"],
    )
    result["outcome_group"] = result["latest_outcome_bucket"]
    result["join_semester_at_school"] = ""
    result["exit_semester_at_school"] = ""
    return result


def write_summary_sheet(wb: Workbook, new_members: pd.DataFrame, source_folder: Path) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)
    metrics = [
        ["Canonical source folder", str(source_folder)],
        ["2015+ new-member rows", len(new_members)],
        ["Resolved outcomes", int(new_members["resolved_outcome_flag"].fillna("").astype(str).eq("Yes").sum()) if not new_members.empty else 0],
        ["Unresolved outcomes", int(new_members["resolved_outcome_flag"].fillna("").astype(str).ne("Yes").sum()) if not new_members.empty else 0],
        ["Important note", "This workbook is now a downstream export from canonical tables and is not an analytical source of truth."],
    ]
    for metric in metrics:
        ws.append(metric)
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_outcome_rates_sheet(wb: Workbook, new_members: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="Outcome Rates")
    headers = [
        "Observed Semester Count",
        "Members",
        "Graduated Count",
        "Graduated Rate",
        "Dropped Count",
        "Dropped Rate",
        "Suspended Count",
        "Suspended Rate",
        "Transfer Count",
        "Transfer Rate",
        "Unresolved Count",
        "Unresolved Rate",
        "Average First-Year GPA",
        "Average Latest Cumulative GPA",
    ]
    ws.append(headers)
    style_header(ws)
    grouped = new_members.groupby("semester_count", dropna=False)
    for semester_count, frame in sorted(grouped, key=lambda item: float(item[0]) if pd.notna(item[0]) else 9999):
        members = len(frame)
        counts = {bucket: int(frame["outcome_group"].fillna("").astype(str).eq(bucket).sum()) for bucket in OUTCOME_ORDER}
        unresolved = counts["Active/Unknown"] + counts["No Further Observation"]
        ws.append(
            [
                semester_count,
                members,
                counts["Graduated"],
                counts["Graduated"] / members if members else "",
                counts["Dropped/Resigned/Revoked/Inactive"],
                counts["Dropped/Resigned/Revoked/Inactive"] / members if members else "",
                counts["Suspended"],
                counts["Suspended"] / members if members else "",
                counts["Transfer"],
                counts["Transfer"] / members if members else "",
                unresolved,
                unresolved / members if members else "",
                coerce_numeric(frame["first_year_avg_term_gpa"]).mean(),
                coerce_numeric(frame["average_cumulative_gpa"]).mean(),
            ]
        )
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_join_hours_outcome_rates_sheet(wb: Workbook, new_members: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="Join Hours Rates")
    headers = [
        "Join Cumulative Hours Bucket",
        "Members",
        "Graduated Count",
        "Graduated Rate",
        "Dropped Count",
        "Dropped Rate",
        "Suspended Count",
        "Suspended Rate",
        "Transfer Count",
        "Transfer Rate",
        "Unresolved Count",
        "Unresolved Rate",
        "Average Join Cumulative Hours",
        "Average First-Year GPA",
        "Average Latest Cumulative GPA",
    ]
    ws.append(headers)
    style_header(ws)
    grouped = new_members.groupby("join_cumulative_hours_bucket", dropna=False)
    for bucket, frame in sorted(grouped, key=lambda item: (1, 999999) if clean_text(item[0]) == "Unknown" else (0, int(clean_text(item[0]).split("-", 1)[0]))):
        members = len(frame)
        counts = {bucket_name: int(frame["outcome_group"].fillna("").astype(str).eq(bucket_name).sum()) for bucket_name in OUTCOME_ORDER}
        unresolved = counts["Active/Unknown"] + counts["No Further Observation"]
        ws.append(
            [
                bucket,
                members,
                counts["Graduated"],
                counts["Graduated"] / members if members else "",
                counts["Dropped/Resigned/Revoked/Inactive"],
                counts["Dropped/Resigned/Revoked/Inactive"] / members if members else "",
                counts["Suspended"],
                counts["Suspended"] / members if members else "",
                counts["Transfer"],
                counts["Transfer"] / members if members else "",
                unresolved,
                unresolved / members if members else "",
                coerce_numeric(frame["entry_cumulative_hours"]).mean(),
                coerce_numeric(frame["first_year_avg_term_gpa"]).mean(),
                coerce_numeric(frame["average_cumulative_gpa"]).mean(),
            ]
        )
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_gpa_by_semester_sheet(wb: Workbook, master: pd.DataFrame, new_members: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="GPA by Semester")
    headers = [
        "Relative Term After Entry",
        "Records",
        "Distinct Members",
        "Average Term GPA",
        "Average Institutional Cumulative GPA",
        "Average Overall Cumulative GPA",
        "Average Total Cumulative Hours",
    ]
    ws.append(headers)
    style_header(ws)
    ids = set(new_members["student_id"].fillna("").astype(str))
    frame = master.loc[master["student_id"].fillna("").astype(str).isin(ids)].copy()
    frame = frame.loc[coerce_numeric(frame["relative_term_index"]).notna() & coerce_numeric(frame["relative_term_index"]).ge(0)]
    grouped = frame.groupby("relative_term_index", dropna=False)
    for relative_term, group in sorted(grouped, key=lambda item: float(item[0]) if pd.notna(item[0]) else 9999):
        ws.append(
            [
                int(float(relative_term)),
                len(group),
                int(group["student_id"].nunique()),
                coerce_numeric(group["term_gpa"]).mean(),
                coerce_numeric(group["institutional_cumulative_gpa"]).mean(),
                coerce_numeric(group["overall_cumulative_gpa"]).mean(),
                coerce_numeric(group["total_cumulative_hours"]).mean(),
            ]
        )
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_new_member_journeys_sheet(wb: Workbook, new_members: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="2015+ New Members")
    headers = [
        "Chapter",
        "Student Name",
        "Student ID",
        "Start Term",
        "Start Basis",
        "First New Member Term",
        "Last Observed Term",
        "Left Term",
        "Exit Reason",
        "Final Status",
        "Outcome Group",
        "Returned Later",
        "Confirmed Join In Window",
        "Semester Count",
        "Semesters From New Member",
        "Join Cumulative Hours",
        "Join Cumulative Hours Bucket",
        "Average First-Year GPA",
        "Latest TxState Cumulative GPA",
        "Latest Overall Cumulative GPA",
        "Term History",
        "Status History",
    ]
    ws.append(headers)
    style_header(ws)
    for _, row in new_members.sort_values(by=["join_year", "chapter", "student_name"], ascending=[True, True, True], na_position="last").iterrows():
        ws.append(
            [
                row.get("chapter", ""),
                row.get("student_name", ""),
                row.get("student_id", ""),
                row.get("start_term", ""),
                row.get("start_basis", ""),
                row.get("first_new_member_term", ""),
                row.get("last_observed_term", ""),
                row.get("left_term", ""),
                row.get("exit_reason_code", ""),
                row.get("final_status", ""),
                row.get("outcome_group", ""),
                row.get("returned_later", ""),
                row.get("confirmed_join_within_window", ""),
                row.get("semester_count", ""),
                row.get("semesters_from_new_member", ""),
                row.get("entry_cumulative_hours", ""),
                row.get("join_cumulative_hours_bucket", ""),
                row.get("first_year_avg_term_gpa", ""),
                row.get("latest_txstate_cumulative_gpa", ""),
                row.get("latest_overall_cumulative_gpa", ""),
                row.get("term_history", ""),
                row.get("status_history", ""),
            ]
        )
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_semester_sheets(wb: Workbook, new_members: pd.DataFrame) -> None:
    grouped: Dict[str, pd.DataFrame] = defaultdict(pd.DataFrame)
    for cohort, frame in new_members.groupby("join_term", dropna=False):
        grouped[clean_text(cohort) or "Unknown"] = frame.copy()
    for cohort in sorted(grouped, key=semester_sort_key):
        ws = wb.create_sheet(title=(cohort or "Unknown")[:31])
        ws.append(["Student Name", "Student ID", "Chapter", "Outcome Group", "Join Cumulative Hours", "Average First-Year GPA"])
        style_header(ws)
        for _, row in grouped[cohort].sort_values(by=["chapter", "student_name"], ascending=[True, True], na_position="last").iterrows():
            ws.append(
                [
                    row.get("student_name", ""),
                    row.get("student_id", ""),
                    row.get("chapter", ""),
                    row.get("outcome_group", ""),
                    row.get("entry_cumulative_hours", ""),
                    row.get("first_year_avg_term_gpa", ""),
                ]
            )
        ws.freeze_panes = "A2"
        autosize_columns(ws)


def build_member_tenure_report(canonical_root: Path, explicit_folder: Path | None, output_path: Path) -> None:
    bundle = load_canonical_bundle(canonical_root=canonical_root, explicit_folder=explicit_folder)
    summary = bundle.tables["student_summary"].copy()
    master = bundle.tables["master_longitudinal"].copy()
    new_members = build_new_member_frame(summary, master)
    if new_members.empty:
        raise FileNotFoundError("No 2015+ canonical new-member rows were available for the tenure export.")

    wb = Workbook()
    write_summary_sheet(wb, new_members, bundle.output_folder)
    write_outcome_rates_sheet(wb, new_members)
    write_join_hours_outcome_rates_sheet(wb, new_members)
    write_gpa_by_semester_sheet(wb, master, new_members)
    write_new_member_journeys_sheet(wb, new_members)
    write_semester_sheets(wb, new_members)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    build_member_tenure_report(
        canonical_root=Path(args.canonical_root).expanduser().resolve(),
        explicit_folder=Path(args.canonical_folder).expanduser().resolve() if args.canonical_folder else None,
        output_path=Path(args.output).expanduser().resolve(),
    )
    print(f"Member tenure report created: {Path(args.output).expanduser().resolve()}")


if __name__ == "__main__":
    main()
