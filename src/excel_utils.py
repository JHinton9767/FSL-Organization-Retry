from __future__ import annotations

from collections import defaultdict

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.shared_utils import clean_text


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def autosize_columns(ws, min_width: int = 12, max_width: int = 32) -> None:
    max_widths = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            width = len(clean_text(value))
            if width > max_widths[idx]:
                max_widths[idx] = width
    for idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, min_width), max_width)


def style_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def safe_sheet_name(value: str, default: str = "Unknown") -> str:
    cleaned = clean_text(value) or default
    for char in "[]:*?/\\":  # Excel-invalid sheet chars
        cleaned = cleaned.replace(char, "")
    return cleaned[:31] or default


def safe_filename(value: str, default: str = "Unknown", max_length: int = 120) -> str:
    text = clean_text(value) or default
    for char in '<>:"/\\|?*':
        text = text.replace(char, "_")
    return text[:max_length].strip(" ._") or default
