from __future__ import annotations

import argparse
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.build_canonical_pipeline import clean_text, coerce_numeric
from src.canonical_bundle import DEFAULT_CANONICAL_ROOT, load_canonical_bundle


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "record_priority"
INDIVIDUAL_PULL_START_YEAR = 2018

TITLE_FILL = "1F4E79"
HEADER_FILL = "DCE6F1"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rank students whose individual academic-history pulls would have the highest analytic impact. "
            f"The ranking assumes the individual pull window begins in {INDIVIDUAL_PULL_START_YEAR}."
        )
    )
    parser.add_argument("--canonical-root", default=str(DEFAULT_CANONICAL_ROOT))
    parser.add_argument("--canonical-folder", default="")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--top", type=int, default=250, help="Top students to include on the priority-only sheet.")
    return parser.parse_args()


def extract_year(term_label: object) -> int | None:
    match = re.search(r"(19\d{2}|20\d{2})", clean_text(term_label))
    return int(match.group(1)) if match else None


def estimate_missing_pre_org_terms(entry_hours: object) -> int:
    if entry_hours in ("", None) or (isinstance(entry_hours, float) and math.isnan(entry_hours)):
        return 0
    return max(int(math.floor(float(entry_hours) / 15.0)), 0)


def latest_cumulative_gpa(row: pd.Series) -> object:
    overall = row.get("latest_overall_cumulative_gpa", "")
    txstate = row.get("latest_txstate_cumulative_gpa", "")
    if overall not in ("", None) and not (isinstance(overall, float) and math.isnan(overall)):
        return overall
    return txstate


def identity_score(identity_basis: str) -> Tuple[int, str]:
    basis = clean_text(identity_basis).lower()
    if not basis:
        return 8, "Identity basis is missing, which increases matching uncertainty."
    if "name" in basis and "email" in basis:
        return 10, "Identity uses name and email matching rather than a clean student ID."
    if "name" in basis:
        return 9, "Identity appears to rely on name matching."
    if "email" in basis:
        return 6, "Identity appears to rely on email matching."
    if "|" in basis:
        return 6, "Identity uses multiple resolution methods."
    return 0, ""


def outcome_score(outcome_bucket: str) -> Tuple[int, str]:
    bucket = clean_text(outcome_bucket)
    mapping = {
        "Graduated": (8, "Student has a terminal graduation outcome, so missing earlier academic detail can materially change time-to-degree interpretation."),
        "Dropped/Resigned/Revoked/Inactive": (8, "Student has a terminal non-graduation outcome, so added term history can change persistence and GPA interpretation."),
        "Suspended": (8, "Student has a suspension outcome, which makes term-level academic history especially useful."),
        "Transfer": (6, "Transfer-related history is more complex and benefits from fuller academic context."),
        "No Further Observation": (5, "Outcome is unresolved because the student disappears from the observed data."),
        "Active/Unknown": (3, "Outcome is still active or unresolved, so additional history may still matter."),
    }
    return mapping.get(bucket, (2, "Outcome is not clearly classified in the current data."))


def standing_score(first_standing: str) -> Tuple[int, str]:
    standing = clean_text(first_standing)
    if standing == "Suspended":
        return 5, "First observed academic standing is suspended."
    if standing == "Probation/Warning":
        return 4, "First observed academic standing is probation or warning."
    return 0, ""


def add_observation_counts(summary: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    if summary.empty or master.empty:
        result = summary.copy()
        for column in [
            "roster_terms_observed",
            "academic_terms_observed",
            "total_terms_observed",
            "roster_terms_observed_2018plus",
            "academic_terms_observed_2018plus",
            "total_terms_observed_2018plus",
        ]:
            if column not in result.columns:
                result[column] = 0
        return result

    master = master.copy()
    master["_in_2018plus_window"] = master["term_code"].map(lambda value: (extract_year(value) or 0) >= INDIVIDUAL_PULL_START_YEAR)
    grouped = master.groupby("student_id", dropna=False)
    counts = pd.DataFrame(
        {
            "student_id": list(grouped.groups.keys()),
            "roster_terms_observed": grouped["roster_present"].apply(lambda values: int(values.fillna("").astype(str).eq("Yes").sum())).tolist(),
            "academic_terms_observed": grouped["academic_present"].apply(lambda values: int(values.fillna("").astype(str).eq("Yes").sum())).tolist(),
            "total_terms_observed": grouped["term_code"].nunique().tolist(),
            "roster_terms_observed_2018plus": grouped.apply(
                lambda frame: int((frame["_in_2018plus_window"] & frame["roster_present"].fillna("").astype(str).eq("Yes")).sum())
            ).tolist(),
            "academic_terms_observed_2018plus": grouped.apply(
                lambda frame: int((frame["_in_2018plus_window"] & frame["academic_present"].fillna("").astype(str).eq("Yes")).sum())
            ).tolist(),
            "total_terms_observed_2018plus": grouped.apply(
                lambda frame: int(frame.loc[frame["_in_2018plus_window"], "term_code"].nunique())
            ).tolist(),
        }
    )
    return summary.merge(counts, on="student_id", how="left")


def build_priority_list(summary: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    summary = summary.copy()
    summary = add_observation_counts(summary, master)
    numeric_fields = [
        "entry_cumulative_hours",
        "roster_terms_observed",
        "academic_terms_observed",
        "total_terms_observed",
        "roster_terms_observed_2018plus",
        "academic_terms_observed_2018plus",
        "total_terms_observed_2018plus",
        "latest_txstate_cumulative_gpa",
        "latest_overall_cumulative_gpa",
    ]
    for field in numeric_fields:
        if field in summary.columns:
            summary[field] = coerce_numeric(summary[field])

    latest_year = max((extract_year(value) or 0) for value in summary["org_entry_cohort"].tolist())
    if latest_year == 0:
        latest_year = datetime.now().year

    rows: List[Dict[str, object]] = []
    for _, row in summary.iterrows():
        student_id = clean_text(row.get("student_id"))
        student_name = clean_text(row.get("student_name"))
        first_name = student_name.split(" ", 1)[0] if student_name else ""
        last_name = student_name.split(" ", 1)[1] if " " in student_name else ""
        if not student_id and not (first_name or last_name):
            continue

        entry_hours = row.get("entry_cumulative_hours", "")
        missing_terms = estimate_missing_pre_org_terms(entry_hours)
        first_org_term = clean_text(row.get("join_term"))
        first_academic_term = clean_text(row.get("first_observed_academic_term"))
        org_year = extract_year(first_org_term)
        first_academic_year = extract_year(first_academic_term)
        cohort_age = max(latest_year - org_year, 0) if org_year is not None else 0
        latest_outcome = clean_text(row.get("latest_outcome_bucket")) or "Unknown"
        identity_basis = clean_text(row.get("org_entry_term_basis"))
        ever_transfer = clean_text(row.get("transfer_flag")) == "Yes"
        total_terms = row.get("total_terms_observed", 0)
        total_terms_2018plus = row.get("total_terms_observed_2018plus", 0)
        academic_terms_2018plus = row.get("academic_terms_observed_2018plus", 0)
        roster_terms_2018plus = row.get("roster_terms_observed_2018plus", 0)
        missing_2018plus_terms = max(int(total_terms_2018plus or 0) - int(academic_terms_2018plus or 0), 0)
        latest_gpa = latest_cumulative_gpa(row)
        same_term_high_hours = (
            first_org_term
            and first_academic_term
            and first_org_term == first_academic_term
            and entry_hours not in ("", None)
            and not (isinstance(entry_hours, float) and math.isnan(entry_hours))
            and float(entry_hours) >= 30
        )

        score = 0
        reasons: List[Tuple[int, str]] = []

        if not first_academic_term and first_org_term:
            score += 30
            reasons.append((30, f"No academic term is currently observed, so the {INDIVIDUAL_PULL_START_YEAR}+ individual history pull could add most of the usable term-level record."))

        if missing_2018plus_terms:
            gap_score = min(missing_2018plus_terms * 7, 28)
            score += gap_score
            reasons.append(
                (
                    gap_score,
                    f"Student has about {missing_2018plus_terms} observed term(s) in the {INDIVIDUAL_PULL_START_YEAR}+ window without matching academic detail, so the individual pull could materially improve GPA and credit-hour trends.",
                )
            )

        if org_year is not None and org_year >= INDIVIDUAL_PULL_START_YEAR:
            score += 12
            reasons.append((12, f"Organization entry begins in or after {INDIVIDUAL_PULL_START_YEAR}, so the individual pull can cover most or all of the organization-era academic timeline."))
        elif total_terms_2018plus not in ("", None) and not (isinstance(total_terms_2018plus, float) and math.isnan(total_terms_2018plus)) and float(total_terms_2018plus) > 0:
            score += 4
            reasons.append((4, f"Student has observed terms in the {INDIVIDUAL_PULL_START_YEAR}+ window, so the individual pull can still improve later longitudinal history even if earlier terms remain unavailable."))

        missing_term_score = min(missing_terms * 4, 32)
        if missing_term_score and org_year is not None and org_year >= INDIVIDUAL_PULL_START_YEAR:
            score += missing_term_score
            reasons.append(
                (
                    missing_term_score,
                    f"Student appears to join with about {missing_terms} prior term(s) already completed, and because the pull window starts in {INDIVIDUAL_PULL_START_YEAR}, some of that early history may now be recoverable.",
                )
            )
        elif missing_terms:
            limited_score = min(missing_terms, 2)
            score += limited_score
            reasons.append((limited_score, f"Student likely has earlier pre-organization history, but because individual records begin in {INDIVIDUAL_PULL_START_YEAR}, only part of that history may be recoverable."))

        if same_term_high_hours and ((org_year is not None and org_year >= INDIVIDUAL_PULL_START_YEAR) or (first_academic_year is not None and first_academic_year >= INDIVIDUAL_PULL_START_YEAR)):
            score += 12
            reasons.append((12, f"Academic history begins in the same term as organization entry despite high cumulative hours, which strongly suggests missing earlier {INDIVIDUAL_PULL_START_YEAR}+ records."))

        age_score = min(cohort_age * 2, 16) if org_year is not None and org_year >= INDIVIDUAL_PULL_START_YEAR else (4 if total_terms_2018plus not in ("", None) and not (isinstance(total_terms_2018plus, float) and math.isnan(total_terms_2018plus)) and float(total_terms_2018plus) > 0 else 0)
        if age_score:
            score += age_score
            reasons.append((age_score, f"Observed cohort timing ({first_org_term}) affects more graduation and trend calculations within the recoverable {INDIVIDUAL_PULL_START_YEAR}+ window."))

        outcome_points, outcome_reason = outcome_score(latest_outcome)
        score += outcome_points
        if outcome_reason:
            reasons.append((outcome_points, outcome_reason))

        identity_points, identity_reason = identity_score(identity_basis)
        score += identity_points
        if identity_reason:
            reasons.append((identity_points, identity_reason))

        if ever_transfer:
            score += 10
            reasons.append((10, "Transfer status appears in the current data, which makes school-start timing more complex."))

        if clean_text(row.get("graduated_4yr_measurable")) == "Yes":
            score += 4
            reasons.append((4, "Student is already contributing to measurable 4-year graduation calculations."))
        if clean_text(row.get("graduated_6yr_measurable")) == "Yes":
            score += 4
            reasons.append((4, "Student is already contributing to measurable 6-year graduation calculations."))

        standing_points, standing_reason = standing_score(clean_text(row.get("first_academic_standing_bucket")))
        score += standing_points
        if standing_reason:
            reasons.append((standing_points, standing_reason))

        if total_terms not in ("", None) and not (isinstance(total_terms, float) and math.isnan(total_terms)):
            span_score = min(int(total_terms // 2), 6)
            score += span_score
            if span_score:
                reasons.append((span_score, "Student has a longer observed timeline, so updated starting records would influence more downstream metrics."))

        priority_tier = "Lower"
        if score >= 60:
            priority_tier = "Very High"
        elif score >= 45:
            priority_tier = "High"
        elif score >= 30:
            priority_tier = "Medium"

        top_reasons = " | ".join(reason for _, reason in sorted(reasons, key=lambda item: (-item[0], item[1]))[:4])

        rows.append(
            {
                "Priority Score": score,
                "Priority Tier": priority_tier,
                "Student ID": student_id,
                "Last Name": last_name,
                "First Name": first_name,
                "Chapter": clean_text(row.get("initial_chapter")),
                "Organization Entry Cohort": first_org_term,
                "First Observed Academic Term": first_academic_term,
                "Individual Pull Starts": str(INDIVIDUAL_PULL_START_YEAR),
                "Entry Cumulative Hours": entry_hours,
                "Estimated Missing Pre-Org Terms": missing_terms,
                "2018+ Roster Terms Observed": roster_terms_2018plus,
                "2018+ Academic Terms Observed": academic_terms_2018plus,
                "2018+ Missing Academic Terms": missing_2018plus_terms,
                "Latest Known Outcome": latest_outcome,
                "Ever Transfer": "Yes" if ever_transfer else "No",
                "Identity Resolution Basis": identity_basis,
                "Latest Cumulative GPA": latest_gpa,
                "Academic Terms Observed": row.get("academic_terms_observed", ""),
                "Total Terms Observed": total_terms,
                "Why Pull Individual History Early": top_reasons,
            }
        )

    priority = pd.DataFrame(rows)
    if priority.empty:
        return priority

    priority["_cohort_year_sort"] = priority["Organization Entry Cohort"].map(lambda value: extract_year(value) or 9999)
    priority = priority.sort_values(
        by=[
            "Priority Score",
            "2018+ Missing Academic Terms",
            "_cohort_year_sort",
            "Organization Entry Cohort",
            "Last Name",
            "First Name",
        ],
        ascending=[False, False, True, True, True, True],
    ).reset_index(drop=True)
    priority = priority.drop(columns=["_cohort_year_sort"])
    priority.insert(0, "Priority Rank", range(1, len(priority) + 1))
    return priority


def write_sheet(ws, title: str, subtitle: str, frame: pd.DataFrame) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A2"].font = Font(italic=True)
    ws.row_dimensions[2].height = 36

    headers = list(frame.columns)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, values in enumerate(frame.itertuples(index=False), start=5):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            header = headers[col_idx - 1]
            if "score" in header.lower() or "terms" in header.lower() or "hours" in header.lower():
                if value not in ("", None):
                    cell.number_format = "0.00" if "GPA" in header else "#,##0.0"

    ws.freeze_panes = "A5"
    for col_idx in range(1, len(headers) + 1):
        column_letter = ws.cell(row=4, column=col_idx).column_letter
        width = 14
        for row in range(1, ws.max_row + 1):
            width = min(max(width, len(clean_text(ws[f"{column_letter}{row}"].value)) + 2), 45)
        ws.column_dimensions[column_letter].width = width


def write_method_sheet(ws) -> None:
    ws["A1"] = "Method"
    ws["A1"].font = Font(bold=True, size=16)
    lines = [
        f"This ranking is meant to help decide which students should have their individual academic histories pulled first, with the usable pull window beginning in {INDIVIDUAL_PULL_START_YEAR}.",
        f"Higher ranks go to students whose missing {INDIVIDUAL_PULL_START_YEAR}+ term-level history is most likely to change semester counts, GPA trends, credit-hour interpretation, or outcome calculations.",
        f"The strongest signals are missing academic detail in the {INDIVIDUAL_PULL_START_YEAR}+ window, high cumulative hours at organization entry, recoverable organization-era history, terminal outcomes, transfer status, and weaker identity resolution.",
        f"Because these individual pulls begin in {INDIVIDUAL_PULL_START_YEAR}, the ranking does not assume you are recovering a student’s full school-start history.",
    ]
    for idx, line in enumerate(lines, start=3):
        ws[f"A{idx}"] = f"- {line}"
        ws[f"A{idx}"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120


def build_full_record_priority_list(
    canonical_root: Path,
    output_root: Path,
    explicit_folder: Path | None,
    top_n: int,
) -> Dict[str, object]:
    bundle = load_canonical_bundle(canonical_root=canonical_root, explicit_folder=explicit_folder)
    summary = bundle.tables["student_summary"].copy()
    master = bundle.tables["master_longitudinal"].copy()
    priority = build_priority_list(summary, master)
    if priority.empty:
        raise ValueError("No students were available to rank from canonical student_summary.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = output_root / f"run_{timestamp}"
    output_folder.mkdir(parents=True, exist_ok=False)

    csv_path = output_folder / "full_academic_record_priority_list.csv"
    workbook_path = output_folder / "Full_Academic_Record_Priority_List.xlsx"
    readme_path = output_folder / "README.md"

    priority.to_csv(csv_path, index=False)

    top_frame = priority.head(top_n).copy()
    reason_counts = (
        priority["Priority Tier"]
        .value_counts()
        .rename_axis("Priority Tier")
        .reset_index(name="Students")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Priority List"
    write_sheet(
        ws,
        "Individual Academic History Priority List",
        f"Students ranked from highest to lowest priority for pulling individual academic history beginning in {INDIVIDUAL_PULL_START_YEAR}.",
        priority,
    )

    top_ws = wb.create_sheet(title="Top Priority")
    write_sheet(
        top_ws,
        f"Top {min(top_n, len(priority))} Students",
        "Start with these students first if you are pulling records one by one.",
        top_frame,
    )

    summary_ws = wb.create_sheet(title="Summary")
    write_sheet(
        summary_ws,
        "Priority Tier Summary",
        "How many students fall into each priority tier.",
        reason_counts,
    )

    method_ws = wb.create_sheet(title="Method")
    write_method_sheet(method_ws)

    wb.save(workbook_path)

    readme = [
        "# Individual Academic History Priority List",
        "",
        f"This folder ranks students from highest to lowest priority for pulling individual academic history beginning in {INDIVIDUAL_PULL_START_YEAR}.",
        "",
        "## Files",
        "",
        f"- `{workbook_path.name}`: Excel version with full list, top-priority subset, tier summary, and method notes.",
        f"- `{csv_path.name}`: Flat ranked list for quick sorting/filtering.",
        "",
        "## Source used",
        "",
        f"- Canonical analytics bundle: `{bundle.output_folder}`",
        "",
        "## Ranking idea",
        "",
        f"- Higher priority means the student is more likely to change important calculations if additional {INDIVIDUAL_PULL_START_YEAR}+ term-level academic history is added.",
        f"- The list emphasizes missing academic detail in the {INDIVIDUAL_PULL_START_YEAR}+ window, high entry cumulative hours, recoverable organization-era history, terminal outcomes, transfer complexity, and weaker identity resolution.",
    ]
    readme_path.write_text("\n".join(readme) + "\n", encoding="utf-8")

    return {
        "output_folder": output_folder,
        "workbook_path": workbook_path,
        "csv_path": csv_path,
        "readme_path": readme_path,
        "top_students": top_frame[["Priority Rank", "Last Name", "First Name", "Student ID", "Priority Score", "Why Pull Individual History Early"]].head(25),
    }


def main() -> None:
    args = parse_args()
    explicit_folder = Path(args.canonical_folder).expanduser().resolve() if args.canonical_folder else None
    result = build_full_record_priority_list(
        canonical_root=Path(args.canonical_root).expanduser().resolve(),
        output_root=Path(args.output_root).expanduser().resolve(),
        explicit_folder=explicit_folder,
        top_n=args.top,
    )
    print(f"Priority list created: {result['workbook_path']}")
    print(f"CSV: {result['csv_path']}")
    print("Top students:")
    print(result["top_students"].to_string(index=False))


if __name__ == "__main__":
    main()
