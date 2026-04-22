from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.build_canonical_pipeline import clean_text, coerce_numeric
from src.canonical_bundle import DEFAULT_CANONICAL_ROOT, load_canonical_bundle


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ENHANCED_ROOT = ROOT / "output" / "enhanced_metrics"
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "presentation_ready"
DEFAULT_SEGMENT_MIN_SIZE = 10
DEFAULT_CHAPTER_MIN_SIZE = 15
DEFAULT_TOP_CHAPTER_COUNT = 10

TITLE_FILL = "1F4E79"
HEADER_FILL = "DCE6F1"
UNRESOLVED = {"Active/Unknown", "No Further Observation", "Unknown", ""}


@dataclass
class ReportBundle:
    source_folder: Path
    summary: pd.DataFrame
    longitudinal: pd.DataFrame
    qa: pd.DataFrame
    kpis: List[Dict[str, object]]
    frames: Dict[str, pd.DataFrame]
    takeaways: List[str]
    limitations: List[str]
    definitions: List[Tuple[str, str]]
    withheld_items: List[str]


@dataclass
class SourceBundle:
    enhanced_folder: Path
    enhanced_workbook: Path
    tables: Dict[str, pd.DataFrame]
    sources_used: List[str]
    sources_ignored: List[str]
    caveats: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build executive-facing reporting outputs from the canonical analytics bundle.")
    parser.add_argument("--canonical-root", default=str(DEFAULT_CANONICAL_ROOT))
    parser.add_argument("--canonical-folder", default="")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--segment-min-size", type=int, default=DEFAULT_SEGMENT_MIN_SIZE)
    parser.add_argument("--chapter-min-size", type=int, default=DEFAULT_CHAPTER_MIN_SIZE)
    parser.add_argument("--top-chapters", type=int, default=DEFAULT_TOP_CHAPTER_COUNT)
    parser.add_argument("--include-charts", action="store_true")
    parser.add_argument("--skip-chart-export", action="store_true")
    return parser.parse_args()


def load_latest_bundle(
    enhanced_root: Path,
    explicit_folder: Path | None = None,
    explicit_workbook: Path | None = None,
) -> SourceBundle:
    folder = explicit_folder.expanduser().resolve() if explicit_folder else None
    if explicit_workbook:
        workbook = explicit_workbook.expanduser().resolve()
        folder = workbook.parent
    else:
        if folder is None:
            root = enhanced_root.expanduser().resolve()
            candidates = [path for path in root.iterdir()] if root.exists() else []
            runs = sorted([path for path in candidates if path.is_dir() and path.name.startswith("run_")])
            if not runs:
                raise FileNotFoundError(f"No enhanced analytics runs found under {root}")
            folder = runs[-1]
        matches = sorted(folder.glob("organization_entry_analytics_enhanced_*.xlsx"))
        workbook = matches[-1] if matches else folder / "organization_entry_analytics_enhanced.xlsx"

    tables: Dict[str, pd.DataFrame] = {}
    for filename, key in [
        ("student_summary.csv", "student_summary"),
        ("cohort_metrics.csv", "cohort_metrics"),
        ("master_longitudinal.csv", "master_longitudinal"),
        ("metric_definitions.csv", "metric_definitions"),
        ("qa_checks.csv", "qa_checks"),
        ("outcome_segments.csv", "outcome_segments"),
        ("status_mapping.csv", "status_mapping"),
        ("change_log.csv", "change_log"),
    ]:
        path = folder / filename
        if path.exists():
            tables[key] = pd.read_csv(path)

    caveats: List[str] = []
    if "master_longitudinal" not in tables:
        caveats.append("Master_Longitudinal was not available, so observed-term trend views are limited.")
    return SourceBundle(
        enhanced_folder=folder,
        enhanced_workbook=workbook,
        tables=tables,
        sources_used=[str(path) for path in folder.glob("*.csv")],
        sources_ignored=[],
        caveats=caveats,
    )


def yes_mask(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.lower().eq("yes")


def percent_text(value: object) -> str:
    if value in ("", None) or pd.isna(value):
        return "Not available"
    return f"{float(value):.1%}"


def decimal_text(value: object) -> str:
    if value in ("", None) or pd.isna(value):
        return "Not available"
    return f"{float(value):.2f}"


def count_text(value: object) -> str:
    if value in ("", None) or pd.isna(value):
        return "0"
    return f"{int(round(float(value))):,}"


def adjusted_grad_rate(frame: pd.DataFrame, numerator_field: str, measurable_field: str | None = None) -> Tuple[object, int]:
    eligible = frame.copy()
    if measurable_field and measurable_field in eligible.columns:
        eligible = eligible.loc[yes_mask(eligible[measurable_field])]
    eligible = eligible.loc[yes_mask(eligible["resolved_outcome_flag"])]
    if "student_id" in eligible.columns:
        eligible = eligible.drop_duplicates(subset=["student_id"], keep="first")
    if eligible.empty:
        return "", 0
    numerator = int(yes_mask(eligible[numerator_field]).sum())
    return float(numerator) / float(len(eligible)), int(len(eligible))


def simple_rate(frame: pd.DataFrame, numerator_field: str, measurable_field: str | None = None) -> Tuple[object, int]:
    eligible = frame.copy()
    if measurable_field and measurable_field in eligible.columns:
        eligible = eligible.loc[yes_mask(eligible[measurable_field])]
    if eligible.empty:
        return "", 0
    numerator = int(yes_mask(eligible[numerator_field]).sum())
    return float(numerator) / float(len(eligible)), int(len(eligible))


def selected_cumulative_gpa(frame: pd.DataFrame) -> pd.Series:
    overall = coerce_numeric(frame["latest_overall_cumulative_gpa"])
    txstate = coerce_numeric(frame["latest_txstate_cumulative_gpa"])
    return overall.where(overall.notna(), txstate)


def cumulative_gpa_band(value: object) -> str:
    if value in ("", None) or pd.isna(value):
        return "Unknown"
    number = float(value)
    if number < 2.0:
        return "Below 2.0"
    if number < 2.5:
        return "2.0 to 2.49"
    if number < 3.0:
        return "2.5 to 2.99"
    if number < 3.5:
        return "3.0 to 3.49"
    return "3.5 to 4.00"


def build_kpis(summary: pd.DataFrame) -> List[Dict[str, object]]:
    total_students = int(summary["student_id"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    cohorts = int(summary["org_entry_cohort"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    grad_rate, grad_n = adjusted_grad_rate(summary, "graduated_eventual")
    next_term, _ = simple_rate(summary, "retained_next_term", "retained_next_term_measurable")
    next_fall, _ = simple_rate(summary, "retained_next_fall", "retained_next_fall_measurable")
    first15 = simple_rate(summary.loc[coerce_numeric(summary["first_term_passed_hours"]).notna()].assign(_flag=coerce_numeric(summary["first_term_passed_hours"]).ge(15).map(lambda v: "Yes" if v else "No")), "_flag")[0]
    year30 = simple_rate(summary.loc[coerce_numeric(summary["first_year_passed_hours"]).notna()].assign(_flag=coerce_numeric(summary["first_year_passed_hours"]).ge(30).map(lambda v: "Yes" if v else "No")), "_flag")[0]
    good = simple_rate(summary.assign(_flag=summary["first_academic_standing_bucket"].fillna("").astype(str).eq("Good Standing").map(lambda v: "Yes" if v else "No")), "_flag")[0]
    unresolved = float(summary["latest_outcome_bucket"].fillna("").astype(str).isin(UNRESOLVED).sum()) / float(len(summary)) if len(summary) else ""
    outcomes = summary["latest_outcome_bucket"].fillna("Unknown").astype(str).value_counts()
    return [
        {"Label": "Students tracked", "Display": count_text(total_students), "Explanation": "Distinct students included in the canonical analytics bundle."},
        {"Label": "Organization-entry cohorts covered", "Display": count_text(cohorts), "Explanation": "Distinct observed organization-entry cohorts in the current analysis."},
        {"Label": "Observed graduation rate after joining, excluding unresolved outcomes", "Display": percent_text(grad_rate), "Explanation": f"Share of resolved students with confirmed graduation evidence after first observed organization entry. Based on {count_text(grad_n)} resolved students."},
        {"Label": "Returned the next term after joining", "Display": percent_text(next_term), "Explanation": "Share of students still observed in the organization in the next measurable term after joining."},
        {"Label": "Returned the following fall after joining", "Display": percent_text(next_fall), "Explanation": "Share of students still observed in the organization the following fall when measurable."},
        {"Label": "Earned 15+ passed hours in the first term after joining", "Display": percent_text(first15), "Explanation": "Share of students who passed at least 15 hours in the first observed academic term after joining."},
        {"Label": "Earned 30+ passed hours in the first year after joining", "Display": percent_text(year30), "Explanation": "Share of students who passed at least 30 hours in the first observed academic year after joining."},
        {"Label": "Average first-term GPA after joining", "Display": decimal_text(coerce_numeric(summary["first_term_gpa"]).mean()), "Explanation": "Average GPA in the first observed academic term after organization entry."},
        {"Label": "In good academic standing in the first term after joining", "Display": percent_text(good), "Explanation": "Share of students in good standing in the first observed academic term after joining."},
        {"Label": "Outcomes still unresolved", "Display": percent_text(unresolved), "Explanation": "Students whose latest outcome remains active/unknown or otherwise unresolved."},
        {"Label": "Latest observed dropped / inactive / resigned / revoked", "Display": percent_text((outcomes.get("Dropped/Resigned/Revoked/Inactive", 0) / len(summary)) if len(summary) else ""), "Explanation": "Share of students whose latest observed outcome is a non-graduate exit."},
        {"Label": "Latest observed suspended", "Display": percent_text((outcomes.get("Suspended", 0) / len(summary)) if len(summary) else ""), "Explanation": "Share of students whose latest observed outcome is suspension."},
        {"Label": "Latest observed transfer", "Display": percent_text((outcomes.get("Transfer", 0) / len(summary)) if len(summary) else ""), "Explanation": "Share of students whose latest observed outcome is transfer."},
    ]


def build_frames(summary: pd.DataFrame, longitudinal: pd.DataFrame, chapter_min_size: int, top_chapters: int) -> Tuple[Dict[str, pd.DataFrame], List[str]]:
    frames: Dict[str, pd.DataFrame] = {}
    withheld: List[str] = []

    cohort = (
        summary.loc[summary["org_entry_cohort"].fillna("").astype(str).str.strip().ne("")]
        .groupby("org_entry_cohort", dropna=False)["student_id"]
        .nunique()
        .reset_index(name="Students")
        .rename(columns={"org_entry_cohort": "Cohort"})
        .sort_values(by="Cohort")
    )
    frames["Cohort Overview"] = cohort.loc[cohort["Cohort"] != "Spring 2015"].copy()

    retention_rows = []
    grad_rows = []
    for cohort_name, frame in summary.groupby("org_entry_cohort", dropna=False):
        if not clean_text(cohort_name):
            continue
        retention_rows.append({
            "Cohort": cohort_name,
            "Returned the next term": simple_rate(frame, "retained_next_term", "retained_next_term_measurable")[0],
            "Returned the following fall": simple_rate(frame, "retained_next_fall", "retained_next_fall_measurable")[0],
            "Still enrolled next term": simple_rate(frame, "continued_next_term", "continued_next_term_measurable")[0],
            "Still enrolled the following fall": simple_rate(frame, "continued_next_fall", "continued_next_fall_measurable")[0],
        })
        grad_rows.append({
            "Cohort": cohort_name,
            "Observed graduation, excluding unresolved outcomes": adjusted_grad_rate(frame, "graduated_eventual")[0],
            "Graduated within 4 years, excluding unresolved outcomes": adjusted_grad_rate(frame, "graduated_4yr", "graduated_4yr_measurable")[0],
            "Graduated within 6 years, excluding unresolved outcomes": adjusted_grad_rate(frame, "graduated_6yr", "graduated_6yr_measurable")[0],
        })
    frames["Retention"] = pd.DataFrame(retention_rows).sort_values(by="Cohort") if retention_rows else pd.DataFrame()
    frames["Graduation Outcomes"] = pd.DataFrame(grad_rows).sort_values(by="Cohort") if grad_rows else pd.DataFrame()

    momentum = []
    for label, field, threshold in [
        ("Passed 12+ hours in first academic term after joining", "first_term_passed_hours", 12),
        ("Passed 15+ hours in first academic term after joining", "first_term_passed_hours", 15),
        ("Passed 24+ hours in first academic year after joining", "first_year_passed_hours", 24),
        ("Passed 30+ hours in first academic year after joining", "first_year_passed_hours", 30),
    ]:
        values = coerce_numeric(summary[field]).dropna()
        momentum.append({"Measure": label, "Eligible students": int(len(values)), "Rate": (float((values >= threshold).sum()) / float(len(values))) if len(values) else ""})
    frames["Credit Momentum"] = pd.DataFrame(momentum)

    gpa_progress = (
        longitudinal.loc[coerce_numeric(longitudinal["relative_term_index"]).notna() & coerce_numeric(longitudinal["relative_term_index"]).ge(0)]
        .groupby("relative_term_index", dropna=False)
        .agg(Records=("student_id", "size"), Distinct_Students=("student_id", "nunique"), Average_Term_GPA=("term_gpa", lambda s: coerce_numeric(s).mean()), Average_Cumulative_GPA=("cumulative_gpa", lambda s: coerce_numeric(s).mean()))
        .reset_index()
        .rename(columns={"relative_term_index": "Relative term after entry", "Distinct_Students": "Distinct students", "Average_Term_GPA": "Average term GPA", "Average_Cumulative_GPA": "Average cumulative GPA"})
        .sort_values(by="Relative term after entry")
    )
    frames["GPA and Academic Progress"] = gpa_progress

    standing = summary["first_academic_standing_bucket"].fillna("Unknown").astype(str).value_counts().reset_index()
    standing.columns = ["Standing group", "Students"]
    standing["Rate"] = standing["Students"] / standing["Students"].sum() if standing["Students"].sum() else ""
    frames["Academic Standing"] = standing

    chapters = []
    for chapter, frame in summary.groupby("initial_chapter", dropna=False):
        size = int(frame["student_id"].nunique())
        if not clean_text(chapter):
            continue
        if size < chapter_min_size:
            continue
        chapters.append({
            "Chapter": chapter,
            "Students": size,
            "Observed graduation, excluding unresolved outcomes": adjusted_grad_rate(frame, "graduated_eventual")[0],
            "Returned the following fall": simple_rate(frame, "retained_next_fall", "retained_next_fall_measurable")[0],
            "Average first-term GPA": coerce_numeric(frame["first_term_gpa"]).mean(),
        })
    if chapters:
        frames["Outcome Breakdown"] = pd.DataFrame(chapters).sort_values(by="Students", ascending=False).head(top_chapters)
    else:
        frames["Outcome Breakdown"] = pd.DataFrame()
        withheld.append(f"Chapter comparison withheld because no chapters met the minimum size of {chapter_min_size}.")

    outcome_gpa = []
    for bucket in ["Graduated", "Dropped/Resigned/Revoked/Inactive", "Suspended", "Transfer"]:
        frame = summary.loc[summary["latest_outcome_bucket"].fillna("").astype(str).eq(bucket)]
        if frame.empty:
            continue
        outcome_gpa.append({"Outcome group": bucket, "Students": int(len(frame)), "Average latest cumulative GPA": coerce_numeric(selected_cumulative_gpa(frame)).mean()})
    frames["Average GPA by Outcome Group"] = pd.DataFrame(outcome_gpa)

    band_frame = summary.copy()
    band_frame["selected_cumulative_gpa"] = selected_cumulative_gpa(band_frame)
    band_frame["Latest cumulative GPA band"] = band_frame["selected_cumulative_gpa"].map(cumulative_gpa_band)
    band_frame = band_frame.loc[yes_mask(band_frame["resolved_outcome_flag"])]
    if "student_id" in band_frame.columns:
        band_frame = band_frame.drop_duplicates(subset=["student_id"], keep="first")
    band_rows = []
    for band, frame in band_frame.groupby("Latest cumulative GPA band", dropna=False):
        if band == "Unknown":
            continue
        band_rows.append({"Latest cumulative GPA band": band, "Students": int(len(frame)), "Observed graduation rate": float(yes_mask(frame["graduated_eventual"]).sum()) / float(len(frame)) if len(frame) else "", "Average latest cumulative GPA": coerce_numeric(frame["selected_cumulative_gpa"]).mean()})
    frames["Graduation Rate by GPA Band"] = pd.DataFrame(band_rows)

    join_rows = []
    for bucket, frame in summary.groupby("entry_hours_bucket", dropna=False):
        if not clean_text(bucket):
            continue
        join_rows.append({"Join hours bucket": bucket, "Students": int(len(frame)), "Observed graduation, excluding unresolved outcomes": adjusted_grad_rate(frame, "graduated_eventual")[0], "Retained In Organization To Next Fall": simple_rate(frame, "retained_next_fall", "retained_next_fall_measurable")[0], "Continued Academically To Next Fall": simple_rate(frame, "continued_next_fall", "continued_next_fall_measurable")[0]})
    frames["Join Hours Comparison"] = pd.DataFrame(join_rows)
    return frames, withheld


def build_takeaways(kpis: List[Dict[str, object]], frames: Dict[str, pd.DataFrame]) -> List[str]:
    lookup = {item["Label"]: item["Display"] for item in kpis}
    takeaways = [
        f"{lookup.get('Students tracked', '0')} students are currently represented in the canonical analytics bundle.",
        f"The adjusted observed graduation rate after joining is {lookup.get('Observed graduation rate after joining, excluding unresolved outcomes', 'Not available')}, excluding students whose outcomes are still unresolved and counting graduation only when confirmed evidence exists.",
        f"Next-fall chapter retention after joining is {lookup.get('Returned the following fall after joining', 'Not available')}.",
        f"First-term 15+ passed hours after joining is {lookup.get('Earned 15+ passed hours in the first term after joining', 'Not available')}.",
        "Recent cohorts should be interpreted cautiously because long-window outcomes are shown only when enough follow-up time exists.",
    ]
    gpa_frame = frames.get("Average GPA by Outcome Group", pd.DataFrame())
    if not gpa_frame.empty and len(gpa_frame) >= 2:
        ordered = gpa_frame.sort_values(by="Average latest cumulative GPA", ascending=False)
        takeaways.append(f"Average latest cumulative GPA is highest for {clean_text(ordered.iloc[0]['Outcome group'])} and lowest for {clean_text(ordered.iloc[-1]['Outcome group'])} among the outcome groups shown.")
    return takeaways


def build_report_bundle(canonical_root: Path, explicit_folder: Path | None, chapter_min_size: int, top_chapters: int) -> ReportBundle:
    bundle = load_canonical_bundle(canonical_root=canonical_root, explicit_folder=explicit_folder)
    summary = bundle.tables["student_summary"].copy()
    longitudinal = bundle.tables["master_longitudinal"].copy()
    qa = bundle.tables["qa_checks"].copy()
    kpis = build_kpis(summary)
    frames, withheld = build_frames(summary, longitudinal, chapter_min_size, top_chapters)
    return ReportBundle(
        source_folder=bundle.output_folder,
        summary=summary,
        longitudinal=longitudinal,
        qa=qa,
        kpis=kpis,
        frames=frames,
        takeaways=build_takeaways(kpis, frames),
        limitations=[
            "Roster tracking begins at observed organization participation, not necessarily true school entry.",
            "Academic data is term-level, so missing terms can reflect either true non-enrollment or missing source coverage.",
            "Some recent cohorts are incomplete and do not yet have enough time for long-window outcomes.",
            "Some exits are explicit, while others are only observed through no further records.",
            "Disappearance without confirmed graduation evidence is treated as unresolved, not graduated.",
            "Some joins or outcomes may still rely on fallback matching when Student ID is missing from a source row.",
        ],
        definitions=[
            ("Organization-entry cohort", "A group of students based on the first observed organization term in the data."),
            ("Retention", "Whether a student is still observed in the organization or the school at a later follow-up point."),
            ("Observed graduation", "Graduation directly supported by available records rather than assumed from disappearance."),
            ("Earned credit momentum", "How quickly students passed credit hours after joining."),
            ("Observed", "Results are limited to what appears in the available data and may not reflect the student's full school history."),
        ],
        withheld_items=withheld,
    )


def style_sheet_title(ws, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(wrap_text=True)


def style_header_row(ws, row_idx: int, columns: int) -> None:
    for cell in ws[row_idx][:columns]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def autosize_columns(ws) -> None:
    widths: Dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = clean_text(cell.value)
            if not text:
                continue
            widths[cell.column_letter] = min(max(widths.get(cell.column_letter, 10), len(text) + 2), 45)
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def write_dataframe(ws, frame: pd.DataFrame, start_row: int) -> int:
    if frame.empty:
        ws.cell(row=start_row, column=1, value="No reliable data was available for this section.")
        return start_row
    headers = list(frame.columns)
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=idx, value=header)
    style_header_row(ws, start_row, len(headers))
    for row_offset, values in enumerate(frame.itertuples(index=False), start=1):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value=value)
            header = headers[col_idx - 1].lower()
            if ("rate" in header or "percent" in header) and value not in ("", None):
                cell.number_format = "0.0%"
            elif ("gpa" in header or "average" in header) and value not in ("", None):
                cell.number_format = "0.00"
    return start_row + len(frame)


def write_slides_data_workbook(output_folder: Path, report: ReportBundle) -> Path:
    path = output_folder / "Executive_Report_Slides_Data.xlsx"
    wb = Workbook()
    first = True
    for sheet_name, frame in report.frames.items():
        ws = wb.active if first else wb.create_sheet(title=sheet_name[:31])
        first = False
        ws.title = sheet_name[:31]
        style_sheet_title(ws, sheet_name, "Slide-ready data table.")
        write_dataframe(ws, frame, 4)
        ws.freeze_panes = "A5"
        autosize_columns(ws)
    wb.save(path)
    return path


def write_executive_workbook(output_folder: Path, report: ReportBundle) -> Path:
    path = output_folder / "Executive_Report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    style_sheet_title(ws, "Executive Summary", "A plain-English summary built directly from the canonical analytics tables.")
    row = 4
    for item in report.kpis:
        ws[f"A{row}"] = item["Label"]
        ws[f"A{row}"].font = Font(bold=True, color=TITLE_FILL)
        ws[f"B{row}"] = item["Display"]
        ws[f"C{row}"] = item["Explanation"]
        ws[f"C{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    autosize_columns(ws)

    takeaways_ws = wb.create_sheet(title="Key Takeaways")
    style_sheet_title(takeaways_ws, "Key Takeaways", "Plain-English summary of the biggest patterns currently visible in the data.")
    for idx, takeaway in enumerate(report.takeaways, start=4):
        takeaways_ws[f"A{idx}"] = f"- {takeaway}"
        takeaways_ws[f"A{idx}"].alignment = Alignment(wrap_text=True)
    autosize_columns(takeaways_ws)

    for sheet_name in ["Cohort Overview", "Retention", "Graduation Outcomes", "Credit Momentum", "GPA and Academic Progress", "Academic Standing", "Outcome Breakdown"]:
        ws = wb.create_sheet(title=sheet_name[:31])
        style_sheet_title(ws, sheet_name, "Section table built from canonical outputs.")
        write_dataframe(ws, report.frames.get(sheet_name, pd.DataFrame()), 4)
        if sheet_name == "GPA and Academic Progress":
            next_row = ws.max_row + 3
            ws[f"A{next_row}"] = "Average cumulative GPA by outcome group"
            write_dataframe(ws, report.frames.get("Average GPA by Outcome Group", pd.DataFrame()), next_row + 1)
            next_row = ws.max_row + 3
            ws[f"A{next_row}"] = "Graduation rate by cumulative GPA band"
            write_dataframe(ws, report.frames.get("Graduation Rate by GPA Band", pd.DataFrame()), next_row + 1)
        if sheet_name == "Outcome Breakdown":
            next_row = ws.max_row + 3
            ws[f"A{next_row}"] = "Join-hours comparison"
            write_dataframe(ws, report.frames.get("Join Hours Comparison", pd.DataFrame()), next_row + 1)
        autosize_columns(ws)

    defs = wb.create_sheet(title="Definitions and Notes")
    style_sheet_title(defs, "Definitions and Notes", "Short plain-language explanations of how to read the measures in this package.")
    defs["A4"] = "Term"
    defs["B4"] = "Meaning"
    style_header_row(defs, 4, 2)
    for idx, (term, meaning) in enumerate(report.definitions, start=5):
        defs[f"A{idx}"] = term
        defs[f"B{idx}"] = meaning
        defs[f"B{idx}"].alignment = Alignment(wrap_text=True)
    autosize_columns(defs)

    limits = wb.create_sheet(title="Data Limitations")
    style_sheet_title(limits, "Data Limitations", "These cautions help keep the results honest and prevent overstatement.")
    for idx, limitation in enumerate(report.limitations, start=4):
        limits[f"A{idx}"] = f"- {limitation}"
        limits[f"A{idx}"].alignment = Alignment(wrap_text=True)
    autosize_columns(limits)

    qa = wb.create_sheet(title="QA Summary")
    style_sheet_title(qa, "QA Summary", "Source checks, caveats, and withheld items.")
    notes = [f"Canonical source folder used: {report.source_folder}"]
    notes.extend(report.withheld_items)
    if report.qa.empty:
        notes.append("No QA table was available.")
    else:
        review_rows = report.qa.loc[report.qa["Status"].fillna("").astype(str).isin(["Review", "Fail"])]
        notes.append("All QA rows currently in the canonical bundle are passing." if review_rows.empty else f"{len(review_rows)} QA rows are marked Review or Fail.")
    for idx, item in enumerate(notes, start=4):
        qa[f"A{idx}"] = f"- {item}"
        qa[f"A{idx}"].alignment = Alignment(wrap_text=True)
    next_row = qa.max_row + 2
    write_dataframe(qa, report.qa, next_row)
    autosize_columns(qa)

    appendix = wb.create_sheet(title="Appendix")
    style_sheet_title(appendix, "Appendix / Technical Detail", "Friendly label map and additional technical detail for advanced readers.")
    appendix["A4"] = "Friendly label"
    appendix["B4"] = "Meaning"
    style_header_row(appendix, 4, 2)
    for idx, row in enumerate([
        ("Student ID", "Institutional student identifier used as the main join key."),
        ("Organization Retention", "Whether a student is still observed in the organization at a later point."),
        ("School Continuation", "Whether a student is still observed in academic records at a later point."),
        ("Earned Credit Momentum", "How quickly students passed hours after joining."),
    ], start=5):
        appendix[f"A{idx}"] = row[0]
        appendix[f"B{idx}"] = row[1]
        appendix[f"B{idx}"].alignment = Alignment(wrap_text=True)
    autosize_columns(appendix)

    wb.save(path)
    return path


def write_markdown_summary(output_folder: Path, report: ReportBundle) -> Path:
    path = output_folder / "Executive_Summary.md"
    lines = ["# Executive Summary", "", "This summary is built directly from the canonical analytics bundle.", "", "## Headline KPIs", ""]
    for item in report.kpis[:10]:
        lines.append(f"- **{item['Label']}**: {item['Display']}")
        lines.append(f"  {item['Explanation']}")
    lines.extend(["", "## Key Takeaways", ""])
    for takeaway in report.takeaways:
        lines.append(f"- {takeaway}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_reporting_readme(output_folder: Path, report: ReportBundle, workbook_path: Path, slides_path: Path) -> Path:
    path = output_folder / "Reporting_README.md"
    lines = [
        "# Executive Reporting Package",
        "",
        "This folder contains the non-technical reporting package built from the canonical analytics outputs.",
        "",
        "## Files",
        "",
        f"- `{workbook_path.name}`: polished executive workbook",
        f"- `{slides_path.name}`: slide-ready data tables",
        "- `Executive_Summary.md`: one-page plain-English summary",
        "- `charts/data/*.csv`: chart-ready data extracts",
        "",
        "## Source of truth",
        "",
        f"- `{report.source_folder}`",
        "",
        "This report reads only from canonical outputs (`roster_term`, `academic_term`, `master_longitudinal`, `student_summary`, `cohort_metrics`, and `qa_checks`).",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_chart_data(output_folder: Path, frames: Dict[str, pd.DataFrame]) -> Path:
    folder = output_folder / "charts" / "data"
    folder.mkdir(parents=True, exist_ok=True)
    for name, frame in frames.items():
        if frame.empty:
            continue
        slug = clean_text(name).lower().replace(" ", "_").replace("/", "_")
        frame.to_csv(folder / f"{slug}.csv", index=False)
    return folder


def make_output_folder(output_root: Path) -> Path:
    timestamp = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    folder = output_root / timestamp
    folder.mkdir(parents=True, exist_ok=False)
    (folder / "charts").mkdir(parents=True, exist_ok=True)
    (folder / "charts" / "data").mkdir(parents=True, exist_ok=True)
    return folder


def build_executive_report(canonical_root: Path, explicit_folder: Path | None, output_root: Path, chapter_min_size: int, top_chapters: int, include_charts: bool, skip_chart_export: bool) -> Dict[str, object]:
    report = build_report_bundle(canonical_root, explicit_folder, chapter_min_size, top_chapters)
    output_folder = make_output_folder(output_root)
    chart_manifest = write_chart_data(output_folder, report.frames)
    slides_path = write_slides_data_workbook(output_folder, report)
    workbook_path = write_executive_workbook(output_folder, report)
    summary_path = write_markdown_summary(output_folder, report)
    readme_path = write_reporting_readme(output_folder, report, workbook_path, slides_path)
    return {
        "output_folder": output_folder,
        "executive_workbook": workbook_path,
        "slides_workbook": slides_path,
        "executive_summary": summary_path,
        "readme": readme_path,
        "chart_manifest": chart_manifest,
        "chart_exported": False,
        "chart_export_message": "Chart PNG export is disabled in the canonical executive report. Chart-ready CSVs were written instead.",
        "withheld_items": report.withheld_items,
    }


def main() -> None:
    args = parse_args()
    result = build_executive_report(
        canonical_root=Path(args.canonical_root).expanduser().resolve(),
        explicit_folder=Path(args.canonical_folder).expanduser().resolve() if args.canonical_folder else None,
        output_root=Path(args.output_root).expanduser().resolve(),
        chapter_min_size=args.chapter_min_size,
        top_chapters=args.top_chapters,
        include_charts=args.include_charts,
        skip_chart_export=args.skip_chart_export,
    )
    print(f"Executive reporting package created: {result['output_folder']}")
    print(f"Executive workbook: {result['executive_workbook']}")
    print(f"Slides workbook: {result['slides_workbook']}")
    print(f"Executive summary: {result['executive_summary']}")
    print(f"README: {result['readme']}")
    print(f"Chart export: {result['chart_export_message']}")
    if result["withheld_items"]:
        print("Withheld items:")
        for item in result["withheld_items"]:
            print(f"- {item}")


if __name__ == "__main__":
    main()
