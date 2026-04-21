from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.build_canonical_pipeline import clean_text
from src.canonical_bundle import DEFAULT_CANONICAL_ROOT, load_canonical_bundle


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_ROOT = ROOT / "output" / "unresolved_outcomes"

TITLE_FILL = "1F4E79"
HEADER_FILL = "DCE6F1"
REVIEW_FILL = "FFF2CC"

UNKNOWN_OUTCOME_GROUP = "Truly Unknown / Unresolved"
ACTIVE_OUTCOME_GROUP = "Still Active"
GRADUATED_OUTCOME_GROUP = "Graduated"
RESOLVED_NON_GRAD_GROUP = "Resolved Non-Graduate Exit"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rank organization-entry years and last-seen years by unresolved outcome counts. "
            "Use this to identify cohorts that need outside records or manual follow-up."
        )
    )
    parser.add_argument("--canonical-root", default=str(DEFAULT_CANONICAL_ROOT))
    parser.add_argument("--canonical-folder", default="")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    return parser.parse_args()


def bool_like(value: object) -> bool:
    text = clean_text(value).strip().lower()
    return text in {"true", "t", "yes", "y", "1"}


def extract_year(*values: object) -> object:
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        if re.fullmatch(r"(19|20)\d{2}", text):
            return int(text)
        match = re.search(r"(19|20)\d{2}", text)
        if match:
            return int(match.group(0))
    return pd.NA


def preferred_column(frame: pd.DataFrame, candidates: Iterable[str], default: str = "") -> pd.Series:
    for column in candidates:
        if column in frame.columns:
            return frame[column]
    return pd.Series([default] * len(frame), index=frame.index)


def classify_population(summary: pd.DataFrame) -> pd.DataFrame:
    result = summary.copy()
    result["student_id"] = preferred_column(result, ["student_id", "Student ID"]).map(clean_text)
    result["student_name"] = preferred_column(result, ["student_name", "Student Name"]).map(clean_text)
    result["chapter"] = preferred_column(result, ["chapter", "initial_chapter", "latest_chapter"]).map(clean_text)
    result["join_term"] = preferred_column(result, ["join_term", "org_entry_cohort", "first_observed_org_term"]).map(clean_text)
    result["join_term_code"] = preferred_column(result, ["join_term_code", "first_observed_org_term_code"]).map(clean_text)
    result["latest_outcome_bucket"] = preferred_column(result, ["latest_outcome_bucket", "final_outcome_bucket"]).map(clean_text)
    result["outcome_resolution_group"] = preferred_column(result, ["outcome_resolution_group"]).map(clean_text)
    result["outcome_evidence_source"] = preferred_column(result, ["outcome_evidence_source"]).map(clean_text)
    result["latest_roster_status_bucket"] = preferred_column(result, ["latest_roster_status_bucket"]).map(clean_text)
    result["last_observed_org_term"] = preferred_column(result, ["last_observed_org_term"]).map(clean_text)
    result["last_observed_org_term_code"] = preferred_column(result, ["last_observed_org_term_code"]).map(clean_text)
    result["last_observed_academic_term"] = preferred_column(result, ["last_observed_academic_term"]).map(clean_text)
    result["last_observed_academic_term_code"] = preferred_column(result, ["last_observed_academic_term_code"]).map(clean_text)

    result["join_year_clean"] = [
        extract_year(year, code, label)
        for year, code, label in zip(
            preferred_column(result, ["join_year"]),
            result["join_term_code"],
            result["join_term"],
        )
    ]
    result["last_seen_year"] = [
        extract_year(ac_code, ac_label, org_code, org_label)
        for ac_code, ac_label, org_code, org_label in zip(
            result["last_observed_academic_term_code"],
            result["last_observed_academic_term"],
            result["last_observed_org_term_code"],
            result["last_observed_org_term"],
        )
    ]

    if "is_unknown_outcome" in result.columns:
        is_unknown = result["is_unknown_outcome"].map(bool_like)
    else:
        is_unknown = result["outcome_resolution_group"].eq(UNKNOWN_OUTCOME_GROUP) | result["latest_outcome_bucket"].isin(
            ["No Further Observation", "Unknown", "Active/Unknown"]
        )

    if "is_active_outcome" in result.columns:
        is_active = result["is_active_outcome"].map(bool_like)
    else:
        is_active = result["outcome_resolution_group"].eq(ACTIVE_OUTCOME_GROUP)

    if "is_resolved_outcome" in result.columns:
        is_resolved = result["is_resolved_outcome"].map(bool_like)
    else:
        is_resolved = result["outcome_resolution_group"].isin([GRADUATED_OUTCOME_GROUP, RESOLVED_NON_GRAD_GROUP])

    if "is_graduated" in result.columns:
        is_graduated = result["is_graduated"].map(bool_like)
    else:
        is_graduated = result["outcome_resolution_group"].eq(GRADUATED_OUTCOME_GROUP) | result["latest_outcome_bucket"].eq("Graduated")

    if "is_known_non_graduate_exit" in result.columns:
        is_known_non_grad = result["is_known_non_graduate_exit"].map(bool_like)
    else:
        is_known_non_grad = result["outcome_resolution_group"].eq(RESOLVED_NON_GRAD_GROUP)

    result["is_unresolved_problem_case"] = is_unknown & ~is_active & ~is_graduated & ~is_resolved
    result["is_active_outcome_clean"] = is_active
    result["is_graduated_clean"] = is_graduated
    result["is_known_non_graduate_exit_clean"] = is_known_non_grad
    result["is_resolved_outcome_clean"] = is_resolved
    return result


def summarize_by_year(frame: pd.DataFrame, year_column: str, label: str) -> pd.DataFrame:
    rows = []
    usable = frame.loc[frame[year_column].notna()].copy()
    for year, group in usable.groupby(year_column, dropna=False):
        total = int(len(group))
        unresolved = int(group["is_unresolved_problem_case"].sum())
        active = int(group["is_active_outcome_clean"].sum())
        graduated = int(group["is_graduated_clean"].sum())
        known_non_grad = int(group["is_known_non_graduate_exit_clean"].sum())
        resolved = int(group["is_resolved_outcome_clean"].sum())
        rows.append(
            {
                label: int(year),
                "Total Students": total,
                "Truly Unresolved Count": unresolved,
                "Truly Unresolved Rate": unresolved / total if total else 0,
                "Resolved Count": resolved,
                "Graduated Count": graduated,
                "Known Non-Graduate Exit Count": known_non_grad,
                "Still Active Count": active,
                "Review Priority Score": unresolved * 1000 + total,
            }
        )
    result = pd.DataFrame(rows)
    if result.empty:
        return pd.DataFrame(
            columns=[
                label,
                "Total Students",
                "Truly Unresolved Count",
                "Truly Unresolved Rate",
                "Resolved Count",
                "Graduated Count",
                "Known Non-Graduate Exit Count",
                "Still Active Count",
                "Review Priority Score",
            ]
        )
    return result.sort_values(["Truly Unresolved Count", "Truly Unresolved Rate", label], ascending=[False, False, True]).reset_index(drop=True)


def summarize_entry_last_seen(frame: pd.DataFrame) -> pd.DataFrame:
    unresolved = frame.loc[frame["is_unresolved_problem_case"]].copy()
    unresolved = unresolved.loc[unresolved["join_year_clean"].notna() & unresolved["last_seen_year"].notna()]
    if unresolved.empty:
        return pd.DataFrame(columns=["Organization Entry Year", "Last Seen Year", "Truly Unresolved Count"])
    grouped = (
        unresolved.groupby(["join_year_clean", "last_seen_year"], dropna=False)
        .size()
        .reset_index(name="Truly Unresolved Count")
        .rename(columns={"join_year_clean": "Organization Entry Year", "last_seen_year": "Last Seen Year"})
    )
    grouped["Organization Entry Year"] = grouped["Organization Entry Year"].astype(int)
    grouped["Last Seen Year"] = grouped["Last Seen Year"].astype(int)
    return grouped.sort_values(["Truly Unresolved Count", "Organization Entry Year", "Last Seen Year"], ascending=[False, True, True]).reset_index(drop=True)


def unresolved_student_list(frame: pd.DataFrame) -> pd.DataFrame:
    unresolved = frame.loc[frame["is_unresolved_problem_case"]].copy()
    columns = [
        "student_id",
        "student_name",
        "chapter",
        "join_year_clean",
        "join_term",
        "last_seen_year",
        "last_observed_org_term",
        "last_observed_academic_term",
        "latest_outcome_bucket",
        "outcome_resolution_group",
        "latest_roster_status_bucket",
        "outcome_evidence_source",
    ]
    unresolved = unresolved.loc[:, [column for column in columns if column in unresolved.columns]].rename(
        columns={
            "student_id": "Student ID",
            "student_name": "Student Name",
            "chapter": "Chapter",
            "join_year_clean": "Organization Entry Year",
            "join_term": "Organization Entry Term",
            "last_seen_year": "Last Seen Year",
            "last_observed_org_term": "Last Observed Org Term",
            "last_observed_academic_term": "Last Observed Academic Term",
            "latest_outcome_bucket": "Latest Outcome Bucket",
            "outcome_resolution_group": "Outcome Resolution Group",
            "latest_roster_status_bucket": "Latest Roster Status",
            "outcome_evidence_source": "Outcome Evidence Source",
        }
    )
    return unresolved.sort_values(["Organization Entry Year", "Last Seen Year", "Chapter", "Student Name"], na_position="last").reset_index(drop=True)


def write_sheet(workbook: Workbook, title: str, frame: pd.DataFrame, description: str) -> None:
    ws = workbook.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A2"] = description
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.append([])
    headers = list(frame.columns)
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True)
    for values in frame.itertuples(index=False, name=None):
        ws.append(list(values))
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(clean_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 45)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            if isinstance(cell.value, float) and "Rate" in clean_text(ws.cell(row=header_row, column=cell.column).value):
                cell.number_format = "0.0%"


def write_workbook(
    path: Path,
    by_entry_year: pd.DataFrame,
    by_last_seen_year: pd.DataFrame,
    entry_last_seen: pd.DataFrame,
    unresolved_students: pd.DataFrame,
    source_folder: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(
        workbook,
        "Entry Year Ranking",
        by_entry_year,
        "Ranks organization-entry years by students who remain truly unresolved, excluding still-active students.",
    )
    write_sheet(
        workbook,
        "Last Seen Year Ranking",
        by_last_seen_year,
        "Ranks the last observed year for unresolved students. This helps identify years where students most often disappear from the data.",
    )
    write_sheet(
        workbook,
        "Entry x Last Seen",
        entry_last_seen,
        "Shows unresolved students by organization-entry year and last-seen year.",
    )
    write_sheet(
        workbook,
        "Unresolved Students",
        unresolved_students,
        "Student-level list for manual follow-up. These students are not counted as confirmed graduates or confirmed non-graduate exits.",
    )
    method = pd.DataFrame(
        [
            {
                "Topic": "Source",
                "Note": f"Built from canonical student_summary in {source_folder}",
            },
            {
                "Topic": "Unresolved definition",
                "Note": "Counts students marked Truly Unknown / Unresolved, excluding Still Active, Graduated, and resolved non-graduate exits.",
            },
            {
                "Topic": "Entry year",
                "Note": "Organization-entry year is based on join_year when available, otherwise join term/code.",
            },
            {
                "Topic": "Last seen year",
                "Note": "Uses last observed academic term first, then last observed organization term when academic term is missing.",
            },
        ]
    )
    write_sheet(workbook, "Method", method, "Plain-English notes for interpreting this report.")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_unresolved_outcome_year_report(
    canonical_root: Path = DEFAULT_CANONICAL_ROOT,
    canonical_folder: Path | None = None,
    output_root: Path = DEFAULT_OUTPUT_ROOT,
) -> dict[str, Path]:
    bundle = load_canonical_bundle(canonical_root, explicit_folder=canonical_folder)
    summary = classify_population(bundle.tables["student_summary"])
    by_entry_year = summarize_by_year(summary, "join_year_clean", "Organization Entry Year")
    by_last_seen_year = summarize_by_year(summary, "last_seen_year", "Last Seen Year")
    entry_last_seen = summarize_entry_last_seen(summary)
    students = unresolved_student_list(summary)

    output_folder = output_root / datetime.now().strftime("run_%Y%m%d_%H%M%S")
    output_folder.mkdir(parents=True, exist_ok=True)
    files = {
        "workbook": output_folder / "Unresolved_Outcome_Problem_Years.xlsx",
        "by_entry_year": output_folder / "unresolved_outcomes_by_entry_year.csv",
        "by_last_seen_year": output_folder / "unresolved_outcomes_by_last_seen_year.csv",
        "entry_last_seen": output_folder / "unresolved_outcomes_entry_x_last_seen.csv",
        "students": output_folder / "unresolved_outcome_students.csv",
    }
    by_entry_year.to_csv(files["by_entry_year"], index=False)
    by_last_seen_year.to_csv(files["by_last_seen_year"], index=False)
    entry_last_seen.to_csv(files["entry_last_seen"], index=False)
    students.to_csv(files["students"], index=False)
    write_workbook(files["workbook"], by_entry_year, by_last_seen_year, entry_last_seen, students, bundle.output_folder)
    return files


def main() -> None:
    args = parse_args()
    explicit = Path(args.canonical_folder).expanduser().resolve() if args.canonical_folder else None
    files = build_unresolved_outcome_year_report(
        canonical_root=Path(args.canonical_root).expanduser().resolve(),
        canonical_folder=explicit,
        output_root=Path(args.output_root).expanduser().resolve(),
    )
    print("Unresolved outcome problem-year report created:")
    for label, path in files.items():
        print(f"{label}: {path}")
    top_years = pd.read_csv(files["by_entry_year"])
    if not top_years.empty:
        print("\nTop organization-entry years by truly unresolved outcomes:")
        print(top_years.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
