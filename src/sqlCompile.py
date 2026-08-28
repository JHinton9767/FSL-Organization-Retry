from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from src.build_canonical_pipeline import (
    parse_term_code,
    roster_file_month_details,
    roster_file_version_details,
    roster_term_label_from_context,
    roster_title_details_from_rows,
    sort_term_code,
)
from src.build_master_roster import (
    SUPPORTED_EXTENSIONS,
    chapter_from_filename,
    detect_inline_chapter_label,
    find_status_column_in_rows,
    get_cell,
    infer_chapter,
    is_excluded_roster_position,
    is_placeholder_sheet_name,
    normalize_banner_id,
    normalize_chapter_name,
    score_header_row,
    source_file_format_priority,
    source_file_label,
)
from src.path_config import ROOT, load_path_config
from src.shared_utils import clean_text


TABLE_NAME = "sqlCompile"
OUTPUT_COLUMNS = ["Semester", "Chapter", "Student ID", "Status"]
STUDENT_NAME_TABLE = "sqlCompile_student_names"
STUDENT_NAME_COLUMNS = ["Student ID", "Student Name"]
ROSTER_INVENTORY_TABLE = "sqlCompile_roster_inventory"
ROSTER_INVENTORY_COLUMNS = [
    "Semester",
    "Chapter",
    "Roster Pass",
    "Roster Pass Priority",
    "Roster Month",
    "Roster Month Priority",
    "Source File",
    "Source Sheet",
    "Student Rows",
]
DEFAULT_OUTPUT_PATH = ROOT / "output" / "sqlCompile" / "sqlCompile.sqlite"
SKIPPED_DIRECTORY_NAMES = {
    ".git",
    ".pytest_cache",
    ".pytest_tmp",
    ".uv-cache",
    ".uv-python",
    "__pycache__",
    "build",
    "cache",
    "dist",
    "exports",
    "output",
    "outputs",
    "pytest_tmp",
    "reports",
}


@dataclass(frozen=True)
class SqlCompileResult:
    output_path: Path
    table_name: str
    row_count: int
    source_file_count: int
    issue_count: int
    issues: pd.DataFrame


def _resolve_path(value: str | Path) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def default_input_roots(config_path: Optional[str | Path] = None) -> List[Path]:
    paths = load_path_config(config_path)
    roots = [paths.rosters_root, paths.roster_inbox_root]
    return _dedupe_paths(roots)


def _dedupe_paths(paths: Iterable[Path]) -> List[Path]:
    unique: List[Path] = []
    seen: set[Path] = set()
    for path in paths:
        resolved = Path(path).resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(resolved)
    return unique


def _path_is_in_skipped_directory(path: Path, root: Path) -> bool:
    try:
        relative_parts = path.relative_to(root).parts[:-1]
    except ValueError:
        relative_parts = path.parts[:-1]
    return any(part in SKIPPED_DIRECTORY_NAMES for part in relative_parts)


def excel_files(roots: Sequence[Path]) -> List[Path]:
    files: List[Path] = []
    seen: set[Path] = set()
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file() or path.name.startswith("~$"):
                continue
            if _path_is_in_skipped_directory(path, root):
                continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(path)
    return sorted(files)


def _source_label(path: Path, roots: Sequence[Path]) -> str:
    for root in roots:
        try:
            return source_file_label(path, root)
        except Exception:
            continue
    return source_file_label(path)


def _find_compile_header_row(table_rows: Sequence[Tuple[object, ...]]) -> Tuple[Optional[int], dict[str, int]]:
    best_score = -1
    best_row_idx: Optional[int] = None
    best_map: dict[str, int] = {}
    for row_idx, row in enumerate(table_rows[:25], start=1):
        score, header_map = score_header_row(list(row))
        if "banner_id" not in header_map:
            continue
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_map = header_map

    if best_row_idx is None:
        return None, {}

    status_row_idx, status_col_idx = find_status_column_in_rows(list(table_rows))
    if "status" not in best_map and status_col_idx is not None:
        best_map["status"] = status_col_idx
    if "status" not in best_map:
        return None, {}

    return max(best_row_idx, status_row_idx or best_row_idx), best_map


def _default_chapter(path: Path, sheet_name: str, table_rows: Sequence[Tuple[object, ...]], header_row_idx: Optional[int]) -> str:
    _, title_chapter, _ = roster_title_details_from_rows(table_rows, header_row_idx)
    if title_chapter:
        return title_chapter
    inferred = infer_chapter(path, sheet_name)
    if inferred:
        return inferred
    sheet_chapter = normalize_chapter_name(sheet_name) if not is_placeholder_sheet_name(sheet_name) else ""
    if sheet_chapter and sheet_chapter != "Unknown":
        return sheet_chapter
    file_chapter = chapter_from_filename(path)
    if file_chapter and file_chapter != "Unknown":
        return file_chapter
    return "Unknown"


def _term_details(path: Path, table_rows: Sequence[Tuple[object, ...]], header_row_idx: Optional[int]) -> Tuple[str, str, int]:
    _, _, title_term_label = roster_title_details_from_rows(table_rows, header_row_idx)
    term_label, _ = roster_term_label_from_context(path, title_term_label)
    term_code, parsed_label, _, _ = parse_term_code(term_label)
    if term_code:
        return parsed_label, term_code, sort_term_code(term_code)
    return clean_text(term_label) or "Unknown", "", 999999


def _status_priority(value: object) -> int:
    text = clean_text(value).upper()
    compact = re.sub(r"[^A-Z0-9]+", "", text)
    if not compact:
        return 0
    if compact in {"A", "ACTIVE", "MEMBER", "ACTIVEMEMBER"}:
        return 1
    if compact in {"N", "NEW", "NEWMEMBER", "NEWMEMBERS", "ASSOCIATEMEMBER", "ASSOCIATEMEMBERS"}:
        return 2
    return 3


def _normalized_header(value: object) -> str:
    text = clean_text(value).lower().replace("_", " ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _full_name_column(header_row: Sequence[object], header_map: dict[str, int]) -> Optional[int]:
    protected_indexes = {
        index
        for field, index in header_map.items()
        if field in {"first_name", "last_name", "banner_id", "email", "status", "semester_joined", "position", "chapter"}
    }
    aliases = {"name", "student name", "member name", "full name", "legal name"}
    for index, header in enumerate(header_row):
        if index in protected_indexes:
            continue
        normalized = _normalized_header(header)
        if normalized in aliases:
            return index
    return None


def _format_student_name(first_name: object = "", last_name: object = "", full_name: object = "") -> str:
    first = clean_text(first_name)
    last = clean_text(last_name)
    if first or last:
        return " ".join(part for part in [first, last] if part)

    full = clean_text(full_name)
    if not full:
        return ""
    if "," in full:
        left, right = [clean_text(part) for part in full.split(",", 1)]
        return " ".join(part for part in [right, left] if part)
    return full


def _student_name_from_row(
    row: Sequence[object],
    header_map: dict[str, int],
    header_row: Sequence[object],
) -> str:
    first_name = get_cell(row, header_map.get("first_name"))
    last_name = get_cell(row, header_map.get("last_name"))
    full_name_index = _full_name_column(header_row, header_map)
    full_name = get_cell(row, full_name_index) if full_name_index is not None else ""
    return _format_student_name(first_name, last_name, full_name)


def _load_sheet_rows(
    path: Path,
    roots: Sequence[Path],
    sheet_name: str,
    table_rows: Sequence[Tuple[object, ...]],
    file_index: int,
    exceptions: List[dict],
) -> List[dict]:
    source_label = _source_label(path, roots)
    header_row_idx, header_map = _find_compile_header_row(table_rows)
    if header_row_idx is None:
        exceptions.append(
            {
                "exception_type": "sql_compile_header_missing",
                "source_file": source_label,
                "source_sheet": sheet_name,
                "details": "Sheet skipped because no Student ID and Status header pair was found.",
            }
        )
        return []

    semester, term_code, term_sort = _term_details(path, table_rows, header_row_idx)
    current_chapter = _default_chapter(path, sheet_name, table_rows, header_row_idx)
    roster_file_version, roster_file_version_priority = roster_file_version_details(" ".join(path.parts))
    roster_file_month, roster_file_month_priority = roster_file_month_details(" ".join(path.parts))
    header_row = table_rows[header_row_idx - 1] if header_row_idx and header_row_idx <= len(table_rows) else ()
    row_results: List[dict] = []

    for source_row_index, row in enumerate(table_rows[header_row_idx:], start=header_row_idx + 1):
        inline_chapter_raw = detect_inline_chapter_label(row, header_map)
        if inline_chapter_raw:
            inline_chapter = normalize_chapter_name(inline_chapter_raw)
            if inline_chapter and inline_chapter != "Unknown":
                current_chapter = inline_chapter
            continue

        student_id_raw = get_cell(row, header_map.get("banner_id"))
        student_id = normalize_banner_id(student_id_raw)
        if not student_id:
            continue

        position_raw = get_cell(row, header_map.get("position"))
        if is_excluded_roster_position(position_raw):
            continue

        chapter_raw = clean_text(get_cell(row, header_map.get("chapter")))
        chapter = normalize_chapter_name(chapter_raw)
        if not chapter or chapter == "Unknown":
            chapter = current_chapter or "Unknown"

        status = clean_text(get_cell(row, header_map.get("status")))
        row_results.append(
            {
                "Semester": semester,
                "Chapter": chapter,
                "Student ID": student_id,
                "Student Name": _student_name_from_row(row, header_map, header_row),
                "Status": status,
                "_term_code": term_code,
                "_term_sort": term_sort,
                "_status_priority": _status_priority(status),
                "_source_version_priority": roster_file_version_priority,
                "_source_month_priority": roster_file_month_priority,
                "_source_format_priority": source_file_format_priority(source_label),
                "_source_file": source_label,
                "_source_sheet": sheet_name,
                "_source_file_index": file_index,
                "_source_row_index": source_row_index,
                "_roster_file_version": roster_file_version,
                "_roster_file_month": roster_file_month,
            }
        )

    return row_results


def load_sql_compile_rows(roots: Sequence[str | Path]) -> Tuple[pd.DataFrame, pd.DataFrame, int]:
    resolved_roots = _dedupe_paths(_resolve_path(root) for root in roots)
    exceptions: List[dict] = []
    rows: List[dict] = []
    paths = excel_files(resolved_roots)

    for file_index, path in enumerate(paths):
        source_label = _source_label(path, resolved_roots)
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            exceptions.append(
                {
                    "exception_type": "sql_compile_open_error",
                    "source_file": source_label,
                    "source_sheet": "",
                    "details": str(exc),
                }
            )
            continue

        try:
            for worksheet in workbook.worksheets:
                table_rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
                rows.extend(_load_sheet_rows(path, resolved_roots, worksheet.title, table_rows, file_index, exceptions))
        finally:
            workbook.close()

    return pd.DataFrame(rows), pd.DataFrame(exceptions), len(paths)


def resolve_semester_statuses(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    ordered = rows.sort_values(
        by=[
            "Student ID",
            "_term_sort",
            "Semester",
            "_status_priority",
            "_source_version_priority",
            "_source_month_priority",
            "_source_format_priority",
            "_source_file",
            "_source_sheet",
            "_source_row_index",
        ],
        ascending=[True, True, True, False, False, False, False, True, True, True],
        na_position="last",
    )
    deduped = ordered.drop_duplicates(subset=["_term_code", "Semester", "Student ID"], keep="first").copy()
    final = deduped.sort_values(["_term_sort", "Semester", "Chapter", "Student ID"], na_position="last")
    return final.loc[:, OUTPUT_COLUMNS].reset_index(drop=True)


def build_roster_inventory(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=ROSTER_INVENTORY_COLUMNS)

    required_columns = [
        "Semester",
        "Chapter",
        "Student ID",
        "_roster_file_version",
        "_source_version_priority",
        "_roster_file_month",
        "_source_month_priority",
        "_source_file",
        "_source_sheet",
    ]
    work = rows.copy()
    for column in required_columns:
        if column not in work.columns:
            work[column] = ""
    for column in ["Semester", "Chapter", "Student ID", "_roster_file_version", "_roster_file_month", "_source_file", "_source_sheet"]:
        work[column] = work[column].fillna("").astype(str).map(clean_text)
    work["_source_version_priority"] = pd.to_numeric(work["_source_version_priority"], errors="coerce").fillna(0)
    work["_source_month_priority"] = pd.to_numeric(work["_source_month_priority"], errors="coerce").fillna(0).astype(int)
    work = work.loc[work["Semester"].ne("") & work["Chapter"].ne("") & work["Chapter"].ne("Unknown")].copy()
    if work.empty:
        return pd.DataFrame(columns=ROSTER_INVENTORY_COLUMNS)

    grouped = (
        work.groupby(
            [
                "Semester",
                "Chapter",
                "_roster_file_version",
                "_source_version_priority",
                "_roster_file_month",
                "_source_month_priority",
                "_source_file",
                "_source_sheet",
            ],
            dropna=False,
        )["Student ID"]
        .nunique()
        .reset_index(name="Student Rows")
    )
    result = grouped.rename(
        columns={
            "_roster_file_version": "Roster Pass",
            "_source_version_priority": "Roster Pass Priority",
            "_roster_file_month": "Roster Month",
            "_source_month_priority": "Roster Month Priority",
            "_source_file": "Source File",
            "_source_sheet": "Source Sheet",
        }
    )
    result["Roster Pass"] = result["Roster Pass"].replace("", "Regular")
    result = result.loc[:, ROSTER_INVENTORY_COLUMNS].copy()
    result["_sort"] = result["Semester"].map(lambda value: sort_term_code(parse_term_code(value)[0]) if parse_term_code(value)[0] else 999999)
    return result.sort_values(["_sort", "Semester", "Chapter", "Roster Pass Priority", "Source File", "Source Sheet"]).drop(columns=["_sort"]).reset_index(drop=True)


def build_student_name_lookup(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=STUDENT_NAME_COLUMNS)

    work = rows.copy()
    required_columns = [
        "Student ID",
        "Student Name",
        "_term_sort",
        "_source_version_priority",
        "_source_month_priority",
        "_source_format_priority",
        "_source_file_index",
        "_source_row_index",
    ]
    for column in required_columns:
        if column not in work.columns:
            work[column] = ""
    for column in ["Student ID", "Student Name"]:
        work[column] = work[column].fillna("").astype(str).map(clean_text)
    work = work.loc[work["Student ID"].ne("") & work["Student Name"].ne("")].copy()
    if work.empty:
        return pd.DataFrame(columns=STUDENT_NAME_COLUMNS)

    work["_term_sort"] = pd.to_numeric(work["_term_sort"], errors="coerce").fillna(0)
    work["_source_version_priority"] = pd.to_numeric(work["_source_version_priority"], errors="coerce").fillna(0)
    work["_source_month_priority"] = pd.to_numeric(work["_source_month_priority"], errors="coerce").fillna(0)
    work["_source_format_priority"] = pd.to_numeric(work["_source_format_priority"], errors="coerce").fillna(0)
    work["_source_file_index"] = pd.to_numeric(work["_source_file_index"], errors="coerce").fillna(0)
    work["_source_row_index"] = pd.to_numeric(work["_source_row_index"], errors="coerce").fillna(0)
    work["_name_count"] = work.groupby(["Student ID", "Student Name"], dropna=False)["Student ID"].transform("size")
    ordered = work.sort_values(
        [
            "Student ID",
            "_name_count",
            "_term_sort",
            "_source_version_priority",
            "_source_month_priority",
            "_source_format_priority",
            "_source_file_index",
            "_source_row_index",
        ],
        ascending=[True, False, False, False, False, False, False, False],
        na_position="last",
    )
    result = ordered.drop_duplicates(subset=["Student ID"], keep="first").loc[:, STUDENT_NAME_COLUMNS].copy()
    return result.sort_values(["Student Name", "Student ID"], na_position="last").reset_index(drop=True)


def build_sql_compile_frame(roots: Sequence[str | Path]) -> Tuple[pd.DataFrame, pd.DataFrame, int]:
    source_rows, issues, source_file_count = load_sql_compile_rows(roots)
    return resolve_semester_statuses(source_rows), issues, source_file_count


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def write_sqlite(
    frame: pd.DataFrame,
    output_path: str | Path,
    table_name: str = TABLE_NAME,
    roster_inventory: Optional[pd.DataFrame] = None,
    roster_inventory_table_name: str = ROSTER_INVENTORY_TABLE,
    student_names: Optional[pd.DataFrame] = None,
    student_name_table_name: str = STUDENT_NAME_TABLE,
) -> Path:
    destination = _resolve_path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    table_identifier = _quote_identifier(table_name)
    column_identifiers = [_quote_identifier(column) for column in OUTPUT_COLUMNS]
    placeholders = ", ".join(["?"] * len(OUTPUT_COLUMNS))

    with sqlite3.connect(destination) as connection:
        connection.execute(f"DROP TABLE IF EXISTS {table_identifier}")
        connection.execute(
            f"CREATE TABLE {table_identifier} ("
            '"Semester" TEXT, '
            '"Chapter" TEXT, '
            '"Student ID" TEXT, '
            '"Status" TEXT'
            ")"
        )
        if not frame.empty:
            connection.executemany(
                f"INSERT INTO {table_identifier} ({', '.join(column_identifiers)}) VALUES ({placeholders})",
                frame.loc[:, OUTPUT_COLUMNS].itertuples(index=False, name=None),
            )
        connection.execute(
            f"CREATE INDEX IF NOT EXISTS {_quote_identifier(f'idx_{table_name}_student_semester')} "
            f"ON {table_identifier} ({_quote_identifier('Student ID')}, {_quote_identifier('Semester')})"
        )
        if roster_inventory is not None:
            inventory = roster_inventory.copy()
            for column in ROSTER_INVENTORY_COLUMNS:
                if column not in inventory.columns:
                    inventory[column] = ""
            inventory.loc[:, ROSTER_INVENTORY_COLUMNS].to_sql(
                roster_inventory_table_name,
                connection,
                if_exists="replace",
                index=False,
            )
            inventory_identifier = _quote_identifier(roster_inventory_table_name)
            connection.execute(
                f"CREATE INDEX IF NOT EXISTS {_quote_identifier(f'idx_{roster_inventory_table_name}_chapter_semester')} "
                f"ON {inventory_identifier} ({_quote_identifier('Chapter')}, {_quote_identifier('Semester')})"
            )
        names = student_names.copy() if student_names is not None else pd.DataFrame(columns=STUDENT_NAME_COLUMNS)
        for column in STUDENT_NAME_COLUMNS:
            if column not in names.columns:
                names[column] = ""
        names.loc[:, STUDENT_NAME_COLUMNS].to_sql(
            student_name_table_name,
            connection,
            if_exists="replace",
            index=False,
        )
        names_identifier = _quote_identifier(student_name_table_name)
        connection.execute(
            f"CREATE INDEX IF NOT EXISTS {_quote_identifier(f'idx_{student_name_table_name}_student')} "
            f"ON {names_identifier} ({_quote_identifier('Student ID')})"
        )
        connection.commit()

    return destination


def sqlCompile(
    input_roots: Optional[Sequence[str | Path]] = None,
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
    config_path: Optional[str | Path] = None,
    table_name: str = TABLE_NAME,
) -> SqlCompileResult:
    roots = list(input_roots) if input_roots else default_input_roots(config_path)
    source_rows, issues, source_file_count = load_sql_compile_rows(roots)
    frame = resolve_semester_statuses(source_rows)
    roster_inventory = build_roster_inventory(source_rows)
    student_names = build_student_name_lookup(source_rows)
    destination = write_sqlite(
        frame,
        output_path,
        table_name=table_name,
        roster_inventory=roster_inventory,
        student_names=student_names,
    )
    return SqlCompileResult(
        output_path=destination,
        table_name=table_name,
        row_count=len(frame),
        source_file_count=source_file_count,
        issue_count=len(issues),
        issues=issues,
    )


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compile Excel roster files into a SQLite table named sqlCompile."
    )
    parser.add_argument("--config", default=None, help="Path to config/local_paths.yaml. Defaults to the standard project path config lookup.")
    parser.add_argument(
        "--input-root",
        action="append",
        dest="input_roots",
        default=None,
        help="Folder of Excel roster files. Repeat for multiple roots. Defaults to configured roster roots.",
    )
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH), help="SQLite database path to write.")
    parser.add_argument("--table", default=TABLE_NAME, help="SQLite table name to replace.")
    parser.add_argument(
        "--cohort-semester",
        action="append",
        dest="cohort_semesters",
        default=None,
        help='After compiling, build a Status N cohort report for this semester, for example "Fall 2025". Repeat for multiple cohorts.',
    )
    parser.add_argument(
        "--all-new-member-cohorts",
        "--all-semesters",
        dest="all_new_member_cohorts",
        action="store_true",
        help="After compiling, build cohort reports for every semester with Status N rows.",
    )
    parser.add_argument("--manual-status-file", default=None, help="CSV of manually researched status rows for cohort reports.")
    parser.add_argument("--cohort-output-dir", default=None, help="Folder where cohort report CSVs are written.")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    result = sqlCompile(
        input_roots=args.input_roots,
        output_path=args.output,
        config_path=args.config,
        table_name=args.table,
    )
    print(f"SQL compile database written to: {result.output_path}")
    print(f"Table: {result.table_name}")
    print(f"Rows: {result.row_count}")
    print(f"Excel source files scanned: {result.source_file_count}")
    if result.issue_count:
        print(f"Sheets/files skipped with issues: {result.issue_count}")
    if args.cohort_semesters or args.all_new_member_cohorts:
        from src.sqlCompile_cohort import (
            DEFAULT_COHORT_OUTPUT_DIR,
            DEFAULT_MANUAL_STATUS_PATH,
            build_new_member_cohort_report,
        )

        report = build_new_member_cohort_report(
            database_path=result.output_path,
            cohort_semesters=args.cohort_semesters,
            all_cohorts=args.all_new_member_cohorts,
            manual_status_file=args.manual_status_file or DEFAULT_MANUAL_STATUS_PATH,
            output_dir=args.cohort_output_dir or DEFAULT_COHORT_OUTPUT_DIR,
            table_name=args.table,
        )
        print(f"New-member cohort report written to: {report.output_dir}")
        print(f"Manual status file: {report.manual_status_path}")
        print(f"Manual form review rows: {report.review_rows}")
        for warning in report.csv_warnings:
            print(f"WARNING: {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
