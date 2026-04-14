from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_ROOT = ROOT / "Copy of Rosters"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SEMESTER_FOLDER_RE = re.compile(r"^(Fall|Spring)\s+(20\d{2})$", re.IGNORECASE)

STANDARD_COLUMNS = [
    "Academic Year",
    "Term",
    "Source File",
    "Chapter",
    "Last Name",
    "First Name",
    "Banner ID",
    "Email",
    "Status",
    "Semester Joined",
    "Position",
]

UNIQUE_BANNER_COLUMNS = [
    "Banner ID",
    "Last Name",
    "First Name",
    "Email",
    "Initial Chapter",
    "Latest Chapter",
    "First Observed Academic Year",
    "First Observed Term",
    "Latest Observed Academic Year",
    "Latest Observed Term",
    "Latest Status",
    "Semester Joined",
    "Latest Position",
    "Terms Observed",
    "Chapters Seen",
    "Statuses Seen",
    "Source Files Seen",
]

HEADER_ALIASES = {
    "last_name": [
        "last name",
        "lastname",
        "surname",
        "member last name",
    ],
    "first_name": [
        "first name",
        "firstname",
        "given name",
        "member first name",
    ],
    "banner_id": [
        "banner id",
        "student id",
        "banner",
        "student number",
        "banner number",
        "z number",
    ],
    "email": [
        "email",
        "e-mail",
        "email address",
        "student email",
    ],
    "status": [
        "status",
        "member status",
        "membership status",
        "roster status",
    ],
    "semester_joined": [
        "semester joined",
        "joined",
        "join term",
        "semester initiated",
        "term joined",
        "semester admitted",
        "initiation term",
    ],
    "position": [
        "position",
        "office",
        "role",
        "member/council",
        "member council",
        "title",
    ],
    "chapter": [
        "chapter",
        "org name",
        "organization",
        "org",
        "group",
        "fraternity/sorority",
        "fsl organization",
    ],
}

CANONICAL_ALIAS_MAP = {
    standard_name: {re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]+", "", alias.lower().replace("_", " "))).strip() for alias in aliases}
    for standard_name, aliases in HEADER_ALIASES.items()
}

STATUS_MAP = {
    "A": "Active",
    "AL": "Alumni",
    "G": "Graduated",
    "I": "Inactive",
    "S": "Suspended",
    "N": "New Member",
    "RS": "Resigned",
    "RV": "Revoked",
    "T": "Transfer",
}

STATUS_PRIORITY = {
    "Graduated": 90,
    "Alumni": 85,
    "Suspended": 80,
    "Revoked": 75,
    "Resigned": 70,
    "Transfer": 65,
    "Inactive": 60,
    "New Member": 55,
    "Active": 50,
    "": 0,
}

GREEK_LETTER_WORDS = {
    "alpha",
    "beta",
    "gamma",
    "delta",
    "epsilon",
    "zeta",
    "eta",
    "theta",
    "iota",
    "kappa",
    "lambda",
    "mu",
    "nu",
    "xi",
    "omicron",
    "pi",
    "rho",
    "sigma",
    "tau",
    "upsilon",
    "phi",
    "chi",
    "psi",
    "omega",
}
ALLOWED_CHAPTER_PHRASES = {
    "order of omega": "Order of Omega",
    "kappa alpha": "Kappa Alpha Order",
}

CHAPTER_JUNK_PATTERNS = [
    r"never responded to email",
    r"greek leadership honor society",
    r"fraternity,\s*inc\.?",
    r"fraternity",
    r"sorority",
    r"roster revised 2",
    r"roster[_\s-]*update",
    r"revised",
    r"updated",
    r"update",
    r"final",
    r"roster",
    r"sept",
    r"nov",
    r"\bfall\s*20\d{2}\b",
    r"\b(19|20)\d{2}\b",
    r"\b2\b",
]


@dataclass(frozen=True)
class ExtractedRow:
    academic_year: str
    term: str
    source_file: str
    source_sheet: str
    chapter: str
    last_name: str
    first_name: str
    banner_id: str
    email: str
    status: str
    semester_joined: str
    position: str

    def as_list(self) -> List[str]:
        return [
            self.academic_year,
            self.term,
            self.source_file,
            self.chapter,
            self.last_name,
            self.first_name,
            self.banner_id,
            self.email,
            self.status,
            self.semester_joined,
            self.position,
        ]


@dataclass(frozen=True)
class FileExtractionStatus:
    academic_year: str
    term: str
    source_file: str
    relative_path: str
    rows_extracted: int
    issue_count: int

    @property
    def extracted_flag(self) -> str:
        return "Yes" if self.rows_extracted > 0 else "No"


@dataclass(frozen=True)
class UniqueBannerRow:
    banner_id: str
    last_name: str
    first_name: str
    email: str
    initial_chapter: str
    latest_chapter: str
    first_observed_academic_year: str
    first_observed_term: str
    latest_observed_academic_year: str
    latest_observed_term: str
    latest_status: str
    semester_joined: str
    latest_position: str
    terms_observed: int
    chapters_seen: str
    statuses_seen: str
    source_files_seen: str

    def as_list(self) -> List[object]:
        return [
            self.banner_id,
            self.last_name,
            self.first_name,
            self.email,
            self.initial_chapter,
            self.latest_chapter,
            self.first_observed_academic_year,
            self.first_observed_term,
            self.latest_observed_academic_year,
            self.latest_observed_term,
            self.latest_status,
            self.semester_joined,
            self.latest_position,
            self.terms_observed,
            self.chapters_seen,
            self.statuses_seen,
            self.source_files_seen,
        ]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def canonical_header(value: object) -> str:
    text = clean_text(value).lower()
    text = text.replace("_", " ")
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def header_matches(standard_name: str, header: str) -> bool:
    aliases = CANONICAL_ALIAS_MAP[standard_name]
    if header in aliases:
        return True

    if standard_name == "status":
        return header.startswith("status") or " status " in f" {header} "

    return any(alias in header for alias in aliases if len(alias) > 4)


def normalize_status(value: str) -> str:
    raw = clean_text(value)
    upper = raw.upper()
    if upper in STATUS_MAP:
        return STATUS_MAP[upper]
    return raw


def normalize_banner_id(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"\.0$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^A-Za-z0-9]", "", text).upper()
    if not text:
        return ""
    if re.fullmatch(r"A0\d{7}", text):
        return text
    if re.fullmatch(r"A\d{8}", text):
        return text
    if re.fullmatch(r"A\d{7}", text):
        return f"A0{text[1:]}"
    if re.fullmatch(r"\d{8}", text):
        return f"A{text}"
    if re.fullmatch(r"\d{7}", text):
        return f"A0{text}"
    return text


def detect_inline_chapter_label(row: Tuple[object, ...], header_map: Dict[str, int]) -> str:
    protected_fields = ["last_name", "first_name", "banner_id", "email", "status", "semester_joined", "position"]
    if any(get_cell(row, header_map.get(field)) for field in protected_fields):
        return ""

    non_empty = [clean_text(value) for value in row if clean_text(value)]
    if not non_empty or len(non_empty) > 2:
        return ""

    for value in non_empty:
        normalized = normalize_chapter_name(value)
        if normalized and normalized != "Unknown" and not is_excluded_chapter(normalized):
            return value
    return ""


def is_new_member_row(row: ExtractedRow) -> bool:
    status = clean_text(row.status).lower()
    position = clean_text(row.position).lower()
    return status == "new member" or "new member" in position


def is_active_member_row(row: ExtractedRow) -> bool:
    return clean_text(row.status).lower() == "active"


def companion_output_path(output_file: Path, suffix: str) -> Path:
    return output_file.with_name(f"{output_file.stem}_{suffix}{output_file.suffix}")


def unique_non_blank(values: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    ordered: List[str] = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(text)
    return ordered


def identity_key(row: ExtractedRow) -> Optional[Tuple[str, ...]]:
    if row.banner_id:
        return ("banner", row.banner_id.lower())
    if row.email:
        return ("email", row.email.lower())
    if row.last_name or row.first_name:
        return ("name", row.chapter.lower(), row.last_name.lower(), row.first_name.lower())
    return None


def row_priority(row: ExtractedRow) -> Tuple[int, int, int, int, str, str]:
    return (
        STATUS_PRIORITY.get(row.status, 10),
        1 if row.banner_id else 0,
        1 if row.email else 0,
        1 if row.semester_joined else 0,
        row.source_file.lower(),
        row.source_sheet.lower(),
    )


def parse_term_from_path(path: Path) -> Tuple[str, str]:
    for part in path.parts:
        match = SEMESTER_FOLDER_RE.fullmatch(part)
        if match:
            return match.group(2), f"{match.group(1).title()} {match.group(2)}"

    for candidate in [path.parent.name, path.stem]:
        match = SEMESTER_FOLDER_RE.search(candidate)
        if match:
            return match.group(2), f"{match.group(1).title()} {match.group(2)}"

    year_match = re.search(r"(20\d{2}|19\d{2})", path.stem)
    if year_match:
        return year_match.group(1), year_match.group(1)
    return "Unknown", "Unknown"


def term_sort_key(academic_year: str, term: str) -> Tuple[int, int, str]:
    year_value = 9999
    if re.fullmatch(r"(19|20)\d{2}", academic_year):
        year_value = int(academic_year)

    term_lower = clean_text(term).lower()
    if term_lower.startswith("spring"):
        season_value = 1
    elif term_lower.startswith("summer"):
        season_value = 2
    elif term_lower.startswith("fall"):
        season_value = 3
    elif term_lower.startswith("winter"):
        season_value = 0
    else:
        season_value = 9

    return year_value, season_value, term_lower


def is_placeholder_sheet_name(value: str) -> bool:
    normalized = re.sub(r"[\s_]+", "", clean_text(value)).lower()
    return normalized in {"sheet1", "sheet2", "sheet3"}


def normalize_chapter_name(value: str) -> str:
    cleaned = clean_text(value)
    if not cleaned:
        return ""

    for pattern in CHAPTER_JUNK_PATTERNS:
        cleaned = re.sub(pattern, " ", cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r"[_.,]+", " ", cleaned)
    cleaned = re.sub(r"[^A-Za-z()\-\s]+", " ", cleaned)

    lowered_cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
    for phrase, canonical in ALLOWED_CHAPTER_PHRASES.items():
        if phrase in lowered_cleaned:
            return canonical

    parts = re.findall(r"[A-Za-z]+|[()-]", cleaned)
    kept_parts: List[str] = []
    for part in parts:
        lower = part.lower()
        if part in {"(", ")", "-"}:
            kept_parts.append(part)
        elif lower in GREEK_LETTER_WORDS:
            kept_parts.append(lower.title())

    normalized = " ".join(kept_parts)
    normalized = re.sub(r"\s*-\s*", "-", normalized)
    normalized = re.sub(r"\(\s+", "(", normalized)
    normalized = re.sub(r"\s+\)", ")", normalized)
    normalized = re.sub(r"\(\)", "", normalized)
    normalized = normalized.replace("Alpha Kappa Alpha (Sigma Epsilon)", "Alpha Kappa Alpha")
    normalized = normalized.replace("Phi Kappa Tau-Gamma Psi", "Phi Kappa Tau")
    normalized = normalized.replace("Sigma Iota Alpha (Sigma Iota Alpha)", "Sigma Iota Alpha")
    normalized = re.sub(r"\s+", " ", normalized).strip(" -")
    return normalized or "Unknown"


def is_order_of_omega(chapter: str) -> bool:
    return normalize_chapter_name(chapter) == "Order of Omega"


def is_excluded_chapter(chapter: str) -> bool:
    normalized = normalize_chapter_name(chapter)
    return normalized in {"Order of Omega", "Epsilon Lambda Alpha"}


def chapter_from_filename(path: Path) -> str:
    stem = clean_text(path.stem)
    if not stem:
        return ""

    if stem.lower() == "raw roster data":
        return "Unknown"

    return normalize_chapter_name(stem)


def infer_chapter(path: Path, sheet_name: str) -> str:
    if is_placeholder_sheet_name(sheet_name):
        return chapter_from_filename(path)

    for candidate in [sheet_name, path.stem, path.parent.name]:
        cleaned = normalize_chapter_name(candidate)
        if not cleaned:
            continue
        if SEMESTER_FOLDER_RE.fullmatch(cleaned):
            continue
        if cleaned.lower() in {"copy of rosters", "rosters", "raw rosters", "master roster", "unknown"}:
            continue
        if re.fullmatch(r"(19|20)\d{2}", cleaned):
            continue
        return cleaned
    return ""


def score_header_row(values: List[object]) -> Tuple[int, Dict[str, int]]:
    matched: Dict[str, int] = {}
    canon = [canonical_header(value) for value in values]
    for idx, header in enumerate(canon):
        for standard_name in CANONICAL_ALIAS_MAP:
            if header_matches(standard_name, header) and standard_name not in matched:
                matched[standard_name] = idx
    return len(matched), matched


def find_status_column(ws, max_scan_rows: int = 30) -> Tuple[Optional[int], Optional[int]]:
    best_match: Tuple[int, Optional[int], Optional[int]] = (0, None, None)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True), start=1):
        for col_idx, value in enumerate(row):
            header = canonical_header(value)
            if not header:
                continue

            score = 0
            if header == "status":
                score = 3
            elif header.startswith("status"):
                score = 2
            elif " status " in f" {header} ":
                score = 1

            if score > best_match[0]:
                best_match = (score, row_idx, col_idx)

    return best_match[1], best_match[2]


def find_header_row(ws) -> Tuple[Optional[int], Dict[str, int]]:
    best_score = 0
    best_row_idx = None
    best_map: Dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), start=1):
        score, header_map = score_header_row(list(row))
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_map = header_map

    required = {"last_name", "first_name"}
    if best_row_idx is None or best_score < 3 or not required.issubset(best_map):
        return None, {}
    return best_row_idx, best_map


def get_cell(row: Tuple[object, ...], index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def row_is_empty(values: Iterable[str]) -> bool:
    return all(not clean_text(value) for value in values)


def extract_rows_from_workbook(path: Path, verbose: bool = False) -> Tuple[List[ExtractedRow], List[str]]:
    rows: List[ExtractedRow] = []
    issues: List[str] = []

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        issues.append(f"FAILED to open {path}: {exc}")
        return rows, issues

    try:
        academic_year, term = parse_term_from_path(path)

        for ws in wb.worksheets:
            header_row_idx, header_map = find_header_row(ws)
            if header_row_idx is None:
                issues.append(f"Skipped {path.name} | sheet '{ws.title}': no usable header row found.")
                continue

            status_row_idx, status_col_idx = find_status_column(ws)
            if "status" not in header_map and status_col_idx is not None:
                header_map["status"] = status_col_idx
            data_start_row = max(header_row_idx, status_row_idx or header_row_idx) + 1
            if "status" not in header_map:
                issues.append(f"Status column not found after full scan in {path.name} | sheet '{ws.title}'.")

            default_chapter = infer_chapter(path, ws.title)
            current_chapter_raw = ws.title
            current_chapter = default_chapter

            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                inline_chapter_raw = detect_inline_chapter_label(row, header_map)
                if inline_chapter_raw:
                    current_chapter_raw = inline_chapter_raw
                    current_chapter = normalize_chapter_name(inline_chapter_raw) or default_chapter
                    continue

                last_name = get_cell(row, header_map.get("last_name"))
                first_name = get_cell(row, header_map.get("first_name"))
                banner_id = normalize_banner_id(get_cell(row, header_map.get("banner_id")))
                email = get_cell(row, header_map.get("email")).lower()
                status = normalize_status(get_cell(row, header_map.get("status")))
                semester_joined = get_cell(row, header_map.get("semester_joined"))
                position = get_cell(row, header_map.get("position"))
                chapter_raw = get_cell(row, header_map.get("chapter")) or current_chapter_raw or ws.title
                chapter = normalize_chapter_name(chapter_raw) or current_chapter or default_chapter

                core_values = [last_name, first_name, banner_id, email, status, semester_joined, position, chapter]
                if row_is_empty(core_values):
                    continue

                if not last_name and not first_name:
                    continue

                rows.append(
                    ExtractedRow(
                        academic_year=academic_year,
                        term=term,
                        source_file=path.name,
                        source_sheet=ws.title,
                        chapter=chapter,
                        last_name=last_name,
                        first_name=first_name,
                        banner_id=banner_id,
                        email=email,
                        status=status,
                        semester_joined=semester_joined,
                        position=position,
                    )
                )
    finally:
        wb.close()

    if verbose:
        print(f"Processed {path}")
    return rows, issues


def dedupe_rows(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    best_rows: Dict[Tuple[str, ...], ExtractedRow] = {}

    for row in rows:
        if row.banner_id:
            key = ("banner", row.academic_year.lower(), row.term.lower(), row.banner_id.lower(), row.chapter.lower())
        elif row.email:
            key = ("email", row.academic_year.lower(), row.term.lower(), row.email.lower(), row.chapter.lower())
        else:
            key = (
                "fallback",
                row.academic_year.lower(),
                row.term.lower(),
                row.chapter.lower(),
                row.last_name.lower(),
                row.first_name.lower(),
                row.semester_joined.lower(),
            )

        existing = best_rows.get(key)
        if existing is None or row_priority(row) > row_priority(existing):
            best_rows[key] = row

    deduped = sorted(
        best_rows.values(),
        key=lambda item: (
            item.academic_year.lower(),
            item.term.lower(),
            item.chapter.lower(),
            item.banner_id.lower() if item.banner_id else "",
            item.last_name.lower(),
            item.first_name.lower(),
        ),
    )
    return deduped, len(rows) - len(deduped)


def dedupe_same_year_banner_ids(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    best_rows: Dict[Tuple[str, str, str], ExtractedRow] = {}

    for row in rows:
        if row.banner_id:
            key = (row.academic_year.lower(), row.term.lower(), row.banner_id.lower())
            existing = best_rows.get(key)
            if existing is None or row_priority(row) > row_priority(existing):
                best_rows[key] = row
            continue
        key = ("no-banner", row.academic_year.lower(), row.term.lower(), row.chapter.lower(), row.last_name.lower(), row.first_name.lower())
        best_rows[key] = row

    deduped = sorted(
        best_rows.values(),
        key=lambda item: (
            item.academic_year.lower(),
            item.term.lower(),
            item.chapter.lower(),
            item.banner_id.lower() if item.banner_id else "",
            item.last_name.lower(),
            item.first_name.lower(),
        ),
    )

    return deduped, len(rows) - len(deduped)


def infer_missing_spring_members(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    fall_rows_by_year_chapter: Dict[Tuple[int, str], List[ExtractedRow]] = defaultdict(list)
    spring_keys_by_year_chapter: Dict[Tuple[int, str], Set[Tuple[str, ...]]] = defaultdict(set)

    for row in rows:
        term_lower = row.term.lower()
        chapter_key = row.chapter.lower()
        key = identity_key(row)
        if term_lower.startswith("fall") and re.fullmatch(r"(19|20)\d{2}", row.academic_year):
            fall_rows_by_year_chapter[(int(row.academic_year), chapter_key)].append(row)
        elif term_lower.startswith("spring") and re.fullmatch(r"(19|20)\d{2}", row.academic_year) and key is not None:
            spring_keys_by_year_chapter[(int(row.academic_year), chapter_key)].add(key)

    inferred_rows: List[ExtractedRow] = []

    fall_years = sorted({year for year, _ in fall_rows_by_year_chapter.keys()})
    for year in fall_years:
        next_year = year + 1
        current_chapters = {chapter for fall_year, chapter in fall_rows_by_year_chapter.keys() if fall_year == year}

        for chapter_key in current_chapters:
            current_fall = fall_rows_by_year_chapter.get((year, chapter_key), [])
            next_fall = fall_rows_by_year_chapter.get((next_year, chapter_key), [])
            if not current_fall or not next_fall:
                continue

            current_keys = {key for key in (identity_key(row) for row in current_fall) if key is not None}
            existing_spring_keys = spring_keys_by_year_chapter[(next_year, chapter_key)]

            for row in next_fall:
                key = identity_key(row)
                if key is None:
                    continue
                if row.status.strip().lower() == "new member":
                    continue
                if key in current_keys:
                    continue
                if key in existing_spring_keys:
                    continue

                spring_term = f"Spring {next_year}"
                inferred_rows.append(
                    ExtractedRow(
                        academic_year=str(next_year),
                        term=spring_term,
                        source_file=row.source_file,
                        source_sheet=f"{row.source_sheet} [Inferred Spring]",
                        chapter=row.chapter,
                        last_name=row.last_name,
                        first_name=row.first_name,
                        banner_id=row.banner_id,
                        email=row.email,
                        status="New Member",
                        semester_joined=spring_term,
                        position=row.position,
                    )
                )
                existing_spring_keys.add(key)

    return rows + inferred_rows, len(inferred_rows)


def remove_order_of_omega_rows(rows: List[ExtractedRow]) -> Tuple[List[ExtractedRow], int]:
    filtered = [row for row in rows if not is_excluded_chapter(row.chapter)]
    return filtered, len(rows) - len(filtered)


def autosize_columns(ws) -> None:
    max_widths = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            width = len(clean_text(value))
            if width > max_widths[idx]:
                max_widths[idx] = width
    for idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 32)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def write_summary_sheet(
    wb: Workbook,
    rows: List[ExtractedRow],
    issues: List[str],
    file_statuses: List[FileExtractionStatus],
    total_files: int,
    duplicates_removed: int,
    same_year_id_removed: int,
    inferred_spring_members: int,
    order_of_omega_removed: int,
) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)

    by_year = defaultdict(int)
    by_term = defaultdict(int)
    with_banner = 0
    missing_banner = 0
    chapters = set()
    new_member_rows = 0
    active_rows = 0

    for row in rows:
        by_year[row.academic_year] += 1
        by_term[row.term] += 1
        if row.banner_id:
            with_banner += 1
        else:
            missing_banner += 1
        if is_new_member_row(row):
            new_member_rows += 1
        if is_active_member_row(row):
            active_rows += 1
        if row.chapter:
            chapters.add(row.chapter)

    metrics = [
        ["Input files processed", total_files],
        ["Total extracted rows", len(rows)],
        ["Rows with Banner ID", with_banner],
        ["Rows missing Banner ID", missing_banner],
        ["Rows classified as New Member", new_member_rows],
        ["Rows classified as Active", active_rows],
        ["Distinct academic years", len(by_year)],
        ["Distinct chapters", len(chapters)],
        ["Duplicate rows removed", duplicates_removed],
        ["Same-semester duplicate Banner IDs removed", same_year_id_removed],
        ["Inferred spring members added", inferred_spring_members],
        ["Order of Omega rows removed", order_of_omega_removed],
    ]
    for item in metrics:
        ws.append(item)

    ws.append([])
    ws.append(["Academic Year", "Row Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for academic_year in sorted(by_year.keys()):
        ws.append([academic_year, by_year[academic_year]])

    ws.append([])
    ws.append(["Term", "Row Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for term in sorted(by_term.keys()):
        ws.append([term, by_term[term]])

    ws.append([])
    ws.append(["Chapter List"])
    ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Chapter"])
    ws[ws.max_row][0].fill = PatternFill("solid", fgColor="D9EAF7")
    ws[ws.max_row][0].font = Font(bold=True)
    for chapter in sorted(chapters):
        ws.append([chapter])

    ws.append([])
    ws.append(["Import Issues"])
    ws[ws.max_row][0].font = Font(bold=True)
    if issues:
        for issue in issues:
            ws.append([issue])
    else:
        ws.append(["None"])

    ws.append([])
    ws.append(["File Extraction Check"])
    ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Academic Year", "Term", "Source File", "Relative Path", "Rows Extracted", "Extracted Data", "Issue Count"])
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)

    for status in sorted(file_statuses, key=lambda item: (term_sort_key(item.academic_year, item.term), item.relative_path.lower())):
        ws.append(
            [
                status.academic_year,
                status.term,
                status.source_file,
                status.relative_path,
                status.rows_extracted,
                status.extracted_flag,
                status.issue_count,
            ]
        )

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_year_sheets(wb: Workbook, rows: List[ExtractedRow], chunk_size: int = 1000) -> None:
    grouped: Dict[Tuple[str, str], List[ExtractedRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.academic_year, row.term)].append(row)

    for academic_year, term in sorted(grouped.keys(), key=lambda item: term_sort_key(item[0], item[1])):
        semester_rows = sorted(
            grouped[(academic_year, term)],
            key=lambda item: (
                item.chapter.lower(),
                item.last_name.lower(),
                item.first_name.lower(),
                item.banner_id.lower() if item.banner_id else "zzzzzzzz",
                item.source_file.lower(),
            ),
        )

        term_label = re.sub(r"[^A-Za-z0-9]+", "_", clean_text(term)).strip("_") or "Unknown"
        for start in range(0, len(semester_rows), chunk_size):
            end = min(start + chunk_size, len(semester_rows))
            chunk_number = (start // chunk_size) + 1
            sheet_name = f"{term_label}_{chunk_number:02d}"
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(STANDARD_COLUMNS)
            style_header(ws)
            for row in semester_rows[start:end]:
                ws.append(row.as_list())
            ws.freeze_panes = "A2"
            autosize_columns(ws)


def write_roster_workbook(
    output_file: Path,
    rows: List[ExtractedRow],
    issues: List[str],
    file_statuses: List[FileExtractionStatus],
    total_files: int,
    duplicates_removed: int,
    same_year_id_removed: int,
    inferred_spring_members: int,
    order_of_omega_removed: int,
    chunk_size: int,
) -> None:
    wb = Workbook()
    write_summary_sheet(
        wb,
        rows,
        issues,
        file_statuses,
        total_files=total_files,
        duplicates_removed=duplicates_removed,
        same_year_id_removed=same_year_id_removed,
        inferred_spring_members=inferred_spring_members,
        order_of_omega_removed=order_of_omega_removed,
    )
    write_year_sheets(wb, rows, chunk_size=chunk_size)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def build_unique_banner_rows(rows: List[ExtractedRow]) -> List[UniqueBannerRow]:
    grouped: Dict[str, List[ExtractedRow]] = defaultdict(list)
    for row in rows:
        if not row.banner_id:
            continue
        grouped[row.banner_id.lower()].append(row)

    unique_rows: List[UniqueBannerRow] = []
    for _, banner_rows in grouped.items():
        ordered_rows = sorted(
            banner_rows,
            key=lambda item: (
                term_sort_key(item.academic_year, item.term),
                item.chapter.lower(),
                item.last_name.lower(),
                item.first_name.lower(),
                item.source_file.lower(),
            ),
        )
        first_row = ordered_rows[0]
        latest_row = max(
            ordered_rows,
            key=lambda item: (
                term_sort_key(item.academic_year, item.term),
                STATUS_PRIORITY.get(item.status, 10),
                1 if item.position else 0,
                item.source_file.lower(),
            ),
        )

        last_name = next((row.last_name for row in reversed(ordered_rows) if row.last_name), first_row.last_name)
        first_name = next((row.first_name for row in reversed(ordered_rows) if row.first_name), first_row.first_name)
        email = next((row.email for row in reversed(ordered_rows) if row.email), first_row.email)
        semester_joined = next((row.semester_joined for row in ordered_rows if row.semester_joined), "")
        latest_position = next((row.position for row in reversed(ordered_rows) if row.position), latest_row.position)

        chapters_seen = unique_non_blank(row.chapter for row in ordered_rows)
        statuses_seen = unique_non_blank(row.status for row in ordered_rows)
        source_files_seen = unique_non_blank(row.source_file for row in ordered_rows)
        distinct_terms = {
            (clean_text(row.academic_year).lower(), clean_text(row.term).lower())
            for row in ordered_rows
            if row.academic_year or row.term
        }

        unique_rows.append(
            UniqueBannerRow(
                banner_id=first_row.banner_id,
                last_name=last_name,
                first_name=first_name,
                email=email,
                initial_chapter=first_row.chapter,
                latest_chapter=latest_row.chapter,
                first_observed_academic_year=first_row.academic_year,
                first_observed_term=first_row.term,
                latest_observed_academic_year=latest_row.academic_year,
                latest_observed_term=latest_row.term,
                latest_status=latest_row.status,
                semester_joined=semester_joined,
                latest_position=latest_position,
                terms_observed=len(distinct_terms),
                chapters_seen=" | ".join(chapters_seen),
                statuses_seen=" | ".join(statuses_seen),
                source_files_seen=" | ".join(source_files_seen),
            )
        )

    return sorted(
        unique_rows,
        key=lambda item: (
            item.banner_id.lower(),
            item.last_name.lower(),
            item.first_name.lower(),
        ),
    )


def write_unique_banner_workbook(output_file: Path, rows: List[UniqueBannerRow], chunk_size: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)

    initial_chapters = {row.initial_chapter for row in rows if row.initial_chapter}
    latest_chapters = {row.latest_chapter for row in rows if row.latest_chapter}
    metrics = [
        ["Distinct Banner IDs", len(rows)],
        ["Rows with Email", sum(1 for row in rows if row.email)],
        ["Distinct initial chapters", len(initial_chapters)],
        ["Distinct latest chapters", len(latest_chapters)],
        ["Workbook purpose", "One row per Banner ID for one-time academic record pulls"],
    ]
    for metric in metrics:
        ws.append(metric)
    ws.freeze_panes = "A2"
    autosize_columns(ws)

    for start in range(0, len(rows), chunk_size):
        end = min(start + chunk_size, len(rows))
        chunk_number = (start // chunk_size) + 1
        chunk_ws = wb.create_sheet(title=f"Banner_IDs_{chunk_number:02d}"[:31])
        chunk_ws.append(UNIQUE_BANNER_COLUMNS)
        style_header(chunk_ws)
        for row in rows[start:end]:
            chunk_ws.append(row.as_list())
        chunk_ws.freeze_panes = "A2"
        autosize_columns(chunk_ws)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def build_master_roster(
    input_root: Path,
    output_file: Path,
    chunk_size: int,
    keep_duplicates: bool,
    verbose: bool,
) -> Dict[str, Path]:
    all_rows: List[ExtractedRow] = []
    issues: List[str] = []
    file_statuses: List[FileExtractionStatus] = []

    files = sorted(path for path in input_root.rglob("*") if path.suffix.lower() in SUPPORTED_EXTENSIONS)
    if not files:
        raise FileNotFoundError(
            f"No Excel files found under {input_root}. Supported types: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    for path in files:
        extracted, file_issues = extract_rows_from_workbook(path, verbose=verbose)
        academic_year, term = parse_term_from_path(path)
        all_rows.extend(extracted)
        issues.extend(file_issues)
        file_statuses.append(
            FileExtractionStatus(
                academic_year=academic_year,
                term=term,
                source_file=path.name,
                relative_path=str(path.relative_to(input_root)),
                rows_extracted=len(extracted),
                issue_count=len(file_issues),
            )
        )

    duplicates_removed = 0
    if not keep_duplicates:
        all_rows, duplicates_removed = dedupe_rows(all_rows)

    all_rows, inferred_spring_members = infer_missing_spring_members(all_rows)
    if not keep_duplicates:
        all_rows, _ = dedupe_rows(all_rows)

    all_rows, order_of_omega_removed = remove_order_of_omega_rows(all_rows)
    all_rows, same_year_id_removed = dedupe_same_year_banner_ids(all_rows)

    write_roster_workbook(
        output_file=output_file,
        rows=all_rows,
        issues=issues,
        file_statuses=file_statuses,
        total_files=len(files),
        duplicates_removed=duplicates_removed,
        same_year_id_removed=same_year_id_removed,
        inferred_spring_members=inferred_spring_members,
        order_of_omega_removed=order_of_omega_removed,
        chunk_size=chunk_size,
    )

    new_member_output = companion_output_path(output_file, "New_Members")
    active_output = companion_output_path(output_file, "Active_Members")
    unique_banner_output = companion_output_path(output_file, "Unique_Banner_IDs")

    new_member_rows = [row for row in all_rows if is_new_member_row(row)]
    active_rows = [row for row in all_rows if is_active_member_row(row)]
    unique_banner_rows = build_unique_banner_rows(all_rows)

    write_roster_workbook(
        output_file=new_member_output,
        rows=new_member_rows,
        issues=issues,
        file_statuses=file_statuses,
        total_files=len(files),
        duplicates_removed=duplicates_removed,
        same_year_id_removed=same_year_id_removed,
        inferred_spring_members=inferred_spring_members,
        order_of_omega_removed=order_of_omega_removed,
        chunk_size=chunk_size,
    )
    write_roster_workbook(
        output_file=active_output,
        rows=active_rows,
        issues=issues,
        file_statuses=file_statuses,
        total_files=len(files),
        duplicates_removed=duplicates_removed,
        same_year_id_removed=same_year_id_removed,
        inferred_spring_members=inferred_spring_members,
        order_of_omega_removed=order_of_omega_removed,
        chunk_size=chunk_size,
    )
    write_unique_banner_workbook(
        output_file=unique_banner_output,
        rows=unique_banner_rows,
        chunk_size=chunk_size,
    )

    return {
        "master": output_file,
        "new_members": new_member_output,
        "active_members": active_output,
        "unique_banner_ids": unique_banner_output,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a single FSL master roster workbook from semester folders of chapter rosters. "
            "If no input path is supplied, the script uses a local 'Copy of Rosters' folder next to the code."
        )
    )
    parser.add_argument(
        "input_root",
        nargs="?",
        default=str(DEFAULT_INPUT_ROOT),
        help="Root folder containing semester folders like 'Fall 2015' and 'Spring 2026'.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="Master_FSL_Roster.xlsx",
        help="Output workbook path. Default: Master_FSL_Roster.xlsx",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=1000,
        help="Number of rows per year sheet. Default: 1000",
    )
    parser.add_argument(
        "--keep-duplicates",
        action="store_true",
        help="Keep cross-file duplicate rows. The same-year duplicate Banner ID pass still runs.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print each workbook as it is processed.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_root = Path(args.input_root).expanduser().resolve()
    output_file = Path(args.output).expanduser().resolve()

    outputs = build_master_roster(
        input_root=input_root,
        output_file=output_file,
        chunk_size=args.chunk_size,
        keep_duplicates=args.keep_duplicates,
        verbose=args.verbose,
    )
    print(f"Master roster created: {outputs['master']}")
    print(f"New-member roster created: {outputs['new_members']}")
    print(f"Active-member roster created: {outputs['active_members']}")
    print(f"Unique Banner ID roster created: {outputs['unique_banner_ids']}")


if __name__ == "__main__":
    main()
