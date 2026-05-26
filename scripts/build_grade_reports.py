from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.config_loader import load_chapter_mapping
from app.io_utils import parse_term_label, safe_slug
from src.path_config import load_path_config


DEFAULT_OUTPUT_DIR = ROOT / "data" / "outgoing" / "grade_reports"
MAROON = "5B0F14"
GOLD = "F4D58D"
LIGHT_BLUE = "BDD7EE"
LIGHT_FILL = "F7F3F0"
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color=MAROON)
SUBTITLE_FONT = Font(bold=True, size=12, italic=True, color=MAROON)
THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)


COMMUNITY_COLUMNS = [
    "Council",
    "Organization Type",
    "Chapter",
    "New Member GPA",
    "New Members",
    "Initiated Member GPA",
    "Initiated Members",
    "Overall Chapter GPA",
    "Total Members",
    "Previous Term Change",
]

CHAPTER_MEMBER_COLUMNS = [
    "Last Name",
    "First Name",
    "Status",
    "Major",
    "Term Hours",
    "Term GPA",
    "TXST GPA",
]


@dataclass(frozen=True)
class ReportBuildResult:
    output_dir: Path
    community_workbook: Path
    chapter_workbooks: List[Path]
    community_summary_csv: Path


def _read_table(folder: Path, table_name: str) -> pd.DataFrame:
    parquet = folder / f"{table_name}.parquet"
    csv = folder / f"{table_name}.csv"
    if parquet.exists():
        return pd.read_parquet(parquet)
    if csv.exists():
        return pd.read_csv(csv)
    return pd.DataFrame()


def _canonical_latest_dir(config_path: Optional[str] = None, canonical_dir: Optional[str | Path] = None) -> Path:
    if canonical_dir:
        return Path(canonical_dir).expanduser().resolve()
    try:
        root = load_path_config(config_path).output_root
    except Exception:
        root = ROOT / "output" / "canonical"
    latest = root / "latest"
    if latest.exists():
        return latest.resolve()
    runs = sorted([path for path in root.glob("run_*") if path.is_dir()], key=lambda path: path.stat().st_mtime, reverse=True)
    return runs[0].resolve() if runs else latest.resolve()


def _term_code_from_label(value: object) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"(19\d{2}|20\d{2})(WI|SP|SU|FA)", text, re.IGNORECASE):
        return text.upper()
    return str(parse_term_label(text).get("code") or "").upper()


def _term_label_from_code(term_code: str) -> str:
    match = re.fullmatch(r"(19\d{2}|20\d{2})(WI|SP|SU|FA)", str(term_code or "").upper())
    if not match:
        return str(term_code or "")
    year, season = match.groups()
    return {"WI": "Winter", "SP": "Spring", "SU": "Summer", "FA": "Fall"}[season] + f" {year}"


def _compact_term(term_code: str) -> str:
    match = re.fullmatch(r"(19\d{2}|20\d{2})(WI|SP|SU|FA)", str(term_code or "").upper())
    if not match:
        return ""
    year, season = match.groups()
    return {"WI": "W", "SP": "S", "SU": "U", "FA": "F"}[season] + year[-2:]


def _previous_term_code(term_code: str) -> str:
    match = re.fullmatch(r"(19\d{2}|20\d{2})(WI|SP|SU|FA)", str(term_code or "").upper())
    if not match:
        return ""
    year = int(match.group(1))
    season = match.group(2)
    if season == "SP":
        return f"{year - 1}FA"
    if season == "FA":
        return f"{year}SP"
    if season == "SU":
        return f"{year}SP"
    return f"{year - 1}FA"


def _latest_term_code(frame: pd.DataFrame) -> str:
    if frame.empty or "term_code" not in frame.columns:
        return ""
    terms = sorted({str(value).upper() for value in frame["term_code"].dropna() if _term_code_from_label(value)})
    if not terms:
        return ""
    order = {"WI": 0, "SP": 1, "SU": 2, "FA": 3}
    return sorted(terms, key=lambda code: (int(code[:4]), order.get(code[-2:], 9)))[-1]


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _yes_mask(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.lower().isin({"yes", "true", "1", "y"})


def _term_frame(frame: pd.DataFrame, term_code: str) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    if term_code and "term_code" in frame.columns:
        return frame.loc[frame["term_code"].fillna("").astype(str).str.upper().eq(term_code)].copy()
    return frame.copy()


def _ensure_columns(frame: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    result = frame.copy()
    for column in columns:
        if column not in result.columns:
            result[column] = ""
    return result


def _student_key(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.upper()


def _dedupe_academic_for_report(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or "student_id" not in frame.columns:
        return pd.DataFrame()
    working = frame.copy()
    working["_student_key"] = _student_key(working["student_id"])
    working = working.loc[working["_student_key"].ne("")]
    if working.empty:
        return working
    working["_has_gpa"] = _numeric(working.get("term_gpa", pd.Series("", index=working.index))).notna().astype(int)
    working["_has_hours"] = _numeric(working.get("attempted_hours_term", pd.Series("", index=working.index))).notna().astype(int)
    working = working.sort_values(["_student_key", "_has_gpa", "_has_hours"], ascending=[True, False, False])
    return working.drop_duplicates(subset=["_student_key"], keep="first").drop(columns=["_has_gpa", "_has_hours"], errors="ignore")


def _status_group(row: pd.Series) -> str:
    text = " ".join(
        str(row.get(column, "") or "")
        for column in ["org_status_bucket", "org_status_raw", "academic_status_raw"]
    ).lower()
    is_new = False
    if "new_member_flag" in row.index:
        is_new = str(row.get("new_member_flag", "") or "").strip().lower() in {"yes", "true", "1", "new member"}
    if is_new or "new member" in text or re.search(r"\bnew\b", text):
        return "New Member"
    if any(token in text for token in ["inactive", "resigned", "revoked", "suspended", "transfer"]):
        return "Inactive Member"
    return "Active Member"


def _roster_logi_report_frame(roster: pd.DataFrame, academic: pd.DataFrame, term_code: str) -> pd.DataFrame:
    roster_term = _term_frame(roster, term_code)
    academic_term = _term_frame(academic, term_code)
    if roster_term.empty or academic_term.empty or "student_id" not in roster_term.columns or "student_id" not in academic_term.columns:
        return pd.DataFrame()

    roster_columns = [
        "student_id",
        "first_name",
        "last_name",
        "chapter",
        "term_code",
        "term_label",
        "term_year",
        "term_season",
        "org_status_bucket",
        "org_status_raw",
        "new_member_flag",
    ]
    academic_columns = [
        "student_id",
        "first_name",
        "last_name",
        "email",
        "major",
        "term_code",
        "term_label",
        "term_year",
        "term_season",
        "academic_status_raw",
        "term_gpa",
        "institutional_cumulative_gpa",
        "overall_cumulative_gpa",
        "attempted_hours_term",
        "source_file",
        "source_sheet",
    ]
    roster_term = _ensure_columns(roster_term, roster_columns)
    academic_term = _ensure_columns(academic_term, academic_columns)
    roster_term = roster_term[roster_columns].copy()
    academic_term = _dedupe_academic_for_report(academic_term[academic_columns].copy())
    if academic_term.empty:
        return pd.DataFrame()

    roster_term["_student_key"] = _student_key(roster_term["student_id"])
    roster_term = roster_term.loc[roster_term["_student_key"].ne("")]
    if roster_term.empty:
        return pd.DataFrame()
    roster_term = roster_term.drop_duplicates(subset=["_student_key", "chapter"], keep="first")

    merged = roster_term.merge(academic_term, on="_student_key", how="left", suffixes=("_roster", "_academic"))
    result = pd.DataFrame(index=merged.index)
    result["student_id"] = merged["student_id_roster"].where(merged["student_id_roster"].fillna("").astype(str).str.strip().ne(""), merged["student_id_academic"])
    for column in ["first_name", "last_name", "term_code", "term_label", "term_year", "term_season"]:
        roster_col = f"{column}_roster"
        academic_col = f"{column}_academic"
        result[column] = merged[academic_col].where(merged[academic_col].fillna("").astype(str).str.strip().ne(""), merged[roster_col])
    for column in [
        "email",
        "major",
        "academic_status_raw",
        "term_gpa",
        "institutional_cumulative_gpa",
        "overall_cumulative_gpa",
        "attempted_hours_term",
        "source_file",
        "source_sheet",
    ]:
        result[column] = merged.get(column, "")
    for column in ["chapter", "org_status_bucket", "org_status_raw", "new_member_flag"]:
        result[column] = merged[column]
    return result


def _legacy_grade_source(canonical_dir: Path, term_code: str, master: pd.DataFrame, academic: pd.DataFrame, roster: pd.DataFrame) -> pd.DataFrame:

    if term_code and not master.empty and "term_code" in master.columns:
        frame = master.loc[master["term_code"].fillna("").astype(str).str.upper().eq(term_code)].copy()
    else:
        frame = master.copy()

    if frame.empty and not academic.empty:
        frame = academic.loc[academic.get("term_code", pd.Series("", index=academic.index)).fillna("").astype(str).str.upper().eq(term_code)].copy()
        if not roster.empty and "student_id" in frame.columns and "student_id" in roster.columns:
            roster_term = roster.loc[roster.get("term_code", pd.Series("", index=roster.index)).fillna("").astype(str).str.upper().eq(term_code)].copy()
            roster_cols = [column for column in ["student_id", "chapter", "org_status_bucket", "org_status_raw", "new_member_flag"] if column in roster_term.columns]
            frame = frame.merge(roster_term[roster_cols].drop_duplicates(subset=["student_id"]), on="student_id", how="left", suffixes=("", "_roster"))
            for column in ["chapter", "org_status_bucket", "org_status_raw", "new_member_flag"]:
                roster_col = f"{column}_roster"
                if roster_col in frame.columns:
                    if column not in frame.columns:
                        frame[column] = frame[roster_col]
                    else:
                        frame[column] = frame[column].where(frame[column].fillna("").astype(str).str.strip().ne(""), frame[roster_col])
    return frame


def _load_grade_source(canonical_dir: Path, term_code: str) -> pd.DataFrame:
    master = _read_table(canonical_dir, "master_longitudinal")
    academic = _read_table(canonical_dir, "academic_term")
    roster = _read_table(canonical_dir, "roster_term")

    frame = _roster_logi_report_frame(roster, academic, term_code)
    if frame.empty:
        frame = _legacy_grade_source(canonical_dir, term_code, master, academic, roster)

    required = [
        "student_id",
        "first_name",
        "last_name",
        "chapter",
        "major",
        "term_gpa",
        "institutional_cumulative_gpa",
        "overall_cumulative_gpa",
        "attempted_hours_term",
        "org_status_bucket",
        "org_status_raw",
        "academic_status_raw",
        "new_member_flag",
    ]
    frame = _ensure_columns(frame, required)
    if not frame.empty:
        frame["status_group"] = frame.apply(_status_group, axis=1)
        frame["term_gpa_num"] = _numeric(frame["term_gpa"])
        frame["txst_gpa_num"] = _numeric(frame["institutional_cumulative_gpa"]).where(
            _numeric(frame["institutional_cumulative_gpa"]).notna(),
            _numeric(frame["overall_cumulative_gpa"]),
        )
        frame["hours_num"] = _numeric(frame["attempted_hours_term"])
    return frame


def _apply_chapter_mapping(frame: pd.DataFrame, chapter_mapping: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in ["council", "org_type", "family", "chapter_group"]:
        if column not in result.columns:
            result[column] = ""
    if chapter_mapping.empty or "chapter" not in result.columns:
        return result
    mapping = chapter_mapping.drop_duplicates(subset=["chapter"]).copy()
    merged = result.merge(mapping, on="chapter", how="left", suffixes=("", "_mapped"))
    for column in ["council", "org_type", "family", "chapter_group"]:
        mapped = f"{column}_mapped"
        if mapped in merged.columns:
            merged[column] = merged[column].where(merged[column].fillna("").astype(str).str.strip().ne(""), merged[mapped])
            merged = merged.drop(columns=[mapped])
    return merged


def _mean_or_na(values: pd.Series) -> object:
    numeric = _numeric(values).dropna()
    return round(float(numeric.mean()), 2) if not numeric.empty else "N/A"


def _count_gpa(values: pd.Series) -> object:
    count = int(_numeric(values).notna().sum())
    return count if count else "N/A"


def _first_nonblank(group: pd.DataFrame, column: str, default: str) -> str:
    if column not in group.columns:
        return default
    values = group[column].dropna().astype(str).str.strip()
    values = values.loc[values.ne("")]
    return values.iloc[0] if not values.empty else default


def _chapter_community_row(chapter: str, group: pd.DataFrame, previous: pd.DataFrame, council: str, org_type: str) -> dict[str, object]:
    gpa_rows = group.loc[group["term_gpa_num"].notna()].copy()
    new_rows = gpa_rows.loc[gpa_rows["status_group"].eq("New Member")]
    active_rows = gpa_rows.loc[gpa_rows["status_group"].eq("Active Member")]
    overall_gpa = _mean_or_na(gpa_rows["term_gpa_num"])
    previous_gpa = "N/A"
    if not previous.empty:
        previous_group = previous.loc[previous["chapter"].fillna("").astype(str).eq(chapter)]
        if not previous_group.empty:
            previous_gpa = _mean_or_na(previous_group["term_gpa_num"])
    change = "N/A"
    if isinstance(overall_gpa, (int, float)) and isinstance(previous_gpa, (int, float)):
        change = round(overall_gpa - previous_gpa, 2)
    return {
        "Council": council or "Unknown",
        "Organization Type": org_type or "Organization",
        "Chapter": chapter,
        "New Member GPA": _mean_or_na(new_rows["term_gpa_num"]),
        "New Members": _count_gpa(new_rows["term_gpa_num"]),
        "Initiated Member GPA": _mean_or_na(active_rows["term_gpa_num"]),
        "Initiated Members": _count_gpa(active_rows["term_gpa_num"]),
        "Overall Chapter GPA": overall_gpa,
        "Total Members": _count_gpa(gpa_rows["term_gpa_num"]),
        "Previous Term Change": change,
    }


def build_community_summary(frame: pd.DataFrame, previous_frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=COMMUNITY_COLUMNS)
    rows: List[dict[str, object]] = []
    for chapter, group in frame.groupby("chapter", dropna=False):
        chapter_name = str(chapter or "").strip() or "Unknown"
        council = _first_nonblank(group, "council", "Unknown")
        org_type = _first_nonblank(group, "org_type", "Organization")
        rows.append(_chapter_community_row(chapter_name, group, previous_frame, council, org_type))
    return pd.DataFrame(rows, columns=COMMUNITY_COLUMNS).sort_values(["Council", "Organization Type", "Chapter"]).reset_index(drop=True)


def _style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A3"


def _write_title(ws, title: str, subtitle: str = "", width: int = 8) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).font = SUBTITLE_FONT
        ws.cell(2, 1).alignment = Alignment(horizontal="center")
        return 4
    return 3


def _write_table(ws, start_row: int, headers: List[str], rows: Iterable[Iterable[object]], title: str = "") -> int:
    row_cursor = start_row
    if title:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=len(headers))
        cell = ws.cell(row_cursor, 1, title)
        cell.fill = PatternFill("solid", fgColor=MAROON)
        cell.font = Font(bold=True, color=GOLD)
        cell.alignment = Alignment(horizontal="center")
        row_cursor += 1
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row_cursor, column_index, header)
        cell.fill = PatternFill("solid", fgColor=MAROON)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row_cursor += 1
    for row in rows:
        for column_index, value in enumerate(row, start=1):
            cell = ws.cell(row_cursor, column_index, value)
            cell.border = THIN_BORDER
        row_cursor += 1
    return row_cursor + 1


def _auto_width(ws, max_width: int = 38) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_community_workbook(summary: pd.DataFrame, term_label: str, output_path: Path) -> Path:
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    _write_title(cover, f"{term_label} Grade Report", "Fraternity and Sorority Life", width=8)
    cover["A5"] = "Texas State University"
    cover["A6"] = "601 University Dr."
    cover["A7"] = "LBJ Student Center #410"
    cover["A8"] = "San Marcos, TX 78666"
    cover["A9"] = "(512) 245-5646"
    cover["A10"] = "FSLife@txstate.edu"
    cover["A11"] = "https://fsl.studentinvolvement.txst.edu/"
    cover["A13"] = "Note"
    cover["B13"] = "Students who have not generated a GPA are not included in GPA averages."
    _auto_width(cover)

    for council, council_frame in summary.groupby("Council", dropna=False):
        ws = wb.create_sheet(str(council or "Unknown")[:31])
        row = _write_title(ws, f"{council} {term_label} Grade Report", width=8)
        change_label = "Previous Term Change"
        table_rows = council_frame[
            [
                "Chapter",
                "New Member GPA",
                "New Members",
                "Initiated Member GPA",
                "Initiated Members",
                "Overall Chapter GPA",
                "Total Members",
                "Previous Term Change",
            ]
        ].values.tolist()
        row = _write_table(
            ws,
            row,
            ["Organizations", "New Members GPA", "New Members", "Initiated GPA", "Initiated Members", "Overall GPA", "Total Members", change_label],
            table_rows,
        )
        averages = []
        for org_type, label in [("Fraternity", "All Fraternity Member Average"), ("Sorority", "All Sorority Member Average")]:
            subset = council_frame.loc[council_frame["Organization Type"].fillna("").astype(str).str.contains(org_type, case=False, na=False)]
            if subset.empty:
                continue
            averages.append([label, int(pd.to_numeric(subset["Total Members"], errors="coerce").fillna(0).sum()), _mean_or_na(subset["Overall Chapter GPA"])])
        averages.append(["All Greek Member Average", int(pd.to_numeric(council_frame["Total Members"], errors="coerce").fillna(0).sum()), _mean_or_na(council_frame["Overall Chapter GPA"])])
        _write_table(ws, row, ["Texas State University Overall Averages", "Students", "GPA"], averages)
        ws["A2"] = "*Students who have not generated a GPA are not included (i.e. student teaching, internships, withdrawal)"
        _style_sheet(ws)
        _auto_width(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _member_rows(frame: pd.DataFrame, status_group: str, include_missing_gpa: bool = False) -> List[List[object]]:
    subset = frame.loc[frame["status_group"].eq(status_group)].copy()
    if include_missing_gpa:
        subset = subset.loc[subset["term_gpa_num"].isna()]
    else:
        subset = subset.loc[subset["term_gpa_num"].notna()]
    subset = subset.sort_values(["last_name", "first_name"], na_position="last")
    rows: List[List[object]] = []
    for _, row in subset.iterrows():
        rows.append(
            [
                row.get("last_name", ""),
                row.get("first_name", ""),
                status_group,
                row.get("major", ""),
                row.get("hours_num", ""),
                row.get("term_gpa_num", ""),
                row.get("txst_gpa_num", ""),
            ]
        )
    return rows


def _summary_rows(frame: pd.DataFrame) -> List[List[object]]:
    active = frame.loc[frame["status_group"].eq("Active Member") & frame["term_gpa_num"].notna()]
    new = frame.loc[frame["status_group"].eq("New Member") & frame["term_gpa_num"].notna()]
    counted = frame.loc[frame["status_group"].isin(["Active Member", "New Member"]) & frame["term_gpa_num"].notna()]
    return [
        ["Active Member Averages", _mean_or_na(active["hours_num"]), _mean_or_na(active["term_gpa_num"]), _mean_or_na(active["txst_gpa_num"])],
        ["New Member Averages", _mean_or_na(new["hours_num"]), _mean_or_na(new["term_gpa_num"]), _mean_or_na(new["txst_gpa_num"])],
        ["Chapter Averages", _mean_or_na(counted["hours_num"]), _mean_or_na(counted["term_gpa_num"]), _mean_or_na(counted["txst_gpa_num"])],
        [],
        ["Membership Numbers", int(frame["status_group"].eq("Active Member").sum()), int(frame["status_group"].eq("New Member").sum()), int(frame["status_group"].isin(["Active Member", "New Member"]).sum())],
    ]


def write_chapter_workbook(chapter: str, chapter_frame: pd.DataFrame, term_label: str, output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Report"
    row = _write_title(ws, f"{chapter} {term_label} Grade Report", "Texas State University Fraternity and Sorority Life", width=10)
    headers = CHAPTER_MEMBER_COLUMNS
    row = _write_table(ws, row, headers, _member_rows(chapter_frame, "Active Member"), f"{chapter} {term_label} Grade Report - Active Members")
    row = _write_table(ws, row, headers, _member_rows(chapter_frame, "New Member"), f"{chapter} {term_label} Grade Report - New Members")
    missing_rows = _member_rows(chapter_frame, "Active Member", include_missing_gpa=True) + _member_rows(chapter_frame, "New Member", include_missing_gpa=True)
    row = _write_table(ws, row, headers, missing_rows, f"{term_label} Grade Report - Members Not Enrolled/GPA Not Counted")

    summary_start_col = 9
    ws.cell(4, summary_start_col, f"{chapter} {term_label} Grade Report")
    ws.cell(4, summary_start_col).fill = PatternFill("solid", fgColor=MAROON)
    ws.cell(4, summary_start_col).font = Font(bold=True, color=GOLD)
    ws.merge_cells(start_row=4, start_column=summary_start_col, end_row=4, end_column=summary_start_col + 3)
    for offset, header in enumerate(["", "Term Hours", "Term GPA", "TXST GPA"]):
        cell = ws.cell(5, summary_start_col + offset, header)
        cell.fill = PatternFill("solid", fgColor=MAROON)
        cell.font = HEADER_FONT
    for row_offset, summary_row in enumerate(_summary_rows(chapter_frame), start=6):
        if not summary_row:
            continue
        for offset, value in enumerate(summary_row):
            cell = ws.cell(row_offset, summary_start_col + offset, value)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.border = THIN_BORDER
    ws.cell(13, summary_start_col, "Membership Numbers")
    ws.cell(13, summary_start_col + 1, "Active Members")
    ws.cell(13, summary_start_col + 2, "New Members")
    ws.cell(13, summary_start_col + 3, "Total Members")

    _style_sheet(ws)
    _auto_width(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_grade_reports(
    term: str = "",
    canonical_dir: Optional[str | Path] = None,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    config_path: Optional[str] = None,
) -> ReportBuildResult:
    canonical = _canonical_latest_dir(config_path=config_path, canonical_dir=canonical_dir)
    base_frame = _load_grade_source(canonical, _term_code_from_label(term) if term else "")
    selected_term_code = _term_code_from_label(term) if term else _latest_term_code(base_frame)
    if selected_term_code and not base_frame.empty and "term_code" in base_frame.columns:
        base_frame = base_frame.loc[base_frame["term_code"].fillna("").astype(str).str.upper().eq(selected_term_code)].copy()
    term_label = _term_label_from_code(selected_term_code) if selected_term_code else "Selected Term"
    previous_code = _previous_term_code(selected_term_code)
    previous_frame = _load_grade_source(canonical, previous_code) if previous_code else pd.DataFrame()

    chapter_mapping = load_chapter_mapping()
    base_frame = _apply_chapter_mapping(base_frame, chapter_mapping)
    previous_frame = _apply_chapter_mapping(previous_frame, chapter_mapping) if not previous_frame.empty else previous_frame

    run_dir = output_dir / safe_slug(term_label)
    chapter_dir = run_dir / "chapter_reports"
    run_dir.mkdir(parents=True, exist_ok=True)
    chapter_dir.mkdir(parents=True, exist_ok=True)

    summary = build_community_summary(base_frame, previous_frame)
    summary_csv = run_dir / f"community_grade_summary_{safe_slug(term_label)}.csv"
    summary.to_csv(summary_csv, index=False)
    community_workbook = write_community_workbook(summary, term_label, run_dir / f"community_grade_report_{safe_slug(term_label)}.xlsx")

    chapter_workbooks: List[Path] = []
    if not base_frame.empty:
        for chapter, chapter_frame in base_frame.groupby("chapter", dropna=False):
            chapter_name = str(chapter or "").strip()
            if not chapter_name:
                continue
            path = write_chapter_workbook(
                chapter_name,
                chapter_frame.copy(),
                term_label,
                chapter_dir / f"{safe_slug(chapter_name)}_grade_report_{safe_slug(term_label)}.xlsx",
            )
            chapter_workbooks.append(path)

    return ReportBuildResult(
        output_dir=run_dir,
        community_workbook=community_workbook,
        chapter_workbooks=chapter_workbooks,
        community_summary_csv=summary_csv,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build community and chapter grade report workbooks from canonical LOGI/roster outputs.")
    parser.add_argument("--config", default=None, help="Optional config/local_paths.yaml path.")
    parser.add_argument("--canonical-dir", default=None, help="Canonical run folder. Defaults to configured output_root/latest.")
    parser.add_argument("--term", default="", help='Term label or code, e.g. "Spring 2025" or "2025SP". Defaults to latest term.')
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = build_grade_reports(
        term=args.term,
        canonical_dir=args.canonical_dir,
        output_dir=Path(args.output_dir).expanduser().resolve(),
        config_path=args.config,
    )
    print(f"Output folder: {result.output_dir}")
    print(f"Community workbook: {result.community_workbook}")
    print(f"Community summary CSV: {result.community_summary_csv}")
    print(f"Chapter workbooks: {len(result.chapter_workbooks)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
