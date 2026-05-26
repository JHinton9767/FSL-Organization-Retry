from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.build_grade_reports import build_community_summary, build_grade_reports


def _canonical_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "student_id": "A00000001",
                "first_name": "Amy",
                "last_name": "Alpha",
                "term_code": "2025SP",
                "chapter": "Alpha Sigma Phi",
                "org_status_bucket": "Active",
                "new_member_flag": "No",
                "major": "Biology",
                "term_gpa": 3.0,
                "institutional_cumulative_gpa": 3.1,
                "overall_cumulative_gpa": 3.1,
                "attempted_hours_term": 12,
            },
            {
                "student_id": "A00000002",
                "first_name": "Nina",
                "last_name": "New",
                "term_code": "2025SP",
                "chapter": "Alpha Sigma Phi",
                "org_status_bucket": "New Member",
                "new_member_flag": "Yes",
                "major": "History",
                "term_gpa": 4.0,
                "institutional_cumulative_gpa": 3.8,
                "overall_cumulative_gpa": 3.8,
                "attempted_hours_term": 15,
            },
            {
                "student_id": "A00000003",
                "first_name": "Missing",
                "last_name": "Gpa",
                "term_code": "2025SP",
                "chapter": "Alpha Sigma Phi",
                "org_status_bucket": "Active",
                "new_member_flag": "No",
                "major": "Math",
                "term_gpa": "",
                "institutional_cumulative_gpa": "",
                "overall_cumulative_gpa": "",
                "attempted_hours_term": "",
            },
            {
                "student_id": "A00000004",
                "first_name": "Prior",
                "last_name": "Term",
                "term_code": "2024FA",
                "chapter": "Alpha Sigma Phi",
                "org_status_bucket": "Active",
                "new_member_flag": "No",
                "major": "Biology",
                "term_gpa": 2.5,
                "institutional_cumulative_gpa": 2.7,
                "overall_cumulative_gpa": 2.7,
                "attempted_hours_term": 12,
            },
        ]
    )


def test_community_summary_splits_new_and_initiated_members() -> None:
    current = _canonical_frame().loc[lambda frame: frame["term_code"].eq("2025SP")].copy()
    previous = _canonical_frame().loc[lambda frame: frame["term_code"].eq("2024FA")].copy()
    current["status_group"] = ["Active Member", "New Member", "Active Member"]
    previous["status_group"] = ["Active Member"]
    current["term_gpa_num"] = pd.to_numeric(current["term_gpa"], errors="coerce")
    previous["term_gpa_num"] = pd.to_numeric(previous["term_gpa"], errors="coerce")
    current["council"] = "IFC"
    current["org_type"] = "Fraternity"

    summary = build_community_summary(current, previous)

    assert summary.iloc[0]["New Member GPA"] == 4.0
    assert summary.iloc[0]["Initiated Member GPA"] == 3.0
    assert summary.iloc[0]["Overall Chapter GPA"] == 3.5
    assert summary.iloc[0]["Previous Term Change"] == 1.0


def test_build_grade_reports_writes_community_and_chapter_workbooks(tmp_path: Path) -> None:
    canonical = tmp_path / "canonical"
    canonical.mkdir()
    _canonical_frame().to_csv(canonical / "master_longitudinal.csv", index=False)
    output = tmp_path / "reports"

    result = build_grade_reports(term="Spring 2025", canonical_dir=canonical, output_dir=output)

    assert result.community_workbook.exists()
    assert result.community_summary_csv.exists()
    assert len(result.chapter_workbooks) == 1
    workbook = load_workbook(result.community_workbook)
    assert "Cover" in workbook.sheetnames
    assert "IFC" in workbook.sheetnames
    chapter_workbook = load_workbook(result.chapter_workbooks[0])
    sheet = chapter_workbook["Grade Report"]
    assert "Alpha Sigma Phi Spring 2025 Grade Report" in sheet["A1"].value
