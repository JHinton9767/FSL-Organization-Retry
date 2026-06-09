from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment

from src.build_canonical_pipeline import (
    ensure_columns,
    build_roster_supplement_from_academic,
    dedupe_table,
    load_schema,
    load_academic_term_table,
    prepare_canonical_sources,
)


def _write_workbook(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_load_academic_term_table_parses_copy_of_grades_logi_and_skips_raw_data(tmp_path: Path) -> None:
    root = tmp_path
    raw_data_path = root / "Copy of Grades" / "2025" / "Fall 2025" / "FSL Raw Data" / "IFC Raw Data" / "Alpha Sigma Phi Raw Data.xlsx"
    historical_raw_data_path = root / "Copy of Grades" / "2024" / "Fall 2024" / "IFC" / "Alpha Sigma Phi Raw Data.xlsx"
    logi_path = root / "Copy of Grades" / "2025" / "Fall 2025" / "LOGI Reports" / "IFC" / "Alpha Sigma Phi LOGI.xlsx"

    _write_workbook(raw_data_path, [["id"], ["A05233818"]])
    _write_workbook(
        historical_raw_data_path,
        [
            ["Last Name", "First Name", "Student Status", "Major", "Semester Hours", "Semester GPA", "TXST GPA", "NetID"],
            ["Brown", "Casey", "Active", "Finance", 16, 3.3, 3.1, "cb123"],
        ],
    )
    _write_workbook(
        logi_path,
        [
            ["Last Name", "First Name", "Student Status", "Major", "Semester Hours", "Semester GPA", "TXST GPA", "Banner ID"],
            ["Smith", "Alex", "Active", "Biology", 15, 3.1, 3.2, "A05233818"],
            ["Jones", "Jamie", "New Member", "Chemistry", 12, 3.5, 3.4, "A05233819"],
            ["Taylor", "Riley", "", "History", "", 2.8, 3.0, "A05233820"],
        ],
    )

    academic, exceptions = load_academic_term_table(root)

    assert exceptions["exception_type"].tolist() == ["academic_missing_or_invalid_student_id"]
    assert len(academic.index) == 3
    assert set(academic["term_code"]) == {"2025FA"}
    assert set(academic["term_source_basis"]) == {"copy_of_grades_logi"}
    assert academic["source_file"].astype(str).str.contains("Copy of Grades").all()
    assert not academic["source_file"].astype(str).str.contains("FSL Raw Data").any()
    assert not academic["source_file"].astype(str).str.contains("2024").any()

    supplement = build_roster_supplement_from_academic(academic)
    alex = supplement.loc[supplement["last_name"].eq("Smith")].iloc[0]
    jamie = supplement.loc[supplement["last_name"].eq("Jones")].iloc[0]
    riley = supplement.loc[supplement["last_name"].eq("Taylor")].iloc[0]

    assert alex["chapter"] == "Alpha Sigma Phi"
    assert alex["org_status_bucket"] == "Active"
    assert jamie["org_status_bucket"] == "New Member"
    assert jamie["new_member_flag"] == "Yes"
    assert riley["org_status_bucket"] != "Graduated"


def test_load_academic_term_table_parses_multi_section_copy_of_grades_report(tmp_path: Path) -> None:
    root = tmp_path
    report_path = root / "Copy of Grades" / "2025" / "Spring 2025" / "IFC" / "Alpha Sigma Phi Spring 2025 Grades.xlsx"

    _write_workbook(
        report_path,
        [
            ["Alpha Sigma Phi Spring 2025 Grades - Active Members"],
            ["Last Name", "First Name", "Status", "Email", "Banner ID", "Major", "Semester Hours", "semester GPA", "Overall GPA"],
            ["Smith", "Alex", "", "alex@example.com", "A05233818", "Biology", 15, 3.1, 3.2],
            [],
            ["Alpha Sigma Phi Spring 2025 Grades - New Members"],
            ["Last Name", "First Name", "Status", "Email", "Banner ID", "Major", "Semester Hours", "semester GPA", "Overall GPA"],
            ["Jones", "Jamie", "", "jamie@example.com", "A05233823", "Chemistry", 12, 3.5, 3.4],
            [],
            ["Alpha Sigma Phi Spring 2025 Grades - Inactive Members"],
            ["Last Name", "First Name", "Status", "Email", "Banner ID", "Major", "Semester Hours", "semester GPA", "Overall GPA"],
            ["Doe", "Riley", "", "riley@example.com", "A05233824", "History", 6, 2.2, 2.5],
            [],
            ["Alpha Sigma Phi Spring 2025 Grades Report"],
            ["Spring Hours", "Spring GPA", "Overall GPA"],
            [13.48, 2.94, 2.93],
        ],
    )

    academic, exceptions = load_academic_term_table(root)

    assert exceptions.empty
    assert len(academic.index) == 3
    assert set(academic["term_code"]) == {"2025SP"}
    assert set(academic["term_source_basis"]) == {"copy_of_grades_section"}

    status_by_last_name = academic.set_index("last_name")["academic_status_raw"].to_dict()
    assert status_by_last_name["Smith"] == "Active"
    assert status_by_last_name["Jones"] == "New Member"
    assert status_by_last_name["Doe"] == "Inactive"


def test_load_academic_term_table_uses_column_k_count_notes(tmp_path: Path) -> None:
    root = tmp_path
    report_path = root / "Copy of Grades" / "2026" / "Spring 2026" / "IFC" / "Alpha Sigma Phi LOGI.xlsx"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Spring 2026"
    worksheet.append(
        [
            "Last Name",
            "First Name",
            "Banner ID",
            "Email",
            "Student Status",
            "Major",
            "Current Academic Standing",
            "Term GPA",
            "Term Passed Hours",
            "Semester Hours",
            "TxState Cumulative GPA",
        ]
    )
    worksheet.append(["Count", "Casey", "A05233818", "casey@example.com", "Active", "Finance", "Good Standing", 3.5, 12, 12, 3.4])
    worksheet["K2"].comment = Comment("Counted", "FSL")
    worksheet.append(["Skipgpa", "Sam", "A05233819", "sam@example.com", "Active", "Finance", "Good Standing", 4.0, 15, 15, 3.9])
    worksheet["K3"].comment = Comment("Not Counted", "FSL")
    worksheet.append(["Notstudent", "Noah", "A05233820", "noah@example.com", "Active", "Finance", "Good Standing", 4.0, 15, 15, "Not a student"])
    worksheet.append(["Lastsemester", "Lee", "A05233821", "lee@example.com", "Active", "Finance", "Good Standing", 2.5, 9, 9, "Last Semester"])
    workbook.save(report_path)
    workbook.close()

    academic, exceptions = load_academic_term_table(root)

    by_last = academic.set_index("last_name")
    assert "Notstudent" not in set(academic["last_name"])
    assert by_last.loc["Count", "term_gpa"] == 3.5
    assert by_last.loc["Skipgpa", "term_gpa"] == 4.0
    assert by_last.loc["Skipgpa", "institutional_cumulative_gpa"] == 3.9
    assert by_last.loc["Lastsemester", "term_gpa"] == 2.5
    assert set(exceptions["exception_type"]) == {"academic_row_excluded_by_count_note"}


def test_load_academic_term_table_parses_spring_2026_council_raw_data_layout(tmp_path: Path) -> None:
    root = tmp_path
    report_path = root / "Spring 2026" / "IFC Raw Data" / "Alpha Sigma Phi Raw Data.xlsx"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IFC"
    worksheet.append(
        [
            "Last Name",
            "First Name",
            "Banner ID",
            "Email",
            "Student Status",
            "Major",
            "Current Academic Standing",
            "Term GPA",
            "Term Passed Hours",
            "Semester Hours",
            "TxState Cumulative GPA",
        ]
    )
    worksheet.append(["Angus", "Alexander", "A05487070", "tcf56@txstate.edu", "AS - Active", "Finance", "GS - Good Standing", 3.2, 12, 12, 3.1])
    worksheet["K2"].comment = Comment("Counted", "FSL")
    worksheet.append(["Later", "Lana", "A05487071", "lana@txstate.edu", "AS - Active", "Finance", "GS - Good Standing", 3.8, 15, 15, "Last Semester"])
    workbook.save(report_path)
    workbook.close()

    academic, exceptions = load_academic_term_table(root)

    assert set(academic["term_code"]) == {"2026SP"}
    assert set(academic["term_source_basis"]) == {"council_raw_grade_report"}
    angus = academic.loc[academic["last_name"].eq("Angus")].iloc[0]
    later = academic.loc[academic["last_name"].eq("Later")].iloc[0]
    assert angus["student_id"] == "A05487070"
    assert angus["academic_standing_bucket"] == "Good Standing"
    assert angus["term_gpa"] == 3.2
    assert later["term_gpa"] == 3.8
    assert exceptions.empty


def test_dedupe_table_prefers_copy_of_grades_academic_rows() -> None:
    frame = {
        "student_id": ["A05233818", "A05233818"],
        "student_id_raw": ["A05233818", "A05233818"],
        "first_name": ["Alex", "Alex"],
        "last_name": ["Smith", "Smith"],
        "email": ["alex@example.com", "alex@example.com"],
        "source_file": [
            r"legacy\Fall 2025\Alpha Sigma Phi grades.csv",
            r"Copy of Grades\2025\Fall 2025\LOGI Reports\IFC\Alpha Sigma Phi LOGI.xlsx",
        ],
        "source_sheet": ["csv", "Sheet1"],
        "term_code": ["2025FA", "2025FA"],
        "term_label": ["Fall 2025", "Fall 2025"],
        "term_year": [2025, 2025],
        "term_season": ["FA", "FA"],
        "term_source_basis": ["filename", "copy_of_grades_logi"],
        "academic_status_raw": ["Active", "Active"],
        "major": ["Biology", "Biology"],
        "term_gpa": [2.7, 3.1],
        "institutional_cumulative_gpa": [2.8, 3.2],
        "overall_cumulative_gpa": [2.8, 3.2],
        "transfer_gpa": ["", ""],
        "attempted_hours_term": [12, 15],
        "earned_hours_term": [12, 15],
        "institutional_cumulative_hours": [60, 63],
        "total_cumulative_hours": [60, 63],
        "academic_standing_raw": ["Good Standing", "Good Standing"],
        "academic_standing_bucket": ["Good Standing", "Good Standing"],
        "graduation_term_code": ["", ""],
        "graduation_term_label": ["", ""],
    }

    deduped, exceptions = dedupe_table(pd.DataFrame(frame), ["student_id", "term_code"], "academic")

    assert len(deduped.index) == 1
    assert len(exceptions.index) == 1
    assert deduped.iloc[0]["source_file"].startswith("Copy of Grades")
    assert deduped.iloc[0]["term_gpa"] == 3.1


def test_load_academic_term_table_reads_utf16_csv_exports(tmp_path: Path) -> None:
    root = tmp_path
    csv_path = root / "2024" / "Fall 2024" / "IFC" / "Alpha Sigma Phi grades.csv"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.write_text(
        "Banner ID,Last Name,First Name,Student Status,Semester Hours,Term GPA,Overall GPA\n"
        "A05233818,Smith,Alex,Active,15,3.1,3.2\n",
        encoding="utf-16",
    )

    academic, exceptions = load_academic_term_table(root)

    assert exceptions.empty
    assert len(academic.index) == 1
    assert academic.iloc[0]["student_id"] == "A05233818"
    assert academic.iloc[0]["term_code"] == "2024FA"
    assert academic.iloc[0]["term_gpa"] == 3.1


def test_prepare_canonical_sources_keeps_roster_membership_authoritative() -> None:
    schema = load_schema()
    roster_term = ensure_columns(
        pd.DataFrame(
            [
                {
                    "student_id": "A05233818",
                    "student_id_raw": "A05233818",
                    "identity_resolution_basis": "source_banner_id",
                    "identity_resolution_notes": "",
                    "first_name": "Alex",
                    "last_name": "Smith",
                    "email": "alex@example.com",
                    "source_file": r"Copy of Rosters\Fall 2025\IFC\Final\Alpha Sigma Phi\roster.xlsx",
                    "source_sheet": "Members",
                    "roster_file_version": "Final",
                    "roster_file_version_priority": 3,
                    "roster_file_month": "",
                    "roster_file_month_priority": 0,
                    "term_code": "2025FA",
                    "term_label": "Fall 2025",
                    "term_year": 2025,
                    "term_season": "FA",
                    "term_source_basis": "folder_or_filename",
                    "chapter": "Alpha Sigma Phi",
                    "chapter_raw": "Alpha Sigma Phi",
                    "chapter_assignment_source": "original",
                    "chapter_assignment_confidence": "high",
                    "chapter_assignment_notes": "",
                    "org_status_raw": "Active",
                    "org_status_bucket": "Active",
                    "org_position_raw": "",
                    "semester_joined_raw": "",
                    "new_member_flag": "No",
                    "org_entry_term_code": "",
                    "org_entry_term_basis": "",
                }
            ]
        ),
        schema["tables"]["roster_term"],
    )
    academic_term = ensure_columns(
        pd.DataFrame(
            [
                {
                    "student_id": "A05233819",
                    "student_id_raw": "A05233819",
                    "identity_resolution_basis": "source_student_id",
                    "identity_resolution_notes": "",
                    "first_name": "Jamie",
                    "last_name": "Jones",
                    "email": "jamie@example.com",
                    "source_file": r"Copy of Grades\2025\Fall 2025\LOGI Reports\IFC\Alpha Sigma Phi LOGI.xlsx",
                    "source_sheet": "Sheet1",
                    "term_code": "2025FA",
                    "term_label": "Fall 2025",
                    "term_year": 2025,
                    "term_season": "FA",
                    "term_source_basis": "copy_of_grades_logi",
                    "academic_status_raw": "Active",
                    "major": "Biology",
                    "term_gpa": 3.2,
                    "institutional_cumulative_gpa": 3.2,
                    "overall_cumulative_gpa": 3.2,
                    "transfer_gpa": "",
                    "attempted_hours_term": 15,
                    "earned_hours_term": 15,
                    "institutional_cumulative_hours": 45,
                    "total_cumulative_hours": 45,
                    "academic_standing_raw": "Good Standing",
                    "academic_standing_bucket": "Good Standing",
                    "graduation_term_code": "",
                    "graduation_term_label": "",
                }
            ]
        ),
        schema["tables"]["academic_term"],
    )

    prepared_roster, prepared_academic, *_ = prepare_canonical_sources(
        roster_term,
        academic_term,
        settings={"secondary_organizations": []},
        manual_chapter_assignments=pd.DataFrame(),
        manual_roster_corrections=pd.DataFrame(),
    )

    assert len(prepared_academic.index) == 1
    assert len(prepared_roster.index) == 1
    assert prepared_roster.iloc[0]["student_id"] == "A05233818"
    assert not prepared_roster["student_id"].astype(str).eq("A05233819").any()
