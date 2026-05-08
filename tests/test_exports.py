from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app.exports import frames_to_excel_bytes


def test_frames_to_excel_bytes_splits_oversized_tables() -> None:
    workbook_bytes = frames_to_excel_bytes(
        {"Filtered Longitudinal": pd.DataFrame({"student_id": range(7), "term": ["Fall 2025"] * 7})},
        max_data_rows_per_sheet=3,
    )

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    data_sheets = [name for name in workbook.sheetnames if name.startswith("Filtered Longitudinal")]
    assert data_sheets == ["Filtered Longitudinal 001", "Filtered Longitudinal 002", "Filtered Longitudinal 003"]
    assert workbook["Filtered Longitudinal 001"].max_row == 4
    assert workbook["Filtered Longitudinal 002"].max_row == 4
    assert workbook["Filtered Longitudinal 003"].max_row == 2
    assert "Export Manifest" in workbook.sheetnames
    assert workbook["Export Manifest"].max_row == 4


def test_frames_to_excel_bytes_keeps_normal_tables_single_sheet() -> None:
    workbook_bytes = frames_to_excel_bytes({"Summary": pd.DataFrame({"metric": ["A"], "value": [1]})})

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    assert "Summary" in workbook.sheetnames
    assert "Export Manifest" in workbook.sheetnames
    assert workbook["Summary"].max_row == 2
